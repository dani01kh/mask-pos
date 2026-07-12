# app.py - Mask POS (Blue Theme + Livelier UI + Scan Anywhere + Delete Products + Shopify-Style Analytics + Cash Drawer + Employees)
# Requirements:
#   pip install pyautogui pygetwindow reportlab python-barcode pillow matplotlib

import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import os
import sys
import json
import shutil
import copy
import math
import random
from pathlib import Path
from datetime import datetime, date, timedelta  # must be before _save_pending_exchange_credit

_SINGLE_INSTANCE_HANDLE = None


def _runtime_base_dir() -> str:
    """Directory where persistent files (DB/config) should live."""
    # If frozen (PyInstaller), use the folder containing the EXE.
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    # Normal python: folder containing this file.
    return os.path.dirname(os.path.abspath(__file__))


BASE_DIR = _runtime_base_dir()
# Ensure we run from BASE_DIR so relative paths (like pos.db) persist next to the EXE.
try:
    os.makedirs(BASE_DIR, exist_ok=True)
    os.chdir(BASE_DIR)
except Exception:
    pass


def _show_startup_message(title: str, message: str) -> None:
    try:
        import ctypes
        ctypes.windll.user32.MessageBoxW(None, str(message), str(title), 0x40)
    except Exception:
        try:
            print(f"{title}: {message}")
        except Exception:
            pass


def _acquire_single_instance() -> None:
    """Prevent two POS windows from using the same register DB at once."""
    global _SINGLE_INSTANCE_HANDLE
    if _SINGLE_INSTANCE_HANDLE is not None:
        return
    try:
        import ctypes
        kernel32 = ctypes.windll.kernel32
        mutex_name = "Global\\MaskPOS_Single_Register_Instance"
        handle = kernel32.CreateMutexW(None, False, mutex_name)
        if not handle:
            raise OSError("CreateMutexW failed")
        if kernel32.GetLastError() == 183:  # ERROR_ALREADY_EXISTS
            kernel32.CloseHandle(handle)
            _show_startup_message("Mask POS already running", "Mask POS is already open. Use the existing window.")
            raise SystemExit(0)
        _SINGLE_INSTANCE_HANDLE = ("mutex", handle)
        return
    except SystemExit:
        raise
    except Exception:
        pass
    try:
        import msvcrt
        lock_path = os.path.join(BASE_DIR, "maskpos.lock")
        handle = open(lock_path, "a+b")
        handle.seek(0)
        msvcrt.locking(handle.fileno(), msvcrt.LK_NBLCK, 1)
        _SINGLE_INSTANCE_HANDLE = handle
    except OSError:
        _show_startup_message("Mask POS already running", "Mask POS is already open. Use the existing window.")
        raise SystemExit(0)
    except Exception:
        pass


def _release_single_instance() -> None:
    global _SINGLE_INSTANCE_HANDLE
    handle = _SINGLE_INSTANCE_HANDLE
    _SINGLE_INSTANCE_HANDLE = None
    if handle is None:
        return
    if isinstance(handle, tuple) and handle and handle[0] == "mutex":
        try:
            import ctypes
            ctypes.windll.kernel32.CloseHandle(handle[1])
        except Exception:
            pass
        return
    try:
        import msvcrt
        handle.seek(0)
        msvcrt.locking(handle.fileno(), msvcrt.LK_UNLCK, 1)
    except Exception:
        pass
    try:
        handle.close()
    except Exception:
        pass


def data_path(filename: str) -> str:
    return os.path.join(BASE_DIR, filename)


def bundled_data_path(filename: str) -> str:
    bundle_dir = getattr(sys, "_MEIPASS", BASE_DIR)
    return os.path.join(bundle_dir, filename)


def resource_path(filename: str) -> str:
    bundled = bundled_data_path(filename)
    if os.path.exists(bundled):
        return bundled
    return data_path(filename)


def _ensure_packaged_runtime_file(filename: str) -> None:
    src = bundled_data_path(filename)
    dst = data_path(filename)
    if src == dst or not os.path.exists(src):
        return
    try:
        if os.path.exists(dst) and os.path.getsize(dst) > 0:
            return
        parent = os.path.dirname(dst)
        if parent:
            os.makedirs(parent, exist_ok=True)
        shutil.copy2(src, dst)
    except Exception:
        pass


for _runtime_file in [
    "pos.db",
    "pos_config.json",
    "config.json",
    "cloudflare_pos_config.json",
    "products.csv",
    "SumatraPDF.exe",
    "SumatraPDF-settings.txt",
    "PdfPreview.dll",
    "PdfFilter.dll",
    "libmupdf.dll",
]:
    _ensure_packaged_runtime_file(_runtime_file)


PENDING_CREDIT_PATH = data_path("pending_exchange_credit.json")


def _load_pending_exchange_credit() -> dict:
    try:
        if not os.path.exists(PENDING_CREDIT_PATH):
            return {"amount": 0.0, "origin_sale_ids": [], "return_ids": [], "bon_codes": []}
        with open(PENDING_CREDIT_PATH, "r", encoding="utf-8") as f:
            data = json.load(f) or {}
        amount = max(0.0, float(data.get("amount") or 0.0))
        origin_sale_ids = []
        for v in data.get("origin_sale_ids", []) or []:
            try:
                iv = int(v)
                if iv not in origin_sale_ids:
                    origin_sale_ids.append(iv)
            except Exception:
                pass
        return_ids = []
        for v in data.get("return_ids", []) or []:
            try:
                iv = int(v)
                if iv not in return_ids:
                    return_ids.append(iv)
            except Exception:
                pass
        bon_codes = []
        for v in data.get("bon_codes", []) or []:
            code = str(v or "").strip().upper()
            if code and code not in bon_codes:
                bon_codes.append(code)
        return {"amount": amount, "origin_sale_ids": origin_sale_ids, "return_ids": return_ids, "bon_codes": bon_codes}
    except Exception:
        return {"amount": 0.0, "origin_sale_ids": [], "return_ids": [], "bon_codes": []}


def _save_pending_exchange_credit(amount: float, origin_sale_ids=None, return_ids=None, bon_codes=None) -> None:
    try:
        amount = round(max(0.0, float(amount or 0.0)), 2)
    except Exception:
        amount = 0.0

    origin_sale_ids = list(origin_sale_ids or [])
    return_ids = list(return_ids or [])
    bon_codes = [str(c or "").strip().upper() for c in list(bon_codes or []) if str(c or "").strip()]

    if amount <= 0.005:
        try:
            if os.path.exists(PENDING_CREDIT_PATH):
                os.remove(PENDING_CREDIT_PATH)
        except Exception:
            pass
        return

    data = {
        "amount": amount,
        "origin_sale_ids": origin_sale_ids,
        "return_ids": return_ids,
        "bon_codes": bon_codes,
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    try:
        tmp = PENDING_CREDIT_PATH + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
        os.replace(tmp, PENDING_CREDIT_PATH)
    except Exception:
        pass


def _local_db_available_or_warn(action_name: str) -> bool:
    """Return True if local pos.db exists; in CONNECT mode, warn and return False.
    Some tools (like raw-SQL exports) only make sense on HOST/Standalone.
    """
    try:
        if backend_mode() == "connect":
            messagebox.showerror("Not available in JOIN mode",
                                 f"{action_name} must be done on the HOST PC (or Standalone).\n\n"
                                 "This JOIN PC uses the host database over the network.")
            return False
    except Exception:
        pass
    return True


import subprocess
import time
import threading
import sqlite3
import calendar
import re
from datetime import datetime, date, timedelta

# --- Ensure local modules import correctly (works for source + PyInstaller) ---
import os as _os, sys as _sys
try:
    _base = getattr(_sys, '_MEIPASS')  # PyInstaller temp dir
except Exception:
    _base = None
if not _base:
    _base = _os.path.dirname(_os.path.abspath(__file__))
if _base and _base not in _sys.path:
    _sys.path.insert(0, _base)
del _os, _sys, _base


def _load_daily_report_module():
    try:
        import daily_report
        return daily_report
    except ModuleNotFoundError as exc:
        import importlib.util
        for candidate in (resource_path("daily_report.py"), bundled_data_path("daily_report.py"), data_path("daily_report.py")):
            if not candidate or not os.path.exists(candidate):
                continue
            spec = importlib.util.spec_from_file_location("daily_report", candidate)
            if spec and spec.loader:
                module = importlib.util.module_from_spec(spec)
                spec.loader.exec_module(module)
                return module  # Bug 5 fix: return module, not a single function
        raise RuntimeError(
            "The daily report module is missing from this app build. "
            "Rebuild/package MaskPOS with daily_report.py included."
        ) from exc


def _load_daily_report_builder():
    return _load_daily_report_module().build_sales_report_excel


from backend import (
    backend_init,
    backend_mode,
    supabase_emergency_enabled,
    get_backend_config,
    set_backend_config,
    verify_mode_admin_password,
    cloud_sync_status,
    cloud_sync_now,
    connection_role, is_connected, last_ok_age_seconds, stop_backend,
    add_product, find_product_by_barcode, adjust_stock,
    list_products, update_product, update_product_details, list_inventory_movements,
    create_sale, get_sale_receipt_data,
    delete_product,
    _range_bounds,
    analytics_kpis_range,
    analytics_breakdown_range,
    analytics_series_in_range,
    analytics_top_products_range,
    analytics_low_stock,
    data_health_summary,
    # shifts and employees
    list_employees, ensure_employee, deactivate_employee,
    employee_pin_required, verify_employee_pin,
    get_open_shift, get_last_closed_shift, open_shift, close_shift, close_shift_with_cash_takeout, shift_summary, list_shifts, reset_next_shift_number,
    record_cash_movement, list_cash_movements,
    list_sales_for_day, search_sales, list_product_sales, list_product_price_history,
    reorder_suggestions, analytics_discount_impact, get_sale_detail, get_sale_detail_with_returns,
    get_sale_by_receipt_scan, create_return, list_returns_for_sale,
    create_bon, get_bon_by_code, list_bons, void_bon,
    delete_sale, void_sale,
    backup_pos_db, open_backups_folder, get_backup_config, set_backup_rclone_remote,
    list_printers,
    clear_printer_queue,
    hard_reset_printing,
    get_printer_config,
    set_printer_config,
    print_configured_receipt,
    print_configured_bon,
    print_configured_gift_receipt,
    print_configured_weekly_receipt,
    print_configured_warehouse_paper,
    test_print_configured,
    get_barcode_printer_config,
    set_barcode_printer_config,
    print_configured_barcodes,
    send_barcode_labels_to_host,
    test_print_barcode_configured,
    get_store_name,
    set_store_name,
    get_store_subtitle,
    set_store_subtitle,
    get_lbp_per_usd,
    set_lbp_per_usd,
    get_daily_report_email_config,
    set_daily_report_email_config,
    send_daily_report_email,
    mark_daily_report_email_sent,
    trigger_daily_report_email_on_host,
    discover_hosts,
    discover_hosts_scan_http,
    get_seasonal_sale_enabled,
    set_seasonal_sale_enabled,
    get_seasonal_sales_map,
    set_seasonal_sale_item,
    remove_seasonal_sale_item,
    clear_seasonal_sales,
    get_bundle_offers_enabled,
    set_bundle_offers_enabled,
    get_bundle_offers_map,
    set_bundle_offer_item,
    remove_bundle_offer_item,
    clear_bundle_offers,
    get_spin_wheel_prizes,
    set_spin_wheel_prizes,
    get_distinct_categories,
    get_data_health_stats,
    list_health_issues,
    bulk_update_products,
    repair_broken_product_links,
    recreate_and_repair_product,
)

try:
    from receipt_pdf import create_temp_receipt_pdf, create_gift_receipt_pdf, create_weekly_selection_receipt_pdf
except ModuleNotFoundError:
    # Fallback receipt generator so the app can start even if receipt_pdf.py isn't present.
    from reportlab.pdfgen import canvas as _rl_canvas
    from reportlab.lib.pagesizes import letter as _rl_letter
    from reportlab.lib.units import mm as _rl_mm
    import tempfile as _tmp, os as _os
    def create_temp_receipt_pdf(lines, title='Receipt'):
        fd, path = _tmp.mkstemp(suffix='.pdf', prefix='receipt_')
        _os.close(fd)
        c = _rl_canvas.Canvas(path, pagesize=_rl_letter)
        y = _rl_letter[1] - 20*_rl_mm
        c.setFont('Helvetica-Bold', 12)
        c.drawString(20*_rl_mm, y, str(title))
        y -= 10*_rl_mm
        c.setFont('Helvetica', 9)
        for ln in lines:
            if y < 15*_rl_mm:
                c.showPage(); y = _rl_letter[1] - 20*_rl_mm; c.setFont('Helvetica', 9)
            c.drawString(20*_rl_mm, y, str(ln)[:120])
            y -= 5*_rl_mm
        c.save()
        return path
    def create_gift_receipt_pdf(lines, title='Gift Receipt'):
        return create_temp_receipt_pdf(lines, title=title)
    def create_weekly_selection_receipt_pdf(shop_name, rows, title='Weekly Receipt'):
        lines = [title] + [f"{r.get('date_label','')} {r.get('printed_time','')} {r.get('receipt_code','')} {r.get('printed_amount','')}" for r in (rows or [])]
        return create_temp_receipt_pdf(lines, title=title)
    del _rl_canvas, _rl_letter, _rl_mm, _tmp, _os
from barcodes_pdf import make_labels_pdf
import app_update

APP_TITLE = "Mask POS"
APP_VERSION = app_update.current_version()
CHROME_PATH = r"C:\Program Files\Google\Chrome\Application\chrome.exe"


# ---------------- PRINT HELPERS ----------------

def open_pdf_in_chrome(pdf_path: str) -> bool:
    try:
        if not os.path.exists(pdf_path):
            print(f"PDF not found: {pdf_path}")
            return False

        if os.path.exists(CHROME_PATH):
            subprocess.Popen([CHROME_PATH, "--new-window", pdf_path])
            return True

        possible_paths = [
            r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
        ]
        for path in possible_paths:
            if os.path.exists(path):
                subprocess.Popen([path, "--new-window", pdf_path])
                return True

        os.startfile(pdf_path)
        return True

    except Exception as e:
        print(f"Error opening PDF: {e}")
        return False




def create_return_receipt_pdf(
    store_name: str,
    return_id: int,
    original_receipt_code: str,
    original_sale_time: str,
    returned_lines: list,
    total_return: float,
) -> str:
    """Create a simple RETURN receipt PDF (standalone, negative transaction).

    This does NOT modify the original sale. It's just a separate receipt for the return record.
    """

    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.units import mm
    except Exception:
        # If reportlab is missing, we can't create a PDF (but the app can still function).
        raise

    import tempfile, os
    fd, path = tempfile.mkstemp(suffix=".pdf", prefix="return_receipt_")
    os.close(fd)

    c = canvas.Canvas(path, pagesize=letter)
    page_w, page_h = letter
    x = 18 * mm
    y = page_h - 18 * mm

    def line(txt, dy=5.2):
        nonlocal y
        c.drawString(x, y, str(txt))
        y -= dy * mm
        if y < 18 * mm:
            c.showPage()
            y = page_h - 18 * mm
            c.setFont("Helvetica", 9)

    c.setFont("Helvetica-Bold", 14)
    line(store_name, dy=7.5)
    c.setFont("Helvetica-Bold", 12)
    line("RETURN RECEIPT", dy=7.0)

    c.setFont("Helvetica", 9)
    line(f"Return #: {return_id}")
    if original_receipt_code:
        line(f"Original receipt: {original_receipt_code}")
    if original_sale_time:
        line(f"Original sale time: {original_sale_time}")

    line("-" * 46, dy=6.5)

    # Items
    for ln in (returned_lines or []):
        name = ln.get("name", "")
        qty = int(ln.get("qty") or 0)
        price = float(ln.get("price") or 0.0)
        # Show as negative per unit
        line(f"{name}  x{qty}")
        line(f"  {money(-price)} each   Line: {money(-(price * qty))}", dy=6.0)

    line("-" * 46, dy=6.5)
    c.setFont("Helvetica-Bold", 12)
    line(f"TOTAL REFUND: {money(-abs(float(total_return or 0.0)))}", dy=7.0)

    c.setFont("Helvetica", 8)
    line("", dy=4.0)
    line("NOTE: This return does not change the original sale receipt.", dy=5.0)

    c.save()
    return path


def trigger_chrome_print_aggressive() -> bool:
    try:
        import pyautogui  # type: ignore
        import pygetwindow as gw  # type: ignore

        time.sleep(3.5)

        for _ in range(20):
            try:
                all_windows = gw.getAllWindows()
                chrome_windows = []
                for window in all_windows:
                    if window.title and any(k in window.title.lower() for k in ["chrome", "pdf", "receipt", "labels"]):
                        chrome_windows.append(window)

                if chrome_windows:
                    for window in chrome_windows:
                        try:
                            if window.isMinimized:
                                window.restore()
                                time.sleep(0.5)

                            window.activate()
                            time.sleep(0.7)
                            window.activate()
                            time.sleep(0.4)

                            pyautogui.hotkey("ctrl", "p")
                            time.sleep(0.3)
                            return True
                        except Exception:
                            continue
            except Exception:
                pass

            time.sleep(0.6)

        try:
            pyautogui.hotkey("ctrl", "p")
        except Exception:
            pass
        return False

    except Exception as e:
        print(f"Critical error in print trigger: {e}")
        return False


def open_and_print_in_thread(pdf_path: str, callback):
    def worker():
        success = open_pdf_in_chrome(pdf_path)
        if success:
            callback(trigger_chrome_print_aggressive())
        else:
            callback(False)

    threading.Thread(target=worker, daemon=True).start()


# ---------------- UI HELPERS ----------------

def money(x) -> str:
    try:
        return f"${float(x):.2f}"
    except Exception:
        return "$0.00"


def money_round_up(x) -> float:
    try:
        cents = float(x) * 100.0
        return int(cents + 0.999999) / 100.0
    except Exception:
        return 0.0


def whole_money_round_up(x) -> float:
    try:
        return float(math.ceil(max(0.0, float(x)) - 1e-9))
    except Exception:
        return 0.0


def lbp_money(x) -> str:
    try:
        return f"LBP {float(x):,.0f}"
    except Exception:
        return "LBP 0"


def drawer_money(x) -> str:
    try:
        return f"${float(x):,.0f}"
    except Exception:
        return "$0"


def parse_whole_money_text(text) -> int:
    raw = str(text or "").strip().replace(",", "")
    if not raw:
        return 0
    return int(round(float(raw)))


def parse_lbp_text(text) -> int:
    digits = re.sub(r"\D", "", str(text or ""))
    return int(digits or "0")


def bind_lbp_grouping(entry):
    def _format(_event=None):
        def _apply():
            raw = entry.get() or ""
            digits = re.sub(r"\D", "", raw)
            formatted = f"{int(digits):,}" if digits else ""
            if raw != formatted:
                entry.delete(0, tk.END)
                entry.insert(0, formatted)
            entry.icursor(tk.END)

        entry.after_idle(_apply)

    entry.bind("<KeyRelease>", _format, add="+")
    entry.bind("<<Paste>>", _format, add="+")
    entry.bind("<FocusOut>", _format, add="+")
    _format()


def fmt_time_ampm(dt: datetime) -> str:
    s = dt.strftime("%I:%M %p")
    return s.lstrip("0")


def row_get(r, key, default=None):
    """Safe getter for sqlite3.Row or dict."""
    try:
        return r[key]
    except Exception:
        try:
            return r.get(key, default)
        except Exception:
            return default


def _note_value(notes: str, key: str) -> str:
    prefix = f"{key}="
    try:
        for part in str(notes or "").split(";"):
            part = part.strip()
            if part.startswith(prefix):
                return part.split("=", 1)[1].strip()
    except Exception:
        pass
    return ""


def _note_float(notes: str, key: str, default: float = 0.0) -> float:
    raw = _note_value(notes, key)
    if raw == "":
        return float(default)
    try:
        return float(raw or 0.0)
    except Exception:
        return float(default)


def payment_method_label(value) -> str:
    pm = str(value or "CASH").strip().upper()
    labels = {
        "CASH": "Cash",
        "WHISH": "Whish",
        "CREDIT_CARD": "Credit Card",
        "CARD": "Credit Card",
        "DEBIT": "Debit Card",
        "EXCHANGE": "Exchange / Store Credit",
        "STORE_CREDIT": "Store Credit",
        "CASH+WHISH": "Cash + Whish",
        "CASH+CARD": "Cash + Credit Card",
        "CASH+CREDIT_CARD": "Cash + Credit Card",
        # UI-internal values before DB normalisation (Bug 4 fix)
        "CASH_WHISH": "Cash + Whish",
        "CASH_CARD": "Cash + Credit Card",
    }
    if pm in labels:
        return labels[pm]
    if "+" in pm:
        return " + ".join(labels.get(part, part.replace("_", " ").title()) for part in pm.split("+"))
    return pm.replace("_", " ").title()


def sale_display_amount(sale_row) -> float:
    """Amount to show in sale history; cash drawer math uses cash_paid separately."""
    pm = str(row_get(sale_row, "payment_method", "") or "").strip().upper()
    for key in ("total_amount", "total"):
        value = row_get(sale_row, key, None)
        if value is None:
            continue
        try:
            amount = float(value or 0.0)
        except Exception:
            continue
        if amount > 0.005:
            return max(0.0, amount)
    if pm in ("EXCHANGE", "STORE_CREDIT"):
        return 0.0
    for key in ("net_sales", "total_sales"):
        value = row_get(sale_row, key, None)
        if value is None:
            continue
        try:
            amount = float(value or 0.0)
        except Exception:
            continue
        if amount > 0.005:
            return max(0.0, amount)
    return 0.0


# ---------------- BARCODE NORMALIZATION (EAN-13) ----------------

def _digits_only(s: str) -> str:
    return "".join(ch for ch in str(s) if ch.isdigit())


def _ean13_check_digit(num12: str) -> str:
    digits = [int(c) for c in num12]
    odd_sum = sum(digits[0::2])
    even_sum = sum(digits[1::2])
    total = odd_sum + 3 * even_sum
    return str((10 - (total % 10)) % 10)


def to_ean13(code: str) -> str:
    """Return a clean 13-digit barcode for storage/printing.

    Rules requested by the cashier:
    - Never auto-create a 13-digit barcode that starts with 0.
    - 13-digit scans stay as-is.
    - 12-digit (or shorter) new barcodes are padded on the RIGHT with zeros.
    - Longer than 13 digits are invalid.

    Lookup compatibility for old leading-zero records is handled separately by
    barcode_candidates(); storage should not blindly prepend 0.
    """
    d = _digits_only(code)
    if not d:
        return ""
    if len(d) > 13:
        return ""
    if len(d) == 13:
        return d
    return d.ljust(13, "0")


def normalize_item_barcode(code: str) -> str:
    """Normalize a barcode for storing a *new* item.

    Storage rules:
    - accept 1..13 digits
    - reject >13 digits
    - never prepend a leading 0 automatically
    - pad short codes on the RIGHT to reach 13 digits
    """
    d = _digits_only(code)
    if not d or len(d) > 13:
        return ""
    return d if len(d) == 13 else d.ljust(13, "0")


def barcode_candidates(code: str) -> list[str]:
    """Candidate barcodes to try when searching DB.

    Behavior requested:
    - If scanner sends 12 digits, try that exact 12-digit barcode first.
    - Then try the legacy leading-zero form (0 + code).
    - Then try the app's current 13-digit storage form (right-padded with zeros).
    - If scanner sends 13 digits that start with 0, also try the matching 12-digit version.
    """
    d = _digits_only(code)
    cands: list[str] = []

    if len(d) == 12:
        cands.append(d)
        cands.append("0" + d)
        cands.append(d + "0")
    elif len(d) == 13:
        cands.append(d)
        if d.startswith("0"):
            cands.append(d[1:])
    elif 1 <= len(d) < 12:
        cands.append(d)
        cands.append(d.ljust(13, "0"))
    elif len(d) > 13:
        return []

    out: list[str] = []
    seen = set()
    for x in cands:
        if x and x not in seen:
            seen.add(x)
            out.append(x)
    return out


class UI:
    APP_BG = "#0f172a"
    SIDEBAR = "#111827"
    SIDEBAR_HOVER = "#1f2937"
    SIDEBAR_ACTIVE = "#2563eb"

    CONTENT_BG = "#f4f7fb"
    CARD = "#ffffff"
    SURFACE = "#f8fafc"
    SURFACE_ALT = "#eef2f7"
    TEXT = "#0f172a"
    TEXT_SOFT = "#334155"
    MUTED = "#64748b"
    BORDER = "#dde5ef"

    PRIMARY = "#2563eb"
    PRIMARY_DARK = "#1d4ed8"
    PRIMARY_SOFT = "#dbeafe"
    SUCCESS = "#16a34a"
    DANGER = "#ef4444"
    DANGER_DARK = "#dc2626"

    FONT = ("Segoe UI", 10)
    FONT_SM = ("Segoe UI", 9)
    FONT_MD = ("Segoe UI", 11)
    FONT_LG = ("Segoe UI", 15, "bold")
    FONT_XL = ("Segoe UI", 20, "bold")

    COMPACT = False

    @staticmethod
    def set_compact(compact: bool):
        """Enable compact UI for smaller screens (reduces font sizes)."""
        UI.COMPACT = bool(compact)
        if UI.COMPACT:
            UI.FONT = ("Segoe UI", 8)
            UI.FONT_SM = ("Segoe UI", 8)
            UI.FONT_MD = ("Segoe UI", 9)
            UI.FONT_LG = ("Segoe UI", 11, "bold")
            UI.FONT_XL = ("Segoe UI", 15, "bold")
        else:
            UI.FONT = ("Segoe UI", 10)
            UI.FONT_SM = ("Segoe UI", 9)
            UI.FONT_MD = ("Segoe UI", 11)
            UI.FONT_LG = ("Segoe UI", 15, "bold")
            UI.FONT_XL = ("Segoe UI", 20, "bold")

    @staticmethod
    def style_ttk(root: tk.Tk):
        style = ttk.Style(root)
        try:
            style.theme_use("clam")
        except Exception:
            pass

        style.configure(".", font=UI.FONT)

        style.configure(
            "Treeview",
            background=UI.CARD,
            fieldbackground=UI.CARD,
            foreground="#111827",
            font=UI.FONT,
            rowheight=(24 if UI.COMPACT else 34),
            bordercolor=UI.BORDER,
            lightcolor=UI.CARD,
            darkcolor=UI.CARD,
            borderwidth=0,
            relief="flat"
        )

        style.map(
            "Treeview",
            background=[("selected !focus", UI.PRIMARY_SOFT), ("selected", UI.PRIMARY_SOFT)],
            foreground=[("selected !focus", "#0f172a"), ("selected", "#0f172a")]
        )

        style.configure(
            "Treeview.Heading",
            font=("Segoe UI", 8 if UI.COMPACT else 10, "bold"),
            background=UI.SURFACE_ALT,
            foreground="#111827",
            bordercolor=UI.BORDER,
            lightcolor=UI.SURFACE_ALT,
            darkcolor=UI.SURFACE_ALT,
            borderwidth=0,
            relief="flat",
            padding=((5, 4) if UI.COMPACT else (8, 7))
        )

        style.map("Treeview.Heading", background=[("active", "#e2e8f0")])

        try:
            style.layout("Treeview", [("Treeview.treearea", {"sticky": "nswe"})])
        except Exception:
            pass

        style.configure(
            "Vertical.TScrollbar",
            gripcount=0,
            background="#cbd5e1",
            darkcolor="#cbd5e1",
            lightcolor="#cbd5e1",
            troughcolor=UI.SURFACE_ALT,
            bordercolor=UI.SURFACE_ALT,
            arrowcolor=UI.MUTED,
            relief="flat",
            width=13,
        )

    @staticmethod
    def style_entry(entry, bg=None):
        try:
            entry.configure(
                bg=(bg or UI.SURFACE),
                fg=UI.TEXT,
                insertbackground=UI.TEXT,
                relief="flat",
                bd=0,
                highlightthickness=1,
                highlightbackground=UI.BORDER,
                highlightcolor=UI.PRIMARY,
            )
        except Exception:
            pass


class Card(tk.Frame):
    def __init__(self, parent, padx=16, pady=16):
        def _compact_pad(value, limit):
            if not UI.COMPACT:
                return value
            if isinstance(value, tuple):
                return tuple(min(int(v), limit) for v in value)
            return min(int(value), limit)

        padx = _compact_pad(padx, 10)
        pady = _compact_pad(pady, 9)
        super().__init__(
            parent,
            bg=UI.CARD,
            bd=0,
            relief="flat",
            highlightthickness=1,
            highlightbackground=UI.BORDER,
            highlightcolor=UI.BORDER,
        )
        self.inner = tk.Frame(self, bg=UI.CARD, padx=padx, pady=pady)
        self.inner.pack(fill="both", expand=True)


class HeaderBar(tk.Frame):
    def __init__(self, parent, title, subtitle=""):
        super().__init__(parent, bg=UI.CARD)
        left = tk.Frame(self, bg=UI.CARD)
        left.pack(side="left", fill="x", expand=True)

        tk.Label(left, text=title, font=UI.FONT_XL, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w")
        if subtitle:
            tk.Label(left, text=subtitle, font=UI.FONT_SM, bg=UI.CARD, fg=UI.MUTED).pack(anchor="w", pady=(2, 0))


class PrimaryButton(tk.Button):
    """
    A tk.Button with consistent styling + hover/press behavior.

    IMPORTANT: we keep per-instance colors, so callers can safely do:
        btn = PrimaryButton(...); btn.config(bg="yellow", activebackground="gold")
    and the hover/leave bindings will respect the new colors.
    """

    def __init__(self, parent, text, command, **kwargs):
        normal_bg = kwargs.pop("bg", UI.PRIMARY)
        hover_bg = kwargs.pop("activebackground", UI.PRIMARY_DARK)
        fg = kwargs.pop("fg", "white")
        active_fg = kwargs.pop("activeforeground", "white")

        super().__init__(
            parent,
            text=text,
            command=command,
            bg=normal_bg,
            fg=fg,
            activebackground=hover_bg,
            activeforeground=active_fg,
            bd=0,
            relief="flat",
            padx=(10 if UI.COMPACT else 16),
            pady=(7 if UI.COMPACT else 11),
            font=("Segoe UI", 8 if UI.COMPACT else 10, "bold"),
            cursor="hand2",
            **kwargs
        )

        self._normal_bg = normal_bg
        self._hover_bg = hover_bg

        # Hover/press effects (use instance colors)
        self.bind("<ButtonPress-1>", lambda e: self.configure(bg=self._hover_bg))
        self.bind("<ButtonRelease-1>", lambda e: self.configure(bg=self._normal_bg))
        self.bind("<Enter>", lambda e: self.configure(bg=self._hover_bg))
        self.bind("<Leave>", lambda e: self.configure(bg=self._normal_bg))

    def configure(self, cnf=None, **kw):
        # Keep internal colors in sync if caller changes them
        if "bg" in kw:
            self._normal_bg = kw["bg"]
        if "activebackground" in kw:
            self._hover_bg = kw["activebackground"]
        return super().configure(cnf, **kw)

    config = configure


class GhostButton(tk.Button):
    def __init__(self, parent, text, command):
        normal_bg = "#ffffff"
        hover_bg = "#f1f5f9"
        border = "#94a3b8"
        text_color = UI.TEXT

        super().__init__(
            parent,
            text=text,
            command=command,
            bg=normal_bg,
            fg=text_color,
            activebackground=hover_bg,
            activeforeground=UI.TEXT,
            bd=1,
            relief="solid",
            padx=(8 if UI.COMPACT else 12),
            pady=(5 if UI.COMPACT else 9),
            font=("Segoe UI", 8 if UI.COMPACT else 10),
            cursor="hand2",
            highlightthickness=1,
            highlightbackground=border,
            highlightcolor=border,
        )
        self.bind("<Enter>", lambda e: self.configure(bg=hover_bg, fg=UI.TEXT))
        self.bind("<Leave>", lambda e: self.configure(bg=normal_bg, fg=text_color))

        # Keyboard accessibility
        self.bind("<Return>", lambda e: self.invoke())


class DangerButton(tk.Button):
    def __init__(self, parent, text, command):
        super().__init__(
            parent,
            text=text,
            command=command,
            bg=UI.DANGER,
            fg="white",
            activebackground=UI.DANGER_DARK,
            activeforeground="white",
            bd=0,
            relief="flat",
            padx=(9 if UI.COMPACT else 12),
            pady=(6 if UI.COMPACT else 9),
            font=("Segoe UI", 8 if UI.COMPACT else 10, "bold"),
            cursor="hand2"
        )
        self.bind("<Enter>", lambda e: self.configure(bg=UI.DANGER_DARK))
        self.bind("<Leave>", lambda e: self.configure(bg=UI.DANGER))


# ---------------- SCROLLABLE FRAME (for small screens) ----------------

class VScrollableFrame(tk.Frame):
    """A vertical scrollable frame (Canvas + inner Frame) that plays nicely on small screens.

    Mousewheel is routed to whichever scroll frame the cursor is currently over.
    """
    _mw_bound = False
    _active = None

    def __init__(self, parent, bg=None, width=None):
        super().__init__(parent, bg=bg)

        self.canvas = tk.Canvas(self, bg=bg, highlightthickness=0, bd=0, width=width)
        self.vsb = tk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.vsb.set)

        self.inner = tk.Frame(self.canvas, bg=bg)
        self._win = self.canvas.create_window((0, 0), window=self.inner, anchor="nw")

        self.canvas.pack(side="left", fill="both", expand=True)
        self.vsb.pack(side="right", fill="y")

        self.inner.bind("<Configure>", self._on_inner_configure)
        self.canvas.bind("<Configure>", self._on_canvas_configure)

        # Activate on hover
        for w in (self, self.canvas, self.inner):
            w.bind("<Enter>", self._activate)
            w.bind("<Leave>", self._deactivate)

        # Bind global mousewheel once
        self.after(0, self._ensure_global_mousewheel)

    def _ensure_global_mousewheel(self):
        if VScrollableFrame._mw_bound:
            return
        VScrollableFrame._mw_bound = True
        top = self.winfo_toplevel()
        top.bind_all("<MouseWheel>", VScrollableFrame._route_mousewheel, add="+")
        top.bind_all("<Button-4>", VScrollableFrame._route_mousewheel, add="+")
        top.bind_all("<Button-5>", VScrollableFrame._route_mousewheel, add="+")

    def _activate(self, _=None):
        VScrollableFrame._active = self

    def _deactivate(self, _=None):
        if VScrollableFrame._active is self:
            VScrollableFrame._active = None

    @staticmethod
    def _route_mousewheel(event):
        """Route mousewheel to the VScrollableFrame under the pointer (when possible),
        otherwise fall back to the last-active frame.

        This does a safe, Python-level walk up the widget 'master' chain instead of
        using Tcl parent lookups so it won't throw while widgets are being created/
        destroyed. It also detects native scrollable widgets (Treeview/Text/Listbox)
        and lets them handle the event themselves.
        """
        try:
            top = None
            w = getattr(event, "widget", None)

            # Best-effort: find widget under pointer if possible
            try:
                top = w.winfo_toplevel() if w is not None else None
                if top is not None:
                    px = top.winfo_pointerx()
                    py = top.winfo_pointery()
                    try:
                        w2 = top.winfo_containing(px, py)
                        if w2 is not None:
                            w = w2
                    except Exception:
                        # If this fails for any reason, fall back to event.widget
                        pass
            except Exception:
                # Keep w as event.widget
                w = getattr(event, "widget", None)

            # Helper: check if widget or any ancestor is a native scrollable
            def _is_native_scrollable(win):
                cur = win
                while cur is not None:
                    try:
                        klass = cur.winfo_class()
                        if klass in ("Treeview", "Text", "Listbox"):
                            return True
                    except Exception:
                        pass
                    cur = getattr(cur, "master", None)
                return False

            try:
                if w is not None and _is_native_scrollable(w):
                    return
            except Exception:
                # If detection fails, do not block scrolling
                pass

            # Walk up via .master to find enclosing VScrollableFrame
            frame = None
            cur = w
            while cur is not None:
                try:
                    if isinstance(cur, VScrollableFrame):
                        frame = cur
                        break
                except Exception:
                    pass
                cur = getattr(cur, "master", None)

            if frame is None:
                frame = VScrollableFrame._active

            if frame is None:
                return

            if hasattr(frame, "_on_mousewheel"):
                try:
                    frame._on_mousewheel(event)
                except Exception:
                    pass
        except Exception:
            # Swallow any unexpected exceptions to avoid crashing the app.
            pass

    def _on_inner_configure(self, _):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        # Keep inner frame width equal to canvas width
        try:
            self.canvas.itemconfigure(self._win, width=event.width)
        except Exception:
            pass

    def _on_mousewheel(self, event):
        """Scroll safely without crashing when widgets are destroyed mid-event."""
        try:
            if not hasattr(self, "canvas") or self.canvas is None:
                return
            try:
                if hasattr(self.canvas, "winfo_exists") and int(self.canvas.winfo_exists()) == 0:
                    return
            except Exception:
                pass

            # Windows / Linux
            if hasattr(event, "delta") and event.delta:
                try:
                    self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
                except Exception:
                    pass
                return

            # Linux scroll
            if getattr(event, "num", None) == 4:
                try:
                    self.canvas.yview_scroll(-3, "units")
                except Exception:
                    pass
            elif getattr(event, "num", None) == 5:
                try:
                    self.canvas.yview_scroll(3, "units")
                except Exception:
                    pass
        except Exception:
            pass


class MaskPOS(tk.Tk):

    def report_callback_exception(self, exc, val, tb):
        """Prevent hard crashes from Tk callback exceptions (especially in frozen EXE)."""
        try:
            import traceback as _tb
            msg = "".join(_tb.format_exception(exc, val, tb))
        except Exception:
            msg = f"{exc}: {val}"
        try:
            # Log to a local file for debugging
            with open("pos_error.log", "a", encoding="utf-8") as f:
                f.write("\n" + msg + "\n")
        except Exception:
            pass
        try:
            messagebox.showerror("App Error", f"{exc.__name__}: {val}")
        except Exception:
            pass
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        try:
            icon_path = resource_path(os.path.join("assets", "maskpos_icon.ico"))
            if os.path.exists(icon_path):
                self.iconbitmap(icon_path)
        except Exception:
            pass
        # Responsive sizing (fixes small laptop screens)
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()

        # Compact mode for smaller screens (prevents clipped buttons)
        UI.set_compact(sw <= 1400 or sh <= 820)

        # Responsive window sizing
        min_w = 740 if UI.COMPACT else 980
        min_h = 500 if UI.COMPACT else 650
        w = max(min_w, min(1450, sw - (10 if UI.COMPACT else 60)))
        h = max(min_h, min(860, sh - (70 if UI.COMPACT else 90)))
        self.geometry(f"{w}x{h}")
        self.minsize(min_w, min_h)
        self.configure(bg=UI.APP_BG)

        UI.style_ttk(self)

        self._active_page = "CashierPage"
        self._scan_buf = ""
        self._scan_last_ts = 0.0
        self._scan_start_widget = None

        # connection UI state
        self._prev_connected = True
        self._disconnect_modal_shown = False

        self._build_layout()
        if UI.COMPACT:
            try:
                self.state("zoomed")
            except Exception:
                pass
        try:
            self.protocol("WM_DELETE_WINDOW", self._on_close)
        except Exception:
            pass
        self.bind_all("<Key>", self._global_key_handler)
        self.bind_all("<F11>", lambda _e: self.toggle_fullscreen())
        self.bind_all("<Escape>", lambda _e: self.exit_fullscreen())

        # crash recovery: prompt if a shift was left open
        self.after(300, self._check_unclosed_shift)

        # auto-refresh UI (no manual Refresh needed)
        self.after(2500, self._ui_tick)
        self.after(700, self._connection_tick)
        self._daily_report_email_last_attempt = ""
        self._daily_report_email_last_attempt_ts = 0.0
        self._daily_report_email_sending = False
        self.after(5000, self._daily_report_email_tick)
        self._update_check_running = False
        self.after(15000, lambda: self.check_for_app_update(silent=True))

    def toggle_fullscreen(self):
        try:
            current = bool(self.attributes("-fullscreen"))
        except Exception:
            current = False
        try:
            self.attributes("-fullscreen", not current)
        except Exception:
            if current:
                try:
                    self.state("zoomed")
                except Exception:
                    pass
            else:
                try:
                    self.state("zoomed")
                except Exception:
                    pass
        return "break"

    def exit_fullscreen(self):
        try:
            if bool(self.attributes("-fullscreen")):
                self.attributes("-fullscreen", False)
                return "break"
        except Exception:
            pass
        return None

    def _build_layout(self):
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        sb_w = 200 if UI.COMPACT else 240
        self.sidebar = tk.Frame(self, bg=UI.SIDEBAR, width=sb_w)
        self.sidebar.grid(row=0, column=0, sticky="nsw")
        self._sidebar_expanded_w = sb_w
        self._sidebar_collapsed_w = 64
        self._sidebar_collapsed = False
        self.sidebar.grid_propagate(False)
        self.sidebar.grid_rowconfigure(0, weight=1)
        self.sidebar.grid_columnconfigure(0, weight=1)

        # Scrollable nav area (small screens)
        self.sidebar_scroll = VScrollableFrame(self.sidebar, bg=UI.SIDEBAR, width=sb_w)
        self.sidebar_scroll.grid(row=0, column=0, sticky="nsew")

        # Footer (always visible)
        self.sidebar_footer = tk.Frame(self.sidebar, bg=UI.SIDEBAR)
        self.sidebar_footer.grid(row=1, column=0, sticky="ew")

        # Main content container.
        # Use page-level scrolling only where needed (Analytics / Shifts / Settings).
        # A global scroll canvas here can make whole pages "fly" vertically on small screens.
        self.content_scroll = None
        self.content_plain = tk.Frame(self, bg=UI.CONTENT_BG)
        self.content_plain.grid(row=0, column=1, sticky="nsew")
        self.content_scroll_inner = self.content_plain
        self.content_plain.grid_rowconfigure(0, weight=1)
        self.content_plain.grid_columnconfigure(0, weight=1)

        self.cashier_page = CashierPage(self.content_plain)
        self.products_page = None
        self.barcodes_page = None
        self.analytics_page = None
        self.shifts_page = None
        self.shift_history_page = None
        self.returns_page = None
        self.offers_page = None
        self.settings_page = None

        # Hotkey: Ctrl+P to reprint receipt from anywhere
        try:
            self.bind_all("<Control-p>", lambda e: self.cashier_page.open_reprint_popup())
        except Exception:
            pass

        self.pages = {"CashierPage": self.cashier_page}
        self._page_factories = {
            "ProductsPage": lambda: ProductsPage(self.content_scroll_inner,
                                                 on_product_added=self.cashier_page.on_product_added),
            "BarcodesPage": lambda: BarcodesPage(self.content_scroll_inner),
            "AnalyticsPage": lambda: AnalyticsPage(self.content_scroll_inner),
            "ShiftsPage": lambda: ShiftsPage(self.content_scroll_inner),
            "ShiftHistoryPage": lambda: ShiftHistoryPage(self.content_scroll_inner),
            "ReturnsPage": lambda: ReturnsPage(self.content_scroll_inner, cashier_page=self.cashier_page),
            "OffersPage": lambda: OffersPage(self.content_scroll_inner),
            "SettingsPage": lambda: SettingsPage(self.content_scroll_inner),
            "DataHealthPage": lambda: DataHealthPage(self.content_scroll_inner),
        }
        self._page_attrs = {
            "ProductsPage": "products_page",
            "BarcodesPage": "barcodes_page",
            "AnalyticsPage": "analytics_page",
            "ShiftsPage": "shifts_page",
            "ShiftHistoryPage": "shift_history_page",
            "ReturnsPage": "returns_page",
            "OffersPage": "offers_page",
            "SettingsPage": "settings_page",
            "DataHealthPage": "data_health_page",
        }

        for p in self.pages.values():
            p.grid(row=0, column=0, sticky="nsew")

        self._build_sidebar()
        try:
            if UI.COMPACT and self.winfo_screenwidth() <= 1366:
                self.toggle_sidebar()
        except Exception:
            pass
        self.show_page("CashierPage")

    def _get_page(self, page_name: str):
        page = self.pages.get(page_name)
        if page is not None:
            return page

        factory = getattr(self, "_page_factories", {}).get(page_name)
        if factory is None:
            raise KeyError(page_name)

        page = factory()
        page.grid(row=0, column=0, sticky="nsew")
        self.pages[page_name] = page

        attr = getattr(self, "_page_attrs", {}).get(page_name)
        if attr:
            setattr(self, attr, page)

        return page

    def _build_sidebar(self):
        top = tk.Frame(self.sidebar_scroll.inner, bg=UI.SIDEBAR)
        top.pack(fill="x", padx=16, pady=(18, 10))

        self._sb_toggle_btn = tk.Button(
            top, text="☰", font=("Segoe UI", 12, "bold"),
            bg=UI.SIDEBAR, fg="#e5e7eb", activebackground=UI.SIDEBAR_HOVER,
            activeforeground="white", bd=0, cursor="hand2",
            command=self.toggle_sidebar
        )
        self._sb_toggle_btn.pack(side="right", padx=(0, 2))

        self._sb_title_label = tk.Label(top, text="MASK POS", font=("Segoe UI", 14, "bold"), bg=UI.SIDEBAR, fg="white")
        self._sb_title_label.pack(anchor="w")
        self._sb_subtitle_label = tk.Label(top, text="Offline Register", font=UI.FONT_SM, bg=UI.SIDEBAR, fg="#94a3b8")
        self._sb_subtitle_label.pack(anchor="w", pady=(2, 0))

        self._sb_sep = tk.Frame(self.sidebar_scroll.inner, bg="#1f2937", height=1)
        self._sb_sep.pack(fill="x", padx=16, pady=12)

        self.nav_buttons = {}
        self._nav_meta = {}
        items = [
            ("🧾", "Cash Register", "CashierPage"),
            ("💵", "Cash Drawer", "ShiftsPage"),
            ("🔁", "Returns / Exchange", "ReturnsPage"),
            ("📦", "Products", "ProductsPage"),
            ("🏷", "Barcodes", "BarcodesPage"),
            ("🩺", "Data Health", "DataHealthPage"),
            ("📈", "Analytics", "AnalyticsPage"),
            ("⚙️", "Settings", "SettingsPage"),
        ]

        items.insert(2, ("#", "Shift History", "ShiftHistoryPage"))
        items.insert(max(0, len(items) - 1), ("$", "Offers", "OffersPage"))

        for icon, label, page in items:
            btn = tk.Frame(
                self.sidebar_scroll.inner,
                bg=UI.SIDEBAR,
                cursor="hand2"
            )
            btn.pack(fill="x", padx=10, pady=4)
            btn.grid_columnconfigure(1, weight=1)

            icon_lbl = tk.Label(
                btn, text=icon, font=("Segoe UI", 11),
                fg="#e5e7eb", bg=UI.SIDEBAR, bd=0
            )
            icon_lbl.grid(row=0, column=0, padx=(16, 12), pady=12, sticky="w")

            text_lbl = tk.Label(
                btn, text=label, font=("Segoe UI", 11),
                fg="#e5e7eb", bg=UI.SIDEBAR, bd=0
            )
            text_lbl.grid(row=0, column=1, pady=12, sticky="w")

            self._nav_meta[page] = {
                'icon': icon,
                'label': label,
                'btn': btn,
                'icon_lbl': icon_lbl,
                'text_lbl': text_lbl
            }

            def make_show_page(p):
                return lambda e=None: self.show_page(p)

            btn.bind("<Button-1>", make_show_page(page))
            icon_lbl.bind("<Button-1>", make_show_page(page))
            text_lbl.bind("<Button-1>", make_show_page(page))

            def make_enter(p):
                return lambda e: on_enter(e, p)

            def make_leave(p):
                return lambda e: on_leave(e, p)

            def on_enter(e, p):
                if p != self._active_page:
                    meta = self._nav_meta[p]
                    meta['btn'].configure(bg=UI.SIDEBAR_HOVER)
                    meta['icon_lbl'].configure(bg=UI.SIDEBAR_HOVER)
                    meta['text_lbl'].configure(bg=UI.SIDEBAR_HOVER)

            def on_leave(e, p):
                if p != self._active_page:
                    meta = self._nav_meta[p]
                    meta['btn'].configure(bg=UI.SIDEBAR)
                    meta['icon_lbl'].configure(bg=UI.SIDEBAR)
                    meta['text_lbl'].configure(bg=UI.SIDEBAR)

            btn.bind("<Enter>", make_enter(page))
            btn.bind("<Leave>", make_leave(page))
            icon_lbl.bind("<Enter>", make_enter(page))
            icon_lbl.bind("<Leave>", make_leave(page))
            text_lbl.bind("<Enter>", make_enter(page))
            text_lbl.bind("<Leave>", make_leave(page))

            self.nav_buttons[page] = btn

            # Footer content
        for w in self.sidebar_footer.winfo_children():
            try:
                w.destroy()
            except Exception:
                pass

        bottom = tk.Frame(self.sidebar_footer, bg=UI.SIDEBAR)
        bottom.pack(fill="x", padx=16, pady=14)

        self.conn_status_label = tk.Label(bottom, text="", font=UI.FONT_SM, bg=UI.SIDEBAR, fg="#94a3b8")
        self.conn_status_label.pack(anchor="w")

        tk.Label(bottom, text="Cash only", font=UI.FONT_SM, bg=UI.SIDEBAR, fg="#94a3b8").pack(anchor="w", pady=(6, 0))


    def toggle_sidebar(self):
        # Collapse/expand sidebar to save space.
        self._sidebar_collapsed = not getattr(self, "_sidebar_collapsed", False)
        collapsed = self._sidebar_collapsed

        new_w = self._sidebar_collapsed_w if collapsed else self._sidebar_expanded_w
        try:
            self.sidebar.configure(width=new_w)
            self.sidebar_scroll.configure(width=new_w)
            try:
                self.sidebar_scroll.canvas.configure(width=new_w)
            except Exception:
                pass
        except Exception:
            pass

        # Show/hide subtitle + separator to save space
        try:
            if collapsed:
                if getattr(self, "_sb_subtitle_label", None):
                    self._sb_subtitle_label.pack_forget()
                if getattr(self, "_sb_sep", None):
                    self._sb_sep.pack_forget()
            else:
                if getattr(self, "_sb_subtitle_label", None):
                    self._sb_subtitle_label.pack(anchor="w", pady=(2, 0))
                if getattr(self, "_sb_sep", None):
                    self._sb_sep.pack(fill="x", padx=16, pady=12)
        except Exception:
            pass

        # Update nav button layout (icon-only in collapsed mode)
        for page, btn in getattr(self, "nav_buttons", {}).items():
            meta = getattr(self, "_nav_meta", {}).get(page, {})
            icon_lbl = meta.get("icon_lbl")
            text_lbl = meta.get("text_lbl")
            if collapsed:
                if text_lbl:
                    text_lbl.grid_remove()
                if icon_lbl:
                    icon_lbl.grid_configure(padx=0)
                    btn.grid_columnconfigure(0, weight=1)
            else:
                if icon_lbl:
                    icon_lbl.grid_configure(padx=(16, 12))
                    btn.grid_columnconfigure(0, weight=0)
                if text_lbl:
                    text_lbl.grid()

        # Keep active highlight
        for name, btn in getattr(self, "nav_buttons", {}).items():
            meta = getattr(self, "_nav_meta", {}).get(name, {})
            icon_lbl = meta.get("icon_lbl")
            text_lbl = meta.get("text_lbl")
            if name == getattr(self, "_active_page", ""):
                btn.configure(bg=UI.SIDEBAR_ACTIVE)
                if icon_lbl: icon_lbl.configure(bg=UI.SIDEBAR_ACTIVE, fg="white")
                if text_lbl: text_lbl.configure(bg=UI.SIDEBAR_ACTIVE, fg="white")
            else:
                btn.configure(bg=UI.SIDEBAR)
                if icon_lbl: icon_lbl.configure(bg=UI.SIDEBAR, fg="#e5e7eb")
                if text_lbl: text_lbl.configure(bg=UI.SIDEBAR, fg="#e5e7eb")

    def show_page(self, page_name: str):
        # Hard lock: Cash Register requires an open shift
        if page_name == "CashierPage":
            try:
                if get_open_shift() is None:
                    messagebox.showwarning("Shift required", "No shift is open. Open a shift first.")
                    page_name = "ShiftsPage"
            except Exception:
                pass

        self._active_page = page_name

        # Content is non-scrollable: keep layouts compact so you never need to scroll to reach actions.
        self.content_plain.tkraise()

        page = self._get_page(page_name)
        page.tkraise()

        for name, btn in self.nav_buttons.items():
            meta = self._nav_meta.get(name, {})
            icon_lbl = meta.get("icon_lbl")
            text_lbl = meta.get("text_lbl")
            if name == page_name:
                btn.configure(bg=UI.SIDEBAR_ACTIVE)
                if icon_lbl: icon_lbl.configure(bg=UI.SIDEBAR_ACTIVE, fg="white")
                if text_lbl: text_lbl.configure(bg=UI.SIDEBAR_ACTIVE, fg="white")
            else:
                btn.configure(bg=UI.SIDEBAR)
                if icon_lbl: icon_lbl.configure(bg=UI.SIDEBAR, fg="#e5e7eb")
                if text_lbl: text_lbl.configure(bg=UI.SIDEBAR, fg="#e5e7eb")

        if page_name == "ShiftsPage":
            try:
                self.shifts_page.refresh_all()
            except Exception:
                pass
        if page_name == "ShiftHistoryPage":
            try:
                self.shift_history_page.load_day()
            except Exception:
                pass
        if page_name == "OffersPage":
            try:
                self.offers_page.refresh_credit_status()
            except Exception:
                pass

    def _ui_tick(self):
        """Periodic UI refresh for pages that show changing DB data."""
        try:
            if self._active_page == "ShiftsPage":
                try:
                    self.shifts_page.refresh_all()
                except Exception:
                    pass
                try:
                    self.shifts_page.load_day(silent=True)
                except Exception:
                    pass
            elif self._active_page == "ShiftHistoryPage":
                try:
                    self.shift_history_page.load_day()
                except Exception:
                    pass
        finally:
            try:
                self.after(5000, self._ui_tick)
            except Exception:
                pass

    def _on_close(self):
        try:
            stop_backend()
        except Exception:
            pass
        try:
            backup_pos_db()
        except Exception:
            pass
        try:
            self.destroy()
        except Exception:
            pass

    def check_for_app_update(self, silent=False, status_var=None):
        if getattr(self, "_update_check_running", False):
            return
        self._update_check_running = True
        if status_var is not None:
            try:
                status_var.set("Checking for updates...")
            except Exception:
                pass

        def worker():
            try:
                info = app_update.check_for_update()
                err = None
            except Exception as exc:
                info = None
                err = exc

            def finish():
                self._update_check_running = False
                if err is not None:
                    if status_var is not None:
                        status_var.set(f"Update check failed: {err}")
                    elif not silent:
                        messagebox.showwarning("App update", f"Could not check for updates.\n{err}")
                    return

                if not info or not info.get("available"):
                    msg = f"Mask POS is up to date. Current version: {APP_VERSION}."
                    if status_var is not None:
                        status_var.set(msg)
                    elif not silent:
                        messagebox.showinfo("App update", msg)
                    return

                text = app_update.describe_update(info)
                if status_var is not None:
                    status_var.set(f"Update available: v{info.get('version')}.")
                if messagebox.askyesno("App update available", text + "\n\nDownload and install now?"):
                    self._download_and_install_update(info, status_var=status_var)

            try:
                self.after(0, finish)
            except Exception:
                pass

        threading.Thread(target=worker, daemon=True).start()

    def _download_and_install_update(self, info, status_var=None):
        if not getattr(sys, "frozen", False):
            messagebox.showinfo(
                "App update",
                "The updater is ready, but automatic install only runs from the packaged Mask POS app.\n\n"
                "After you manually install the next packaged version, future updates can install from inside the app."
            )
            return

        def set_status(text):
            if status_var is not None:
                try:
                    status_var.set(text)
                except Exception:
                    pass

        def progress(done, total):
            if total:
                pct = int((done / max(1, total)) * 100)
                try:
                    self.after(0, lambda: set_status(f"Downloading update... {pct}%"))
                except Exception:
                    pass
            else:
                mb = done / (1024 * 1024)
                try:
                    self.after(0, lambda: set_status(f"Downloading update... {mb:.1f} MB"))
                except Exception:
                    pass

        def worker():
            try:
                zip_path = app_update.download_update(info, progress=progress)
            except Exception as exc:
                try:
                    self.after(0, lambda: messagebox.showwarning("App update", f"Download failed.\n{exc}"))
                    self.after(0, lambda: set_status(f"Download failed: {exc}"))
                except Exception:
                    pass
                return

            def install_now():
                if not messagebox.askyesno(
                    "Install update",
                    "The update has downloaded.\n\nMask POS will close, install the update, and reopen. Continue?"
                ):
                    set_status("Update downloaded but not installed.")
                    return
                try:
                    set_status("Installing update...")
                    app_update.launch_installer(
                        zip_path,
                        BASE_DIR,
                        restart_exe=sys.executable,
                        parent_pid=os.getpid(),
                    )
                    try:
                        stop_backend()
                    except Exception:
                        pass
                    try:
                        backup_pos_db()
                    except Exception:
                        pass
                    try:
                        _release_single_instance()
                    except Exception:
                        pass
                    os._exit(0)
                except Exception as exc:
                    messagebox.showwarning("App update", f"Could not start installer.\n{exc}")
                    set_status(f"Install failed: {exc}")

            try:
                self.after(0, install_now)
            except Exception:
                pass

        threading.Thread(target=worker, daemon=True).start()

    def _check_unclosed_shift(self):
        """On launch: if a previous shift is still open (e.g., crash), force a cashier-proof decision."""
        try:
            s = get_open_shift()
        except Exception:
            s = None

        if not s:
            return

        emp = row_get(s, "employee_name", "") or ""
        opened = row_get(s, "opened_at", "") or ""
        opening_cash = row_get(s, "opening_cash", "")

        msg = "Previous shift was not closed.\n\n"
        if emp:
            msg += f"Employee: {emp}\n"
        if opened:
            msg += f"Opened at: {opened}\n"
        if opening_cash != "":
            msg += f"Opening cash: {opening_cash}\n"
        msg += "\nChoose an action:"

        win = tk.Toplevel(self)
        win.title("Unclosed Shift")
        win.transient(self)
        win.grab_set()
        win.configure(bg=UI.CONTENT_BG)
        win.resizable(False, False)

        body = Card(win, padx=16, pady=14)
        body.pack(fill="both", expand=True, padx=12, pady=12)
        tk.Label(body.inner, text=msg, bg=UI.CARD, fg=UI.TEXT, justify="left", wraplength=380).pack(anchor="w")

        btns = tk.Frame(body.inner, bg=UI.CARD)
        btns.pack(fill="x", pady=(12, 0))

        def _close():
            try:
                win.grab_release()
            except Exception:
                pass
            win.destroy()

        def _resume():
            _close()
            try:
                self.show_page("CashierPage")
            except Exception:
                pass

        def _go_shift_page():
            _close()
            try:
                self.show_page("ShiftsPage")
            except Exception:
                pass

        def _close_now():
            # PIN check (crash recovery)
            emp_name = (emp or "").strip()
            if emp_name and employee_pin_required(emp_name):
                pin = simpledialog.askstring("PIN Required", f"Enter PIN for {emp_name}:", parent=win)
                if pin is None:
                    return
                if not verify_employee_pin(emp_name, pin):
                    messagebox.showerror("Wrong PIN", "Incorrect PIN.", parent=win)
                    return

            closing_str = simpledialog.askstring("Close Shift", "Enter closing cash amount:", parent=win)
            if closing_str is None:
                return
            try:
                closing_cash = float(closing_str)
            except Exception:
                messagebox.showerror("Invalid", "Please enter a valid number for closing cash.", parent=win)
                return

            try:
                sid = int(row_get(s, "id"))
                close_shift(sid, closing_cash, "Closed after crash recovery prompt")
            except Exception as e:
                messagebox.showerror("Error", f"{type(e).__name__}: {e}", parent=win)
                return

            messagebox.showinfo("Closed", "Shift closed successfully.", parent=win)
            _close()
            try:
                self.show_page("ShiftsPage")
            except Exception:
                pass

        PrimaryButton(btns, "Resume Shift", _resume).pack(side="left")
        GhostButton(btns, "Go to Cash Drawer", _go_shift_page).pack(side="left", padx=8)
        DangerButton(btns, "Close Shift Now", _close_now).pack(side="left", padx=8)
        GhostButton(btns, "Ignore", _close).pack(side="right")

        win.protocol("WM_DELETE_WINDOW", _close)

        try:
            win.update_idletasks()
            x = self.winfo_rootx() + (self.winfo_width() // 2) - (win.winfo_width() // 2)
            y = self.winfo_rooty() + (self.winfo_height() // 2) - (win.winfo_height() // 2)
            win.geometry(f"+{max(10, x)}+{max(10, y)}")
        except Exception:
            pass

    def _connection_tick(self):
        """Updates connection indicator + blocks sales when JOIN disconnects."""
        try:
            role = connection_role()
            if role == "HOST":
                if hasattr(self, "conn_status_label") and self.conn_status_label:
                    self.conn_status_label.config(text="HOST ● Online", fg=UI.SUCCESS)
                self._prev_connected = True
                self._disconnect_modal_shown = False

            elif role == "JOIN":
                connected = bool(is_connected())
                if hasattr(self, "conn_status_label") and self.conn_status_label:
                    if connected:
                        self.conn_status_label.config(text="JOIN ● Connected", fg=UI.SUCCESS)
                    else:
                        self.conn_status_label.config(text="JOIN ● DISCONNECTED", fg=UI.DANGER)

                # notify cashier + disable Complete Sale
                try:
                    self.cashier_page.set_connection_state(connected)
                except Exception:
                    pass

                # modal on transition to disconnected (once)
                if (self._prev_connected is True) and (connected is False) and (not self._disconnect_modal_shown):
                    self._disconnect_modal_shown = True
                    try:
                        messagebox.showwarning("Disconnected", "Connection to host lost")
                    except Exception:
                        pass

                if connected:
                    self._disconnect_modal_shown = False

                self._prev_connected = connected

            elif role == "CLOUD":
                now = time.time()
                cached = getattr(self, "_cloud_status_cache", None)
                last = float(getattr(self, "_cloud_status_last", 0.0) or 0.0)
                if cached is None or (now - last) >= 5.0:
                    try:
                        cached = cloud_sync_status(probe=True)
                    except Exception as e:
                        cached = {"online": False, "pending": 0, "last_error": str(e)}
                    self._cloud_status_cache = cached
                    self._cloud_status_last = now

                online = cached.get("online")
                pending = int(cached.get("pending") or 0)
                if hasattr(self, "conn_status_label") and self.conn_status_label:
                    if online is True:
                        self.conn_status_label.config(text=f"CLOUD Online  Pending {pending}", fg=UI.SUCCESS)
                    elif online is False:
                        self.conn_status_label.config(text=f"CLOUD Offline cache  Pending {pending}", fg=UI.DANGER)
                    else:
                        self.conn_status_label.config(text=f"CLOUD Local cache  Pending {pending}", fg="#94a3b8")
                self._prev_connected = True
                self._disconnect_modal_shown = False

            else:
                # Standalone: hide indicator text
                if hasattr(self, "conn_status_label") and self.conn_status_label:
                    self.conn_status_label.config(text="", fg="#94a3b8")
                self._prev_connected = True
                self._disconnect_modal_shown = False

        finally:
            try:
                self.after(700, self._connection_tick)
            except Exception:
                pass

    def _email_report_body(self, day: str, summary: dict, drawer_summary: dict | None = None) -> str:
        drawer_net = float(summary.get("drawer_net_change", 0.0) or 0.0)
        return (
            f"Mask POS Daily Report - {day}\n\n"
            f"Total sales / new money: {money(summary.get('merchandise_sales', 0.0))}\n"
            f"Cash collected: {money(summary.get('cash_collected', 0.0))}\n"
            f"Returns / refunds: {money(summary.get('returns', 0.0))} total, {money(summary.get('cash_refunds', 0.0))} cash\n"
            f"Drawer movement: +{money(summary.get('cash_added', 0.0))} in, -{money(summary.get('cash_removed', 0.0))} out\n"
            f"Drawer net cash change: {money(drawer_net)}\n"
            f"Orders: {int(summary.get('orders', 0) or 0)}\n\n"
            "Attached: Excel details + PDF with sales list and drawer summary."
        )

    def send_daily_report_email_for_day(self, day: str, source: str = "manual", silent: bool = False, force: bool = False) -> bool:
        day = str(day or "").strip() or datetime.now().strftime("%Y-%m-%d")
        try:
            datetime.strptime(day, "%Y-%m-%d")
        except Exception:
            if not silent:
                messagebox.showerror("Daily report", "Invalid report date.")
            return False

        try:
            if backend_mode() == "connect":
                self._daily_report_email_sending = True
                try:
                    ok, msg = trigger_daily_report_email_on_host(day, source, force)
                    if ok:
                        if not silent:
                            messagebox.showinfo("Daily report emailed", msg)
                        return True
                    else:
                        if not silent:
                            messagebox.showwarning("Daily report email", msg)
                        return False
                finally:
                    self._daily_report_email_sending = False
        except Exception as ex:
            if not silent:
                messagebox.showerror("Daily report", f"Could not trigger email from host.\n\n{ex}")
            return False

        cfg = get_daily_report_email_config()
        if not bool(cfg.get("enabled", True)):
            if not silent:
                messagebox.showinfo("Daily report", "Daily report email is disabled in Settings.")
            return False
        if (not force) and source == "schedule" and str(cfg.get("last_auto_sent_date") or "") == day:
            return True
        if (not force) and source not in ("close", "schedule") and str(cfg.get("last_sent_date") or "") == day:
            if not silent:
                messagebox.showinfo("Daily report", f"The daily report for {day} was already emailed.")
            return True

        missing = []
        if not str(cfg.get("sender_email") or "").strip():
            missing.append("sender email")
        if not str(cfg.get("smtp_username") or "").strip():
            missing.append("SMTP username")
        if not str(cfg.get("smtp_password") or "").strip():
            missing.append("SMTP app password")
        if missing:
            if not silent:
                messagebox.showwarning(
                    "Daily report email",
                    "The Excel report can be generated, but email is not configured yet.\n\n"
                    "Missing: " + ", ".join(missing) + "\n\n"
                    "Go to Settings > Daily Report Email."
                )
            return False

        if (not force) and bool(getattr(self, "_daily_report_email_sending", False)):
            if (not silent) and source not in ("close", "schedule"):
                messagebox.showinfo("Daily report", "The daily report email is already being sent.")
            return True

        self._daily_report_email_sending = True
        try:
            daily_report_mod = _load_daily_report_module()
            d = datetime.strptime(day, "%Y-%m-%d")
            reports_folder = os.path.join(BASE_DIR, "reports")
            result = daily_report_mod.build_sales_report_excel(data_path("pos.db"), reports_folder, d.year, d.month, str(d.day))
            drawer_pdf = daily_report_mod.build_cash_drawer_pdf(data_path("pos.db"), reports_folder, d.year, d.month, str(d.day))
        except Exception as ex:
            self._daily_report_email_sending = False
            if not silent:
                messagebox.showerror("Daily report", f"Could not generate the daily reports.\n\n{ex}")
            return False

        try:
            summary = result.get("summary", {}) or {}
            drawer_summary = drawer_pdf.get("summary", {}) or {}
            subject = f"Mask POS Daily Sales Report - {day}"
            body = self._email_report_body(day, summary, drawer_summary)
            attachments = [result.get("path", ""), drawer_pdf.get("path", "")]
            ok, msg = send_daily_report_email(subject, body, attachments, cfg.get("recipients"))
            if ok:
                try:
                    mark_daily_report_email_sent(day, source=source)
                except Exception:
                    pass
                if not silent:
                    messagebox.showinfo("Daily report emailed", msg)
                return True

            if not silent:
                messagebox.showwarning("Daily report email", msg)
            return False
        finally:
            self._daily_report_email_sending = False

    def _daily_report_email_tick(self):
        try:
            cfg = get_daily_report_email_config()
            if not bool(cfg.get("enabled", True)):
                return

            now = datetime.now()
            day = now.strftime("%Y-%m-%d")
            if str(cfg.get("last_auto_sent_date") or "") == day:
                return

            send_time = str(cfg.get("send_time") or "19:50").strip()
            try:
                hh, mm = [int(p) for p in send_time.split(":", 1)]
            except Exception:
                hh, mm = 19, 50
            target = now.replace(hour=max(0, min(23, hh)), minute=max(0, min(59, mm)), second=0, microsecond=0)
            # Send once the configured time has passed. The old exact-minute
            # check could miss the schedule if this tick ran at :49:59 then :50:59.
            if now < target:
                return

            attempt_key = f"{day}-{target.strftime('%H:%M')}"
            now_ts = time.time()
            last_attempt_ts = float(getattr(self, "_daily_report_email_last_attempt_ts", 0.0) or 0.0)
            if getattr(self, "_daily_report_email_last_attempt", "") == attempt_key and (now_ts - last_attempt_ts) < 600:
                return
            self._daily_report_email_last_attempt = attempt_key
            self._daily_report_email_last_attempt_ts = now_ts
            self.send_daily_report_email_for_day(day, source="schedule", silent=True)
        finally:
            try:
                self.after(30000, self._daily_report_email_tick)
            except Exception:
                pass

    def _looks_like_scanner_code(self, code: str) -> bool:
        text = str(code or "").strip()
        if len(text) < 6:
            return False
        digits = _digits_only(text)
        if len(digits) >= 6:
            return True
        upper = text.upper()
        return upper.startswith("R-") or upper.startswith("MASKPOS")

    def _clear_accidental_scan_widget(self, widget, code: str):
        if not isinstance(widget, tk.Entry):
            return
        try:
            if hasattr(self.cashier_page, "scan_entry") and widget is self.cashier_page.scan_entry:
                return
        except Exception:
            pass

        page_name = getattr(self, "_active_page", "")
        try:
            if page_name == "BarcodesPage" and getattr(self, "barcodes_page", None):
                if widget is getattr(self.barcodes_page, "qty_e", None):
                    widget.delete(0, tk.END)
                    widget.insert(0, "1")
                    return
                if widget is getattr(self.barcodes_page, "search_entry", None):
                    self.barcodes_page.search_var.set("")
                    widget.delete(0, tk.END)
                    self.barcodes_page.refresh()
                    return
            if page_name == "ProductsPage" and getattr(self, "products_page", None):
                product_entries = (
                    getattr(self.products_page, "barcode_e", None),
                    getattr(self.products_page, "search_e", None),
                    getattr(self.products_page, "stock_e", None),
                    getattr(self.products_page, "edit_stock", None),
                )
                if widget in product_entries:
                    widget.delete(0, tk.END)
                    return
        except Exception:
            pass

        try:
            current = widget.get()
            if str(code or "") in str(current or ""):
                widget.delete(0, tk.END)
        except Exception:
            pass

    def _global_key_handler(self, event):
        key = event.keysym
        ch = event.char or ""

        if key == "F1":
            try:
                self.cashier_page.show_shortcut_help()
            except Exception:
                pass
            return "break"

        if key.lower() == "z" and (int(getattr(event, "state", 0) or 0) & 0x4):
            if getattr(self, "_active_page", "") == "CashierPage":
                try:
                    self.cashier_page.undo_last_cart_action()
                except Exception:
                    pass
                return "break"

        # --- Keyboard-first Cash Register shortcuts (only when on CashierPage) ---
        try:
            if getattr(self, "_active_page", "") == "CashierPage" and hasattr(self, "cashier_page"):
                w = event.widget

                # If typing in an Entry, let that Entry own the keystrokes.
                # This is especially important for scan_entry: scanners type the
                # barcode into the field and press Enter, so the global scan
                # buffer must not process the same scan a second time.
                if isinstance(w, tk.Entry):
                    if hasattr(self.cashier_page, "scan_entry") and w is self.cashier_page.scan_entry:
                        if key == "Escape":
                            try:
                                self.cashier_page.cancel_sale()
                            except Exception:
                                pass
                        return
                    elif key != "Escape":
                        return

                if key in ("plus", "KP_Add"):
                    try:
                        self.cashier_page.change_selected_qty(+1)
                    except Exception:
                        pass
                    return

                if key in ("minus", "KP_Subtract"):
                    try:
                        self.cashier_page.change_selected_qty(-1)
                    except Exception:
                        pass
                    return

                if key == "Delete":
                    try:
                        self.cashier_page.remove_selected_line()
                    except Exception:
                        pass
                    return

                # Enter: if we have a scan buffer, let scan logic handle it below.
                if key in ("Return", "KP_Enter"):
                    if (getattr(self, "_scan_buf", "") or "").strip():
                        pass
                    else:
                        try:
                            self.cashier_page.invoke_primary_action()
                        except Exception:
                            pass
                        return

                if key == "Escape":
                    try:
                        self.cashier_page.cancel_sale()
                    except Exception:
                        pass
                    return
        except Exception:
            pass

        try:
            if (
                getattr(self, "_active_page", "") == "ReturnsPage"
                and getattr(self, "returns_page", None)
                and isinstance(event.widget, tk.Entry)
                and event.widget is self.returns_page.scan_entry
            ):
                self._scan_buf = ""
                self._scan_start_widget = None
                return
        except Exception:
            pass

        now = time.time()
        dt = now - self._scan_last_ts
        self._scan_last_ts = now

        if dt > 0.25:
            self._scan_buf = ""
            self._scan_start_widget = event.widget

        if key in ("Return", "KP_Enter"):
            code = self._scan_buf.strip()
            start_widget = getattr(self, "_scan_start_widget", None)
            self._scan_buf = ""
            self._scan_start_widget = None
            if code and self._looks_like_scanner_code(code):
                self._clear_accidental_scan_widget(start_widget or event.widget, code)
                if self._active_page == "ReturnsPage":
                    try:
                        self.returns_page.on_scan_code(code)
                    except Exception:
                        pass
                else:
                    self.show_page("CashierPage")
                    self.cashier_page.on_product_added(code)
                return

        if key in ("Shift_L", "Shift_R", "Control_L", "Control_R", "Alt_L", "Alt_R", "Caps_Lock", "Tab", "Escape"):
            return

        if ch and ch.isprintable():
            # If the user scans a barcode while not focused on any text input,
            # auto-focus the scan entry on CashierPage / ReturnsPage and input the character.
            active_page_obj = None
            if getattr(self, "_active_page", "") == "CashierPage" and hasattr(self, "cashier_page"):
                active_page_obj = self.cashier_page
            elif getattr(self, "_active_page", "") == "ReturnsPage" and hasattr(self, "returns_page"):
                active_page_obj = self.returns_page

            if active_page_obj and hasattr(active_page_obj, "scan_entry"):
                w = event.widget
                if not isinstance(w, (tk.Entry, tk.Text)):
                    entry = active_page_obj.scan_entry
                    if entry and entry.winfo_exists():
                        entry.focus_set()
                        entry.insert(tk.END, ch)
                        self._scan_buf = ""
                        self._scan_start_widget = None
                        return

            if not self._scan_buf:
                self._scan_start_widget = event.widget
            self._scan_buf += ch


# ---------------- CASHIER PAGE ----------------

class CashierPage(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg=UI.CONTENT_BG)
        self.cart_lines = []
        self.held_orders = []
        self._held_order_seq = 1
        self._cashier_recovery_path = Path(data_path("cashier_recovery.json"))
        self._recovered_cart = False
        self._undo_stack = []
        self._pending_scan_qty = 1
        pending_credit = _load_pending_exchange_credit()
        self.exchange_credit = float(pending_credit.get("amount") or 0.0)
        self.exchange_origin_sale_ids = list(pending_credit.get("origin_sale_ids") or [])
        self.exchange_return_ids = list(pending_credit.get("return_ids") or [])
        self.exchange_bon_codes = list(pending_credit.get("bon_codes") or [])
        self.exchange_original_sale_id = (self.exchange_origin_sale_ids[0] if self.exchange_origin_sale_ids else None)
        self._connected = True
        self._complete_btn = None
        self._wheel_spinning = False
        self._wheel_prize_claimed = False
        self._wheel_discount_pct = 0.0
        self._wheel_receipt_label = ""

        self._processing_sale = False
        self._shift_locked = False
        self._primary_enabled_requested = True
        self._primary_label = "Checkout"
        self._primary_cmd = self.checkout
        self._load_cashier_recovery()
        self._build()
        self.refresh_cart()
        self._update_held_order_button()
        if self._recovered_cart:
            self.after(350, lambda: messagebox.showinfo(
                "Cart recovered",
                "The unfinished cart from the previous session was restored.",
                parent=self,
            ))

    def _load_cashier_recovery(self):
        try:
            raw = json.loads(self._cashier_recovery_path.read_text(encoding="utf-8"))
        except Exception:
            return
        try:
            active = raw.get("active_cart") or {}
            lines = [dict(x) for x in (active.get("lines") or []) if isinstance(x, dict)]
            lines = [x for x in lines if int(x.get("qty") or 0) > 0][:250]
            self.cart_lines = lines
            self._recovered_cart = bool(lines)
            self._wheel_prize_claimed = bool(active.get("wheel_prize_claimed", False))
            self._wheel_discount_pct = float(active.get("wheel_discount_pct") or 0.0)
            self._wheel_receipt_label = str(active.get("wheel_receipt_label") or "")

            held = []
            for raw_order in (raw.get("held_orders") or [])[:50]:
                if not isinstance(raw_order, dict):
                    continue
                order = dict(raw_order)
                order["lines"] = [dict(x) for x in (order.get("lines") or []) if isinstance(x, dict)][:250]
                if order["lines"]:
                    held.append(order)
            self.held_orders = held
            highest = max([int(x.get("id") or 0) for x in held] or [0])
            self._held_order_seq = max(highest + 1, int(raw.get("next_held_order_id") or 1))
        except Exception:
            self.cart_lines = []
            self.held_orders = []
            self._held_order_seq = 1
            self._recovered_cart = False

    def _save_cashier_recovery(self, include_active=True):
        active = {
            "lines": copy.deepcopy(self.cart_lines) if include_active else [],
            "wheel_prize_claimed": bool(getattr(self, "_wheel_prize_claimed", False)) if include_active else False,
            "wheel_discount_pct": float(getattr(self, "_wheel_discount_pct", 0.0) or 0.0) if include_active else 0.0,
            "wheel_receipt_label": str(getattr(self, "_wheel_receipt_label", "") or "") if include_active else "",
        }
        payload = {
            "version": 1,
            "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "active_cart": active,
            "held_orders": copy.deepcopy(self.held_orders),
            "next_held_order_id": int(self._held_order_seq or 1),
        }
        path = self._cashier_recovery_path
        tmp = path.with_name(f".{path.name}.{os.getpid()}.tmp")
        try:
            path.parent.mkdir(parents=True, exist_ok=True)
            tmp.write_text(json.dumps(payload, indent=2), encoding="utf-8")
            os.replace(tmp, path)
        except Exception:
            try:
                tmp.unlink(missing_ok=True)
            except Exception:
                pass

    def _update_held_order_button(self):
        button = getattr(self, "_btn_held_orders", None)
        if button is not None:
            count = len(getattr(self, "held_orders", []) or [])
            button.config(text=(f"Held Orders ({count})" if count else "Held Orders"))

    def _remember_cart_state(self):
        snapshot = copy.deepcopy(self.cart_lines)
        if not self._undo_stack or self._undo_stack[-1] != snapshot:
            self._undo_stack.append(snapshot)
            del self._undo_stack[:-20]

    def undo_last_cart_action(self):
        if not self._undo_stack:
            messagebox.showinfo("Undo", "There is no cart change to undo.", parent=self)
            return
        self.cart_lines = self._undo_stack.pop()
        self.refresh_cart()
        self.refresh_picker()

    def show_shortcut_help(self):
        messagebox.showinfo(
            "Cash Register Shortcuts",
            "F1  Show this help\nCtrl+Z  Undo last cart change\n+ / -  Change selected quantity\n"
            "Delete  Remove selected item\nEnter  Continue checkout\nEscape  Cancel/back\n\n"
            "Smart quantity: type 5* and press Enter, then scan the product.",
            parent=self,
        )

    def _build(self):

        # Fixed layout for Cash Register:
        # - Header at top
        # - Cart + Find Product area in the middle (tables scroll internally)
        # - Bottom action bar pinned (no page scrolling needed)

        # This page must not grow taller than the visible viewport.
        # We clamp the middle area height on every resize.
        self.grid_rowconfigure(0, weight=0)
        self.grid_rowconfigure(1, weight=1)
        self.grid_rowconfigure(2, weight=0)
        self.grid_columnconfigure(0, weight=1)
        self.grid_propagate(False)

        # --- Header ---
        header = Card(self, padx=(10 if UI.COMPACT else 22), pady=(7 if UI.COMPACT else 14))
        header.grid(row=0, column=0, sticky="ew", padx=(8 if UI.COMPACT else 16), pady=((8 if UI.COMPACT else 14), (6 if UI.COMPACT else 12)))
        header.inner.grid_columnconfigure(0, weight=1)
        header.inner.grid_columnconfigure(1, weight=0)

        title_block = tk.Frame(header.inner, bg=UI.CARD)
        title_block.grid(row=0, column=0, sticky="w")
        tk.Label(title_block, text="Cash Register", font=UI.FONT_XL, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w")
        tk.Label(title_block, text="Ready to sell", font=UI.FONT_SM, bg=UI.CARD, fg=UI.MUTED).pack(anchor="w", pady=(2, 0))

        scan_row = tk.Frame(header.inner, bg=UI.CARD)
        scan_row.grid(row=0, column=1, sticky="e")
        scan_row.grid_columnconfigure(1, weight=1)

        tk.Label(scan_row, text="Barcode", font=UI.FONT_MD, bg=UI.CARD, fg=UI.TEXT_SOFT).grid(row=0, column=0, sticky="w")
        self.scan_entry = tk.Entry(scan_row, font=("Segoe UI", 11 if UI.COMPACT else 15), width=(22 if UI.COMPACT else 28))
        UI.style_entry(self.scan_entry, bg="#ffffff")
        self.scan_entry.grid(row=0, column=1, sticky="ew", padx=(8, 8), ipady=(4 if UI.COMPACT else 7))
        self.scan_entry.bind("<Return>", self._scan_entry_return)

        GhostButton(scan_row, "Reprint", self.open_reprint_popup).grid(row=0, column=2, sticky="e", padx=(0, 8))
        GhostButton(scan_row, "Quick item", self.add_quick_item).grid(row=0, column=3, sticky="e")

        def _wrap_scan_row(_evt=None):
            try:
                w = header.inner.winfo_width()
                if w and w < 820:
                    scan_row.grid_configure(row=1, column=0, columnspan=2, sticky="ew", pady=(14, 0))
                    self.scan_entry.config(width=(16 if UI.COMPACT else 18))
                else:
                    scan_row.grid_configure(row=0, column=1, columnspan=1, sticky="e", pady=0)
                    self.scan_entry.config(width=(22 if UI.COMPACT else 28))
            except Exception:
                pass

        header.inner.bind("<Configure>", _wrap_scan_row)
        _wrap_scan_row()

        # --- Middle area (clamped height, internal scroll in tables) ---
        body = tk.Frame(self, bg=UI.CONTENT_BG)
        body.grid(row=1, column=0, sticky="nsew", padx=(8 if UI.COMPACT else 16))
        body.grid_columnconfigure(0, weight=3)
        body.grid_columnconfigure(1, weight=2)
        body.grid_rowconfigure(0, weight=1)
        body.grid_propagate(False)

        left = Card(body, padx=18, pady=18)
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 8 if UI.COMPACT else 14))
        right = Card(body, padx=18, pady=18)
        right.grid(row=0, column=1, sticky="nsew")

        # Left: Cart
        cart_head = tk.Frame(left.inner, bg=UI.CARD)
        cart_head.pack(fill="x")
        tk.Label(cart_head, text="Cart", font=UI.FONT_LG, bg=UI.CARD, fg=UI.TEXT).pack(side="left")
        self.cart_count_lbl = tk.Label(cart_head, text="0 items", font=UI.FONT_SM, bg=UI.CARD, fg=UI.MUTED)
        self.cart_count_lbl.pack(side="right", pady=(3, 0))

        # --- Cart area uses grid so action buttons never get squished ---
        cart_area = tk.Frame(left.inner, bg=UI.CARD)
        cart_area.pack(fill="both", expand=True, pady=(10, 0))
        cart_area.grid_rowconfigure(0, weight=1)  # table grows/shrinks
        cart_area.grid_rowconfigure(1, weight=0)  # shortcuts fixed
        cart_area.grid_rowconfigure(2, weight=0)  # total/gift fixed
        cart_area.grid_columnconfigure(0, weight=1)

        # Cart table (scrolls internally)
        cart_table = tk.Frame(cart_area, bg=UI.CARD, highlightthickness=1, highlightbackground=UI.BORDER)
        cart_table.grid(row=0, column=0, sticky="nsew")

        cols = ("item", "price", "qty", "disc", "subtotal")
        self.tree = ttk.Treeview(cart_table, columns=cols, show="headings", height=12)
        self.tree.heading("item", text="Item")
        self.tree.heading("price", text="Price")
        self.tree.heading("qty", text="Qty")
        self.tree.heading("disc", text="Disc %")
        self.tree.heading("subtotal", text="Subtotal")
        self.tree.column("item", width=360, minwidth=180, anchor="w", stretch=True)
        self.tree.column("price", width=96, minwidth=78, anchor="e", stretch=False)
        self.tree.column("qty", width=72, minwidth=54, anchor="center", stretch=False)
        self.tree.column("disc", width=82, minwidth=62, anchor="center", stretch=False)
        self.tree.column("subtotal", width=118, minwidth=92, anchor="e", stretch=False)
        self.tree.tag_configure("odd", background=UI.SURFACE)
        self.tree.tag_configure("even", background=UI.CARD)
        self.tree.tag_configure("discounted", foreground="#047857")

        cart_scroll = ttk.Scrollbar(cart_table, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=cart_scroll.set)
        self.tree.pack(side="left", fill="both", expand=True)
        cart_scroll.pack(side="right", fill="y")
        self.tree.bind("<Double-1>", lambda e: self.set_line_discount())

        # Shortcuts row (never collapses; wraps to 2 lines if needed)
        shortcut_outer = tk.Frame(cart_area, bg=UI.CARD)
        shortcut_outer.grid(row=1, column=0, sticky="ew", pady=(8, 0))
        shortcut_outer.grid_columnconfigure(0, weight=1)

        shortcut_row = tk.Frame(shortcut_outer, bg=UI.CARD)
        shortcut_row.grid(row=0, column=0, sticky="ew")

        # grid so buttons keep a minimum width and don't compress
        for i in range(4):
            shortcut_row.grid_columnconfigure(i, weight=1, uniform="shortcut")

        self._btn_qty_plus = GhostButton(shortcut_row, "+ Qty", lambda: self.change_selected_qty(+1))
        self._btn_qty_minus = GhostButton(shortcut_row, "- Qty", lambda: self.change_selected_qty(-1))
        self._btn_remove = GhostButton(shortcut_row, "Remove", self.remove_selected_line)
        self._btn_cancel = DangerButton(shortcut_row, "Cancel Sale", self.cancel_sale)

        self._btn_qty_plus.grid(row=0, column=0, sticky="ew", padx=(0, 6), pady=(0, 6))
        self._btn_qty_minus.grid(row=0, column=1, sticky="ew", padx=(0, 6), pady=(0, 6))
        self._btn_remove.grid(row=0, column=2, sticky="ew", padx=(0, 6), pady=(0, 6))
        self._btn_cancel.grid(row=0, column=3, sticky="ew", padx=(0, 0), pady=(0, 6))

        # If the window is too narrow, move cancel to second row (so text isn't cut)
        def _wrap_shortcuts(_evt=None):
            try:
                w = shortcut_outer.winfo_width()
                # threshold tuned for small laptops
                if w and w < 520:
                    self._btn_cancel.grid_configure(row=1, column=0, columnspan=4, sticky="ew", padx=0, pady=(0, 0))
                else:
                    self._btn_cancel.grid_configure(row=0, column=3, columnspan=1, sticky="ew", padx=(0, 0),
                                                    pady=(0, 6))
            except Exception:
                pass

        shortcut_outer.bind("<Configure>", _wrap_shortcuts)
        _wrap_shortcuts()

        # Total + gift row (fixed)
        bottom_row = tk.Frame(cart_area, bg=UI.CARD)
        bottom_row.grid(row=2, column=0, sticky="ew", pady=(8, 0))
        bottom_row.grid_columnconfigure(0, weight=1)

        self.total_lbl = tk.Label(
            bottom_row,
            text="Total: $0.00",
            font=("Segoe UI", 13 if UI.COMPACT else 18, "bold"),
            bg=UI.CARD,
            fg=UI.TEXT
        )
        self.total_label = self.total_lbl  # alias for older code paths
        self.total_lbl.grid(row=0, column=0, sticky="w")

        self.pending_credit_lbl = tk.Label(
            bottom_row,
            text="",
            font=UI.FONT_SM,
            bg=UI.CARD,
            fg=UI.PRIMARY
        )
        self.pending_credit_lbl.grid(row=1, column=0, sticky="w", pady=(4, 0))

        self.clear_credit_btn = GhostButton(bottom_row, "Clear Credit", self.clear_exchange_credit_prompt)
        self.clear_credit_btn.grid(row=1, column=1, sticky="e", pady=(4, 0))
        self.clear_credit_btn.grid_remove()

        self.gift_var = tk.BooleanVar(value=False)
        self.gift_receipt_var = self.gift_var  # alias for older code paths
        tk.Checkbutton(
            bottom_row,
            text="Gift receipt",
            variable=self.gift_var,
            bg=UI.CARD,
            fg=UI.TEXT_SOFT,
            activebackground=UI.CARD,
            selectcolor=UI.CARD,
        ).grid(row=0, column=1, sticky="e")
        # Seasonal Sale toggle removed from Cashier screen (admin-only in Settings)
        # Right: Find Product / Right Stack (Picker, Review, Cash)
        self.right_title_lbl = tk.Label(right.inner, text="Find Product", font=UI.FONT_LG, bg=UI.CARD, fg=UI.TEXT)
        self.right_title_lbl.pack(anchor="w")

        # This stack will host: picker_frame (default), review_frame, cash_frame
        self.right_stack = tk.Frame(right.inner, bg=UI.CARD)
        self.right_stack.pack(fill="both", expand=True)

        self.picker_frame = tk.Frame(self.right_stack, bg=UI.CARD)
        self.picker_frame.pack(fill="both", expand=True)

        # Use grid so the Add Selected button stays visible on small windows
        self.picker_frame.grid_rowconfigure(0, weight=0)  # search
        self.picker_frame.grid_rowconfigure(1, weight=1)  # table
        self.picker_frame.grid_rowconfigure(2, weight=0)  # button
        self.picker_frame.grid_columnconfigure(0, weight=1)

        search_row = tk.Frame(self.picker_frame, bg=UI.CARD)
        search_row.grid(row=0, column=0, sticky="ew", pady=(12, 10))
        search_row.grid_columnconfigure(0, weight=1)

        self.search_entry = tk.Entry(search_row, font=UI.FONT_MD)
        UI.style_entry(self.search_entry, bg=UI.SURFACE)
        self.search_entry.grid(row=0, column=0, sticky="ew", padx=(0, 10), ipady=6)
        self.search_entry.bind("<KeyRelease>", lambda e: self.refresh_picker())

        GhostButton(search_row, "Search", self.refresh_picker).grid(
            row=0, column=1, sticky="e")

        picker_table = tk.Frame(self.picker_frame, bg=UI.CARD, highlightthickness=1, highlightbackground=UI.BORDER)
        picker_table.grid(row=1, column=0, sticky="nsew")
        picker_table.grid_columnconfigure(0, weight=1)
        picker_table.grid_rowconfigure(0, weight=1)

        pcols = ("barcode", "name", "price", "stock")
        self.picker_tree = ttk.Treeview(picker_table, columns=pcols, show="headings", height=10)
        self.picker_tree.heading("barcode", text="Barcode")
        self.picker_tree.heading("name", text="Product")
        self.picker_tree.heading("price", text="Price")
        self.picker_tree.heading("stock", text="Stock")
        self.picker_tree.column("barcode", width=116, minwidth=92, stretch=False)
        self.picker_tree.column("name", width=220, minwidth=130, anchor="w", stretch=True)
        self.picker_tree.column("price", width=82, minwidth=68, anchor="e", stretch=False)
        self.picker_tree.column("stock", width=60, minwidth=48, anchor="center", stretch=False)
        self.picker_tree.tag_configure("odd", background=UI.SURFACE)
        self.picker_tree.tag_configure("even", background=UI.CARD)

        picker_scroll = ttk.Scrollbar(picker_table, orient="vertical", command=self.picker_tree.yview)
        self.picker_tree.configure(yscrollcommand=picker_scroll.set)
        self.picker_tree.grid(row=0, column=0, sticky="nsew")
        picker_scroll.grid(row=0, column=1, sticky="ns")

        self.picker_tree.bind("<Double-1>", lambda e: self.add_selected_product())

        btn_row = tk.Frame(self.picker_frame, bg=UI.CARD)
        btn_row.grid(row=2, column=0, sticky="ew", pady=(10, 0))
        btn_row.grid_columnconfigure(0, weight=1)

        self._btn_add_selected = PrimaryButton(btn_row, "Add Selected", self.add_selected_product)
        # Make it full-width on small screens, right-aligned on wide screens
        self._btn_add_selected.grid(row=0, column=0, sticky="ew")

        def _wrap_add_selected(_evt=None):
            try:
                w = btn_row.winfo_width()
                self._btn_add_selected.grid_configure(sticky="ew")
            except Exception:
                pass

        btn_row.bind("<Configure>", _wrap_add_selected)
        _wrap_add_selected()

        # Frames created later:
        self.review_frame = None
        self.cash_frame = None

        self.refresh_picker()
        # --- Bottom action bar (pinned and bigger) ---
        action = Card(self, padx=(10 if UI.COMPACT else 18), pady=(8 if UI.COMPACT else 14))
        action.grid(row=2, column=0, sticky="ew", padx=(8 if UI.COMPACT else 16), pady=((6 if UI.COMPACT else 12), (8 if UI.COMPACT else 16)))
        action.inner.grid_columnconfigure(0, weight=1)

        bar = tk.Frame(action.inner, bg=UI.CARD)
        bar.grid(row=0, column=0, sticky="ew")
        bar.grid_columnconfigure(0, weight=1)

        left_actions = tk.Frame(bar, bg=UI.CARD)
        left_actions.pack(side="left", fill="x", expand=True)

        right_actions = tk.Frame(bar, bg=UI.CARD)
        right_actions.pack(side="right")

        self._btn_line_disc = GhostButton(left_actions, "Line Discount", self.set_line_discount)
        self._btn_hold_order = GhostButton(left_actions, "Hold Order", self.hold_current_order)
        self._btn_held_orders = GhostButton(left_actions, "Held Orders", self.open_held_orders)
        self._btn_load_bon = GhostButton(left_actions, "Load Bon", self.open_bon_popup)
        self._btn_spin_wheel = GhostButton(left_actions, "Spin Wheel", self.open_spin_wheel)
        self._btn_undo = GhostButton(left_actions, "Undo", self.undo_last_cart_action)
        self._btn_clear_cart = GhostButton(left_actions, "Clear Cart", self.clear_cart)
        self._btn_line_disc.pack(side="left")
        self._btn_hold_order.pack(side="left", padx=(12, 0))
        self._btn_held_orders.pack(side="left", padx=(12, 0))
        self._btn_load_bon.pack(side="left", padx=(12, 0))
        self._btn_spin_wheel.pack(side="left", padx=(12, 0))
        self._btn_undo.pack(side="left", padx=(12, 0))
        self._btn_clear_cart.pack(side="left", padx=(12, 0))

        self.checkout_amount_lbl = tk.Label(
            right_actions,
            text="$0.00",
            font=("Segoe UI", 11 if UI.COMPACT else 15, "bold"),
            bg=UI.CARD,
            fg=UI.TEXT,
        )
        self.checkout_amount_lbl.pack(side="left", padx=(0, 12))

        self.primary_action_btn = PrimaryButton(right_actions, "Checkout", self.checkout)
        try:
            self.primary_action_btn.config(width=22)
        except Exception:
            pass
        self.primary_action_btn.pack(side="right")

        def _wrap_action_bar(_evt=None):
            try:
                w = bar.winfo_width()
                if w and w < 720:
                    right_actions.pack_forget()
                    left_actions.pack_forget()
                    right_actions.pack(side="top", fill="x", expand=False, pady=(0, 8))
                    left_actions.pack(side="top", fill="x", expand=True)
                    self.checkout_amount_lbl.pack_configure(side="top", anchor="e", padx=(0, 0), pady=(0, 4 if UI.COMPACT else 8))
                    self.primary_action_btn.pack_configure(side="top", fill="x")
                    self._btn_line_disc.pack_configure(side="left", fill="x", expand=True)
                    self._btn_clear_cart.pack_configure(side="left", fill="x", expand=True)
                else:
                    right_actions.pack_forget()
                    left_actions.pack_forget()
                    left_actions.pack(side="left", fill="x", expand=True)
                    right_actions.pack(side="right")
                    self.checkout_amount_lbl.pack_configure(side="left", anchor="center", padx=(0, 12), pady=0)
                    self.primary_action_btn.pack_configure(side="right", fill="none")
                    self._btn_line_disc.pack_configure(side="left", fill="none", expand=False)
                    self._btn_clear_cart.pack_configure(side="left", fill="none", expand=False)
            except Exception:
                pass

        bar.bind("<Configure>", _wrap_action_bar)
        _wrap_action_bar()

        self._stage = "cart"
        self._primary_label = "Checkout"
        self._primary_cmd = self.checkout
        self._primary_enabled_requested = True
        self._apply_primary_state()

        # Clamp middle area height so the action bar is always visible.
        def _clamp_layout(_evt=None):
            try:
                self.update_idletasks()
                # Make this page match its container size
                try:
                    self.config(width=self.master.winfo_width(), height=self.master.winfo_height())
                except Exception:
                    pass

                total_h = max(400, self.winfo_height())
                header_h = max(0, header.winfo_height())
                action_h = max(0, action.winfo_height())
                # account for paddings between blocks
                available = max(160, total_h - header_h - action_h - (26 if UI.COMPACT else 44))
                body.config(height=available)

                body_w = int(body.winfo_width() or self.winfo_width())
                if body_w and body_w < 860:
                    body.grid_columnconfigure(0, weight=1)
                    body.grid_columnconfigure(1, weight=0)
                    body.grid_rowconfigure(0, weight=3)
                    body.grid_rowconfigure(1, weight=2)
                    left.grid_configure(row=0, column=0, sticky="nsew", padx=0, pady=(0, 8))
                    right.grid_configure(row=1, column=0, sticky="nsew", padx=0, pady=0)
                else:
                    body.grid_columnconfigure(0, weight=3)
                    body.grid_columnconfigure(1, weight=2)
                    body.grid_rowconfigure(0, weight=1)
                    body.grid_rowconfigure(1, weight=0)
                    left.grid_configure(row=0, column=0, sticky="nsew", padx=(0, 14), pady=0)
                    right.grid_configure(row=0, column=1, sticky="nsew", padx=0, pady=0)

                # Adjust tree row counts based on available pixels
                if UI.COMPACT:
                    rows_left = max(4, min(9, int((available - 92) / 28)))
                    rows_right = max(3, min(7, int((available - 100) / 30)))
                elif body_w and body_w < 860:
                    rows_left = max(4, min(7, int((available - 120) / 42)))
                    rows_right = max(3, min(5, int((available - 130) / 48)))
                else:
                    rows_left = max(6, min(13, int((available - 120) / 36)))
                    rows_right = max(5, min(10, int((available - 130) / 38)))
                self.tree.configure(height=rows_left)
                self.picker_tree.configure(height=rows_right)
            except Exception:
                return

        self.bind("<Configure>", _clamp_layout)
        try:
            self.master.bind("<Configure>", _clamp_layout)
        except Exception:
            pass
        self.after(60, _clamp_layout)

        # Make the tables adapt so everything fits without scrolling.
        def _fit_tables(_evt=None):
            try:
                # IMPORTANT: This page lives inside the app's scroll canvas.
                # Using self.winfo_height() can reflect the full scroll region,
                # not the visible viewport. That makes the tables too tall and
                # pushes the bottom action bar off-screen.
                top = self.winfo_toplevel()
                viewport_h = None
                try:
                    if hasattr(top, "content_scroll") and hasattr(top.content_scroll, "canvas"):
                        viewport_h = int(top.content_scroll.canvas.winfo_height())
                except Exception:
                    viewport_h = None

                total_h = viewport_h or int(self.winfo_height())
                if total_h < 260:
                    return

                # Estimate heights already used by header + action + paddings
                header_h = int(header.winfo_height())
                action_h = int(action.winfo_height())

                # Remaining vertical pixels for the two big panels (cart + picker)
                remaining = max(160, total_h - header_h - action_h - (34 if UI.COMPACT else 56))
                body_w = int(body.winfo_width() or self.winfo_width())

                # Convert pixels to tree rows (roughly 24-26px per row)
                # Clamp smaller so everything fits on 768p screens.
                if UI.COMPACT:
                    rows_left = max(4, min(9, int((remaining - 92) / 28)))
                    rows_right = max(3, min(7, int((remaining - 100) / 30)))
                elif body_w and body_w < 860:
                    rows_left = max(4, min(7, int((remaining - 120) / 42)))
                    rows_right = max(3, min(5, int((remaining - 130) / 48)))
                else:
                    rows_left = max(6, min(13, int((remaining - 120) / 36)))
                    rows_right = max(5, min(10, int((remaining - 130) / 38)))
                self.tree.configure(height=rows_left)
                self.picker_tree.configure(height=rows_right)
            except Exception:
                pass

        # Re-fit on both page and viewport resizes
        self.bind("<Configure>", _fit_tables, add="+")
        try:
            top = self.winfo_toplevel()
            if hasattr(top, "content_scroll") and hasattr(top.content_scroll, "canvas"):
                top.content_scroll.canvas.bind("<Configure>", _fit_tables, add="+")
        except Exception:
            pass
        self.after(100, _fit_tables)

        self._shift_lock_tick()

    def _persist_exchange_credit(self):
        try:
            amount = round(max(0.0, float(getattr(self, "exchange_credit", 0.0) or 0.0)), 2)
        except Exception:
            amount = 0.0
        origins = []
        for v in getattr(self, "exchange_origin_sale_ids", []) or []:
            try:
                iv = int(v)
                if iv not in origins:
                    origins.append(iv)
            except Exception:
                pass
        returns = []
        for v in getattr(self, "exchange_return_ids", []) or []:
            try:
                iv = int(v)
                if iv not in returns:
                    returns.append(iv)
            except Exception:
                pass
        bon_codes = []
        for v in getattr(self, "exchange_bon_codes", []) or []:
            code = str(v or "").strip().upper()
            if code and code not in bon_codes:
                bon_codes.append(code)

        self.exchange_credit = amount
        self.exchange_origin_sale_ids = origins if amount > 0 else []
        self.exchange_return_ids = returns if amount > 0 else []
        self.exchange_bon_codes = bon_codes if amount > 0 else []
        self.exchange_original_sale_id = (self.exchange_origin_sale_ids[0] if self.exchange_origin_sale_ids else None)
        _save_pending_exchange_credit(amount, self.exchange_origin_sale_ids, self.exchange_return_ids, self.exchange_bon_codes)

    def add_exchange_credit(self, amount: float, origin_sale_id=None, return_id=None):
        try:
            amount = round(max(0.0, float(amount or 0.0)), 2)
        except Exception:
            amount = 0.0
        if amount <= 0:
            return

        try:
            self.exchange_credit = round(max(0.0, float(getattr(self, "exchange_credit", 0.0) or 0.0)) + amount, 2)
        except Exception:
            self.exchange_credit = amount

        if not hasattr(self, "exchange_origin_sale_ids"):
            self.exchange_origin_sale_ids = []
        if not hasattr(self, "exchange_return_ids"):
            self.exchange_return_ids = []

        try:
            sid = int(origin_sale_id)
            if sid not in self.exchange_origin_sale_ids:
                self.exchange_origin_sale_ids.append(sid)
        except Exception:
            pass
        try:
            rid = int(return_id)
            if rid not in self.exchange_return_ids:
                self.exchange_return_ids.append(rid)
        except Exception:
            pass

        self._persist_exchange_credit()
        try:
            self.refresh_cart()
        except Exception:
            pass

    def set_exchange_credit_balance(self, amount: float):
        try:
            self.exchange_credit = round(max(0.0, float(amount or 0.0)), 2)
        except Exception:
            self.exchange_credit = 0.0
        self._persist_exchange_credit()
        try:
            self.refresh_cart()
        except Exception:
            pass

    def clear_exchange_credit_prompt(self):
        try:
            current = round(max(0.0, float(getattr(self, "exchange_credit", 0.0) or 0.0)), 2)
        except Exception:
            current = 0.0
        if current <= 0.005:
            return
        if not messagebox.askyesno(
                "Clear store credit",
                f"Remove the pending store credit of {money(current)}?\n\nThis only clears the unused credit waiting in this register."
        ):
            return
        self.exchange_credit = 0.0
        self.exchange_origin_sale_ids = []
        self.exchange_return_ids = []
        self.exchange_bon_codes = []
        self.exchange_original_sale_id = None
        self._persist_exchange_credit()
        self.refresh_cart()

    def _looks_like_bon_code(self, raw: str) -> bool:
        compact = re.sub(r"[^A-Za-z0-9]", "", str(raw or "")).upper()
        return compact.startswith("BON") and len(compact) >= 12

    def _loaded_bon_total(self) -> tuple[float, list[str]]:
        total = 0.0
        active_codes = []
        for raw in list(getattr(self, "exchange_bon_codes", []) or []):
            code = str(raw or "").strip().upper()
            if not code:
                continue
            try:
                bon = get_bon_by_code(code)
            except Exception:
                bon = None
            if not bon:
                continue
            status = str(row_get(bon, "status", "ACTIVE") or "ACTIVE").upper()
            try:
                remaining = round(max(0.0, float(row_get(bon, "remaining_amount", 0.0) or 0.0)), 2)
            except Exception:
                remaining = 0.0
            if status == "ACTIVE" and remaining > 0.005:
                total += remaining
                if code not in active_codes:
                    active_codes.append(code)
        return round(total, 2), active_codes

    def _sync_bon_credit_balance(self):
        if not getattr(self, "exchange_bon_codes", []):
            return
        total, active_codes = self._loaded_bon_total()
        self.exchange_bon_codes = active_codes
        self.exchange_credit = total
        self._persist_exchange_credit()

    def open_bon_popup(self):
        code = simpledialog.askstring("Load Bon", "Scan or enter bon code:", parent=self)
        if code is None:
            return
        self.load_bon_credit(code)

    def load_bon_credit(self, raw_code: str, silent: bool = False) -> bool:
        raw_code = str(raw_code or "").strip()
        if not raw_code:
            return False

        try:
            bon = get_bon_by_code(raw_code)
        except Exception as e:
            if not silent:
                messagebox.showerror("Bon", f"Could not load bon.\n{e}")
            return True

        if not bon:
            if not silent:
                messagebox.showerror("Bon", "Bon not found.")
            return True

        code = str(row_get(bon, "code", raw_code) or raw_code).strip().upper()
        status = str(row_get(bon, "status", "ACTIVE") or "ACTIVE").upper()
        try:
            remaining = round(max(0.0, float(row_get(bon, "remaining_amount", 0.0) or 0.0)), 2)
        except Exception:
            remaining = 0.0

        if status != "ACTIVE" or remaining <= 0.005:
            if not silent:
                messagebox.showwarning("Bon", f"Bon {code} has no active balance.")
            return True

        current = round(max(0.0, float(getattr(self, "exchange_credit", 0.0) or 0.0)), 2)
        if current > 0.005 and not getattr(self, "exchange_bon_codes", []):
            if not messagebox.askyesno(
                    "Replace pending credit",
                    f"The register already has pending exchange credit of {money(current)}.\n\n"
                    f"Replace it with bon {code} ({money(remaining)})?"):
                return True
            self.exchange_credit = 0.0
            self.exchange_origin_sale_ids = []
            self.exchange_return_ids = []
            self.exchange_original_sale_id = None

        if not hasattr(self, "exchange_bon_codes"):
            self.exchange_bon_codes = []
        if code in self.exchange_bon_codes:
            if not silent:
                messagebox.showinfo("Bon", f"Bon {code} is already loaded.")
            return True

        self.exchange_bon_codes.append(code)
        try:
            sid = int(row_get(bon, "original_sale_id", 0) or 0)
            if sid:
                if not hasattr(self, "exchange_origin_sale_ids"):
                    self.exchange_origin_sale_ids = []
                if sid not in self.exchange_origin_sale_ids:
                    self.exchange_origin_sale_ids.append(sid)
        except Exception:
            pass
        try:
            rid = int(row_get(bon, "return_id", 0) or 0)
            if rid:
                if not hasattr(self, "exchange_return_ids"):
                    self.exchange_return_ids = []
                if rid not in self.exchange_return_ids:
                    self.exchange_return_ids.append(rid)
        except Exception:
            pass

        self._sync_bon_credit_balance()
        self.refresh_cart()
        if not silent:
            messagebox.showinfo("Bon loaded", f"Loaded {code}: {money(remaining)}")
        try:
            self.scan_entry.focus_set()
        except Exception:
            pass
        return True

    def remove_exchange_credit_for_return(self, return_id: int, amount: float) -> bool:
        try:
            rid = int(return_id)
            amount = round(max(0.0, float(amount or 0.0)), 2)
            current = round(max(0.0, float(getattr(self, "exchange_credit", 0.0) or 0.0)), 2)
        except Exception:
            return False

        pending_returns = []
        for v in getattr(self, "exchange_return_ids", []) or []:
            try:
                iv = int(v)
                if iv not in pending_returns:
                    pending_returns.append(iv)
            except Exception:
                pass

        if rid not in pending_returns or current + 0.005 < amount:
            return False

        self.exchange_return_ids = [x for x in pending_returns if x != rid]
        self.exchange_credit = round(max(0.0, current - amount), 2)
        self._persist_exchange_credit()
        try:
            self.refresh_cart()
        except Exception:
            pass
        return True

    def _shift_lock_tick(self):
        """Disable checkout if no shift is open (cashier-proof hard lock)."""
        try:
            open_shift_row = None
            try:
                open_shift_row = get_open_shift()
            except Exception:
                open_shift_row = None

            locked = (open_shift_row is None)
            self._shift_locked = bool(locked)

            # Primary button is the final checkout action
            if hasattr(self, "primary_action_btn") and self.primary_action_btn is not None:
                try:
                    self._apply_primary_state()
                except Exception:
                    pass

            # If you have a separate 'Proceed' button, lock it too
            if hasattr(self, "proceed_btn") and self.proceed_btn is not None:
                if locked:
                    self.proceed_btn.config(state="disabled")
                else:
                    self.proceed_btn.config(state="normal")
        finally:
            try:
                self.after(900, self._shift_lock_tick)
            except Exception:
                pass

    def _prompt_quick_item(self, title: str = "Quick item", scanned_barcode: str = ""):
        """Prompt for a one-off item (name + price) and optionally keep the scanned barcode."""
        win = tk.Toplevel(self)
        win.title(title)
        try:
            win.transient(self.winfo_toplevel())
        except Exception:
            pass
        win.grab_set()

        outer = tk.Frame(win, padx=12, pady=12)
        outer.pack(fill="both", expand=True)

        normalized_barcode = normalize_item_barcode(scanned_barcode)

        # Fields
        row = 0
        if normalized_barcode:
            tk.Label(outer, text="Barcode:", font=UI.FONT_MD).grid(row=row, column=0, sticky="w", pady=(0, 6))
            bc_entry = tk.Entry(outer, font=UI.FONT_MD, width=28)
            bc_entry.grid(row=row + 1, column=0, sticky="ew", pady=(0, 10))
            bc_entry.insert(0, normalized_barcode)
            row += 2
        else:
            bc_entry = None

        tk.Label(outer, text="Item name:", font=UI.FONT_MD).grid(row=row, column=0, sticky="w", pady=(0, 6))
        name_entry = tk.Entry(outer, font=UI.FONT_MD, width=28)
        name_entry.grid(row=row + 1, column=0, sticky="ew", pady=(0, 10))
        row += 2

        tk.Label(outer, text="Item price:", font=UI.FONT_MD).grid(row=row, column=0, sticky="w", pady=(0, 6))
        price_entry = tk.Entry(outer, font=UI.FONT_MD, width=28)
        price_entry.grid(row=row + 1, column=0, sticky="ew")
        row += 2

        outer.grid_columnconfigure(0, weight=1)

        # Buttons
        btns = tk.Frame(outer)
        btns.grid(row=row, column=0, sticky="e", pady=(12, 0))

        result = {"ok": False, "name": "", "price": 0.0, "barcode": normalized_barcode}

        def _ok():
            barcode_value = normalized_barcode
            if bc_entry is not None:
                barcode_value = _digits_only(bc_entry.get() or "")
                if not barcode_value:
                    messagebox.showerror("Quick item", "Barcode must be 13 digits or fewer. Short barcodes are saved as 13 digits by adding zeros on the right. The app still checks the legacy leading-zero match during lookup.")
                    bc_entry.focus_set()
                    return

            name = (name_entry.get() or "").strip()
            price_str = (price_entry.get() or "").strip()
            if not name:
                messagebox.showerror("Quick item", "Please enter an item name.")
                name_entry.focus_set()
                return
            try:
                price = float(price_str)
            except Exception:
                messagebox.showerror("Quick item", "Please enter a valid price.")
                price_entry.focus_set()
                return
            result.update({"ok": True, "name": name, "price": float(price), "barcode": barcode_value})
            win.destroy()

        def _cancel():
            win.destroy()

        ttk.Button(btns, text="OK", command=_ok).pack(side="left", padx=(0, 8))
        ttk.Button(btns, text="Cancel", command=_cancel).pack(side="left")

        # Enter/Esc shortcuts
        win.bind("<Return>", lambda e: _ok())
        win.bind("<Escape>", lambda e: _cancel())

        # Position near center of parent
        try:
            win.update_idletasks()
            pw = self.winfo_toplevel()
            x = pw.winfo_rootx() + (pw.winfo_width() // 2) - (win.winfo_width() // 2)
            y = pw.winfo_rooty() + (pw.winfo_height() // 2) - (win.winfo_height() // 2)
            win.geometry(f"+{max(x, 10)}+{max(y, 10)}")
        except Exception:
            pass

        if bc_entry is not None:
            name_entry.focus_set()
        else:
            name_entry.focus_set()
        win.wait_window()

        if not result["ok"]:
            return None
        return result

    def add_or_increment_quick_item(self, name: str, price: float, barcode: str | None = None):
        name_clean = str(name).strip()
        price_val = float(price)
        bc = normalize_item_barcode(barcode or "")
        self._remember_cart_state()

        for line in self.cart_lines:
            if (line.get("product_id") is None) or bool(line.get("is_quick", False)):
                existing_name = str(line.get("name") or "").strip().lower()
                existing_price = float(line.get("price") or 0.0)
                if existing_name == name_clean.lower() and abs(existing_price - price_val) < 0.0001:
                    line["qty"] += 1
                    if bc and not line.get("barcode"):
                        line["barcode"] = bc
                    self._recalc_line(line)
                    self.refresh_cart()
                    self.refresh_picker()
                    self.scan_entry.focus_set()
                    return

        # Not found, add new quick line
        line = {
            "product_id": None,
            "name": name_clean,
            "price": price_val,
            "original_price": price_val,
            "qty": 1,
            "discount_pct": 0.0,
            "barcode": bc or None,
            "is_quick": True,
        }
        self._recalc_line(line)
        self.cart_lines.append(line)
        self.refresh_cart()
        self.refresh_picker()
        self.scan_entry.focus_set()

    def add_quick_item(self):
        # Add a one-off/manual product directly in the register (no barcode required).
        res = self._prompt_quick_item("Quick item")
        if not res:
            return
        self.add_or_increment_quick_item(res["name"], res["price"], res.get("barcode"))


    def _materialize_quick_line(self, line: dict) -> dict:
        """Ensure a 'quick item' line has a real product_id for backend sales.

        Quick/manual items are not part of the catalog, but the backend expects a valid product_id
        in every cart line. We therefore create a real product entry (category='Quick') with stock 0,
        then attach its product_id to this cart line.

        If this fails, we raise an exception so checkout can stop cleanly (instead of breaking later).
        """
        if line.get("product_id"):
            return line

        name = (line.get("name") or "").strip() or "Quick item"
        try:
            price = float(line.get("price") or 0.0)
        except Exception:
            price = 0.0
        if price < 0:
            price = 0.0

        # Creating products requires host connectivity in JOIN mode
        if connection_role() == "JOIN" and getattr(self, "_connected", True) is False:
            raise RuntimeError("Not connected to host, cannot save quick item.")

        preferred_barcode = normalize_item_barcode(line.get("barcode") or "")

        # Check if identical quick item already exists in database
        prod = None
        try:
            candidates = list_products(name) or []
            matches = []
            for p in candidates:
                try:
                    p_name = str(row_get(p, "name") or "").strip()
                    p_cat = str(row_get(p, "category") or "").strip()
                    p_price = float(row_get(p, "sell_price") or 0.0)
                    p_id = int(row_get(p, "id") or row_get(p, "product_id") or 0)
                except Exception:
                    continue
                if p_id and p_name.lower() == name.lower() and p_cat.lower() == "quick" and abs(p_price - price) < 0.0001:
                    matches.append(p)
            if matches:
                prod = max(matches, key=lambda p: int(row_get(p, "id") or row_get(p, "product_id") or 0))
        except Exception:
            pass

        if prod:
            pid = row_get(prod, "id") or row_get(prod, "product_id")
            line["product_id"] = int(pid)
            line["barcode"] = row_get(prod, "barcode") or preferred_barcode or None
            line["is_quick"] = True
            return line

        # Create product (stock-neutral). Try to preserve the scanned barcode when possible.
        bc_or_id = add_product(
            name=name,
            category="Quick",
            brand="",
            sell_price=price,
            stock_qty=0,
            low_stock_level=0,
            barcode=(preferred_barcode or None),
        )

        # If local logic returns a numeric product id directly, use it.
        if isinstance(bc_or_id, int):
            line["product_id"] = bc_or_id
            line["barcode"] = preferred_barcode or None
            line["is_quick"] = True
            return line

        bc = str(bc_or_id)

        # Otherwise, fetch the created product to obtain its real product id
        prod = find_product_by_barcode(bc)
        if not isinstance(prod, dict) and preferred_barcode and preferred_barcode != bc:
            prod = find_product_by_barcode(preferred_barcode)
        if not isinstance(prod, dict):
            try:
                candidates = list_products(name) or []
                matches = []
                for p in candidates:
                    try:
                        p_name = str(row_get(p, "name") or "").strip()
                        p_cat = str(row_get(p, "category") or "").strip()
                        p_price = float(row_get(p, "sell_price") or 0.0)
                        p_id = int(row_get(p, "id") or row_get(p, "product_id") or 0)
                    except Exception:
                        continue
                    if p_id and p_name == name and p_cat == "Quick" and abs(p_price - price) < 0.0001:
                        matches.append(p)
                if matches:
                    prod = max(matches, key=lambda p: int(row_get(p, "id") or row_get(p, "product_id") or 0))
            except Exception:
                pass
        pid = None
        if isinstance(prod, dict):
            pid = prod.get("id") or prod.get("product_id")
        elif prod is not None:
            try:
                pid = row_get(prod, "id") or row_get(prod, "product_id")
            except Exception:
                pid = None
        if not pid:
            raise RuntimeError("Quick item was created but could not be fetched back from the database.")

        line["product_id"] = int(pid)
        line["barcode"] = normalize_item_barcode((row_get(prod, "barcode") if prod is not None else bc) or preferred_barcode or bc)

        line["is_quick"] = True
        return line

    def scan_barcode(self):
        self.on_scan(None)

    def _scan_entry_return(self, event=None):
        try:
            top = self.winfo_toplevel()
            top._scan_buf = ""
        except Exception:
            pass
        raw_scan = self.scan_entry.get().strip()
        qty_match = re.fullmatch(r"(\d{1,3})\*", raw_scan)
        if qty_match:
            qty = max(1, min(999, int(qty_match.group(1))))
            self._pending_scan_qty = qty
            self.scan_entry.delete(0, tk.END)
            try:
                self.right_title_lbl.configure(text=f"Quantity {qty}: scan or choose a product")
            except Exception:
                pass
            return "break"
        if not raw_scan:
            try:
                if getattr(self, "_stage", "cart") == "cart" and self.cart_lines:
                    self._show_review_panel()
                    if getattr(self, "_stage", "") == "review":
                        self.invoke_primary_action()
                else:
                    self.invoke_primary_action()
            except Exception:
                pass
            return "break"
        self.on_scan(None)
        return "break"

    def on_product_added(self, barcode: str):
        self.scan_entry.delete(0, tk.END)
        self.scan_entry.insert(0, barcode)
        self.on_scan(None)

    def on_scan(self, event):
        barcode = self.scan_entry.get().strip()
        self.scan_entry.delete(0, tk.END)
        if not barcode:
            return

        if self._looks_like_bon_code(barcode):
            self.load_bon_credit(barcode)
            return

        normalized_barcode = normalize_item_barcode(barcode)
        if not normalized_barcode:
            messagebox.showwarning("Invalid barcode", "Item barcode must be 13 digits or fewer. Short codes are saved as 13 digits by adding zeros on the right.")
            self.scan_entry.focus_set()
            return

        row = None
        for cand in barcode_candidates(barcode):
            row = find_product_by_barcode(cand)
            if row:
                break
        if not row:
            display_barcode = normalized_barcode or _digits_only(barcode)
            if messagebox.askyesno("Not found", f"Barcode not found: {display_barcode}\n\nAdd as a quick item?"):
                res = self._prompt_quick_item("Quick item", scanned_barcode=(normalized_barcode or _digits_only(barcode)))
                if not res:
                    return
                self.add_or_increment_quick_item(
                    res["name"],
                    res["price"],
                    res.get("barcode") or normalized_barcode or _digits_only(barcode)
                )
                return
            else:
                # Bug 7 fix: user already saw the "Not found" question; no second popup needed
                return

        product_id = row["id"]
        name = row["name"]
        price = float(row["sell_price"])
        qty_to_add = max(1, min(999, int(getattr(self, "_pending_scan_qty", 1) or 1)))
        self._pending_scan_qty = 1
        self._remember_cart_state()

        for line in self.cart_lines:
            if line["product_id"] == product_id:
                line["qty"] += qty_to_add
                if "discount_pct" not in line:
                    line["discount_pct"] = 0.0
                # Keep barcode on the line (needed for seasonal sale lookups)
                try:
                    if not line.get("barcode"):
                        line["barcode"] = str(row.get("barcode") or barcode).strip()
                except Exception:
                    pass

                # Re-apply automatic offers (does nothing if no matching offer exists)
                try:
                    self._apply_auto_offers_to_cart()
                except Exception:
                    pass
                self._recalc_line(line)
                self.refresh_cart()
                self.refresh_picker()
                return

        # Seasonal sale auto-discount (barcode-based)
        try:
            bc_key = str(row.get("barcode") or barcode).strip()
        except Exception:
            bc_key = str(barcode).strip()

        sale_pct = 0.0
        try:
            if bool(get_seasonal_sale_enabled()):
                sale_pct = float(get_seasonal_sales_map().get(bc_key) or 0.0)
        except Exception:
            sale_pct = 0.0
        sale_pct = max(0.0, min(100.0, sale_pct))

        line = {
            "product_id": product_id,
            "barcode": bc_key,
            "name": name,
            "price": price,
            "original_price": price,
            "qty": qty_to_add,
            "discount_pct": float(sale_pct if sale_pct > 0 else 0.0),
            "sale_pct": float(sale_pct if sale_pct > 0 else 0.0),
            "sale_applied": bool(sale_pct > 0),
            "manual_override": False,
        }
        self._recalc_line(line)
        self.cart_lines.append(line)
        try:
            self._apply_auto_offers_to_cart()
        except Exception:
            pass
        self.refresh_cart()
        self.refresh_picker()

    def _recalc_line(self, line):
        try:
            price = float(line.get("price") or 0)
        except Exception:
            price = 0.0
        try:
            qty = int(line.get("qty") or 0)
        except Exception:
            qty = 0
        try:
            disc = float(line.get("discount_pct") or 0)
        except Exception:
            disc = 0.0
        disc = max(0.0, min(100.0, disc))
        line["discount_pct"] = disc
        line_total = whole_money_round_up(price * qty * (1.0 - disc / 100.0))

        if bool(line.get("bundle_offer_applied", False)):
            try:
                offer_qty = int(line.get("bundle_offer_qty") or 0)
                offer_price = float(line.get("bundle_offer_price") or 0.0)
            except Exception:
                offer_qty = 0
                offer_price = 0.0
            if qty >= offer_qty >= 2 and offer_price > 0:
                line_total = whole_money_round_up(qty * (offer_price / offer_qty))
            else:
                line["bundle_offer_applied"] = False
                line["bundle_offer_qty"] = 0
                line["bundle_offer_price"] = 0.0

        line["line_total"] = line_total

    def set_line_discount(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Select item", "Select an item in the cart first.")
            return
        idx = self.tree.index(sel[0])
        if idx < 0 or idx >= len(self.cart_lines):
            return
        line = self.cart_lines[idx]

        win = tk.Toplevel(self)
        win.title("Line Discount / Price")
        win.geometry("500x360")
        win.minsize(440, 330)
        win.configure(bg=UI.CONTENT_BG)
        win.grab_set()

        box = Card(win, padx=18, pady=18)
        box.pack(fill="both", expand=True, padx=14, pady=14)

        tk.Label(
            box.inner,
            text=f"Edit line: {line.get('name', '')}",
            font=UI.FONT_LG,
            bg=UI.CARD,
            fg=UI.TEXT
        ).pack(anchor="w")

        tk.Label(
            box.inner,
            text="Set a new unit price, apply a discount, or use a quick discount.",
            font=UI.FONT_MD,
            bg=UI.CARD,
            fg="#334155"
        ).pack(anchor="w", pady=(8, 14))

        form = tk.Frame(box.inner, bg=UI.CARD)
        form.pack(fill="x")

        row1 = tk.Frame(form, bg=UI.CARD)
        row1.pack(fill="x", pady=6)
        tk.Label(row1, text="Unit price", width=14, anchor="w", bg=UI.CARD, fg=UI.TEXT).pack(side="left")
        price_e = tk.Entry(row1, font=("Segoe UI", 14), width=12, bd=1, relief="solid")
        price_e.insert(0, str(float(line.get("price") or 0)))
        price_e.pack(side="left")

        row2 = tk.Frame(form, bg=UI.CARD)
        row2.pack(fill="x", pady=6)
        tk.Label(row2, text="Discount %", width=14, anchor="w", bg=UI.CARD, fg=UI.TEXT).pack(side="left")
        disc_e = tk.Entry(row2, font=("Segoe UI", 14), width=12, bd=1, relief="solid")
        disc_e.insert(0, str(float(line.get("discount_pct") or 0)))
        disc_e.pack(side="left")

        quick = tk.Frame(box.inner, bg=UI.CARD)
        quick.pack(fill="x", pady=(10, 0))
        tk.Label(quick, text="Quick discount", width=14, anchor="w", bg=UI.CARD, fg=UI.TEXT).pack(side="left")
        quick_btns = tk.Frame(quick, bg=UI.CARD)
        quick_btns.pack(side="left", fill="x", expand=True)

        def set_quick_discount(pct):
            try:
                disc_e.delete(0, tk.END)
                disc_e.insert(0, str(float(pct)))
                disc_e.focus_set()
                disc_e.select_range(0, tk.END)
            except Exception:
                pass

        for pct in (0, 10, 20, 30, 50):
            GhostButton(quick_btns, f"{pct}%", lambda p=pct: set_quick_discount(p)).pack(side="left", padx=(0, 6))

        hint = tk.Label(
            box.inner,
            text="Tip: double-click any cart item to open this window.",
            font=UI.FONT_SM,
            bg=UI.CARD,
            fg=UI.MUTED
        )
        hint.pack(anchor="w", pady=(10, 0))

        def apply():
            try:
                new_price = float(price_e.get().strip() or "0")
            except Exception:
                messagebox.showerror("Invalid", "Unit price must be a number.")
                return
            if new_price < 0:
                new_price = 0.0

            try:
                val = float(disc_e.get().strip() or "0")
            except Exception:
                messagebox.showerror("Invalid", "Discount must be a number.")
                return
            val = max(0.0, min(100.0, val))

            # If cashier edits price/discount, treat it as a manual override (do not auto-overwrite with seasonal sale)
            try:
                old_price = float(line.get("price") or 0.0)
            except Exception:
                old_price = 0.0
            try:
                old_disc = float(line.get("discount_pct") or 0.0)
            except Exception:
                old_disc = 0.0

            self._remember_cart_state()
            if "original_price" not in line:
                line["original_price"] = float(old_price)
            line["price"] = float(new_price)
            line["discount_pct"] = float(val)

            if (abs(float(new_price) - old_price) > 1e-9) or (abs(float(val) - old_disc) > 1e-9):
                line["manual_override"] = True
                # Once manually edited, we stop considering it a seasonal-sale-managed line
                line["sale_applied"] = False
                line["sale_pct"] = 0.0
                line["bundle_offer_applied"] = False
                line["bundle_offer_qty"] = 0
                line["bundle_offer_price"] = 0.0

            self._recalc_line(line)
            self.refresh_cart()
            win.destroy()
            try:
                if self.review_frame is not None:
                    self._show_review_panel()
            except Exception:
                pass

        btns = tk.Frame(box.inner, bg=UI.CARD)
        btns.pack(fill="x", pady=(16, 0))
        GhostButton(btns, "Cancel", win.destroy).pack(side="right")
        PrimaryButton(btns, "Apply", apply).pack(side="right", padx=(0, 10))

    def refresh_picker(self):
        if not hasattr(self, "picker_tree"):
            return
        # The search box on the picker is `self.search_entry`; keep fallback for
        # older constructors that might have used `picker_search`.
        if hasattr(self, "search_entry"):
            q = (self.search_entry.get() or "").strip()
        elif hasattr(self, "picker_search"):
            q = (self.picker_search.get() or "").strip()
        else:
            q = ""

        rows = list_products(q)

        for i in self.picker_tree.get_children():
            self.picker_tree.delete(i)

        for idx, r in enumerate(rows):
            try:
                p_cat = str(row_get(r, "category") or "").strip()
                if p_cat.lower() == "quick":
                    continue
            except Exception:
                pass
            tag = "odd" if idx % 2 else "even"
            self.picker_tree.insert("", tk.END, tags=(tag,), values=(
                row_get(r, "barcode") or "",
                row_get(r, "name") or "",
                money(float(row_get(r, "sell_price") or 0)),
                int(row_get(r, "stock_qty") or 0),
            ))

    def add_selected_product(self):
        if not hasattr(self, "picker_tree"):
            return
        sel = self.picker_tree.selection()
        if not sel:
            messagebox.showinfo("Select product", "Select a product first.")
            return
        vals = self.picker_tree.item(sel[0], "values")
        if not vals:
            return
        barcode = str(vals[0]).strip()
        if barcode:
            barcode13 = to_ean13(barcode)
            self.on_product_added(barcode13)

    def refresh_cart(self):
        try:
            self._apply_auto_offers_to_cart()
        except Exception:
            pass

        for i in self.tree.get_children():
            self.tree.delete(i)

        total = 0.0
        for idx, line in enumerate(self.cart_lines):
            if "discount_pct" not in line:
                line["discount_pct"] = 0.0
            if "original_price" not in line:
                line["original_price"] = float(line.get("price") or 0.0)
            self._recalc_line(line)

            total += float(line.get("line_total") or 0)

            # Show sale discounts as Sxx% so cashier can tell it is seasonal-sale driven
            try:
                disc_val = float(line.get("discount_pct") or 0)
            except Exception:
                disc_val = 0.0
            disc_text = f"{disc_val:.0f}%"
            if bool(line.get("sale_applied", False)) and disc_val > 0:
                disc_text = f"S{disc_val:.0f}%"
            if bool(line.get("bundle_offer_applied", False)):
                try:
                    offer_qty = int(line.get("bundle_offer_qty") or 0)
                    offer_price = float(line.get("bundle_offer_price") or 0.0)
                    disc_text = f"{offer_qty}/{money(offer_price)}"
                except Exception:
                    disc_text = "Offer"

            qty = max(1, int(line.get("qty") or 1))
            original_unit = float(line.get("original_price", line.get("price", 0)) or 0.0)
            final_unit = float(line.get("line_total") or 0.0) / qty
            unit_discount = max(0.0, original_unit - final_unit)
            is_discounted = unit_discount > 0.005
            item_text = str(line.get("name", ""))
            if is_discounted:
                effective_pct = (unit_discount / original_unit * 100.0) if original_unit > 0 else 0.0
                item_text += f"  [WAS {money(original_unit)} → {money(final_unit)}]"
                disc_text = f"-{money(unit_discount)} ({effective_pct:.0f}%)"
            tag = "odd" if idx % 2 else "even"
            tags = (tag, "discounted") if is_discounted else (tag,)
            self.tree.insert("", tk.END, tags=tags, values=(
                item_text,
                money(final_unit),
                int(line.get("qty") or 0),
                disc_text,
                money(float(line.get("line_total") or 0)),
            ))

        self.total_label.configure(text=f"Total: {money(total)}")
        try:
            item_count = sum(int(l.get("qty") or 0) for l in self.cart_lines)
            label = "item" if item_count == 1 else "items"
            self.cart_count_lbl.configure(text=f"{item_count} {label}")
        except Exception:
            pass
        try:
            self.checkout_amount_lbl.configure(text=money(total))
        except Exception:
            pass
        try:
            pending_credit = max(0.0, float(getattr(self, "exchange_credit", 0.0) or 0.0))
            if pending_credit > 0.005:
                bon_codes = list(getattr(self, "exchange_bon_codes", []) or [])
                if bon_codes:
                    shown = ", ".join(str(c) for c in bon_codes[:2])
                    if len(bon_codes) > 2:
                        shown += f" +{len(bon_codes) - 2}"
                    self.pending_credit_lbl.configure(text=f"Pending bon credit: {money(pending_credit)} ({shown})")
                else:
                    self.pending_credit_lbl.configure(text=f"Pending exchange credit: {money(pending_credit)}")
                try:
                    self.clear_credit_btn.grid()
                except Exception:
                    pass
            else:
                self.pending_credit_lbl.configure(text="")
                try:
                    self.clear_credit_btn.grid_remove()
                except Exception:
                    pass
        except Exception:
            pass
        if getattr(self, "_stage", "cart") == "cart":
            self._primary_enabled_requested = True
        self._apply_primary_state()
        self._save_cashier_recovery()
        self._update_held_order_button()

    def _bundle_offer_total(self, price: float, qty: int, offer: dict) -> float | None:
        try:
            offer_qty = int((offer or {}).get("qty") or 0)
            offer_price = float((offer or {}).get("price") or 0.0)
        except Exception:
            return None
        if offer_qty < 2 or offer_price <= 0 or qty < offer_qty:
            return None
        return whole_money_round_up(int(qty) * (offer_price / offer_qty))

    def _apply_auto_offers_to_cart(self):
        """Apply/clear configured item offers on current cart lines.

        Rules:
        - Only affects catalog items (product_id set). Quick/manual items are never auto-discounted.
        - If a line was manually edited (manual_override=True), automatic offers will NOT overwrite it.
        - Seasonal percent discounts and bundle offers are compared; the best customer price wins.
        """
        try:
            seasonal_enabled = bool(get_seasonal_sale_enabled())
        except Exception:
            seasonal_enabled = False
        try:
            sale_map = get_seasonal_sales_map()
        except Exception:
            sale_map = {}
        try:
            bundle_enabled = bool(get_bundle_offers_enabled())
        except Exception:
            bundle_enabled = True
        try:
            bundle_map = get_bundle_offers_map()
        except Exception:
            bundle_map = {}

        for line in self.cart_lines:
            # Skip quick/manual lines
            if not line.get("product_id"):
                continue
            if bool(line.get("manual_override", False)):
                continue

            bc = str(line.get("barcode") or "").strip()
            try:
                price = float(line.get("price") or 0.0)
                qty = int(line.get("qty") or 0)
            except Exception:
                price = 0.0
                qty = 0

            pct = 0.0
            if seasonal_enabled and bc:
                try:
                    pct = float(sale_map.get(bc) or 0.0)
                except Exception:
                    pct = 0.0
            pct = max(0.0, min(100.0, pct))

            pct_total = whole_money_round_up(price * qty * (1.0 - pct / 100.0)) if pct > 0 else None
            bundle_offer = bundle_map.get(bc) if (bundle_enabled and bc) else None
            bundle_total = self._bundle_offer_total(price, qty, bundle_offer) if bundle_offer else None
            regular_total = round(price * qty, 2)

            use_bundle = (
                bundle_total is not None
                and bundle_total < regular_total - 0.005
                and (pct_total is None or bundle_total <= pct_total + 0.005)
            )

            use_pct = (
                pct_total is not None
                and pct_total < regular_total - 0.005
                and not use_bundle
            )

            line["sale_applied"] = bool(use_pct)
            line["sale_pct"] = float(pct if use_pct else 0.0)
            line["discount_pct"] = float(pct if use_pct else 0.0)

            line["bundle_offer_applied"] = bool(use_bundle)
            if use_bundle:
                line["bundle_offer_qty"] = int(bundle_offer.get("qty") or 0)
                line["bundle_offer_price"] = float(bundle_offer.get("price") or 0.0)
            else:
                line["bundle_offer_qty"] = 0
                line["bundle_offer_price"] = 0.0

            self._recalc_line(line)

    def _apply_seasonal_sale_to_cart(self):
        """Backward-compatible wrapper for older call sites."""
        self._apply_auto_offers_to_cart()

    def _toggle_seasonal_sale(self):
        try:
            enabled = bool(self.seasonal_sale_var.get())
        except Exception:
            enabled = False
        try:
            set_seasonal_sale_enabled(enabled)
        except Exception:
            pass
        self._apply_seasonal_sale_to_cart()
        self.refresh_cart()
        self.refresh_picker()

    def _apply_spin_wheel_prize(self, prize: dict) -> str:
        kind = str(prize.get("type") or "none").strip().lower()
        label = str(prize.get("label") or "Prize").strip() or "Prize"
        if kind == "discount":
            pct = max(0.0, min(100.0, float(prize.get("value") or 0.0)))
            self._wheel_discount_pct = max(float(self._wheel_discount_pct or 0.0), pct)
            self._wheel_receipt_label = f"{pct:g}% discount"
            return f"You won {pct:g}% off!"
        if kind == "free_item":
            barcode = str(prize.get("barcode") or "").strip()
            row = None
            for candidate in barcode_candidates(barcode):
                row = find_product_by_barcode(candidate)
                if row:
                    break
            if not row:
                return f"{label}\nThe configured free item could not be found. Please ask a manager."
            if int(row_get(row, "stock_qty") or 0) <= 0:
                return f"{label}\nThe configured free item is out of stock. Please ask a manager."
            line = {
                "product_id": int(row["id"]),
                "barcode": str(row_get(row, "barcode") or barcode).strip(),
                "name": f"{row['name']} (FREE PRIZE)",
                "price": float(row_get(row, "sell_price") or 0.0),
                "qty": 1,
                "discount_pct": 100.0,
                "manual_override": True,
                "wheel_prize": True,
            }
            self._recalc_line(line)
            self.cart_lines.append(line)
            self._wheel_receipt_label = f"Free item - {row['name']}"
            self.refresh_cart()
            return f"You won a free {row['name']}!"
        self._wheel_receipt_label = ""
        return "No prize this time. Thanks for playing!"

    def open_spin_wheel(self):
        if not self.cart_lines:
            messagebox.showinfo("Spin Wheel", "Add at least one item to the cart before spinning.")
            return
        if self._wheel_prize_claimed:
            messagebox.showinfo("Spin Wheel", "A wheel prize has already been claimed for this cart.")
            return
        prizes = []
        for prize in get_spin_wheel_prizes():
            if not bool(prize.get("enabled", True)) or float(prize.get("weight") or 0) <= 0:
                continue
            if str(prize.get("type") or "").strip().lower() == "free_item":
                product = None
                for candidate in barcode_candidates(str(prize.get("barcode") or "")):
                    product = find_product_by_barcode(candidate)
                    if product:
                        break
                if not product or int(row_get(product, "stock_qty") or 0) <= 0:
                    continue
            prizes.append(prize)
        if not prizes:
            messagebox.showinfo(
                "Spin Wheel",
                "No wheel prizes are currently available. Check Offers > Manage Spin Wheel."
            )
            return

        win = tk.Toplevel(self)
        win.title("Spin Wheel")
        try:
            win.state("zoomed")
        except Exception:
            win.geometry("1000x900")
        win.minsize(760, 720)
        win.configure(bg=UI.CONTENT_BG)
        win.grab_set()

        card = Card(win, padx=22, pady=18)
        card.pack(fill="both", expand=True, padx=14, pady=14)
        HeaderBar(card.inner, "Spin to Win", "Tap the wheel or press Spin to play.").pack(fill="x")

        screen_w = max(760, int(win.winfo_screenwidth() or 1000))
        screen_h = max(720, int(win.winfo_screenheight() or 900))
        canvas_size = max(580, min(820, screen_w - 180, screen_h - 230))
        canvas = tk.Canvas(card.inner, width=canvas_size, height=canvas_size, bg="#FFF7ED",
                           highlightthickness=0, cursor="hand2")
        canvas.pack(pady=(10, 4))
        status_var = tk.StringVar(value="Tap the wheel or press Spin")
        tk.Label(card.inner, textvariable=status_var, font=UI.FONT_LG, bg=UI.CARD, fg=UI.TEXT).pack(pady=(2, 8))
        tk.Label(card.inner, text="Prizes are selected randomly. Odds vary by prize.", font=UI.FONT_SM,
                 bg=UI.CARD, fg=UI.MUTED).pack(pady=(0, 6))
        colors = ("#E11D48", "#F59E0B", "#0EA5E9", "#10B981", "#7C3AED", "#F97316", "#EC4899", "#14B8A6")
        rotation = {"value": 0.0}
        light_phase = {"value": 0}
        confetti = []
        reveal = {"kind": "", "title": "", "subtitle": ""}
        size = canvas_size - 130
        center = canvas_size / 2
        left = (canvas_size - size) / 2
        top = (canvas_size - size) / 2 + 14
        weights = [float(p.get("weight") or 0) for p in prizes]
        total_weight = sum(weights)
        extents = [(360.0 * weight / total_weight) for weight in weights]
        prize_starts = []
        running = 0.0
        for wedge_extent in extents:
            prize_starts.append(running)
            running += wedge_extent

        def draw_wheel():
            canvas.delete("all")
            canvas.create_oval(left - 31, top - 31, left + size + 31, top + size + 31,
                               fill="#FB923C", outline="#F59E0B", width=5)
            canvas.create_oval(left - 20, top - 20, left + size + 20, top + size + 20,
                               fill="#FFF7ED", outline="#FDE68A", width=3)
            for i, prize in enumerate(prizes):
                start = rotation["value"] + prize_starts[i]
                wedge_extent = extents[i]
                canvas.create_arc(left, top, left + size, top + size, start=start, extent=wedge_extent,
                                  fill=colors[i % len(colors)], outline="#FFFFFF", width=3)
                canvas.create_arc(left + 12, top + 12, left + size - 12, top + size - 12, start=start + 2,
                                  extent=max(0.0, wedge_extent - 4), style="arc", outline="#FFFFFF", width=2)
                angle = math.radians(start + wedge_extent / 2)
                radius = size * 0.32
                x = center + math.cos(angle) * radius
                y = top + size / 2 - math.sin(angle) * radius
                if wedge_extent >= 12:
                    canvas.create_text(x, y, text=str(prize.get("label") or "Prize"), fill="#FFFFFF",
                                       font=("Segoe UI", 14, "bold"), width=max(90, int(size * 0.20)),
                                       justify="center")
            for angle in range(0, 360, 15):
                rad = math.radians(angle)
                dot_radius = size / 2 + 25
                x = center + math.cos(rad) * dot_radius
                y = top + size / 2 - math.sin(rad) * dot_radius
                lit = ((angle // 15) + light_phase["value"]) % 2 == 0
                bulb = "#FEF3C7" if lit else "#F59E0B"
                canvas.create_oval(x - 6, y - 6, x + 6, y + 6, fill=bulb, outline="#FFFFFF", width=1)
            hub = max(100, int(size * 0.18))
            canvas.create_oval(center - hub / 2 - 12, top + size / 2 - hub / 2 - 12,
                               center + hub / 2 + 12, top + size / 2 + hub / 2 + 12,
                               fill="#F59E0B", outline="#FDE68A", width=4)
            canvas.create_oval(center - hub / 2, top + size / 2 - hub / 2,
                               center + hub / 2, top + size / 2 + hub / 2,
                               fill="#FFFFFF", outline="#FFF7ED", width=4)
            canvas.create_text(center, top + size / 2, text="SPIN", fill="#BE123C", font=("Segoe UI", 21, "bold"))
            canvas.create_polygon(center - 28, top - 14, center + 28, top - 14, center, top + 48,
                                  fill="#E11D48", outline="#FFFFFF", width=3)
            for particle in confetti:
                canvas.create_rectangle(
                    particle["x"], particle["y"], particle["x"] + particle["size"],
                    particle["y"] + particle["size"] * 0.6, fill=particle["color"], outline=""
                )
            if reveal["kind"]:
                box_w = min(520, int(canvas_size * 0.72))
                box_h = 230
                box_left = center - box_w / 2
                box_top = top + size / 2 - box_h / 2
                canvas.create_rectangle(box_left, box_top, box_left + box_w, box_top + box_h,
                                        fill="#FFFFFF", outline=("#FDBA74" if reveal["kind"] == "lose" else "#F59E0B"),
                                        width=6)
                face_y = box_top + 62
                if reveal["kind"] == "lose":
                    canvas.create_oval(center - 31, face_y - 31, center + 31, face_y + 31,
                                       fill="#FDE68A", outline="#F59E0B", width=3)
                    canvas.create_oval(center - 15, face_y - 8, center - 8, face_y - 1, fill="#78350F", outline="")
                    canvas.create_oval(center + 8, face_y - 8, center + 15, face_y - 1, fill="#78350F", outline="")
                    canvas.create_arc(center - 15, face_y + 8, center + 15, face_y + 28,
                                      start=20, extent=140, style="arc", outline="#78350F", width=3)
                else:
                    canvas.create_oval(center - 34, face_y - 34, center + 34, face_y + 34,
                                       fill="#FDE047", outline="#F59E0B", width=4)
                    canvas.create_text(center, face_y, text="WIN", fill="#BE123C", font=("Segoe UI", 15, "bold"))
                canvas.create_text(center, box_top + 135, text=reveal["title"],
                                   fill=("#C2410C" if reveal["kind"] == "lose" else "#BE123C"),
                                   font=("Segoe UI", 24, "bold"))
                canvas.create_text(center, box_top + 180, text=reveal["subtitle"],
                                   fill=("#9A3412" if reveal["kind"] == "lose" else "#7C2D12"),
                                   font=("Segoe UI", 15, "bold"))

        def spin():
            if self._wheel_spinning:
                return
            self._wheel_spinning = True
            close_btn.configure(state="disabled")
            spin_btn.configure(state="disabled")
            status_var.set("Spinning...")
            selected = random.choices(prizes, weights=[float(p.get("weight") or 0) for p in prizes], k=1)[0]
            idx = prizes.index(selected)
            selected_center = prize_starts[idx] + extents[idx] / 2
            target = rotation["value"] + (360.0 * random.randint(6, 8)) + ((90.0 - selected_center - rotation["value"]) % 360.0)
            start_rotation = rotation["value"]
            frames = 110

            def animate(frame=0):
                progress = min(1.0, frame / frames)
                eased = 1.0 - pow(1.0 - progress, 4)
                rotation["value"] = start_rotation + (target - start_rotation) * eased
                light_phase["value"] = frame // 4
                draw_wheel()
                if frame < frames:
                    win.after(18, lambda: animate(frame + 1))
                    return
                self._wheel_spinning = False
                self._wheel_prize_claimed = True
                result = self._apply_spin_wheel_prize(selected)
                status_var.set(result.replace("\n", " - "))
                selected_kind = str(selected.get("type") or "none").strip().lower()
                is_loser = selected_kind == "none"
                unavailable = selected_kind == "free_item" and "\n" in result
                if is_loser:
                    reveal.update({"kind": "lose", "title": "So close!", "subtitle": "Maybe next time"})
                elif unavailable:
                    reveal.update({"kind": "lose", "title": "Manager help needed", "subtitle": "Prize unavailable"})
                else:
                    reveal.update({"kind": "win", "title": "You won!", "subtitle": result})
                confetti.extend({
                    "x": random.randint(20, max(21, canvas_size - 20)),
                    "y": random.randint(-180, -10),
                    "size": random.randint(5, 10) if is_loser else random.randint(6, 13),
                    "speed": random.randint(5, 11) if is_loser else random.randint(8, 17),
                    "drift": random.randint(-3, 3),
                    "color": random.choice(("#FDBA74", "#FDE68A", "#FCA5A5", "#BFDBFE") if is_loser else
                                           ("#FDE047", "#FB7185", "#38BDF8", "#34D399", "#C084FC", "#FB923C")),
                } for _ in range(52 if is_loser else 95))

                def celebrate(tick=0):
                    light_phase["value"] += 1
                    for particle in confetti:
                        particle["x"] += particle["drift"]
                        particle["y"] += particle["speed"]
                    draw_wheel()
                    if tick < 32:
                        win.after(24, lambda: celebrate(tick + 1))
                        return
                    win.after(1500, win.destroy)

                celebrate()

            animate()

        draw_wheel()
        buttons = tk.Frame(card.inner, bg=UI.CARD)
        buttons.pack(fill="x")
        close_btn = GhostButton(buttons, "Close", win.destroy)
        close_btn.pack(side="right")
        spin_btn = PrimaryButton(buttons, "Spin", spin)
        spin_btn.pack(side="right", padx=(0, 10))
        win.protocol("WM_DELETE_WINDOW", lambda: None if self._wheel_spinning else win.destroy())
        canvas.bind("<Button-1>", lambda _event: spin(), add="+")
        win.bind("<space>", lambda _event: spin(), add="+")

    def clear_cart(self):
        self.cart_lines = []
        self._wheel_prize_claimed = False
        self._wheel_discount_pct = 0.0
        self._wheel_receipt_label = ""
        self.refresh_cart()
        self.refresh_picker()
        self.scan_entry.focus()

    def _held_order_snapshot(self, label: str) -> dict:
        total = 0.0
        try:
            for line in self.cart_lines:
                total += float(line.get("line_total") or 0.0)
        except Exception:
            total = 0.0
        return {
            "id": self._held_order_seq,
            "label": (label or "").strip() or f"Order {self._held_order_seq}",
            "created_at": datetime.now().strftime("%H:%M"),
            "lines": copy.deepcopy(self.cart_lines),
            "exchange_credit": float(getattr(self, "exchange_credit", 0.0) or 0.0),
            "exchange_origin_sale_ids": list(getattr(self, "exchange_origin_sale_ids", []) or []),
            "exchange_return_ids": list(getattr(self, "exchange_return_ids", []) or []),
            "exchange_bon_codes": list(getattr(self, "exchange_bon_codes", []) or []),
            "exchange_original_sale_id": getattr(self, "exchange_original_sale_id", None),
            "wheel_prize_claimed": bool(getattr(self, "_wheel_prize_claimed", False)),
            "wheel_discount_pct": float(getattr(self, "_wheel_discount_pct", 0.0) or 0.0),
            "wheel_receipt_label": str(getattr(self, "_wheel_receipt_label", "") or ""),
            "total": total,
        }

    def hold_current_order(self):
        if not self.cart_lines:
            try:
                self.scan_entry.focus()
            except Exception:
                pass
            return

        default_label = f"Order {self._held_order_seq}"
        label = simpledialog.askstring("Hold order", "Name this held order:", initialvalue=default_label, parent=self)
        if label is None:
            return

        snap = self._held_order_snapshot(label)
        self.held_orders.append(snap)
        self._held_order_seq += 1
        self.cart_lines = []
        self._wheel_prize_claimed = False
        self._wheel_discount_pct = 0.0
        self._wheel_receipt_label = ""
        self.set_exchange_credit_balance(0.0)
        try:
            self._hide_right_panels()
        except Exception:
            pass
        self.refresh_cart()
        self.refresh_picker()
        self._apply_primary_state()
        try:
            self.scan_entry.focus()
        except Exception:
            pass

    def open_held_orders(self):
        win = tk.Toplevel(self)
        win.title("Held Orders")
        win.geometry("620x420")
        win.minsize(540, 340)
        win.configure(bg=UI.CONTENT_BG)
        try:
            win.transient(self.winfo_toplevel())
        except Exception:
            pass
        win.grab_set()

        outer = Card(win, padx=14, pady=14)
        outer.pack(fill="both", expand=True, padx=12, pady=12)
        HeaderBar(outer.inner, "Held Orders", "Resume a paused cart. Held orders are saved across restarts.").pack(fill="x")

        cols = ("label", "items", "total", "time")
        tree = ttk.Treeview(outer.inner, columns=cols, show="headings", height=10)
        tree.heading("label", text="Order")
        tree.heading("items", text="Items")
        tree.heading("total", text="Total")
        tree.heading("time", text="Held")
        tree.column("label", width=260, anchor="w")
        tree.column("items", width=70, anchor="center")
        tree.column("total", width=100, anchor="e")
        tree.column("time", width=80, anchor="center")
        tree.pack(fill="both", expand=True, pady=(12, 10))

        def refresh():
            tree.delete(*tree.get_children())
            for order in self.held_orders:
                oid = str(order.get("id"))
                item_count = sum(int(l.get("qty") or 0) for l in (order.get("lines") or []))
                tree.insert("", tk.END, iid=oid, values=(
                    order.get("label") or oid,
                    item_count,
                    money(order.get("total") or 0.0),
                    order.get("created_at") or "",
                ))

        def selected_order():
            sel = tree.selection()
            if not sel:
                return None
            oid = int(sel[0])
            for order in self.held_orders:
                if int(order.get("id") or 0) == oid:
                    return order
            return None

        def resume():
            order = selected_order()
            if not order:
                return
            if self.cart_lines and not messagebox.askyesno(
                    "Resume held order",
                    "Replace the current cart with this held order?",
                    parent=win):
                return
            self.cart_lines = copy.deepcopy(order.get("lines") or [])
            self.exchange_credit = float(order.get("exchange_credit") or 0.0)
            self.exchange_origin_sale_ids = list(order.get("exchange_origin_sale_ids") or [])
            self.exchange_return_ids = list(order.get("exchange_return_ids") or [])
            self.exchange_bon_codes = list(order.get("exchange_bon_codes") or [])
            self.exchange_original_sale_id = order.get("exchange_original_sale_id")
            # Keep resumed store credit crash-safe as soon as it becomes active.
            self._persist_exchange_credit()
            self._wheel_prize_claimed = bool(order.get("wheel_prize_claimed", False))
            self._wheel_discount_pct = float(order.get("wheel_discount_pct") or 0.0)
            self._wheel_receipt_label = str(order.get("wheel_receipt_label") or "")
            try:
                self.held_orders.remove(order)
            except Exception:
                pass
            try:
                self._hide_right_panels()
            except Exception:
                pass
            self.refresh_cart()
            self.refresh_picker()
            self._apply_primary_state()
            win.destroy()
            try:
                self.scan_entry.focus()
            except Exception:
                pass

        def delete():
            order = selected_order()
            if not order:
                return
            if not messagebox.askyesno("Delete held order", "Remove this held order?", parent=win):
                return
            try:
                self.held_orders.remove(order)
            except Exception:
                pass
            self._save_cashier_recovery()
            self._update_held_order_button()
            refresh()

        btns = tk.Frame(outer.inner, bg=UI.CARD)
        btns.pack(fill="x")
        GhostButton(btns, "Close", win.destroy).pack(side="right")
        GhostButton(btns, "Delete", delete).pack(side="right", padx=(0, 8))
        PrimaryButton(btns, "Resume", resume).pack(side="right", padx=(0, 8))
        tree.bind("<Double-1>", lambda _e: resume())

        refresh()

    # ---------------- Keyboard-first helpers ----------------

    def _selected_cart_index(self):
        """Return selected cart line index, or last line index if nothing selected."""
        try:
            sel = self.tree.selection()
            if sel:
                return int(self.tree.index(sel[0]))
        except Exception:
            pass
        try:
            if len(self.cart_lines) > 0:
                return len(self.cart_lines) - 1
        except Exception:
            pass
        return None

    def change_selected_qty(self, delta: int):
        """+ increases qty, - decreases qty, reaching 0 removes the line."""
        idx = self._selected_cart_index()
        if idx is None:
            return
        try:
            line = self.cart_lines[idx]
        except Exception:
            return

        try:
            q = int(line.get("qty") or 0)
        except Exception:
            q = 0

        self._remember_cart_state()
        q2 = q + int(delta)
        if q2 <= 0:
            try:
                self.cart_lines.pop(idx)
            except Exception:
                pass
        else:
            line["qty"] = q2

        try:
            self.refresh_cart()
        except Exception:
            pass

        # Keep selection stable
        try:
            children = self.tree.get_children("")
            if children:
                pick = children[min(max(idx, 0), len(children) - 1)]
                self.tree.selection_set(pick)
                self.tree.focus(pick)
        except Exception:
            pass

    def remove_selected_line(self):
        idx = self._selected_cart_index()
        if idx is None:
            return
        self._remember_cart_state()
        try:
            self.cart_lines.pop(idx)
        except Exception:
            return
        try:
            self.refresh_cart()
        except Exception:
            pass

    def invoke_primary_action(self):
        """Enter triggers the main button (Checkout/Proceed/Complete) when available."""
        try:
            # Only invoke if Tk considers it enabled
            if hasattr(self, "primary_action_btn") and self.primary_action_btn:
                state = str(self.primary_action_btn.cget("state"))
                if state != "disabled":
                    self.primary_action_btn.invoke()
        except Exception:
            pass

    def cancel_sale(self):
        # ESC should safely back out without breaking the primary button flow.
        # If cart is empty, just restore focus and ensure the primary action is in a sane state.
        if len(self.cart_lines) == 0:
            try:
                self.scan_entry.focus()
            except Exception:
                pass
            try:
                # Keep the fixed bottom-right action consistent when nothing is in cart.
                self._set_stage("cart", "Checkout", self.checkout, enabled=False)
            except Exception:
                pass
            return

        if not messagebox.askyesno("Cancel sale", "Cancel the current sale and clear the cart?"):
            return

        # Clear everything and return to the cart stage (not review).
        self.clear_cart()
        try:
            self._hide_right_panels()
        except Exception:
            pass
        try:
            self._set_stage("cart", "Checkout", self.checkout, enabled=False)
        except Exception:
            pass

    def _hide_right_panels(self):
        try:
            if self.review_frame is not None:
                self.review_frame.destroy()
        except Exception:
            pass
        self.review_frame = None

        try:
            if self.cash_frame is not None:
                self.cash_frame.destroy()
        except Exception:
            pass
        self.cash_frame = None

        try:
            self.picker_frame.pack(fill="both", expand=True)
        except Exception:
            pass

        try:
            self.right_title_lbl.configure(text="Find Product")
        except Exception:
            pass

        try:
            self.refresh_picker()
        except Exception:
            pass

        try:
            self.scan_entry.focus()
        except Exception:
            pass
        # Reset the fixed primary action back to checkout in the SAME bottom-right spot
        try:
            self._set_stage("cart", "Checkout", self.checkout, enabled=(len(self.cart_lines) > 0))
        except Exception:
            pass

    def _show_review_panel(self):
        if not self.cart_lines:
            messagebox.showinfo("Empty", "Cart is empty.")
            return

        for l in self.cart_lines:
            self._recalc_line(l)

        try:
            self.right_title_lbl.configure(text="Review Sale")
        except Exception:
            pass

        try:
            self.picker_frame.pack_forget()
        except Exception:
            pass

        try:
            if self.cash_frame is not None:
                self.cash_frame.destroy()
        except Exception:
            pass
        self.cash_frame = None

        try:
            if self.review_frame is not None:
                self.review_frame.destroy()
        except Exception:
            pass

        self.review_frame = tk.Frame(self.right_stack, bg=UI.CARD)
        self.review_frame.pack(fill="both", expand=True)
        # --- Make Review Sale panel vertically scrollable (so buttons are reachable on small screens) ---
        self._review_canvas = tk.Canvas(self.review_frame, bg=UI.CARD, highlightthickness=0)
        self._review_vsb = ttk.Scrollbar(self.review_frame, orient="vertical", command=self._review_canvas.yview)
        self._review_canvas.configure(yscrollcommand=self._review_vsb.set)

        self._review_canvas.pack(side="left", fill="both", expand=True)
        self._review_vsb.pack(side="right", fill="y")

        self.review_inner = tk.Frame(self._review_canvas, bg=UI.CARD)
        self._review_window = self._review_canvas.create_window((0, 0), window=self.review_inner, anchor="nw")

        def _review_reflow(_evt=None):
            try:
                self._review_canvas.configure(scrollregion=self._review_canvas.bbox("all"))
            except Exception:
                pass
            # Keep inner width equal to the visible canvas width
            try:
                self._review_canvas.itemconfigure(self._review_window, width=self._review_canvas.winfo_width())
            except Exception:
                pass

        self.review_inner.bind("<Configure>", _review_reflow)
        self._review_canvas.bind("<Configure>", _review_reflow)

        # Mouse wheel scroll only when cursor is over the review panel
        def _review_on_mousewheel(e):
            try:
                # Windows / macOS (Tk uses delta)
                if hasattr(e, "delta") and e.delta:
                    delta = int(-1 * (e.delta / 120))
                # Linux
                elif getattr(e, "num", None) == 4:
                    delta = -3
                elif getattr(e, "num", None) == 5:
                    delta = 3
                else:
                    delta = 0

                if delta:
                    self._review_canvas.yview_scroll(delta, "units")
                return "break"
            except Exception:
                return "break"

        # Bind only to the review widgets (do NOT use bind_all/unbind_all, it breaks global scrolling)
        for _w in (self._review_canvas, self.review_inner):
            try:
                _w.bind("<MouseWheel>", _review_on_mousewheel, add="+")
                _w.bind("<Button-4>", _review_on_mousewheel, add="+")
                _w.bind("<Button-5>", _review_on_mousewheel, add="+")
            except Exception:
                pass

        rf = self.review_inner


        top_actions = tk.Frame(rf, bg=UI.CARD)
        top_actions.pack(fill="x", pady=(2, 0))

        review_table = tk.Frame(rf, bg=UI.CARD, highlightthickness=1, highlightbackground=UI.BORDER)
        review_table.pack(fill="both", expand=True, pady=(12, 12))

        cols = ("item", "price", "qty", "disc", "total")
        self.review_tree = ttk.Treeview(review_table, columns=cols, show="headings", height=8)
        self.review_tree.heading("item", text="Item")
        self.review_tree.heading("price", text="$")
        self.review_tree.heading("qty", text="Qty")
        self.review_tree.heading("disc", text="Disc")
        self.review_tree.heading("total", text="Total")
        self.review_tree.column("item", width=160, minwidth=110, anchor="w", stretch=True)
        self.review_tree.column("price", width=68, minwidth=56, anchor="e", stretch=False)
        self.review_tree.column("qty", width=46, minwidth=40, anchor="center", stretch=False)
        self.review_tree.column("disc", width=54, minwidth=46, anchor="center", stretch=False)
        self.review_tree.column("total", width=76, minwidth=62, anchor="e", stretch=False)
        self.review_tree.tag_configure("odd", background=UI.SURFACE)
        self.review_tree.tag_configure("even", background=UI.CARD)
        self.review_tree.pack(fill="both", expand=True)
        self.review_tree.bind("<Double-1>", lambda e: set_line_disc_selected())

        self.order_disc_var = tk.DoubleVar(value=float(getattr(self, "_wheel_discount_pct", 0.0) or 0.0))

        # Total discount mode:
        # - PCT: percentage off the (already line-discounted) subtotal
        # - AMT: fixed amount removed from the (already line-discounted) subtotal
        self.order_disc_mode_var = tk.StringVar(value="PCT")
        self.order_disc_amt_var = tk.DoubleVar(value=0.0)
        self.review_total_var = tk.StringVar(value="Total: $0.00")
        tk.Label(rf, textvariable=self.review_total_var, font=("Segoe UI", 18, "bold"), bg=UI.CARD,
                 fg=UI.TEXT).pack(anchor="w", pady=(0, 8))

        self.review_credit_var = tk.StringVar(value="")
        self.review_due_var = tk.StringVar(value="")
        tk.Label(rf, textvariable=self.review_credit_var, font=UI.FONT_MD, bg=UI.CARD, fg=UI.MUTED).pack(
            anchor="w", pady=(0, 2))
        tk.Label(rf, textvariable=self.review_due_var, font=("Segoe UI", 14, "bold"), bg=UI.CARD,
                 fg=UI.TEXT).pack(anchor="w", pady=(0, 10))

        def compute_sale_lines(materialize: bool = False):
            """Build sale lines from current cart.

            - materialize=False: preview only (do NOT create products for quick items).
            - materialize=True: ensure quick/manual items are saved as real products (product_id required).
            """
            sale_lines = []
            for l in self.cart_lines:
                if materialize:
                    # Ensure quick/manual items are backed by a real product entry
                    try:
                        self._materialize_quick_line(l)
                    except Exception as e:
                        messagebox.showerror(
                            "Error",
                            f"Could not save quick item for checkout: {e}\n\n"
                            "Tip: If you are in JOIN mode, make sure you are connected to the host."
                        )
                        return None
                    if not l.get("product_id"):
                        messagebox.showerror(
                            "Error",
                            "Could not save quick item for checkout (missing product id).\n"
                            "Try removing the quick item and adding it again."
                        )
                        return None

                # Add a clear sale label on the RECEIPT (without changing product catalog name)
                nm = l.get("name")
                try:
                    if bool(l.get("sale_applied", False)) and float(l.get("sale_pct") or 0) > 0:
                        nm = f"{nm} (SALE -{float(l.get('sale_pct') or 0):.0f}%)"
                except Exception:
                    pass
                try:
                    if bool(l.get("bundle_offer_applied", False)):
                        oq = int(l.get("bundle_offer_qty") or 0)
                        op = float(l.get("bundle_offer_price") or 0.0)
                        if oq >= 2 and op > 0:
                            nm = f"{nm} (OFFER {oq}/{money(op)})"
                except Exception:
                    pass

                sale_lines.append({
                    "product_id": l.get("product_id"),
                    "barcode": str(l.get("barcode") or ""),
                    "name": nm,
                    "price": float(l.get("price") or 0),
                    "original_price": float(l.get("original_price", l.get("price", 0)) or 0),
                    "qty": int(l.get("qty") or 0),
                    "discount_pct": float(l.get("discount_pct") or 0),
                    "line_total": float(l.get("line_total") or 0),
                    "is_quick": bool(l.get("is_quick", False)),
                })

            mode = "PCT"
            try:
                mode = str(self.order_disc_mode_var.get() if hasattr(self, "order_disc_mode_var") else "PCT").upper()
            except Exception:
                mode = "PCT"

            pct = max(0.0, min(100.0, float(self.order_disc_var.get() or 0)))
            amt = 0.0
            try:
                amt = float(self.order_disc_amt_var.get() if hasattr(self, "order_disc_amt_var") else 0.0) or 0.0
            except Exception:
                amt = 0.0
            if amt < 0:
                amt = 0.0

            # Apply AFTER line discounts
            base_totals = [float(x.get("line_total") or 0) for x in sale_lines]
            base_sum = sum(base_totals)

            apply_factor = None
            if base_sum > 0:
                if mode == "AMT":
                    if amt > 0.0:
                        if amt > base_sum:
                            amt = base_sum
                        apply_factor = (base_sum - amt) / base_sum
                else:
                    # default PCT
                    if pct > 0.0:
                        apply_factor = 1.0 - pct / 100.0

            if apply_factor is not None:
                target_total = whole_money_round_up(base_sum * apply_factor)
                new_sum = 0.0
                for i, ln in enumerate(sale_lines):
                    if i == len(sale_lines) - 1:
                        ln["line_total"] = round(max(0.0, target_total - new_sum), 2)
                    else:
                        ln["line_total"] = round(target_total * (base_totals[i] / base_sum), 2)
                        new_sum += ln["line_total"]
            return sale_lines

        def refresh_review():
            try:
                self.review_tree.delete(*self.review_tree.get_children())
            except Exception:
                pass

            sale_lines = compute_sale_lines(materialize=False)
            if not sale_lines:
                return
            total = 0.0
            for idx, ln in enumerate(sale_lines):
                total += float(ln.get("line_total") or 0)
                disc_text = f"{float(ln.get('discount_pct') or 0):.0f}%"
                try:
                    name_text = str(ln.get("name", "") or "")
                    m = re.search(r"\(OFFER\s+([^)]+)\)", name_text)
                    if m:
                        disc_text = m.group(1)
                except Exception:
                    pass
                tag = "odd" if idx % 2 else "even"
                self.review_tree.insert("", tk.END, tags=(tag,), values=(
                    ln.get("name", ""),
                    money(float(ln.get("price") or 0)),
                    int(ln.get("qty") or 0),
                    disc_text,
                    money(float(ln.get("line_total") or 0)),
                ))

            self.review_total_var.set(f"Subtotal: {money(total)}")

            try:
                self._sync_bon_credit_balance()
            except Exception:
                pass
            credit = float(getattr(self, "exchange_credit", 0.0) or 0.0)
            if credit < 0:
                credit = 0.0
            applied = min(credit, total)
            if applied > 0:
                label = "Bon credit" if getattr(self, "exchange_bon_codes", []) else "Exchange credit"
                self.review_credit_var.set(f"{label}: -{money(applied)}")
            else:
                self.review_credit_var.set("")
            self.review_due_var.set(f"Amount due: {money(total - applied)}")

        def set_line_disc_selected():
            sel = self.review_tree.selection()
            if not sel:
                messagebox.showinfo("Select item", "Select an item first.")
                return
            idx = self.review_tree.index(sel[0])
            if idx < 0 or idx >= len(self.cart_lines):
                return

            try:
                cart_iids = self.tree.get_children()
                if idx < len(cart_iids):
                    self.tree.selection_set(cart_iids[idx])
            except Exception:
                pass

            self.set_line_discount()
            refresh_review()
            self.refresh_cart()

        def set_order_discount():
            win = tk.Toplevel(self)
            win.title("Total Discount")
            win.geometry("460x310")
            win.configure(bg=UI.CONTENT_BG)
            win.grab_set()

            c = Card(win, padx=18, pady=18)
            c.pack(fill="both", expand=True, padx=14, pady=14)

            tk.Label(c.inner, text="Total Discount", font=UI.FONT_LG, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w")
            tk.Label(c.inner, text="Applies after line discounts.", font=UI.FONT_MD, bg=UI.CARD, fg="#334155").pack(
                anchor="w", pady=(6, 12))

            mode_var = tk.StringVar()
            try:
                mode_var.set(str(self.order_disc_mode_var.get() or "PCT").upper())
            except Exception:
                mode_var.set("PCT")

            # Mode selector
            modes = tk.Frame(c.inner, bg=UI.CARD)
            modes.pack(anchor="w", pady=(0, 10))
            tk.Radiobutton(modes, text="Percent (%)", variable=mode_var, value="PCT", bg=UI.CARD).pack(side="left")
            tk.Radiobutton(modes, text="Amount off total", variable=mode_var, value="AMT", bg=UI.CARD).pack(side="left",
                                                                                                            padx=(
                                                                                                            16, 0))

            # Entry
            hint_var = tk.StringVar(value="")
            tk.Label(c.inner, textvariable=hint_var, font=UI.FONT_SM, bg=UI.CARD, fg=UI.MUTED).pack(anchor="w",
                                                                                                    pady=(0, 6))

            e = tk.Entry(c.inner, font=("Segoe UI", 16), width=12, bd=1, relief="solid")
            e.pack(anchor="w", pady=(0, 14))

            def _sync_entry():
                m = str(mode_var.get() or "PCT").upper()
                if m == "AMT":
                    hint_var.set("Enter amount to remove from the total (example: 10.00)")
                    try:
                        e.delete(0, tk.END)
                        e.insert(0, str(float(self.order_disc_amt_var.get() or 0)))
                    except Exception:
                        pass
                else:
                    hint_var.set("Enter percent off the total (0 to 100)")
                    try:
                        e.delete(0, tk.END)
                        e.insert(0, str(float(self.order_disc_var.get() or 0)))
                    except Exception:
                        pass
                try:
                    e.focus()
                    e.select_range(0, tk.END)
                except Exception:
                    pass

            def apply():
                m = str(mode_var.get() or "PCT").upper()
                try:
                    val = float(e.get().strip() or "0")
                except Exception:
                    messagebox.showerror("Invalid", "Discount must be a number.")
                    return

                if m == "AMT":
                    if val < 0:
                        val = 0.0
                    self.order_disc_mode_var.set("AMT")
                    self.order_disc_amt_var.set(val)
                    # Keep pct at 0 when using amount mode
                    self.order_disc_var.set(0.0)
                else:
                    val = max(0.0, min(100.0, val))
                    self.order_disc_mode_var.set("PCT")
                    self.order_disc_var.set(val)
                    # Keep amount at 0 when using percent mode
                    self.order_disc_amt_var.set(0.0)

                refresh_review()
                win.destroy()

            def clear_all():
                try:
                    self.order_disc_var.set(0.0)
                    self.order_disc_amt_var.set(0.0)
                    self.order_disc_mode_var.set("PCT")
                except Exception:
                    pass
                refresh_review()
                win.destroy()

            mode_var.trace_add("write", lambda *_: _sync_entry())
            _sync_entry()

            btns = tk.Frame(c.inner, bg=UI.CARD)
            btns.pack(fill="x", pady=(6, 0))
            GhostButton(btns, "Cancel", win.destroy).pack(side="right")
            GhostButton(btns, "Clear", clear_all).pack(side="right", padx=(0, 10))
            PrimaryButton(btns, "Apply", apply).pack(side="right", padx=(0, 10))

        def proceed():
            sale_lines = compute_sale_lines(materialize=True)
            if not sale_lines:
                return
            self._show_cash_panel(sale_lines)

        controls_left = tk.Frame(top_actions, bg=UI.CARD)
        controls_left.pack(side="left", fill="x", expand=True)
        controls_right = tk.Frame(top_actions, bg=UI.CARD)
        controls_right.pack(side="right")

        GhostButton(controls_left, "Back", self._hide_right_panels).pack(side="left", padx=(0, 8), pady=(0, 6))
        GhostButton(controls_left, "Line Discount", set_line_disc_selected).pack(side="left", padx=(0, 8), pady=(0, 6))
        GhostButton(controls_left, "Order Discount", set_order_discount).pack(side="left", padx=(0, 8), pady=(0, 6))

        def _wrap_review_actions(_evt=None):
            try:
                w = top_actions.winfo_width()
                if w and w < 520:
                    controls_right.pack_forget()
                    controls_left.pack_configure(side="top", fill="x", expand=True)
                    for child in controls_left.winfo_children():
                        child.pack_configure(side="top", fill="x", padx=0, pady=(0, 6))
                else:
                    controls_left.pack_configure(side="left", fill="x", expand=True)
                    controls_right.pack_configure(side="right")
                    for child in controls_left.winfo_children():
                        child.pack_configure(side="left", fill="none", padx=(0, 8), pady=(0, 6))
            except Exception:
                pass

        top_actions.bind("<Configure>", _wrap_review_actions)
        _wrap_review_actions()

        # Use the fixed bottom-right primary action button (same position always)
        try:
            self._set_stage("review", "Proceed", proceed, enabled=True)
        except Exception:
            pass

        refresh_review()

    def _show_cash_panel(self, sale_lines):
        try:
            if self.review_frame is not None:
                self.review_frame.destroy()
        except Exception:
            pass
        self.review_frame = None

        try:
            self.picker_frame.pack_forget()
        except Exception:
            pass

        try:
            if self.cash_frame is not None:
                self.cash_frame.destroy()
        except Exception:
            pass

        self.cash_frame = tk.Frame(self.right_stack, bg=UI.CARD)
        self.cash_frame.pack(fill="both", expand=True)

        try:
            self.right_title_lbl.configure(text="Checkout")
        except Exception:
            pass

        total = sum(float(l.get("line_total") or 0) for l in sale_lines)
        try:
            self._sync_bon_credit_balance()
        except Exception:
            pass

        credit = float(getattr(self, "exchange_credit", 0.0) or 0.0)
        if credit < 0:
            credit = 0.0
        applied_credit = min(credit, total)
        due = max(total - applied_credit, 0.0)

        payment_method_var = tk.StringVar(value=("EXCHANGE" if due <= 0.005 else "CASH"))

        tk.Label(self.cash_frame, text=f"Subtotal: {money(total)}", font=("Segoe UI", 14, "bold"), bg=UI.CARD,
                 fg=UI.TEXT).pack(anchor="w", pady=(8, 2))
        if applied_credit > 0:
            credit_label = "Bon credit" if getattr(self, "exchange_bon_codes", []) else "Exchange credit"
            tk.Label(self.cash_frame, text=f"{credit_label}: -{money(applied_credit)}", font=UI.FONT_MD, bg=UI.CARD,
                     fg=UI.MUTED).pack(anchor="w", pady=(0, 2))
        tk.Label(self.cash_frame, text=f"Amount due: {money(due)}", font=("Segoe UI", 20, "bold"), bg=UI.CARD,
                 fg=UI.TEXT).pack(anchor="w", pady=(0, 12))

        tk.Label(self.cash_frame, text="Payment method", font=UI.FONT_MD, bg=UI.CARD, fg=UI.TEXT_SOFT).pack(anchor="w", pady=(0, 4))
        method_row = tk.Frame(self.cash_frame, bg=UI.CARD)
        method_row.pack(anchor="w", fill="x", pady=(0, 10))
        payment_methods = [
            ("Cash", "CASH"),
            ("Whish", "WHISH"),
            ("Credit Card", "CREDIT_CARD"),
            ("Cash + Whish", "CASH_WHISH"),
            ("Cash + Card", "CASH_CARD"),
        ]
        for idx, (label, value) in enumerate(payment_methods):
            tk.Radiobutton(
                method_row,
                text=label,
                variable=payment_method_var,
                value=value,
                bg=UI.CARD,
                fg=UI.TEXT,
                selectcolor=UI.CARD,
                activebackground=UI.CARD,
                command=lambda: _update_payment_method_state(),
            ).grid(row=idx // 2, column=idx % 2, sticky="w", padx=(0, 14), pady=(0, 4))

        cash_lbl = tk.Label(self.cash_frame, text="Cash received", font=UI.FONT_MD, bg=UI.CARD, fg=UI.TEXT_SOFT)
        cash_lbl.pack(anchor="w")
        cash_e = tk.Entry(self.cash_frame, font=("Segoe UI", 16), width=14)
        UI.style_entry(cash_e, bg=UI.SURFACE)
        cash_e.pack(anchor="w", pady=(6, 10), ipady=6)
        try:
            cash_e.insert(0, f"{due:.2f}")
            cash_e.select_range(0, tk.END)
        except Exception:
            pass
        cash_e.focus()

        change_var = tk.StringVar(value="Change: $0.00")
        tk.Label(self.cash_frame, textvariable=change_var, font=("Segoe UI", 14, "bold"), bg=UI.CARD,
                 fg=UI.SUCCESS).pack(anchor="w", pady=(4, 6))

        split_balance_var = tk.StringVar(value="")
        tk.Label(self.cash_frame, textvariable=split_balance_var, font=UI.FONT_MD, bg=UI.CARD,
                 fg=UI.TEXT_SOFT).pack(anchor="w", pady=(0, 4))

        status_label = tk.Label(self.cash_frame, text="", font=UI.FONT_SM, bg=UI.CARD, fg=UI.PRIMARY)
        status_label.pack(anchor="w", pady=(6, 0))

        def _entry_amount():
            try:
                return float(cash_e.get().strip() or "0")
            except Exception:
                return 0.0

        def _set_entry_amount(value):
            try:
                cash_e.configure(state="normal")
                cash_e.delete(0, tk.END)
                cash_e.insert(0, f"{float(value or 0.0):.2f}")
            except Exception:
                pass

        def _split_label(pm: str) -> str:
            return "Whish" if pm == "CASH_WHISH" else "Credit Card"

        def _update_payment_method_state():
            pm = str(payment_method_var.get() or "CASH").upper()
            if due <= 0.005:
                try:
                    _set_entry_amount(0.0)
                    cash_e.configure(state="disabled")
                    cash_lbl.configure(text="Store credit covered")
                except Exception:
                    pass
                split_balance_var.set("")
                change_var.set("Change: $0.00")
                return
            if pm == "CASH":
                try:
                    cash_e.configure(state="normal")
                    cash_lbl.configure(text="Cash received")
                except Exception:
                    pass
                split_balance_var.set("")
                calc_change()
                return
            if pm in ("CASH_WHISH", "CASH_CARD"):
                try:
                    cash_e.configure(state="normal")
                    cash_lbl.configure(text="Cash portion")
                    current = _entry_amount()
                    if current >= due - 0.005:
                        cash_e.delete(0, tk.END)
                        cash_e.insert(0, "0.00")
                    cash_e.focus_set()
                    cash_e.select_range(0, tk.END)
                except Exception:
                    pass
                calc_change()
                return
            try:
                cash_e.configure(state="normal")
                cash_e.delete(0, tk.END)
                cash_e.insert(0, f"{due:.2f}")
                cash_e.configure(state="disabled")
                cash_lbl.configure(text=f"{'Whish' if pm == 'WHISH' else 'Credit Card'} amount")
            except Exception:
                pass
            split_balance_var.set("")
            change_var.set("Change: $0.00")

        def calc_change(*_):
            pm = str(payment_method_var.get() or "CASH").upper()
            if pm in ("CASH_WHISH", "CASH_CARD"):
                cash_part = max(0.0, _entry_amount())
                balance = max(0.0, due - cash_part)
                split_balance_var.set(f"{_split_label(pm)} portion: {money(balance)}")
                change_var.set("Change: $0.00")
                return
            if pm != "CASH":
                split_balance_var.set("")
                change_var.set("Change: $0.00")
                return
            try:
                cash = float(cash_e.get().strip() or "0")
            except Exception:
                cash = 0.0
            split_balance_var.set("")
            change_var.set(f"Change: {money(max(0.0, cash - due))}")

        cash_e.bind("<KeyRelease>", calc_change)
        payment_method_var.trace_add("write", lambda *_: _update_payment_method_state())
        calc_change()
        _update_payment_method_state()

        def complete():
            if not self._connected and connection_role() == "JOIN":
                try:
                    messagebox.showwarning("Disconnected", "Connection to host lost")
                except Exception:
                    pass
                return
            chosen_ui_pm = str(payment_method_var.get() or "CASH").strip().upper()
            chosen_pm = chosen_ui_pm
            cash_paid_amount = 0.0
            whish_paid_amount = 0.0
            card_paid_amount = 0.0
            if due <= 0.005:
                chosen_pm = "EXCHANGE"
            elif chosen_ui_pm == "CASH":
                try:
                    cash = float(cash_e.get().strip() or "0")
                except Exception:
                    messagebox.showerror("Invalid", "Cash received must be a number.")
                    return
                if cash + 0.005 < due:
                    messagebox.showerror("Cash received", "Cash received is less than the amount due.")
                    return
                cash_paid_amount = float(due)
                chosen_pm = "CASH"
            elif chosen_ui_pm in ("WHISH", "CREDIT_CARD"):
                chosen_pm = chosen_ui_pm
                if chosen_ui_pm == "WHISH":
                    whish_paid_amount = float(due)
                else:
                    card_paid_amount = float(due)
            elif chosen_ui_pm in ("CASH_WHISH", "CASH_CARD"):
                try:
                    cash_part = float(cash_e.get().strip() or "0")
                except Exception:
                    messagebox.showerror("Invalid", "Cash portion must be a number.")
                    return
                cash_part = round(max(0.0, cash_part), 2)
                if cash_part <= 0.005:
                    messagebox.showerror("Split payment", "Enter the cash amount for the split payment.")
                    return
                if cash_part >= due - 0.005:
                    messagebox.showerror("Split payment", "The non-cash portion must be greater than $0.00.")
                    return
                non_cash_part = round(max(0.0, due - cash_part), 2)
                cash_paid_amount = cash_part
                if chosen_ui_pm == "CASH_WHISH":
                    whish_paid_amount = non_cash_part
                    chosen_pm = "CASH+WHISH"
                else:
                    card_paid_amount = non_cash_part
                    chosen_pm = "CASH+CREDIT_CARD"
            else:
                chosen_pm = "CASH"
                cash_paid_amount = float(due)

            # Prevent double-click / double-Enter glitches
            if getattr(self, "_processing_sale", False):
                return
            self._processing_sale = True
            # Once checkout starts, do not offer the same cart after a hard crash;
            # stock and sale persistence may already be partially in progress.
            self._save_cashier_recovery(include_active=False)

            self._apply_primary_state()
            status_label.config(text="Processing sale...")
            self.update()

            stock_committed = []
            sale_saved = False

            def rollback_stock():
                for product_id, qty in reversed(stock_committed):
                    try:
                        adjust_stock(product_id, qty, reason="Sale checkout rolled back", movement_type="SALE_ROLLBACK")
                    except Exception:
                        pass
                stock_committed.clear()

            try:
                for l in sale_lines:
                    pid = l.get("product_id")
                    qty = int(l.get("qty") or 0)
                    if pid and qty and not l.get('is_quick'):
                        if not adjust_stock(pid, -qty, reason="Sale completed", movement_type="SALE"):
                            raise RuntimeError(
                                f"Could not update stock for {l.get('name') or 'an item'}. "
                                "Refresh products and try again."
                            )
                        stock_committed.append((pid, qty))

                # If exchange credit covers the entire amount due, record this as an EXCHANGE sale
                # (so it still counts as a real sale, but doesn't inflate cash-in-drawer totals).
                pm = chosen_pm
                # For exchanges we treat the return credit as a payment/credit, NOT as a discount on the original sale.
                # Exchange credit reduces amount due but must NOT be applied as an order discount.
                odt = 0.0  # exchange credit is stored separately (notes) so item prices stay correct
                notes_parts = []
                notes = ""
                try:
                    if float(cash_paid_amount or 0.0) > 0.005:
                        notes_parts.append(f"PAYMENT_CASH={float(cash_paid_amount):.2f}")
                    if float(whish_paid_amount or 0.0) > 0.005:
                        notes_parts.append(f"PAYMENT_WHISH={float(whish_paid_amount):.2f}")
                    if float(card_paid_amount or 0.0) > 0.005:
                        notes_parts.append(f"PAYMENT_CARD={float(card_paid_amount):.2f}")
                    if float(applied_credit) > 0:
                        notes_parts.append(f"EXCHANGE_CREDIT_APPLIED={float(applied_credit):.2f}")
                        origin_ids = []
                        for v in getattr(self, "exchange_origin_sale_ids", []) or []:
                            try:
                                iv = int(v)
                                if iv not in origin_ids:
                                    origin_ids.append(iv)
                            except Exception:
                                pass
                        return_ids = []
                        for v in getattr(self, "exchange_return_ids", []) or []:
                            try:
                                iv = int(v)
                                if iv not in return_ids:
                                    return_ids.append(iv)
                            except Exception:
                                pass
                        if origin_ids:
                            notes_parts.append(f"ORIG_SALE_ID={origin_ids[0]}")
                            notes_parts.append("ORIG_SALE_IDS=" + ",".join(str(x) for x in origin_ids))
                        if return_ids:
                            notes_parts.append("RETURN_IDS=" + ",".join(str(x) for x in return_ids))
                        bon_codes = []
                        bon_available = 0.0
                        try:
                            bon_available, bon_codes = self._loaded_bon_total()
                        except Exception:
                            bon_available, bon_codes = 0.0, []
                        bon_credit_applied = round(min(float(applied_credit), float(bon_available or 0.0)), 2)
                        if bon_codes and bon_credit_applied > 0.005:
                            notes_parts.append("BON_CODES=" + ",".join(str(x) for x in bon_codes))
                            notes_parts.append(f"BON_CREDIT_APPLIED={bon_credit_applied:.2f}")
                        # If credit fully covers this sale, mark payment method as EXCHANGE so it won't
                        # inflate cash-in-drawer totals.
                        if float(due) <= 0.005:
                            pm = "EXCHANGE"

                except Exception:
                    pass
                try:
                    wheel_receipt_label = str(getattr(self, "_wheel_receipt_label", "") or "").strip()
                    if wheel_receipt_label:
                        wheel_receipt_label = wheel_receipt_label.replace(";", " ").replace("\r", " ").replace("\n", " ")
                        notes_parts.append(f"WHEEL_PRIZE={wheel_receipt_label}")
                    notes = ";".join([p for p in notes_parts if p])
                except Exception:
                    notes = ""


                try:
                    sale_id = create_sale(sale_lines, payment_method=pm, customer_name="",
                                          order_discount_total=odt, notes=notes)
                    sale_saved = True
                except KeyError as e:
                    raise RuntimeError(
                        "Could not complete sale because the server response was incomplete. "
                        "Check the host connection and try again."
                    ) from e

                # Bug 8 fix: quick items are now kept in the catalog so they don't
                # need to be re-entered each time. Deletion removed intentionally.

                try:
                    sale, items = get_sale_receipt_data(sale_id)
                except Exception as e:
                    sale, items = None, []
                    messagebox.showwarning(
                        "Sale saved",
                        f"Sale #{sale_id} was saved, but its receipt could not be loaded.\n\n"
                        f"Use Reprint after checking the connection.\n{e}"
                    )

                status_label.config(text="Printing receipt...")
                self.update()

                printed_normal = False
                printed_gift = False
                if sale is not None:
                    try:
                        printed_normal = bool(print_configured_receipt(get_store_name(), sale, items))
                    except Exception:
                        printed_normal = False

                # If Gift receipt checkbox is ON, print the gift copy too
                user_cancelled_gift = False
                if sale is not None and bool(getattr(self, "gift_var", None).get() if getattr(self, "gift_var",
                                                                                              None) is not None else False):
                    try:
                        gift_items = self._select_gift_items(items)
                        if gift_items is not None:
                            if len(gift_items) > 0:
                                printed_gift = bool(print_configured_gift_receipt(get_store_name(), sale, gift_items))
                            else:
                                printed_gift = True
                        else:
                            user_cancelled_gift = True
                            printed_gift = True
                    except Exception:
                        printed_gift = False

                # Do not open the PDF viewer as a print fallback; report the print failure clearly.
                if sale is not None and not printed_normal:
                    messagebox.showwarning(
                        "Print failed",
                        "The sale was saved, but the receipt was not sent to the printer.\n\n"
                        "Check Settings > Receipt Printer and use Test Print."
                    )

                if sale is not None and (bool(getattr(self, "gift_var", None).get() if getattr(self, "gift_var",
                                                                                               None) is not None else False)) and (
                not printed_gift) and not user_cancelled_gift:
                    messagebox.showwarning(
                        "Gift print failed",
                        "The gift receipt was not sent to the printer.\n\n"
                        "Check Settings > Receipt Printer and use Test Print."
                    )
                self.cart_lines = []
                self._undo_stack.clear()
                self._wheel_prize_claimed = False
                self._wheel_discount_pct = 0.0
                self._wheel_receipt_label = ""
                remaining_credit = round(max(0.0, credit - applied_credit), 2)
                self.set_exchange_credit_balance(remaining_credit)
                try:
                    self._sync_bon_credit_balance()
                except Exception:
                    pass
                self._processing_sale = False
                self.refresh_cart()
                self.refresh_picker()
                self._hide_right_panels()

            except Exception as e:
                if not sale_saved:
                    rollback_stock()
                    messagebox.showerror("Error", f"Could not complete sale: {e}")
                else:
                    self.cart_lines = []
                    self._undo_stack.clear()
                    self._wheel_prize_claimed = False
                    self._wheel_discount_pct = 0.0
                    self._wheel_receipt_label = ""
                    remaining_credit = round(max(0.0, credit - applied_credit), 2)
                    self.set_exchange_credit_balance(remaining_credit)
                    try:
                        self._sync_bon_credit_balance()
                    except Exception:
                        pass
                    messagebox.showwarning(
                        "Sale saved",
                        f"The sale was saved, but checkout could not finish cleanly.\n\n"
                        f"Do not submit it again. Use Reprint if a receipt is needed.\n{e}"
                    )
                self._processing_sale = False
                self._save_cashier_recovery()
                self._apply_primary_state()
                status_label.config(text="")

        # Bottom row inside the cash panel: keep Back here, but use the FIXED primary action button for completing the sale
        bottom = tk.Frame(self.cash_frame, bg=UI.CARD)
        bottom.pack(side="bottom", fill="x", pady=(6, 0))

        GhostButton(bottom, "Back", self._show_review_panel).pack(side="left")

        # Fixed bottom-right primary action (same position as Checkout/Proceed)
        try:
            can_complete = True
            if connection_role() == "JOIN" and not self._connected:
                can_complete = False
            self._set_stage("cash", "Complete Sale", complete, enabled=can_complete)
        except Exception:
            pass

        cash_e.bind("<Return>", lambda e: complete())

    def _select_gift_items(self, items):
        """Prompt user with a dialog to select which items and what quantities to include in the gift receipt.
        
        Returns:
            list[dict]: A list of items to include in the gift receipt (cloned with updated qty),
                        or None if the user cancelled the dialog.
        """
        top = self.winfo_toplevel()
        win = tk.Toplevel(top)
        win.title("Select Gift Items")
        win.geometry("500x480")
        win.minsize(440, 380)
        win.configure(bg=UI.CONTENT_BG)
        win.transient(top)
        win.grab_set()

        box = Card(win, padx=18, pady=18)
        box.pack(fill="both", expand=True, padx=14, pady=14)

        tk.Label(
            box.inner,
            text="Select Gift Items",
            font=UI.FONT_LG,
            bg=UI.CARD,
            fg=UI.TEXT
        ).pack(anchor="w")

        tk.Label(
            box.inner,
            text="Uncheck items to exclude them, or adjust their quantities for the gift receipt.",
            font=UI.FONT_MD,
            bg=UI.CARD,
            fg="#475569"
        ).pack(anchor="w", pady=(4, 10))

        # Scrollable container for the items list
        scroll = VScrollableFrame(box.inner, bg=UI.CARD)
        scroll.pack(fill="both", expand=True, pady=(0, 10))

        item_rows = []
        for it in items:
            row = tk.Frame(scroll.inner, bg=UI.CARD)
            row.pack(fill="x", pady=4)

            # Checkbox
            var_checked = tk.BooleanVar(value=True)
            chk = tk.Checkbutton(
                row,
                text=str(it.get("name") or ""),
                variable=var_checked,
                bg=UI.CARD,
                fg=UI.TEXT,
                activebackground=UI.CARD,
                activeforeground=UI.TEXT,
                selectcolor=UI.CARD,
                anchor="w",
                font=UI.FONT_MD
            )
            chk.pack(side="left", fill="x", expand=True)

            qty_val = int(it.get("qty", 1))
            var_spin = None
            spin = None
            if qty_val > 1:
                var_spin = tk.IntVar(value=qty_val)
                qty_f = tk.Frame(row, bg=UI.CARD)
                qty_f.pack(side="right")
                
                tk.Label(qty_f, text="Gift Qty:", bg=UI.CARD, fg=UI.TEXT_SOFT, font=UI.FONT_SM).pack(side="left", padx=(5, 2))
                spin = tk.Spinbox(
                    qty_f,
                    from_=1,
                    to=qty_val,
                    textvariable=var_spin,
                    width=4,
                    font=UI.FONT_MD,
                    bd=1,
                    relief="solid"
                )
                spin.pack(side="left")
                
                def make_toggle_cmd(cv, sp):
                    return lambda: sp.config(state="normal" if cv.get() else "disabled")
                chk.config(command=make_toggle_cmd(var_checked, spin))

            item_rows.append({
                "item_data": it,
                "var_checked": var_checked,
                "var_spin": var_spin,
                "spin_widget": spin
            })

        # Bottom buttons
        btns = tk.Frame(box.inner, bg=UI.CARD)
        btns.pack(fill="x", pady=(10, 0))

        result = {"cancelled": True, "selected_items": []}

        def _print_gift():
            result["cancelled"] = False
            for r in item_rows:
                if r["var_checked"].get():
                    it_copy = dict(r["item_data"])
                    if r["var_spin"] is not None:
                        try:
                            chosen_qty = int(r["var_spin"].get())
                            max_qty = int(r["item_data"].get("qty", 1))
                            chosen_qty = max(1, min(max_qty, chosen_qty))
                            it_copy["qty"] = chosen_qty
                        except Exception:
                            pass
                    result["selected_items"].append(it_copy)
            win.grab_release()
            win.destroy()

        def _cancel():
            win.grab_release()
            win.destroy()

        PrimaryButton(btns, "Print Gift Receipt", _print_gift).pack(side="left")
        GhostButton(btns, "Cancel", _cancel).pack(side="right")

        win.protocol("WM_DELETE_WINDOW", _cancel)
        win.wait_window(win)

        if result["cancelled"]:
            return None
        return result["selected_items"]

    def open_reprint_popup(self):
        """Receipt reprint: Sale ID, receipt barcode scan, or paste."""
        top = self.winfo_toplevel()
        win = tk.Toplevel(top)
        win.title("Reprint Receipt")
        win.transient(top)
        win.grab_set()
        win.configure(bg=UI.CONTENT_BG)
        win.resizable(False, False)

        body = Card(win, padx=16, pady=14)
        body.pack(fill="both", expand=True, padx=12, pady=12)

        tk.Label(body.inner, text="Enter Sale ID or scan/paste receipt barcode:", bg=UI.CARD, fg=UI.TEXT).pack(
            anchor="w")
        v = tk.StringVar(value="")
        ent = tk.Entry(body.inner, textvariable=v, font=("Segoe UI", 12), width=34, bd=1, relief="solid")
        ent.pack(anchor="w", pady=(8, 0))
        ent.focus_set()

        hint = tk.Label(body.inner, text="Tip: you can also browse sales by date in Cash Drawer.", bg=UI.CARD,
                        fg="#64748b")
        hint.pack(anchor="w", pady=(8, 0))

        btns = tk.Frame(body.inner, bg=UI.CARD)
        btns.pack(fill="x", pady=(12, 0))

        def _close():
            try:
                win.grab_release()
            except Exception:
                pass
            win.destroy()

        def _resolve_sale_id(raw: str):
            s = str(raw or "").strip()
            if not s:
                return None
            if s.isdigit():
                try:
                    return int(s)
                except Exception:
                    return None
            try:
                sale, _items = get_sale_by_receipt_scan(s)
                if sale:
                    sid = row_get(sale, "id", None)
                    if sid is not None:
                        return int(sid)
            except Exception:
                pass
            m = re.search(r"(\d+)", s)
            if m:
                try:
                    return int(m.group(1))
                except Exception:
                    return None
            return None

        def _do_reprint():
            sale_id = _resolve_sale_id(v.get())
            if not sale_id:
                messagebox.showerror("Missing", "Please enter a Sale ID or scan a receipt barcode.", parent=win)
                return
            try:
                sale, items = get_sale_detail_with_returns(int(sale_id))
            except Exception as e:
                messagebox.showerror("Error", f"{type(e).__name__}: {e}", parent=win)
                return
            if not sale:
                messagebox.showerror("Not found", "Sale not found.", parent=win)
                return
            try:
                ok = bool(print_configured_receipt(get_store_name(), sale, items))
                if not ok:
                    messagebox.showwarning(
                        "Print", "Receipt was not sent. Check Settings > Receipt Printer and use Test Print.",
                        parent=win,
                    )
                    return
            except Exception as e:
                messagebox.showerror("Error", f"{type(e).__name__}: {e}", parent=win)
                return
            _close()

        def _do_gift_reprint():
            sale_id = _resolve_sale_id(v.get())
            if not sale_id:
                messagebox.showerror("Missing", "Please enter a Sale ID or scan a receipt barcode.", parent=win)
                return
            sale, items = get_sale_detail_with_returns(int(sale_id))
            if not sale:
                messagebox.showerror("Not found", "Sale not found.", parent=win)
                return
            selected = self._select_gift_items(items)
            if selected is None:
                return
            if not selected:
                messagebox.showinfo("Gift receipt", "Select at least one item.", parent=win)
                return
            if not print_configured_gift_receipt(get_store_name(), sale, selected):
                messagebox.showwarning(
                    "Print", "Gift receipt was not sent. Check Settings > Receipt Printer and use Test Print.",
                    parent=win,
                )
                return
            _close()

        PrimaryButton(btns, "Reprint", _do_reprint).pack(side="left")
        GhostButton(btns, "Gift Receipt", _do_gift_reprint).pack(side="left", padx=(8, 0))
        GhostButton(btns, "Go to Cash Drawer", self.go_to_sales_today).pack(side="left", padx=8)
        GhostButton(btns, "Close", _close).pack(side="right")

        ent.bind("<Return>", lambda e: _do_reprint())
        win.protocol("WM_DELETE_WINDOW", _close)

    def go_to_sales_today(self):
        """Jump to Cash Drawer page and show today's sales list."""
        top = self.winfo_toplevel()
        try:
            top.show_page("ShiftsPage")
            try:
                top.shifts_page.date_var.set(datetime.now().strftime("%Y-%m-%d"))
                top.shifts_page.refresh_all()
            except Exception:
                pass
        except Exception:
            pass

    def go_to_sales_date(self):
        """Jump to Cash Drawer page and show sales for the selected date (dropdowns)."""
        y = (self.sales_year.get() or "").strip()
        m = (self.sales_month.get() or "").strip()
        d = (self.sales_day.get() or "").strip()
        if not (y and m and d):
            day = datetime.now().strftime("%Y-%m-%d")
        else:
            day = f"{y}-{m}-{d}"

        top = self.winfo_toplevel()
        try:
            top.show_page("ShiftsPage")
            try:
                top.shifts_page.date_var.set(day)
                top.shifts_page.refresh_all()
            except Exception:
                pass
        except Exception:
            pass

    def checkout(self):
        self._show_review_panel()

    def set_connection_state(self, connected: bool):
        """Called by main app when JOIN connection changes."""
        self._connected = bool(connected)

        # Legacy: if an old in-panel Complete button exists, keep it in sync
        if self._complete_btn is not None:
            try:
                legacy_enabled = (
                    self._connected
                    and not getattr(self, "_processing_sale", False)
                    and not getattr(self, "_shift_locked", False)
                )
                self._complete_btn.config(state=("normal" if legacy_enabled else "disabled"))
            except Exception:
                pass

        # Ensure the fixed primary action button is updated too
        self._apply_primary_state()

    def _set_stage(self, stage: str, label: str, cmd, enabled: bool = True):
        """Update the fixed bottom-right primary action button."""
        self._stage = stage
        self._primary_label = label
        self._primary_cmd = cmd
        self._primary_enabled_requested = bool(enabled)
        try:
            self.primary_action_btn.config(text=label, command=cmd)
        except Exception:
            pass
        self._apply_primary_state()

    def _apply_primary_state(self):
        """Enable/disable primary action based on stage, cart content, and connection."""
        if not hasattr(self, "primary_action_btn") or self.primary_action_btn is None:
            return

        if getattr(self, "_processing_sale", False):
            try:
                self.primary_action_btn.config(state="disabled", text="Processing...")
            except Exception:
                pass
            return

        if getattr(self, "_shift_locked", False):
            try:
                self.primary_action_btn.config(state="disabled", text="Open Shift First")
            except Exception:
                pass
            return

        enabled = bool(getattr(self, "_primary_enabled_requested", True))
        label = getattr(self, "_primary_label", "Checkout")

        # If cart is empty, don't allow checkout/proceed/complete.
        try:
            if len(self.cart_lines) == 0:
                enabled = False
        except Exception:
            pass

        # If JOIN disconnected, hard-block completing sale.
        if connection_role() == "JOIN" and getattr(self, "_connected", True) is False:
            if getattr(self, "_stage", "cart") == "cash":
                enabled = False

        try:
            self.primary_action_btn.config(text=label, state=("normal" if enabled else "disabled"))
        except Exception:
            pass


# ---------------- PRODUCTS PAGE ----------------

def open_product_sales_history_window(parent, product_id, product_name=""):
    win = tk.Toplevel(parent)
    win.title(f"Sales History - {product_name or 'Product'}")
    win.geometry("820x500")
    win.minsize(620, 390)
    win.configure(bg=UI.CONTENT_BG)
    win.transient(parent.winfo_toplevel())
    card = Card(win, padx=14, pady=14)
    card.pack(fill="both", expand=True, padx=12, pady=12)
    HeaderBar(card.inner, product_name or "Product Sales", "Selling price, quantities, returns, and voided sales.").pack(fill="x")
    cols = ("date", "receipt", "qty", "returned", "price", "total", "status")
    tree = ttk.Treeview(card.inner, columns=cols, show="headings", height=15)
    specs = [("date","Date / time",145),("receipt","Receipt",85),("qty","Qty",55),("returned","Returned",70),("price","Sold at",85),("total","Line total",95),("status","Status",70)]
    for col, title, width in specs:
        tree.heading(col, text=title); tree.column(col, width=width, anchor=("center" if col in {"receipt","qty","returned","status"} else "w"))
    tree.pack(fill="both", expand=True, pady=(10, 8))
    tree.tag_configure("voided", foreground="#991b1b", background="#fee2e2")
    sale_ids = {}
    for idx, row in enumerate(list_product_sales(int(product_id), 500, True) or []):
        sid = int(row_get(row, "sale_id", 0) or 0)
        iid = f"sale_{sid}_{idx}"
        sale_ids[iid] = sid
        voided = bool(int(row_get(row, "is_voided", 0) or 0))
        tree.insert("", tk.END, iid=iid, tags=(("voided",) if voided else ()), values=(
            row_get(row,"created_at",""), row_get(row,"receipt_code","") or sid,
            row_get(row,"qty",0), row_get(row,"returned_qty",0), money(row_get(row,"price",0)),
            money(row_get(row,"line_total",0)), "VOID" if voided else "Active"))
    def reprint():
        sel = tree.selection()
        if not sel:
            messagebox.showinfo("Reprint", "Select a sale first.", parent=win); return
        sale, items = get_sale_receipt_data(sale_ids.get(sel[0]))
        if not sale or not print_configured_receipt(get_store_name(), sale, items):
            messagebox.showwarning("Reprint", "Receipt was not sent. Check the configured printer.", parent=win)
    row = tk.Frame(card.inner, bg=UI.CARD); row.pack(fill="x")
    PrimaryButton(row, "Reprint Selected", reprint).pack(side="left")
    GhostButton(row, "Close", win.destroy).pack(side="right")


def open_product_price_history_window(parent, product_id, product_name=""):
    win = tk.Toplevel(parent)
    win.title(f"Price History - {product_name or 'Product'}")
    win.geometry("650x430")
    win.minsize(520, 340)
    win.configure(bg=UI.CONTENT_BG)
    win.transient(parent.winfo_toplevel())
    card = Card(win, padx=14, pady=14); card.pack(fill="both", expand=True, padx=12, pady=12)
    HeaderBar(card.inner, product_name or "Price History", "Every saved catalog price change.").pack(fill="x")
    cols=("time","old","new","change","reason")
    tree=ttk.Treeview(card.inner,columns=cols,show="headings",height=14)
    for col,title,width in [("time","Changed",150),("old","Old",80),("new","New",80),("change","Change",80),("reason","Reason",190)]:
        tree.heading(col,text=title); tree.column(col,width=width,anchor=("e" if col in {"old","new","change"} else "w"))
    tree.pack(fill="both",expand=True,pady=(10,8))
    for row in list_product_price_history(int(product_id),500) or []:
        change=float(row_get(row,"price_change",0) or 0)
        tree.insert("",tk.END,values=(row_get(row,"changed_at",""),money(row_get(row,"old_price",0)),money(row_get(row,"new_price",0)),f"{change:+.2f}",row_get(row,"reason","")))
    if not tree.get_children():
        tk.Label(card.inner,text="No price changes recorded yet.",bg=UI.CARD,fg=UI.MUTED).pack(anchor="w")
    GhostButton(card.inner,"Close",win.destroy).pack(anchor="e")


class ProductsPage(tk.Frame):
    def __init__(self, parent, on_product_added):
        super().__init__(parent, bg=UI.CONTENT_BG)
        self.on_product_added = on_product_added
        self.selected_product_id = None
        self.selected_barcode = ""
        self._build()
        self.refresh_list()

    def _build(self):
        scroll = VScrollableFrame(self, bg=UI.CONTENT_BG)
        scroll.pack(fill="both", expand=True)
        wrap = tk.Frame(scroll.inner, bg=UI.CONTENT_BG)
        wrap.pack(fill="both", expand=True, padx=(10 if UI.COMPACT else 18), pady=(10 if UI.COMPACT else 18))

        header = Card(wrap, padx=18, pady=14)
        header.pack(fill="x")
        HeaderBar(header.inner, "Products", "Add, edit, search, delete, and print labels directly.").pack(fill="x")

        body = tk.Frame(wrap, bg=UI.CONTENT_BG)
        body.pack(fill="both", expand=True, pady=((8 if UI.COMPACT else 14), 0))
        body.grid_columnconfigure(0, weight=0, minsize=(360 if UI.COMPACT else 430))
        body.grid_columnconfigure(1, weight=1)
        body.grid_rowconfigure(0, weight=1)

        left = Card(body, padx=14, pady=14)
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 6 if UI.COMPACT else 10))

        right = Card(body, padx=14, pady=14)
        right.grid(row=0, column=1, sticky="nsew", padx=((6 if UI.COMPACT else 10), 0))

        tk.Label(left.inner, text="Add Product", font=UI.FONT_LG, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w")

        def field(parent, label):
            row = tk.Frame(parent, bg=UI.CARD)
            row.pack(fill="x", pady=6)
            tk.Label(row, text=label, bg=UI.CARD, fg="#334155", width=16, anchor="w").pack(side="left")
            e = tk.Entry(row, bd=1, relief="solid")
            e.pack(side="left", fill="x", expand=True)
            return e

        self.name_e = field(left.inner, "Product name")

        row_price = tk.Frame(left.inner, bg=UI.CARD)
        row_price.pack(fill="x", pady=6)
        tk.Label(row_price, text="Sell price", bg=UI.CARD, fg="#334155", width=16, anchor="w").pack(side="left")
        self.price_e = tk.Entry(row_price, bd=1, relief="solid")
        self.price_e.pack(side="left", fill="x", expand=True)

        row_stock = tk.Frame(left.inner, bg=UI.CARD)
        row_stock.pack(fill="x", pady=6)
        tk.Label(row_stock, text="Stock qty", bg=UI.CARD, fg="#334155", width=16, anchor="w").pack(side="left")
        self.stock_e = tk.Entry(row_stock, bd=1, relief="solid")
        self.stock_e.pack(side="left", fill="x", expand=True)

        btn_row = tk.Frame(left.inner, bg=UI.CARD)
        btn_row.pack(fill="x", pady=(12, 0))
        PrimaryButton(btn_row, "Add", self.add_clicked).pack(side="left")
        PrimaryButton(
            btn_row,
            "Add + Print Stock",
            lambda: self.add_clicked(print_stock=True),
            bg=UI.SUCCESS,
            activebackground="#15803d",
        ).pack(side="left", padx=(8, 0))
        GhostButton(
            btn_row,
            "Add + Send Print to Host",
            lambda: self.add_clicked(send_host_print=True),
        ).pack(side="left", padx=(8, 0))

        # Optional purchasing data stays below the print actions so workers can
        # ignore it during normal product entry.
        self.cost_e = field(left.inner, "Cost price (optional)")
        self.supplier_e = field(left.inner, "Supplier (optional)")

        tk.Frame(left.inner, height=(12 if UI.COMPACT else 18), bg=UI.CARD).pack(fill="x")
        self.barcode_e = field(left.inner, "Barcode (optional)")
        self.location_e = field(left.inner, "Location / section (optional)")

        row_low = tk.Frame(left.inner, bg=UI.CARD)
        row_low.pack(fill="x", pady=6)
        tk.Label(row_low, text="Low stock", bg=UI.CARD, fg="#334155", width=16, anchor="w").pack(side="left")
        self.low_e = tk.Entry(row_low, bd=1, relief="solid")
        self.low_e.insert(0, "0")
        self.low_e.pack(side="left", fill="x", expand=True)

        row_cat = tk.Frame(left.inner, bg=UI.CARD)
        row_cat.pack(fill="x", pady=6)
        tk.Label(row_cat, text="Category", bg=UI.CARD, fg="#334155", width=16, anchor="w").pack(side="left")
        self.category_cb = ttk.Combobox(row_cat, values=[], state="normal")
        self.category_cb.pack(side="left", fill="x", expand=True)

        row_brand = tk.Frame(left.inner, bg=UI.CARD)
        row_brand.pack(fill="x", pady=6)
        tk.Label(row_brand, text="Brand", bg=UI.CARD, fg="#334155", width=16, anchor="w").pack(side="left")
        self.brand_e = tk.Entry(row_brand, bd=1, relief="solid")
        self.brand_e.pack(side="left", fill="x", expand=True)

        self.last_barcode = tk.StringVar(value="")
        tk.Label(left.inner, textvariable=self.last_barcode, bg=UI.CARD, fg=UI.MUTED).pack(anchor="w", pady=(10, 0))

        tk.Label(right.inner, text="All Products", font=UI.FONT_LG, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w")

        search_row = tk.Frame(right.inner, bg=UI.CARD)
        search_row.pack(fill="x", pady=(10, 10))
        tk.Label(search_row, text="Search", bg=UI.CARD, fg="#334155").pack(side="left")
        self.search_e = tk.Entry(search_row, bd=1, relief="solid")
        self.search_e.pack(side="left", fill="x", expand=True, padx=8)
        self.search_e.bind("<KeyRelease>", lambda e: self.refresh_list())
        GhostButton(search_row, "Search", self.refresh_list).pack(side="left", padx=(0, 8))
        GhostButton(search_row, "Refresh", self.refresh_list).pack(side="left")
        GhostButton(search_row, "Reorder", self.open_reorder_suggestions).pack(side="left", padx=(8, 0))

        cols = ("name", "price", "stock", "barcode", "location", "low")
        self.prod_tree = ttk.Treeview(right.inner, columns=cols, show="headings", height=12)
        self.prod_tree.heading("name", text="Name")
        self.prod_tree.heading("price", text="Price")
        self.prod_tree.heading("stock", text="Stock")
        self.prod_tree.heading("barcode", text="Barcode")
        self.prod_tree.heading("location", text="Location")
        self.prod_tree.heading("low", text="Low stock")
        self.prod_tree.column("name", width=240)
        self.prod_tree.column("price", width=90, anchor="e")
        self.prod_tree.column("stock", width=70, anchor="center")
        self.prod_tree.column("barcode", width=135)
        self.prod_tree.column("location", width=140)
        self.prod_tree.column("low", width=80, anchor="center")
        self.prod_tree.pack(fill="both", expand=True)
        self.prod_tree.bind("<<TreeviewSelect>>", self.on_select)

        # Stock highlighting (yellow for low, red for out)
        try:
            self.prod_tree.tag_configure("low_stock", background="#fff3b0")
            self.prod_tree.tag_configure("out_stock", background="#ffb4b4")
        except Exception:
            pass

        edit = tk.Frame(right.inner, bg=UI.CARD)
        edit.pack(fill="x", pady=(12, 0))

        tk.Label(edit, text="Edit selected", font=("Segoe UI", 11, "bold"), bg=UI.CARD, fg=UI.TEXT).grid(
            row=0, column=0, sticky="w", pady=(0, 8), columnspan=8
        )

        def edit_field(label, col, width=16):
            tk.Label(edit, text=label, bg=UI.CARD, fg="#334155").grid(row=1, column=col, sticky="w", padx=(0, 6))
            e = tk.Entry(edit, width=width, bd=1, relief="solid")
            e.grid(row=2, column=col, sticky="w", padx=(0, 12))
            return e

        self.edit_name = edit_field("Name", 0, width=28)
        self.edit_price = edit_field("Price", 1, width=10)
        self.edit_stock = edit_field("Stock", 2, width=10)
        self.edit_location = edit_field("Location", 3, width=16)
        self.edit_low = edit_field("Low", 4, width=10)

        # Row 3 & 4: Category and Brand
        tk.Label(edit, text="Category", bg=UI.CARD, fg="#334155").grid(row=3, column=0, sticky="w", padx=(0, 6), pady=(8, 0))
        self.edit_category = ttk.Combobox(edit, width=26, values=[], state="normal")
        self.edit_category.grid(row=4, column=0, sticky="w", padx=(0, 12))

        tk.Label(edit, text="Brand", bg=UI.CARD, fg="#334155").grid(row=3, column=1, sticky="w", padx=(0, 6), pady=(8, 0))
        self.edit_brand = tk.Entry(edit, width=12, bd=1, relief="solid")
        self.edit_brand.grid(row=4, column=1, sticky="w", padx=(0, 12))

        btns = tk.Frame(edit, bg=UI.CARD)
        btns.grid(row=4, column=3, columnspan=3, sticky="e", padx=(10, 0), pady=(8, 0))

        PrimaryButton(btns, "Save", self.save_changes).pack(side="left", padx=(0, 8))
        GhostButton(btns, "Details", self.open_product_details).pack(side="left", padx=(0, 8))
        DangerButton(btns, "Delete", self.delete_selected).pack(side="left")

        self.selected_detail_lbl = tk.Label(
            edit, text="Select an item for cost, supplier, and stock history.",
            bg=UI.CARD, fg=UI.MUTED, anchor="w", justify="left",
        )
        self.selected_detail_lbl.grid(row=5, column=0, columnspan=8, sticky="ew", pady=(12, 0))

    def open_reorder_suggestions(self):
        win = tk.Toplevel(self)
        win.title("Reorder Suggestions")
        win.geometry("900x560")
        win.minsize(680, 430)
        win.configure(bg=UI.CONTENT_BG)
        win.transient(self.winfo_toplevel())
        card = Card(win, padx=14, pady=14); card.pack(fill="both", expand=True, padx=12, pady=12)
        HeaderBar(card.inner, "Reorder Suggestions", "Optional suggestions based on recent sales speed and current stock.").pack(fill="x")
        controls=tk.Frame(card.inner,bg=UI.CARD); controls.pack(fill="x",pady=(10,8))
        days_var=tk.StringVar(value="30"); target_var=tk.StringVar(value="14"); supplier_var=tk.StringVar(value="")
        for label,var,width in [("Sales days",days_var,6),("Target days",target_var,6),("Supplier",supplier_var,20)]:
            tk.Label(controls,text=label,bg=UI.CARD,fg=UI.TEXT).pack(side="left",padx=(0,4))
            tk.Entry(controls,textvariable=var,width=width,bd=1,relief="solid").pack(side="left",padx=(0,10))
        cols=("supplier","product","sold","daily","stock","cover","suggested","cost")
        tree=ttk.Treeview(card.inner,columns=cols,show="headings",height=16)
        for col,title,width in [("supplier","Supplier",125),("product","Product",210),("sold","Sold",60),("daily","/ day",65),("stock","Stock",60),("cover","Days cover",75),("suggested","Order",70),("cost","Est. cost",85)]:
            tree.heading(col,text=title); tree.column(col,width=width,anchor=("w" if col in {"supplier","product"} else "center"))
        tree.pack(fill="both",expand=True)
        summary=tk.StringVar(value=""); tk.Label(card.inner,textvariable=summary,bg=UI.CARD,fg=UI.MUTED).pack(anchor="w",pady=(8,0))
        def load():
            try: rows=reorder_suggestions(int(days_var.get()),int(target_var.get()),supplier_var.get().strip(),2000) or []
            except Exception as exc: messagebox.showerror("Reorder",str(exc),parent=win); return
            tree.delete(*tree.get_children()); total_cost=0.0
            for idx,row in enumerate(rows):
                cost=float(row_get(row,"estimated_cost",0) or 0); total_cost+=cost
                cover=row_get(row,"days_cover",None)
                tree.insert("",tk.END,iid=str(idx),values=(row_get(row,"supplier","") or "Unassigned",row_get(row,"name",""),row_get(row,"net_qty_sold",0),f"{float(row_get(row,'avg_daily_units',0) or 0):.2f}",row_get(row,"stock_qty",0),(f"{float(cover):.1f}" if cover is not None else "—"),row_get(row,"suggested_qty",0),money(cost)))
            summary.set(f"{len(rows)} products suggested   |   Estimated cost where costs are entered: {money(total_cost)}")
        PrimaryButton(controls,"Refresh",load).pack(side="left")
        GhostButton(card.inner,"Close",win.destroy).pack(anchor="e",pady=(8,0)); load()

    def refresh_list(self):
        try:
            cats = get_distinct_categories() or []
            self.category_cb['values'] = cats
            self.edit_category['values'] = cats
        except Exception:
            pass

        query = self.search_e.get().strip() if hasattr(self, "search_e") else ""
        rows = list_products(query)

        # Some backends return the full list even when a query is provided.
        # To make search always work, we also filter client-side.
        q = (query or "").strip().lower()
        if q:
            terms = [t for t in re.split(r"\s+", q) if t]
            def _match(r):
                name = str(row_get(r, "name", "") or "").lower()
                bc = str(row_get(r, "barcode", "") or "").lower()
                cat = str(row_get(r, "category", "") or "").lower()
                loc = str(row_get(r, "location", "") or "").lower()
                return all((t in name) or (t in bc) or (t in cat) or (t in loc) for t in terms)
            rows = [r for r in rows if _match(r)]

        for i in self.prod_tree.get_children():
            self.prod_tree.delete(i)

        self._product_rows_by_id = {}
        for r in rows:
            try:
                p_cat = str(row_get(r, "category") or "").strip()
                if p_cat.lower() == "quick":
                    continue
            except Exception:
                pass

            try:
                stock = int(row_get(r, "stock_qty", 0) or 0)
            except Exception:
                stock = 0
            try:
                low = int(row_get(r, "low_stock_level", 0) or 0)
            except Exception:
                low = 0

            tags = ()
            if stock <= 0:
                tags = ("out_stock",)
            elif low > 0 and stock <= low:
                tags = ("low_stock",)

            pid = row_get(r, "id", "")
            try:
                self._product_rows_by_id[int(pid)] = dict(r)
            except Exception:
                pass
            name = row_get(r, "name", "")
            sell_price = row_get(r, "sell_price", 0)
            barcode = row_get(r, "barcode", "")
            location = row_get(r, "location", "")

            try:
                price_str = f"{float(sell_price):.2f}"
            except Exception:
                price_str = "0.00"

            self.prod_tree.insert("", tk.END, iid=str(pid), tags=tags, values=(
                name,
                price_str,
                stock,
                str(barcode or ""),
                str(location or ""),
                low
            ))

    def on_select(self, event):
        sel = self.prod_tree.selection()
        if not sel:
            return
        self.selected_product_id = int(sel[0])
        vals = self.prod_tree.item(sel[0], "values")

        self.edit_name.delete(0, tk.END)
        self.edit_name.insert(0, vals[0])

        self.edit_price.delete(0, tk.END)
        self.edit_price.insert(0, vals[1])

        self.edit_stock.delete(0, tk.END)
        self.edit_stock.insert(0, vals[2])

        self.edit_location.delete(0, tk.END)
        self.edit_location.insert(0, vals[4])

        self.edit_low.delete(0, tk.END)
        self.edit_low.insert(0, vals[5])

        self.selected_barcode = vals[3]
        product = getattr(self, "_product_rows_by_id", {}).get(self.selected_product_id, {})
        
        cat = str(row_get(product, "category", "") or "").strip()
        brand = str(row_get(product, "brand", "") or "").strip()
        self.edit_category.set(cat)
        self.edit_brand.delete(0, tk.END)
        self.edit_brand.insert(0, brand)

        cost = float(row_get(product, "cost_price", 0.0) or 0.0)
        supplier = str(row_get(product, "supplier", "") or "").strip() or "Not set"
        margin = float(vals[1] or 0.0) - cost
        self.selected_detail_lbl.config(
            text=f"Cost: {money(cost)}   |   Supplier: {supplier}   |   Unit margin: {money(margin)}"
        )

    def open_product_details(self):
        if not self.selected_product_id:
            messagebox.showinfo("Select", "Select a product first.")
            return
        product = getattr(self, "_product_rows_by_id", {}).get(self.selected_product_id, {})
        win = tk.Toplevel(self)
        win.title("Product Details")
        win.geometry("720x520")
        win.minsize(560, 420)
        win.configure(bg=UI.CONTENT_BG)
        win.transient(self.winfo_toplevel())
        win.grab_set()

        card = Card(win, padx=16, pady=16)
        card.pack(fill="both", expand=True, padx=12, pady=12)
        tk.Label(card.inner, text=str(row_get(product, "name", "Product")), font=UI.FONT_LG,
                 bg=UI.CARD, fg=UI.TEXT).pack(anchor="w")
        tk.Label(card.inner, text=f"Barcode: {row_get(product, 'barcode', '')}   |   Sell price: {money(row_get(product, 'sell_price', 0))}",
                 bg=UI.CARD, fg=UI.MUTED).pack(anchor="w", pady=(3, 10))

        fields = tk.Frame(card.inner, bg=UI.CARD)
        fields.pack(fill="x")
        tk.Label(fields, text="Cost price (optional)", bg=UI.CARD, fg=UI.TEXT).grid(row=0, column=0, sticky="w")
        cost_var = tk.StringVar(value=(f"{float(row_get(product, 'cost_price', 0) or 0):.2f}"))
        tk.Entry(fields, textvariable=cost_var, width=18, bd=1, relief="solid").grid(row=1, column=0, sticky="w", padx=(0, 12))
        tk.Label(fields, text="Supplier (optional)", bg=UI.CARD, fg=UI.TEXT).grid(row=0, column=1, sticky="w")
        supplier_var = tk.StringVar(value=str(row_get(product, "supplier", "") or ""))
        tk.Entry(fields, textvariable=supplier_var, width=34, bd=1, relief="solid").grid(row=1, column=1, sticky="ew")
        fields.grid_columnconfigure(1, weight=1)

        def load_history():
            tree.delete(*tree.get_children())
            for movement in list_inventory_movements(self.selected_product_id, 500) or []:
                tree.insert("", tk.END, values=(
                    row_get(movement, "created_at", ""), row_get(movement, "movement_type", ""),
                    row_get(movement, "qty_change", 0), row_get(movement, "qty_before", 0),
                    row_get(movement, "qty_after", 0), row_get(movement, "reason", ""),
                ))

        def save_details():
            try:
                cost = max(0.0, float(cost_var.get().strip() or "0"))
            except Exception:
                messagebox.showerror("Invalid", "Cost price must be a number.", parent=win)
                return
            try:
                update_product_details(self.selected_product_id, cost, supplier_var.get().strip())
                self.refresh_list()
                if str(self.selected_product_id) in self.prod_tree.get_children():
                    self.prod_tree.selection_set(str(self.selected_product_id))
                    self.on_select(None)
                messagebox.showinfo("Saved", "Optional product details saved.", parent=win)
            except Exception as exc:
                messagebox.showerror("Could not save", str(exc), parent=win)

        # Pack buttons at the bottom so they are always visible
        buttons = tk.Frame(card.inner, bg=UI.CARD)
        buttons.pack(side="bottom", fill="x", pady=(12, 0))
        PrimaryButton(buttons, "Save Details", save_details).pack(side="left")
        GhostButton(buttons, "Refresh History", load_history).pack(side="left", padx=(8, 0))
        GhostButton(buttons, "Sales", lambda: open_product_sales_history_window(
            win, self.selected_product_id, str(row_get(product, "name", ""))
        )).pack(side="left", padx=(8, 0))
        GhostButton(buttons, "Price Changes", lambda: open_product_price_history_window(
            win, self.selected_product_id, str(row_get(product, "name", ""))
        )).pack(side="left", padx=(8, 0))
        GhostButton(buttons, "Close", win.destroy).pack(side="right")

        tk.Label(card.inner, text="Inventory movement history", font=UI.FONT_MD,
                 bg=UI.CARD, fg=UI.TEXT).pack(anchor="w", pady=(16, 6))

        cols = ("time", "type", "change", "before", "after", "reason")
        tree_wrap = tk.Frame(card.inner, bg=UI.CARD)
        tree_wrap.pack(fill="both", expand=True)
        tree = ttk.Treeview(tree_wrap, columns=cols, show="headings", height=6)
        for col, title, width in [
            ("time", "Time", 140), ("type", "Type", 110), ("change", "+/-", 55),
            ("before", "Before", 60), ("after", "After", 60), ("reason", "Reason", 220),
        ]:
            tree.heading(col, text=title)
            tree.column(col, width=width, anchor=("center" if col in {"change", "before", "after"} else "w"))
        vsb = ttk.Scrollbar(tree_wrap, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        load_history()

    def add_selected_to_cart(self):
        if not self.selected_barcode:
            messagebox.showinfo("Select", "Select a product first.")
            return
        self.on_product_added(self.selected_barcode)

    def save_changes(self):
        if not self.selected_product_id:
            messagebox.showinfo("Select", "Select a product first.")
            return
        try:
            name = self.edit_name.get().strip()
            price = float(self.edit_price.get().strip())
            stock = int(float(self.edit_stock.get().strip()))
            low = int(float(self.edit_low.get().strip() or "0"))
            location = self.edit_location.get().strip()
            category = self.edit_category.get().strip()
            brand = self.edit_brand.get().strip()
        except Exception:
            messagebox.showerror("Invalid", "Check name, price, stock, location, low values.")
            return

        update_product(self.selected_product_id, name, price, stock, low, location, category, brand)
        messagebox.showinfo("Saved", "Product updated.")
        self.refresh_list()

    def delete_selected(self):
        if not self.selected_product_id:
            messagebox.showinfo("Select", "Select a product first.")
            return

        name = self.edit_name.get().strip() or "this product"
        if not messagebox.askyesno("Delete product",
                                   f"Delete {name}?\n\nThis will hide it from the POS. Sales history stays."):
            return

        delete_product(self.selected_product_id)
        self.selected_product_id = None
        self.selected_barcode = ""
        messagebox.showinfo("Deleted", "Product deleted.")
        self.refresh_list()

    def _selected_label_payload(self, qty: int = 1) -> list[dict]:
        if not self.selected_product_id or not self.selected_barcode:
            return []
        try:
            qty = max(1, int(float(qty or 1)))
        except Exception:
            qty = 1
        try:
            price = float(self.edit_price.get().strip() or 0)
        except Exception:
            price = 0.0
        return [{
            "name": self.edit_name.get().strip(),
            "price": price,
            "barcode": str(self.selected_barcode or "").strip(),
            "location": self.edit_location.get().strip() if hasattr(self, "edit_location") else "",
            "qty": qty,
        }]

    def send_selected_label_to_host(self):
        labels = self._selected_label_payload(1)
        if not labels:
            messagebox.showinfo("Select", "Select a product first.")
            return
        ok, msg = send_barcode_labels_to_host(labels, title=f"{labels[0]['name']} Label")
        if ok:
            messagebox.showinfo("Sent to Host", "Label print job was sent to Host mode.")
        else:
            messagebox.showwarning("Not sent", f"Could not send label to Host.\n\n{msg}")

    def add_clicked(self, print_stock: bool = False, send_host_print: bool = False):
        name = self.name_e.get().strip()
        if not name:
            messagebox.showerror("Missing", "Product name is required.")
            return

        try:
            price = float(self.price_e.get().strip())
        except Exception:
            messagebox.showerror("Invalid", "Sell price must be a number.")
            return

        try:
            stock = int(float(self.stock_e.get().strip() or "0"))
        except Exception:
            messagebox.showerror("Invalid", "Stock qty must be a number.")
            return

        try:
            low = int(float(self.low_e.get().strip() or "0"))
        except Exception:
            messagebox.showerror("Invalid", "Low stock level must be a number.")
            return

        location = self.location_e.get().strip() if hasattr(self, "location_e") else ""
        try:
            cost_price = max(0.0, float(self.cost_e.get().strip() or "0"))
        except Exception:
            messagebox.showerror("Invalid", "Cost price must be a number or left blank.")
            return
        supplier = self.supplier_e.get().strip() if hasattr(self, "supplier_e") else ""

        manual_barcode_raw = ""
        try:
            manual_barcode_raw = self.barcode_e.get().strip()
        except Exception:
            manual_barcode_raw = ""
        manual_barcode = ""
        if manual_barcode_raw:
            manual_barcode = _digits_only(manual_barcode_raw)
            if not manual_barcode:
                messagebox.showerror("Invalid", "Barcode must contain digits only.")
                self.barcode_e.focus()
                return
            if len(manual_barcode) > 13:
                messagebox.showerror("Invalid", "Barcode must be 13 digits or fewer.")
                self.barcode_e.focus()
                return

        category = self.category_cb.get().strip()
        brand = self.brand_e.get().strip()

        barcode = add_product(
            name=name,
            category=category,
            brand=brand,
            sell_price=price,
            stock_qty=stock,
            low_stock_level=low,
            barcode=(manual_barcode or None),
            location=location,
            cost_price=cost_price,
            supplier=supplier,
        )

        barcode_display = str(barcode or "")

        printed_ok = False
        host_sent = False
        host_msg = ""
        if print_stock and stock > 0:
            labels = [{
                "name": name,
                "price": price,
                "barcode": barcode,
                "location": location,
                "qty": stock,
            }]
            printed_ok = bool(print_configured_barcodes(labels, title=f"{name} Labels"))
        elif send_host_print and stock > 0:
            labels = [{
                "name": name,
                "price": price,
                "barcode": barcode,
                "location": location,
                "qty": stock,
            }]
            host_sent, host_msg = send_barcode_labels_to_host(labels, title=f"{name} Labels")

        if send_host_print:
            if stock <= 0:
                self.last_barcode.set(f"Added {name}    Barcode {barcode_display}    No host print sent (stock is 0)")
            elif host_sent:
                self.last_barcode.set(f"Added {name}    Barcode {barcode_display}    Sent {stock} label(s) to Host")
            else:
                self.last_barcode.set(f"Added {name}    Barcode {barcode_display}    Host print not sent: {host_msg}")
        elif print_stock:
            if stock <= 0:
                self.last_barcode.set(f"Added {name}    Barcode {barcode_display}    No labels printed (stock is 0)")
            elif printed_ok:
                self.last_barcode.set(f"Added {name}    Barcode {barcode_display}    Printed {stock} label(s)")
            else:
                self.last_barcode.set(f"Added {name}    Barcode {barcode_display}    Print failed / printer not set")
        else:
            self.last_barcode.set(f"Added {name}    Barcode {barcode_display}")

        self.refresh_list()

        self.name_e.delete(0, tk.END)
        self.barcode_e.delete(0, tk.END)
        self.location_e.delete(0, tk.END)
        self.cost_e.delete(0, tk.END)
        self.supplier_e.delete(0, tk.END)
        self.price_e.delete(0, tk.END)
        self.stock_e.delete(0, tk.END)
        self.low_e.delete(0, tk.END)
        self.category_cb.set("")
        self.brand_e.delete(0, tk.END)
        self.low_e.insert(0, "0")
        self.name_e.focus()


# ---------------- BARCODES PAGE ----------------

class BarcodesPage(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg=UI.CONTENT_BG)
        self._build()
        self.refresh()

    def _build(self):
        scroll = VScrollableFrame(self, bg=UI.CONTENT_BG)
        scroll.pack(fill="both", expand=True)
        wrap = tk.Frame(scroll.inner, bg=UI.CONTENT_BG)
        wrap.pack(fill="both", expand=True, padx=(10 if UI.COMPACT else 18), pady=(10 if UI.COMPACT else 18))

        header = Card(wrap, padx=18, pady=14)
        header.pack(fill="x")
        HeaderBar(header.inner, "Barcodes", "Select a product, set quantity, print labels, and replenish stock.").pack(fill="x")

        body = Card(wrap, padx=14, pady=14)
        body.pack(fill="both", expand=True, pady=((8 if UI.COMPACT else 14), 0))

        top = tk.Frame(body.inner, bg=UI.CARD)
        top.pack(fill="x")
        GhostButton(top, "Send to Host + Add Stock", self.send_selected_to_host).pack(side="left", padx=(10, 10))
        GhostButton(top, "Refresh", self.refresh).pack(side="left")

        # Search (name / barcode / category / brand)
        self.search_var = tk.StringVar(value="")
        tk.Label(top, text="Search:", bg=UI.CARD, fg=UI.TEXT).pack(side="left", padx=(18, 6))
        self.search_entry = tk.Entry(top, textvariable=self.search_var, width=28, bd=1, relief="solid")
        self.search_entry.pack(side="left", padx=(0, 8))
        GhostButton(top, "Search", self.search_clicked).pack(side="left")
        GhostButton(top, "Clear", self.clear_search).pack(side="left", padx=(8, 0))
        try:
            self.search_entry.bind("<Return>", lambda e: self.search_clicked())
        except Exception:
            pass

        cols = ("name", "price", "barcode", "location", "stock", "qty")
        self.tree = ttk.Treeview(body.inner, columns=cols, show="headings", height=14, selectmode="extended")
        self.tree.heading("name", text="Name")
        self.tree.heading("price", text="Price")
        self.tree.heading("barcode", text="Barcode")
        self.tree.heading("location", text="Location")
        self.tree.heading("stock", text="Stock")
        self.tree.heading("qty", text="Qty")
        self.tree.column("name", width=300)
        self.tree.column("price", width=100, anchor="e")
        self.tree.column("barcode", width=170)
        self.tree.column("location", width=170)
        self.tree.column("stock", width=80, anchor="center")
        self.tree.column("qty", width=80, anchor="center")
        self.tree.pack(fill="both", expand=True, pady=(12, 0))

        qty_row = tk.Frame(body.inner, bg=UI.CARD)
        qty_row.pack(fill="x", pady=(12, 0))

        tk.Label(qty_row, text="New labels / stock to add", bg=UI.CARD, fg="#334155").pack(side="left")
        self.qty_e = tk.Entry(qty_row, width=10, bd=1, relief="solid")
        self.qty_e.insert(0, "1")
        self.qty_e.pack(side="left", padx=8)
        self.qty_e.bind("<Return>", self._qty_entry_return)
        PrimaryButton(qty_row, "Print Labels + Add Stock", self.generate_pdf_selected).pack(side="left", padx=(4, 0))
        GhostButton(qty_row, "Print Only", self.print_selected_only).pack(side="left", padx=(8, 0))
        GhostButton(qty_row, "Send Host Only", self.send_selected_to_host_only).pack(side="left", padx=(8, 0))
        GhostButton(qty_row, "Stop Label Printer", self.stop_label_printer).pack(side="left", padx=(8, 0))
        GhostButton(qty_row, "Details", self.open_product_details).pack(side="right")

        self.selected_detail_lbl = tk.Label(
            body.inner, text="Select an item for cost, supplier, and stock history.",
            bg=UI.CARD, fg=UI.MUTED, anchor="w", justify="left"
        )
        self.selected_detail_lbl.pack(fill="x", pady=(8, 0))

        self.tree.bind("<<TreeviewSelect>>", self.on_select)

    def refresh(self):
        """Reload product list (honors current search query if present)."""
        q = ""
        try:
            q = (self.search_var.get() or "").strip()
        except Exception:
            q = ""
        self.tree.delete(*self.tree.get_children())
        rows = list_products(q)
        self._product_rows_by_id = {}
        for r in rows:
            pid = row_get(r, "id", "")
            try:
                self._product_rows_by_id[int(pid)] = dict(r)
            except Exception:
                pass
            self.tree.insert("", tk.END, iid=str(pid), values=(
                row_get(r, "name", ""),
                f"{float(row_get(r, 'sell_price', 0) or 0):.2f}",
                row_get(r, "barcode", ""),
                row_get(r, "location", ""),
                int(row_get(r, "stock_qty", 0) or 0),
                1,
            ))

    def search_clicked(self):
        self.refresh()

    def clear_search(self):
        try:
            self.search_var.set("")
        except Exception:
            pass
        self.refresh()

    def _qty_entry_return(self, event=None):
        try:
            top = self.winfo_toplevel()
            pending = str(getattr(top, "_scan_buf", "") or "").strip()
            if pending and hasattr(top, "_looks_like_scanner_code") and top._looks_like_scanner_code(pending):
                return None
        except Exception:
            pass
        self.generate_pdf_selected()
        return "break"

    def apply_qty(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Select", "Select a product row first.")
            return
        try:
            q = int(self.qty_e.get().strip())
            if q < 1:
                q = 1
        except Exception:
            messagebox.showerror("Invalid", "Qty must be a number.")
            return

        iid = sel[0]
        vals = list(self.tree.item(iid, "values"))
        vals[5] = q
        self.tree.item(iid, values=vals)

    def generate_pdf_selected(self):
        picked = self._selected_labels()
        if not picked:
            return
        product_id, qty, labels = picked

        # Try direct print to configured barcode printer
        ok = False
        try:
            ok = bool(print_configured_barcodes(labels, title="Mask POS Labels"))
        except Exception:
            ok = False

        if ok:
            if self._replenish_selected_stock(product_id, qty):
                self._reset_qty_after_action(product_id)
                messagebox.showinfo("Printed", f"Barcode labels were sent and stock increased by {qty}.")
            else:
                messagebox.showwarning("Stock", "Labels were sent, but stock was not increased. Refresh and try stock update again.")
            return

        messagebox.showwarning(
            "Not printed",
            "The label was not sent to the barcode printer.\n\n"
            "Check Settings > Barcode Printer and use Test Label."
        )

    def print_selected_only(self):
        picked = self._selected_labels()
        if not picked:
            return
        product_id, qty, labels = picked

        ok = False
        try:
            ok = bool(print_configured_barcodes(labels, title="Mask POS Labels"))
        except Exception:
            ok = False

        if ok:
            self._reset_qty_after_action(product_id)
            messagebox.showinfo("Printed", f"Printed {qty} barcode label(s). Stock was not changed.")
            return

        messagebox.showwarning(
            "Not printed",
            "The label was not sent to the barcode printer.\n\n"
            "Check Settings > Barcode Printer and use Test Label."
        )

    def send_selected_to_host(self):
        picked = self._selected_labels()
        if not picked:
            return
        product_id, qty, labels = picked
        ok, msg = send_barcode_labels_to_host(labels, title="Mask POS Labels")
        if ok:
            if self._replenish_selected_stock(product_id, qty):
                self._reset_qty_after_action(product_id)
                messagebox.showinfo("Sent to Host", f"Label print job was sent and stock increased by {qty}.")
            else:
                messagebox.showwarning("Stock", "Label job was sent, but stock was not increased. Refresh and try stock update again.")
        else:
            messagebox.showwarning("Not sent", f"Could not send label to Host.\n\n{msg}")

    def send_selected_to_host_only(self):
        picked = self._selected_labels()
        if not picked:
            return
        product_id, qty, labels = picked
        ok, msg = send_barcode_labels_to_host(labels, title="Mask POS Labels")
        if ok:
            self._reset_qty_after_action(product_id)
            messagebox.showinfo("Sent to Host", f"Sent {qty} barcode label(s) to Host. Stock was not changed.")
        else:
            messagebox.showwarning("Not sent", f"Could not send label to Host.\n\n{msg}")

    def stop_label_printer(self):
        cfg = get_barcode_printer_config()
        prn = str(cfg.get("barcode_printer_name") or "").strip()
        if not prn:
            messagebox.showwarning("Stop Label Printer", "Select a barcode printer in Settings first.")
            return
        if not messagebox.askyesno(
            "Stop Label Printer",
            "This will cancel pending barcode label jobs and reset the Windows print queue. Continue?",
        ):
            return
        ok = False
        try:
            ok = bool(hard_reset_printing(prn))
        except Exception:
            ok = False
        if not ok:
            try:
                ok = bool(clear_printer_queue(prn))
            except Exception:
                ok = False
        if ok:
            messagebox.showinfo("Stop Label Printer", "Barcode printer queue was cleared/reset.")
        else:
            messagebox.showwarning("Stop Label Printer", "Could not clear the barcode printer queue. Try running as Administrator.")

    def _selected_labels(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Select", "Select a product row first.")
            return []

        iid = sel[0]
        name, price, barcode, location, stock, qty = self.tree.item(iid, "values")

        try:
            qty = int(self.qty_e.get().strip())
        except Exception:
            try:
                qty = int(qty)
            except Exception:
                qty = 1
        if qty < 1:
            qty = 1
        vals = list(self.tree.item(iid, "values"))
        vals[5] = qty
        self.tree.item(iid, values=vals)

        labels = [{
            "name": name,
            "price": float(price),
            "barcode": str(barcode),
            "location": str(location or ""),
            "qty": qty
        }]
        return int(iid), qty, labels

    def _reset_qty_after_action(self, product_id: int | None = None):
        try:
            self.qty_e.delete(0, tk.END)
        except Exception:
            pass
        try:
            target = str(product_id) if product_id is not None else (self.tree.selection()[0] if self.tree.selection() else "")
            if target:
                vals = list(self.tree.item(target, "values"))
                if len(vals) >= 6:
                    vals[5] = 1
                    self.tree.item(target, values=vals)
        except Exception:
            pass

    def _replenish_selected_stock(self, product_id: int, qty: int) -> bool:
        try:
            ok = bool(adjust_stock(
                int(product_id), int(qty),
                reason="Stock received / labels printed", movement_type="STOCK_IN",
            ))
        except Exception:
            ok = False
        if ok:
            try:
                self.refresh()
                self.tree.selection_set(str(product_id))
                self.tree.focus(str(product_id))
                self.tree.see(str(product_id))
            except Exception:
                pass
        return ok

    def on_select(self, event):
        sel = self.tree.selection()
        if not sel:
            return
        try:
            self.selected_product_id = int(sel[0])
            vals = self.tree.item(sel[0], "values")
            product = getattr(self, "_product_rows_by_id", {}).get(self.selected_product_id, {})
            cost = float(row_get(product, "cost_price", 0.0) or 0.0)
            supplier = str(row_get(product, "supplier", "") or "").strip() or "Not set"
            margin = float(vals[1] or 0.0) - cost
            self.selected_detail_lbl.config(
                text=f"Cost: {money(cost)}   |   Supplier: {supplier}   |   Unit margin: {money(margin)}"
            )
        except Exception:
            pass

    def open_product_details(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Select", "Select a product first.")
            return
        self.selected_product_id = int(sel[0])
        product = getattr(self, "_product_rows_by_id", {}).get(self.selected_product_id, {})
        win = tk.Toplevel(self)
        win.title("Product Details")
        win.geometry("720x520")
        win.minsize(560, 420)
        win.configure(bg=UI.CONTENT_BG)
        win.transient(self.winfo_toplevel())
        win.grab_set()

        card = Card(win, padx=16, pady=16)
        card.pack(fill="both", expand=True, padx=12, pady=12)
        tk.Label(card.inner, text=str(row_get(product, "name", "Product")), font=UI.FONT_LG,
                 bg=UI.CARD, fg=UI.TEXT).pack(anchor="w")
        tk.Label(card.inner, text=f"Barcode: {row_get(product, 'barcode', '')}   |   Sell price: {money(row_get(product, 'sell_price', 0))}",
                 bg=UI.CARD, fg=UI.MUTED).pack(anchor="w", pady=(3, 10))

        fields = tk.Frame(card.inner, bg=UI.CARD)
        fields.pack(fill="x")
        tk.Label(fields, text="Cost price (optional)", bg=UI.CARD, fg=UI.TEXT).grid(row=0, column=0, sticky="w")
        cost_var = tk.StringVar(value=(f"{float(row_get(product, 'cost_price', 0) or 0):.2f}"))
        tk.Entry(fields, textvariable=cost_var, width=18, bd=1, relief="solid").grid(row=1, column=0, sticky="w", padx=(0, 12))
        tk.Label(fields, text="Supplier (optional)", bg=UI.CARD, fg=UI.TEXT).grid(row=0, column=1, sticky="w")
        supplier_var = tk.StringVar(value=str(row_get(product, "supplier", "") or ""))
        tk.Entry(fields, textvariable=supplier_var, width=34, bd=1, relief="solid").grid(row=1, column=1, sticky="ew")
        fields.grid_columnconfigure(1, weight=1)

        def load_history():
            tree.delete(*tree.get_children())
            for movement in list_inventory_movements(self.selected_product_id, 500) or []:
                tree.insert("", tk.END, values=(
                    row_get(movement, "created_at", ""), row_get(movement, "movement_type", ""),
                    row_get(movement, "qty_change", 0), row_get(movement, "qty_before", 0),
                    row_get(movement, "qty_after", 0), row_get(movement, "reason", ""),
                ))

        def save_details():
            try:
                cost = max(0.0, float(cost_var.get().strip() or "0"))
            except Exception:
                messagebox.showerror("Invalid", "Cost price must be a number.", parent=win)
                return
            try:
                update_product_details(self.selected_product_id, cost, supplier_var.get().strip())
                self.refresh()
                if str(self.selected_product_id) in self.tree.get_children():
                    self.tree.selection_set(str(self.selected_product_id))
                    self.on_select(None)
                messagebox.showinfo("Saved", "Optional product details saved.", parent=win)
            except Exception as exc:
                messagebox.showerror("Could not save", str(exc), parent=win)

        # Pack buttons at the bottom so they are always visible
        buttons = tk.Frame(card.inner, bg=UI.CARD)
        buttons.pack(side="bottom", fill="x", pady=(12, 0))
        PrimaryButton(buttons, "Save Details", save_details).pack(side="left")
        GhostButton(buttons, "Refresh History", load_history).pack(side="left", padx=(8, 0))
        GhostButton(buttons, "Sales", lambda: open_product_sales_history_window(
            win, self.selected_product_id, str(row_get(product, "name", ""))
        )).pack(side="left", padx=(8, 0))
        GhostButton(buttons, "Price Changes", lambda: open_product_price_history_window(
            win, self.selected_product_id, str(row_get(product, "name", ""))
        )).pack(side="left", padx=(8, 0))
        GhostButton(buttons, "Close", win.destroy).pack(side="right")

        tk.Label(card.inner, text="Inventory movement history", font=UI.FONT_MD,
                 bg=UI.CARD, fg=UI.TEXT).pack(anchor="w", pady=(16, 6))

        cols = ("time", "type", "change", "before", "after", "reason")
        tree_wrap = tk.Frame(card.inner, bg=UI.CARD)
        tree_wrap.pack(fill="both", expand=True)
        tree = ttk.Treeview(tree_wrap, columns=cols, show="headings", height=6)
        for col, title, width in [
            ("time", "Time", 140), ("type", "Type", 110), ("change", "+/-", 55),
            ("before", "Before", 60), ("after", "After", 60), ("reason", "Reason", 220),
        ]:
            tree.heading(col, text=title)
            tree.column(col, width=width, anchor=("center" if col in {"change", "before", "after"} else "w"))
        vsb = ttk.Scrollbar(tree_wrap, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")
        load_history()


# ---------------- ANALYTICS PAGE ----------------

class AnalyticsPage(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg=UI.CONTENT_BG)
        self.period_var = tk.StringVar(value="Today")
        self.start_var = tk.StringVar(value="")
        self.end_var = tk.StringVar(value="")
        self.product_sort_var = tk.StringVar(value="Revenue")
        self.product_search_var = tk.StringVar(value="")
        self.product_rows = []
        self._build()
        self.apply_period()

    def _build(self):
        wrap = VScrollableFrame(self, bg=UI.CONTENT_BG)
        wrap.pack(fill="both", expand=True)
        content = tk.Frame(wrap.inner, bg=UI.CONTENT_BG)
        content.pack(fill="both", expand=True, padx=18, pady=18)

        header = Card(content, padx=18, pady=14)
        header.pack(fill="x")

        topbar = tk.Frame(header.inner, bg=UI.CARD)
        topbar.pack(fill="x")

        tk.Label(topbar, text="Analytics", font=UI.FONT_XL, bg=UI.CARD, fg=UI.TEXT).pack(side="left")

        self.last_refresh_lbl = tk.Label(topbar, text="", font=UI.FONT_SM, bg=UI.CARD, fg=UI.MUTED)
        self.last_refresh_lbl.pack(side="left", padx=12)

        controls = tk.Frame(header.inner, bg=UI.CARD)
        controls.pack(fill="x", pady=(12, 0))

        self.period_cb = ttk.Combobox(
            controls,
            textvariable=self.period_var,
            values=["Today", "This week", "This month", "This year", "All time", "Custom"],
            state="readonly",
            width=12
        )
        self.period_cb.pack(side="left")
        self.period_cb.bind("<<ComboboxSelected>>", lambda e: self.apply_period())

        self.start_e = tk.Entry(controls, textvariable=self.start_var, width=12, bd=1, relief="solid")
        self.start_e.pack(side="left", padx=8)

        self.end_e = tk.Entry(controls, textvariable=self.end_var, width=12, bd=1, relief="solid")
        self.end_e.pack(side="left", padx=(0, 10))

        GhostButton(controls, "Refresh", self.refresh).pack(side="left")
        GhostButton(controls, "Discount Impact", self.open_discount_impact).pack(side="left", padx=(8, 0))

        self.start_e.bind("<Return>", lambda e: self.refresh())
        self.end_e.bind("<Return>", lambda e: self.refresh())

        # ---------------- Reports (Excel) ----------------
        # Exports are copies only. Nothing is deleted from the database.
        today = datetime.now()
        self.rep_year = tk.IntVar(value=today.year)
        self.rep_month = tk.IntVar(value=today.month)
        self.rep_day = tk.StringVar(value="All")

        report_card = Card(content, padx=18, pady=14)
        report_card.pack(fill="x", pady=(12, 0))

        tk.Label(report_card.inner, text="Reports", font=UI.FONT_LG, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w")
        tk.Label(
            report_card.inner,
            text="Export sales and inventory to Excel. This never deletes anything.",
            font=UI.FONT_SM,
            bg=UI.CARD,
            fg=UI.MUTED
        ).pack(anchor="w", pady=(2, 10))

        row = tk.Frame(report_card.inner, bg=UI.CARD)
        row.pack(fill="x")

        years = [str(y) for y in range(max(2020, today.year - 5), today.year + 6)]
        months = [(i, calendar.month_name[i]) for i in range(1, 13)]
        month_labels = [m[1] for m in months]
        days = ["All"] + [str(i) for i in range(1, 32)]

        tk.Label(row, text="Year", bg=UI.CARD, fg=UI.MUTED).pack(side="left", padx=(0, 6))
        year_cb = ttk.Combobox(row, values=years, width=8, state="readonly")
        year_cb.set(str(self.rep_year.get()))
        year_cb.pack(side="left", padx=(0, 12))

        tk.Label(row, text="Month", bg=UI.CARD, fg=UI.MUTED).pack(side="left", padx=(0, 6))
        month_cb = ttk.Combobox(row, values=month_labels, width=12, state="readonly")
        month_cb.set(calendar.month_name[self.rep_month.get()])
        month_cb.pack(side="left", padx=(0, 12))

        tk.Label(row, text="Day", bg=UI.CARD, fg=UI.MUTED).pack(side="left", padx=(0, 6))
        day_cb = ttk.Combobox(row, values=days, width=6, state="readonly")
        day_cb.set(self.rep_day.get())
        day_cb.pack(side="left", padx=(0, 12))

        def _sync_report_vars_from_widgets():
            try:
                self.rep_year.set(int(year_cb.get()))
            except Exception:
                self.rep_year.set(today.year)

            m_label = (month_cb.get() or "").strip()
            m_num = self.rep_month.get()
            for n, lab in months:
                if lab == m_label:
                    m_num = int(n)
                    break
            self.rep_month.set(m_num)
            self.rep_day.set((day_cb.get() or "All").strip() or "All")

        GhostButton(
            row,
            "Sales Report (Excel)",
            lambda: (_sync_report_vars_from_widgets(),
                     self._export_sales_report_excel(self.rep_year.get(), self.rep_month.get(), self.rep_day.get()))
        ).pack(side="left", padx=(8, 8))

        GhostButton(
            row,
            "Inventory Report (Excel)",
            lambda: self._export_inventory_report_excel()
        ).pack(side="left", padx=(0, 8))

        GhostButton(
            row,
            "Sales PDF (Day)",
            lambda: (_sync_report_vars_from_widgets(),
                     self._export_sales_report_pdf(self.rep_year.get(), self.rep_month.get(), self.rep_day.get()))
        ).pack(side="left", padx=(0, 8))

        GhostButton(
            row,
            "Sales PDF (Range)",
            self._export_sales_report_pdf_range
        ).pack(side="left", padx=(0, 8))

        GhostButton(
            row,
            "Email Today",
            lambda: self.winfo_toplevel().send_daily_report_email_for_day(
                datetime.now().strftime("%Y-%m-%d"),
                source="manual",
                silent=False,
                force=True,
            )
        ).pack(side="left", padx=(0, 8))

        GhostButton(
            row,
            "Open Reports Folder",
            self._open_reports_folder
        ).pack(side="left")

        self.report_status_lbl = tk.Label(report_card.inner, text="", bg=UI.CARD, fg=UI.MUTED)
        self.report_status_lbl.pack(anchor="w", pady=(8, 0))

        self.range_lbl = tk.Label(controls, text="", bg=UI.CARD, fg=UI.MUTED)
        self.range_lbl.pack(side="right")

        kpi_row = tk.Frame(content, bg=UI.CONTENT_BG)
        kpi_row.pack(fill="x", pady=(14, 10))
        for i in range(4):
            kpi_row.grid_columnconfigure(i, weight=1)

        self.kpi_gross = self._kpi_card(kpi_row, 0, "Total sales", "$0.00")
        self.kpi_cash = self._kpi_card(kpi_row, 1, "Cash collected", "$0.00")
        self.kpi_items = self._kpi_card(kpi_row, 2, "Net items sold", "0")
        self.kpi_aov = self._kpi_card(kpi_row, 3, "Average order value", "$0.00")

        mid = tk.Frame(content, bg=UI.CONTENT_BG)
        mid.pack(fill="both", expand=True)
        mid.grid_columnconfigure(0, weight=3)
        mid.grid_columnconfigure(1, weight=2)
        mid.grid_rowconfigure(0, weight=1)

        self.card_sales_over_time = Card(mid, padx=14, pady=14)
        self.card_sales_over_time.grid(row=0, column=0, sticky="nsew", padx=(0, 10))

        tk.Label(self.card_sales_over_time.inner, text="Total sales over time", font=UI.FONT_LG,
                 bg=UI.CARD, fg=UI.TEXT).pack(anchor="w")

        self.sales_chart = self._make_chart(self.card_sales_over_time.inner)

        self.card_breakdown = Card(mid, padx=14, pady=14)
        self.card_breakdown.grid(row=0, column=1, sticky="nsew", padx=(10, 0))

        tk.Label(self.card_breakdown.inner, text="Total sales breakdown", font=UI.FONT_LG,
                 bg=UI.CARD, fg=UI.TEXT).pack(anchor="w")

        self.breakdown_box = tk.Frame(self.card_breakdown.inner, bg=UI.CARD)
        self.breakdown_box.pack(fill="both", expand=True, pady=(10, 0))

        tk.Label(self.card_breakdown.inner, text="Payment Methods", font=UI.FONT_MD,
                 bg=UI.CARD, fg=UI.MUTED).pack(anchor="w", pady=(14, 4))
        self.pm_breakdown_box = tk.Frame(self.card_breakdown.inner, bg=UI.CARD)
        self.pm_breakdown_box.pack(fill="both", expand=True)

        tk.Label(self.card_breakdown.inner, text="Sales by Category", font=UI.FONT_MD,
                 bg=UI.CARD, fg=UI.MUTED).pack(anchor="w", pady=(14, 4))
        self.cat_breakdown_box = tk.Frame(self.card_breakdown.inner, bg=UI.CARD)
        self.cat_breakdown_box.pack(fill="both", expand=True)

        bottom = tk.Frame(content, bg=UI.CONTENT_BG)
        bottom.pack(fill="both", expand=True, pady=(10, 0))
        bottom.grid_columnconfigure(0, weight=2)
        bottom.grid_columnconfigure(1, weight=3)
        bottom.grid_rowconfigure(0, weight=1)

        self.card_aov = Card(bottom, padx=14, pady=14)
        self.card_aov.grid(row=0, column=0, sticky="nsew", padx=(0, 10))

        tk.Label(self.card_aov.inner, text="Average order value over time", font=UI.FONT_LG,
                 bg=UI.CARD, fg=UI.TEXT).pack(anchor="w")

        self.aov_chart = self._make_chart(self.card_aov.inner)

        self.card_top_sellers = Card(bottom, padx=14, pady=14)
        self.card_top_sellers.grid(row=0, column=1, sticky="nsew", padx=(10, 0))

        tk.Label(self.card_top_sellers.inner, text="Top sellers", font=UI.FONT_LG,
                 bg=UI.CARD, fg=UI.TEXT).pack(anchor="w")
        self.top_sellers_canvas = tk.Canvas(
            self.card_top_sellers.inner,
            height=220,
            bg=UI.CARD,
            highlightthickness=0,
        )
        self.top_sellers_canvas.pack(fill="both", expand=True, pady=(10, 0))
        self.top_sellers_canvas.bind("<Configure>", lambda e: self._draw_top_sellers())

        self.card_by_product = Card(content, padx=14, pady=14)
        self.card_by_product.pack(fill="both", expand=True, pady=(10, 0))

        prod_head = tk.Frame(self.card_by_product.inner, bg=UI.CARD)
        prod_head.pack(fill="x", pady=(0, 10))
        tk.Label(prod_head, text="Product performance", font=UI.FONT_LG,
                 bg=UI.CARD, fg=UI.TEXT).pack(side="left")
        self.product_summary_lbl = tk.Label(prod_head, text="", font=UI.FONT_SM, bg=UI.CARD, fg=UI.MUTED)
        self.product_summary_lbl.pack(side="left", padx=12)

        prod_tools = tk.Frame(self.card_by_product.inner, bg=UI.CARD)
        prod_tools.pack(fill="x", pady=(0, 10))
        tk.Label(prod_tools, text="Sort", bg=UI.CARD, fg=UI.MUTED).pack(side="left", padx=(0, 6))
        sort_cb = ttk.Combobox(
            prod_tools,
            textvariable=self.product_sort_var,
            values=["Revenue", "Qty sold", "Net qty", "Name"],
            state="readonly",
            width=12,
        )
        sort_cb.pack(side="left", padx=(0, 12))
        sort_cb.bind("<<ComboboxSelected>>", lambda e: self._render_product_rows())

        tk.Label(prod_tools, text="Search", bg=UI.CARD, fg=UI.MUTED).pack(side="left", padx=(0, 6))
        search_e = tk.Entry(prod_tools, textvariable=self.product_search_var, width=28, bd=1, relief="solid")
        search_e.pack(side="left")
        search_e.bind("<KeyRelease>", lambda e: self._render_product_rows())
        GhostButton(prod_tools, "Clear", lambda: (self.product_search_var.set(""), self._render_product_rows())).pack(side="left", padx=(8, 0))

        cols = ("rank", "name", "sold", "returned", "net_qty", "revenue", "avg", "share", "stock", "barcode", "category")
        table_wrap = tk.Frame(self.card_by_product.inner, bg=UI.CARD)
        table_wrap.pack(fill="both", expand=True)
        table_wrap.grid_rowconfigure(0, weight=1)
        table_wrap.grid_columnconfigure(0, weight=1)

        self.prod_tree = ttk.Treeview(table_wrap, columns=cols, show="headings", height=14)
        headings = {
            "rank": "#",
            "name": "Product",
            "sold": "Sold",
            "returned": "Returned",
            "net_qty": "Net",
            "revenue": "Revenue",
            "avg": "Avg",
            "share": "% Sales",
            "stock": "Stock",
            "barcode": "Barcode",
            "category": "Category",
        }
        widths = {
            "rank": 45,
            "name": 260,
            "sold": 70,
            "returned": 80,
            "net_qty": 70,
            "revenue": 110,
            "avg": 90,
            "share": 80,
            "stock": 70,
            "barcode": 130,
            "category": 130,
        }
        numeric_cols = {"rank", "sold", "returned", "net_qty", "revenue", "avg", "share", "stock"}
        for col in cols:
            self.prod_tree.heading(col, text=headings[col])
            self.prod_tree.column(col, width=widths[col], minwidth=45, anchor=("e" if col in numeric_cols else "w"), stretch=(col in {"name", "category"}))
        prod_vsb = ttk.Scrollbar(table_wrap, orient="vertical", command=self.prod_tree.yview)
        prod_hsb = ttk.Scrollbar(table_wrap, orient="horizontal", command=self.prod_tree.xview)
        self.prod_tree.configure(yscrollcommand=prod_vsb.set, xscrollcommand=prod_hsb.set)
        self.prod_tree.grid(row=0, column=0, sticky="nsew")
        prod_vsb.grid(row=0, column=1, sticky="ns")
        prod_hsb.grid(row=1, column=0, sticky="ew")

        health_card = Card(content, padx=14, pady=14)
        health_card.pack(fill="both", expand=True, pady=(10, 0))
        health_head = tk.Frame(health_card.inner, bg=UI.CARD)
        health_head.pack(fill="x")
        tk.Label(health_head, text="Data Health", font=UI.FONT_LG, bg=UI.CARD, fg=UI.TEXT).pack(side="left")
        self.data_health_summary_lbl = tk.Label(health_head, text="", font=UI.FONT_SM, bg=UI.CARD, fg=UI.MUTED)
        self.data_health_summary_lbl.pack(side="left", padx=12)
        GhostButton(health_head, "Run Checks", self._refresh_data_health).pack(side="right")
        GhostButton(health_head, "Go to Products", self._go_products_for_health).pack(side="right", padx=(0, 8))
        tk.Label(
            health_card.inner,
            text="Only real integrity problems are checked. Blank location and low-stock settings are intentionally ignored.",
            bg=UI.CARD, fg=UI.MUTED, font=UI.FONT_SM,
        ).pack(anchor="w", pady=(4, 8))

        health_wrap = tk.Frame(health_card.inner, bg=UI.CARD)
        health_wrap.pack(fill="both", expand=True)
        health_wrap.grid_columnconfigure(0, weight=2)
        health_wrap.grid_columnconfigure(1, weight=3)
        health_wrap.grid_rowconfigure(0, weight=1)
        self.data_health_tree = ttk.Treeview(
            health_wrap, columns=("status", "check", "count"), show="headings", height=9,
        )
        self.data_health_tree.heading("status", text="Status")
        self.data_health_tree.heading("check", text="Check")
        self.data_health_tree.heading("count", text="Count")
        self.data_health_tree.column("status", width=75, anchor="center")
        self.data_health_tree.column("check", width=300)
        self.data_health_tree.column("count", width=70, anchor="center")
        self.data_health_tree.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        self.data_health_tree.tag_configure("OK", foreground="#166534")
        self.data_health_tree.tag_configure("MEDIUM", foreground="#92400e")
        self.data_health_tree.tag_configure("HIGH", foreground="#b91c1c")
        self.data_health_tree.tag_configure("CRITICAL", foreground="#7f1d1d")
        self.data_health_tree.bind("<<TreeviewSelect>>", self._show_data_health_detail)

        self.data_health_detail = tk.Text(
            health_wrap, height=9, wrap="word", bg="#f8fafc", fg=UI.TEXT,
            relief="solid", bd=1, font=UI.FONT_SM, padx=10, pady=8,
        )
        self.data_health_detail.grid(row=0, column=1, sticky="nsew")
        self.data_health_detail.configure(state="disabled")
        self._data_health_by_key = {}
        self.after(150, self._refresh_data_health)

    def _refresh_data_health(self):
        try:
            health = data_health_summary(8) or {}
            issues = list(health.get("issues") or [])

            backup = get_backup_config() or {}
            remote = str(backup.get("backup_rclone_remote") or "").strip()
            offsite = backup.get("offsite") or {}
            if remote and offsite and not bool(offsite.get("ok")):
                issues.append({
                    "key": "offsite_backup", "title": "Off-site backup needs attention", "count": 1,
                    "severity": "HIGH", "explanation": "The latest configured off-site backup upload did not succeed.",
                    "samples": [str(offsite.get("message") or "Open Settings > Backups for details.")[:240]],
                })

            sync = cloud_sync_status(probe=False) or {}
            pending = int(sync.get("pending", 0) or 0)
            last_error = str(sync.get("last_error") or "").strip()
            if pending or last_error:
                issues.append({
                    "key": "cloud_sync", "title": "Cloud sync queue / error", "count": pending or 1,
                    "severity": "HIGH" if last_error else "MEDIUM",
                    "explanation": "Pending or failed cloud records can make another register show older data.",
                    "samples": [last_error[:240]] if last_error else [f"Pending records: {pending}"],
                })

            self._data_health_by_key = {str(i.get("key")): i for i in issues}
            self.data_health_tree.delete(*self.data_health_tree.get_children())
            problem_count = 0
            for issue in issues:
                count = int(issue.get("count", 0) or 0)
                status = str(issue.get("severity") or ("OK" if count == 0 else "MEDIUM")).upper()
                if count > 0:
                    problem_count += 1
                self.data_health_tree.insert(
                    "", tk.END, iid=str(issue.get("key")), tags=(status,),
                    values=(status, issue.get("title", ""), count),
                )
            checked = str(health.get("checked_at") or "")
            self.data_health_summary_lbl.config(
                text=(f"{problem_count} check(s) need attention  |  {checked}" if problem_count else f"All checks OK  |  {checked}"),
                fg=(UI.DANGER if problem_count else UI.SUCCESS),
            )
            children = self.data_health_tree.get_children()
            if children:
                first_problem = next((x for x in children if int(self._data_health_by_key.get(x, {}).get("count", 0) or 0) > 0), children[0])
                self.data_health_tree.selection_set(first_problem)
                self._show_data_health_detail()
        except Exception as exc:
            self.data_health_summary_lbl.config(text=f"Checks could not run: {str(exc)[:120]}", fg=UI.DANGER)

    def _show_data_health_detail(self, _event=None):
        selected = self.data_health_tree.selection()
        if not selected:
            return
        issue = self._data_health_by_key.get(str(selected[0]), {})
        lines = [str(issue.get("title") or ""), "", str(issue.get("explanation") or "")]
        samples = list(issue.get("samples") or [])
        if samples:
            lines.extend(["", "Examples:"] + [f"• {sample}" for sample in samples])
        self.data_health_detail.configure(state="normal")
        self.data_health_detail.delete("1.0", tk.END)
        self.data_health_detail.insert("1.0", "\n".join(lines))
        self.data_health_detail.configure(state="disabled")

    def _go_products_for_health(self):
        top = self.winfo_toplevel()
        if hasattr(top, "show_page"):
            top.show_page("ProductsPage")

    def _kpi_card(self, parent, col, title, value):
        c = Card(parent, padx=14, pady=12)
        c.grid(row=0, column=col, sticky="nsew", padx=(0 if col == 0 else 10, 0))
        tk.Label(c.inner, text=title, font=UI.FONT_SM, bg=UI.CARD, fg=UI.MUTED).pack(anchor="w")
        v = tk.Label(c.inner, text=value, font=("Segoe UI", 18, "bold"), bg=UI.CARD, fg=UI.TEXT)
        v.pack(anchor="w", pady=(6, 0))
        return v

    def _make_chart(self, parent):
        # Native Tk chart: fast, dependency-free, and works in the 32-bit build.
        canvas = tk.Canvas(parent, height=235, bg=UI.CARD, highlightthickness=0)
        canvas.pack(fill="both", expand=True, pady=(10, 0))
        chart = {"canvas": canvas, "labels": [], "values": [], "ylabel": ""}
        canvas.bind("<Configure>", lambda _e, c=chart: self._draw_native_line(c), add="+")
        return chart

    def apply_period(self):
        p = self.period_var.get().strip().lower()

        if p == "today":
            s, e = _range_bounds("today")
        elif p == "this week":
            s, e = _range_bounds("week")
        elif p == "this month":
            s, e = _range_bounds("month")
        elif p == "this year":
            s, e = _range_bounds("year")
        elif p == "all time":
            s = "2000-01-01"
            e = date.today().strftime("%Y-%m-%d")
        else:
            s = self.start_var.get().strip()
            e = self.end_var.get().strip()
            if not s or not e:
                s, e = _range_bounds("week")

        self.start_var.set(s)
        self.end_var.set(e)
        self.refresh()

    def _breakdown_row(self, label, amount, bold=False):
        row = tk.Frame(self.breakdown_box, bg=UI.CARD)
        row.pack(fill="x", pady=4)

        f = ("Segoe UI", 10, "bold") if bold else UI.FONT
        tk.Label(row, text=label, bg=UI.CARD, fg="#1f2937", font=f).pack(side="left")
        tk.Label(row, text=money(amount), bg=UI.CARD, fg="#1f2937", font=f).pack(side="right")

        tk.Frame(self.breakdown_box, bg=UI.BORDER, height=1).pack(fill="x", pady=2)

    def _sub_breakdown_row(self, container, label, value, is_money=True):
        row = tk.Frame(container, bg=UI.CARD)
        row.pack(fill="x", pady=2)
        tk.Label(row, text=label, bg=UI.CARD, fg="#4b5563", font=UI.FONT_SM).pack(side="left")
        val_text = money(value) if is_money else str(value)
        tk.Label(row, text=val_text, bg=UI.CARD, fg="#1f2937", font=("Segoe UI", 9, "bold")).pack(side="right")
        tk.Frame(container, bg=UI.BORDER, height=1).pack(fill="x", pady=1)

    def _plot_line(self, chart, labels, values, ylabel):
        chart["labels"] = list(labels or [])
        chart["values"] = [float(v or 0.0) for v in (values or [])]
        chart["ylabel"] = str(ylabel or "")
        self._draw_native_line(chart)

    def _draw_native_line(self, chart):
        canvas = chart.get("canvas")
        if canvas is None:
            return
        canvas.delete("all")
        values = list(chart.get("values") or [])
        labels = list(chart.get("labels") or [])
        width = max(300, int(canvas.winfo_width() or 300))
        height = max(170, int(canvas.winfo_height() or 235))
        left, right, top, bottom = 54, 14, 16, 34
        plot_w, plot_h = max(1, width - left - right), max(1, height - top - bottom)
        if not values:
            canvas.create_text(width / 2, height / 2, text="No sales in this range", fill=UI.MUTED, font=UI.FONT_SM)
            return
        low = min(0.0, min(values))
        high = max(values)
        if abs(high - low) < 0.001:
            high = low + 1.0
        for i in range(5):
            y = top + (plot_h * i / 4)
            value = high - ((high - low) * i / 4)
            canvas.create_line(left, y, width - right, y, fill="#e2e8f0")
            canvas.create_text(left - 6, y, text=(f"${value:,.0f}" if chart.get("ylabel") != "AOV" else f"${value:,.0f}"),
                               anchor="e", fill=UI.MUTED, font=("Segoe UI", 8))
        count = len(values)
        points = []
        for i, value in enumerate(values):
            x = left + (plot_w * i / max(1, count - 1))
            y = top + ((high - value) / (high - low) * plot_h)
            points.extend((x, y))
        if len(points) >= 4:
            canvas.create_line(*points, fill=UI.PRIMARY, width=3, smooth=(count > 3))
        else:
            canvas.create_oval(points[0] - 4, points[1] - 4, points[0] + 4, points[1] + 4, fill=UI.PRIMARY, outline="")
        step = max(1, count // 6)
        for i in range(0, count, step):
            x = left + (plot_w * i / max(1, count - 1))
            label = str(labels[i] if i < len(labels) else i)
            if len(label) > 10:
                label = label[-8:]
            canvas.create_text(x, height - 12, text=label, fill=UI.MUTED, font=("Segoe UI", 8))

    def _product_sort_key(self, row):
        sort = str(self.product_sort_var.get() or "Revenue").lower()
        if sort == "qty sold":
            return (-int(row.get("qty_sold") or 0), -float(row.get("net_revenue") or 0.0), str(row.get("name") or "").lower())
        if sort == "net qty":
            return (-int(row.get("net_qty") or 0), -float(row.get("net_revenue") or 0.0), str(row.get("name") or "").lower())
        if sort == "name":
            return (str(row.get("name") or "").lower(),)
        return (-float(row.get("net_revenue") or 0.0), -int(row.get("qty_sold") or 0), str(row.get("name") or "").lower())

    def _filtered_product_rows(self):
        q = str(self.product_search_var.get() or "").strip().lower()
        rows = list(getattr(self, "product_rows", []) or [])
        if q:
            rows = [
                r for r in rows
                if q in str(r.get("name") or "").lower()
                or q in str(r.get("barcode") or "").lower()
                or q in str(r.get("category") or "").lower()
                or q in str(r.get("brand") or "").lower()
            ]
        rows.sort(key=self._product_sort_key)
        return rows

    def _render_product_rows(self):
        if not hasattr(self, "prod_tree"):
            return
        rows = self._filtered_product_rows()
        total_qty = sum(int(r.get("qty_sold") or 0) for r in rows)
        total_revenue = sum(float(r.get("net_revenue") or 0.0) for r in rows)
        if hasattr(self, "product_summary_lbl"):
            self.product_summary_lbl.config(text=f"{len(rows)} products | {total_qty} sold | {money(total_revenue)} net")

        self.prod_tree.delete(*self.prod_tree.get_children())
        for idx, r in enumerate(rows[:500], start=1):
            qty_sold = int(r.get("qty_sold") or 0)
            qty_returned = int(r.get("qty_returned") or 0)
            net_qty = int(r.get("net_qty") or (qty_sold - qty_returned))
            revenue = float(r.get("net_revenue") if r.get("net_revenue") is not None else r.get("revenue") or 0.0)
            avg = float(r.get("avg_unit") or ((revenue / qty_sold) if qty_sold else 0.0))
            share = float(r.get("sales_share") or 0.0)
            stock_raw = r.get("current_stock", "")
            try:
                stock_display = str(int(float(stock_raw))) if stock_raw not in ("", None) else ""
            except Exception:
                stock_display = str(stock_raw or "")
            self.prod_tree.insert("", tk.END, values=(
                idx,
                r.get("name", ""),
                qty_sold,
                qty_returned,
                net_qty,
                money(revenue),
                money(avg),
                f"{share * 100:.1f}%",
                stock_display,
                r.get("barcode", ""),
                r.get("category", ""),
            ))

        self._top_seller_rows = rows[:10]
        self._draw_top_sellers()

    def _draw_top_sellers(self):
        canvas = getattr(self, "top_sellers_canvas", None)
        if canvas is None:
            return
        canvas.delete("all")
        rows = list(getattr(self, "_top_seller_rows", []) or [])
        width = max(320, int(canvas.winfo_width() or 520))
        height = max(180, int(canvas.winfo_height() or 220))
        pad = 12
        if not rows:
            canvas.create_text(
                width // 2,
                height // 2,
                text="No item sales in this range",
                fill=UI.MUTED,
                font=UI.FONT_SM,
            )
            return

        sort = str(self.product_sort_var.get() or "Revenue").lower()
        if sort == "qty sold":
            metric_key = "qty_sold"
            value_label = lambda v: str(int(v))
        elif sort == "net qty":
            metric_key = "net_qty"
            value_label = lambda v: str(int(v))
        else:
            metric_key = "net_revenue"
            value_label = money

        values = [max(0.0, float(r.get(metric_key) or 0.0)) for r in rows]
        max_value = max(values) if values else 0.0
        if max_value <= 0:
            max_value = 1.0

        label_w = max(110, min(210, int(width * 0.36)))
        value_w = 78
        bar_x = label_w + pad
        bar_max_w = max(70, width - bar_x - value_w - pad)
        row_h = max(18, min(28, int((height - (pad * 2)) / max(1, len(rows)))))
        char_limit = max(12, int(label_w / 7))

        for i, (row, value) in enumerate(zip(rows, values)):
            y = pad + i * row_h
            name = str(row.get("name") or "")
            if len(name) > char_limit:
                name = name[:max(1, char_limit - 3)] + "..."
            bar_w = int(bar_max_w * (value / max_value))
            canvas.create_text(
                pad,
                y + row_h / 2,
                text=name,
                anchor="w",
                fill=UI.TEXT,
                font=("Segoe UI", 8),
            )
            canvas.create_rectangle(
                bar_x,
                y + 4,
                bar_x + bar_w,
                y + row_h - 4,
                fill="#2F80ED",
                outline="",
            )
            canvas.create_rectangle(
                bar_x + bar_w,
                y + 4,
                bar_x + bar_max_w,
                y + row_h - 4,
                fill="#E8EEF6",
                outline="",
            )
            canvas.create_text(
                width - pad,
                y + row_h / 2,
                text=value_label(value),
                anchor="e",
                fill=UI.TEXT,
                font=("Segoe UI", 8, "bold"),
            )

    def open_discount_impact(self):
        start_date=self.start_var.get().strip(); end_date=self.end_var.get().strip()
        try: data=analytics_discount_impact(start_date,end_date,250) or {}
        except Exception as exc: messagebox.showerror("Discount Impact",str(exc),parent=self); return
        win=tk.Toplevel(self); win.title("Discount Impact"); win.geometry("900x540"); win.minsize(680,420)
        win.configure(bg=UI.CONTENT_BG); win.transient(self.winfo_toplevel())
        card=Card(win,padx=14,pady=14); card.pack(fill="both",expand=True,padx=12,pady=12)
        HeaderBar(card.inner,"Discount Impact",f"{start_date} to {end_date} — products sold below their original value.").pack(fill="x")
        summary=tk.Frame(card.inner,bg=UI.CARD); summary.pack(fill="x",pady=(10,8))
        for title,key in [("Before discounts","before_discount"),("After discounts","after_discount"),("Discount given","discount_amount"),("Est. profit after","estimated_profit_after")]:
            box=tk.Frame(summary,bg="#f8fafc",highlightthickness=1,highlightbackground=UI.BORDER); box.pack(side="left",fill="x",expand=True,padx=(0,8))
            tk.Label(box,text=title,bg="#f8fafc",fg=UI.MUTED).pack(anchor="w",padx=8,pady=(6,0)); tk.Label(box,text=money(data.get(key,0)),bg="#f8fafc",fg=UI.TEXT,font=("Segoe UI",11,"bold")).pack(anchor="w",padx=8,pady=(0,6))
        cols=("product","qty","before","discount","after","profit")
        tree=ttk.Treeview(card.inner,columns=cols,show="headings",height=15)
        for col,title,width in [("product","Product",260),("qty","Qty",55),("before","Before",90),("discount","Discount",90),("after","After",90),("profit","Est. profit",90)]:
            tree.heading(col,text=title); tree.column(col,width=width,anchor=("w" if col=="product" else "e"))
        tree.pack(fill="both",expand=True)
        for idx,row in enumerate(data.get("items") or []):
            after=float(row_get(row,"after_discount",0) or 0); cost=float(row_get(row,"estimated_cost",0) or 0)
            tree.insert("",tk.END,iid=str(idx),values=(row_get(row,"name",""),row_get(row,"qty",0),money(row_get(row,"before_discount",0)),money(row_get(row,"discount_amount",0)),money(after),money(after-cost)))
        GhostButton(card.inner,"Close",win.destroy).pack(anchor="e",pady=(8,0))

    def refresh(self):
        start_date = self.start_var.get().strip()
        end_date = self.end_var.get().strip()

        self.range_lbl.config(text=f"{start_date} to {end_date}")
        self.last_refresh_lbl.config(text=f"Last refreshed: {fmt_time_ampm(datetime.now())}")

        k = analytics_kpis_range(start_date, end_date)
        self.kpi_gross.config(text=money(k.get("net_sales", 0)))
        self.kpi_cash.config(text=money(k.get("cash_sales_total", 0)))
        self.kpi_items.config(text=str(k.get("items_sold", 0)))
        self.kpi_aov.config(text=money(k.get("avg_order_value", 0)))
        self.range_lbl.config(text=f"{start_date} to {end_date}  |  {int(k.get('orders', 0) or 0)} orders")

        for w in self.breakdown_box.winfo_children():
            w.destroy()

        b = analytics_breakdown_range(start_date, end_date) or {}
        # Never crash Analytics if the server is unreachable or an older host returns different fields.
        # Accept either a plain breakdown dict OR {"ok":False,"error":...}.
        if isinstance(b, dict) and ("gross_sales" not in b) and ("breakdown" in b):
            b = b.get("breakdown") or {}
        if not isinstance(b, dict):
            b = {}

        # If backend returned an error payload, show zeros (and keep UI alive).
        if b.get("ok") is False and "error" in b:
            try:
                self._breakdown_row("Server error", 0.0, bold=True)
                self._breakdown_row(str(b.get("error") or "Unknown error")[:80], 0.0)
            except Exception:
                pass
            b = {}


        gs = float(b.get("gross_sales", 0) or 0.0)
        disc = float(b.get("discounts", 0) or 0.0)
        ret = float(b.get("returns", 0) or 0.0)
        net = float(b.get("net_sales", (gs - disc - ret)) or 0.0)
        ship = float(b.get("shipping", 0) or 0.0)
        tax = float(b.get("taxes", 0) or 0.0)
        total = float(b.get("total_sales", (net + ship + tax)) or 0.0)

        self._breakdown_row("Item value before discounts", gs)
        if disc > 0.005:
            self._breakdown_row("Discounts", -disc)
        self._breakdown_row("Sales after discounts", net + ret)
        if ret > 0.005:
            self._breakdown_row("Returns", -ret)
        self._breakdown_row("Total sales", total, bold=True)
        if float(k.get("cogs", 0) or 0.0) > 0.005:
            self._breakdown_row("Product cost entered", -float(k.get("cogs", 0) or 0.0))
            self._breakdown_row("Gross profit", float(k.get("gross_profit", 0) or 0.0), bold=True)

        for w in self.pm_breakdown_box.winfo_children():
            w.destroy()
        for w in self.cat_breakdown_box.winfo_children():
            w.destroy()

        pm_data = b.get("pm_breakdown") or {}
        cat_data = b.get("cat_breakdown") or {}

        if pm_data:
            for pm_name, pm_val in sorted(pm_data.items(), key=lambda x: x[1], reverse=True):
                self._sub_breakdown_row(self.pm_breakdown_box, pm_name, pm_val)
        else:
            tk.Label(self.pm_breakdown_box, text="No payment data available", font=UI.FONT_SM, bg=UI.CARD, fg=UI.MUTED).pack(anchor="w")

        if cat_data:
            for cat_name, cat_val in sorted(cat_data.items(), key=lambda x: x[1], reverse=True):
                self._sub_breakdown_row(self.cat_breakdown_box, cat_name, cat_val)
        else:
            tk.Label(self.cat_breakdown_box, text="No category data available", font=UI.FONT_SM, bg=UI.CARD, fg=UI.MUTED).pack(anchor="w")

        group = "day"
        try:
            sd = datetime.strptime(start_date, "%Y-%m-%d")
            ed = datetime.strptime(end_date, "%Y-%m-%d")
            days = (ed - sd).days + 1
            if days == 1:
                group = "hour"
            elif days > 70:
                group = "month"
            else:
                group = "day"
        except Exception:
            group = "day"

        series = analytics_series_in_range(start_date, end_date, group=group) or []
        labels = [str(p.get("label","")) for p in series if isinstance(p, dict)]
        revs = [float((p.get("revenue") or 0.0)) for p in series if isinstance(p, dict)]
        orders = [int((p.get("orders") or 0)) for p in series if isinstance(p, dict)]
        order_sales = [float((p.get("order_value_sales") or p.get("revenue") or 0.0)) for p in series if isinstance(p, dict)]
        aov = [(order_sales[i] / orders[i]) if orders[i] > 0 else 0.0 for i in range(len(series))]

        self._plot_line(self.sales_chart, labels, revs, "Sales")
        self._plot_line(self.aov_chart, labels, aov, "AOV")

        top = analytics_top_products_range(start_date, end_date, limit=500) or []
        self.product_rows = []
        for r in top:
            qty_sold = int(row_get(r, "qty_sold", 0) or 0)
            qty_returned = int(row_get(r, "qty_returned", 0) or 0)
            net_qty = int(row_get(r, "net_qty", qty_sold - qty_returned) or 0)
            revenue = float(row_get(r, "net_revenue", row_get(r, "revenue", 0)) or 0.0)
            avg = float(row_get(r, "avg_unit", ((revenue / qty_sold) if qty_sold else 0.0)) or 0.0)
            self.product_rows.append({
                "name": row_get(r, "name", ""),
                "barcode": row_get(r, "barcode", ""),
                "category": row_get(r, "category", ""),
                "brand": row_get(r, "brand", ""),
                "qty_sold": qty_sold,
                "qty_returned": qty_returned,
                "net_qty": net_qty,
                "net_revenue": revenue,
                "revenue": revenue,
                "avg_unit": avg,
                "sales_share": float(row_get(r, "sales_share", 0.0) or 0.0),
                "current_stock": row_get(r, "current_stock", ""),
            })
        self._render_product_rows()

    # ---------------- CASH DRAWER PAGE (SHIFTS + DAILY HISTORY) ----------------

    def _reports_folder(self) -> str:
        folder = os.path.join(os.path.dirname(__file__), "reports")
        os.makedirs(folder, exist_ok=True)
        return folder

    def _open_reports_folder(self):
        folder = self._reports_folder()
        try:
            os.startfile(folder)  # Windows
        except Exception:
            try:
                subprocess.Popen(["explorer", folder])
            except Exception:
                messagebox.showinfo("Reports folder", f"Reports are saved in:\n\n{folder}")

    def _lazy_import_openpyxl(self):
        try:
            from openpyxl import Workbook  # type: ignore
            from openpyxl.utils import get_column_letter  # type: ignore
            from openpyxl.chart import LineChart, BarChart, Reference  # type: ignore
            return Workbook, get_column_letter, LineChart, BarChart, Reference
        except Exception:
            messagebox.showerror(
                "Missing dependency",
                "Excel export needs openpyxl.\n\nFix:\n1) Open Command Prompt\n2) cd into your project folder\n3) venv\\Scripts\\activate\n4) pip install openpyxl"
            )
            return None

    def _export_sales_report_pdf(self, year: int, month: int, day_str: str):
        try:
            if not _local_db_available_or_warn("Sales PDF export"):
                return
            daily_report = _load_daily_report_module()
            result = daily_report.build_cash_drawer_pdf(
                data_path("pos.db"),
                self._reports_folder(),
                int(year),
                int(month),
                str(day_str),
            )
            out_path = result.get("path", "")
            if hasattr(self, "report_status_lbl"):
                self.report_status_lbl.config(text=f"Saved: {out_path}")
            if out_path and os.path.exists(out_path):
                os.startfile(out_path)
            messagebox.showinfo("Report generated", f"Saved and opened:\n\n{out_path}")
            return result
        except Exception as ex:
            messagebox.showerror("Export failed", f"Could not generate the PDF report.\n\n{ex}")
            return

    def _export_sales_report_pdf_range(self):
        try:
            if not _local_db_available_or_warn("Sales PDF export"):
                return
            start_str = self.start_var.get().strip()
            end_str = self.end_var.get().strip()
            if not start_str or not end_str:
                messagebox.showerror("Invalid Range", "Please enter start and end dates (YYYY-MM-DD) in the custom range entry boxes.")
                return
            
            try:
                start_date = datetime.strptime(start_str, "%Y-%m-%d")
                end_date = datetime.strptime(end_str, "%Y-%m-%d")
            except Exception:
                messagebox.showerror("Invalid Range", "Dates must be in YYYY-MM-DD format.")
                return
            
            end_date_bounds = end_date + timedelta(days=1)
            
            daily_report = _load_daily_report_module()
            result = daily_report.build_cash_drawer_pdf(
                data_path("pos.db"),
                self._reports_folder(),
                start_date=start_date,
                end_date=end_date_bounds,
                custom_stamp=f"range_{start_date.strftime('%Y%m%d')}_to_{end_date.strftime('%Y%m%d')}"
            )
            out_path = result.get("path", "")
            if hasattr(self, "report_status_lbl"):
                self.report_status_lbl.config(text=f"Saved: {out_path}")
            if out_path and os.path.exists(out_path):
                os.startfile(out_path)
            messagebox.showinfo("Report generated", f"Range PDF Saved and opened:\n\n{out_path}")
        except Exception as ex:
            messagebox.showerror("Export failed", f"Could not generate the Range PDF report.\n\n{ex}")

    def _export_sales_report_excel(self, year: int, month: int, day_str: str):
        """Export sales + shifts to an Excel file.

        What you get:
          - Summary sheet (monthly or single day)
          - One sheet per day (YYYY-MM-DD) with:
              * Big daily total box (green)
              * Shifts listed in order with opening/closing + employee
              * Sales under each shift
              * Unassigned sales section (if any)
          - RawSales sheet (flat table)
          - RawItems sheet (flat table)
          - Charts sheet (kept from the old export, based on raw tables)

        This never deletes anything from the database.
        """
        imported = self._lazy_import_openpyxl()
        if not imported:
            return
        Workbook, get_column_letter, LineChart, BarChart, Reference = imported

        try:
            if not _local_db_available_or_warn("Sales report export"):
                return
            build_sales_report_excel = _load_daily_report_builder()
            result = build_sales_report_excel(
                data_path("pos.db"),
                self._reports_folder(),
                int(year),
                int(month),
                str(day_str),
            )
            out_path = result.get("path", "")
            if hasattr(self, "report_status_lbl"):
                self.report_status_lbl.config(text=f"Saved: {out_path}")
            messagebox.showinfo("Report generated", f"Saved:\n\n{out_path}")
            return result
        except Exception as ex:
            messagebox.showerror("Export failed", f"Could not generate the Excel report.\n\n{ex}")
            return

        # styles (safe even if openpyxl styles are missing in some environments)
        try:
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side  # type: ignore
        except Exception:
            Font = Alignment = PatternFill = Border = Side = None

        def _style(cell, font=None, align=None, fill=None, border=None, number_format=None):
            try:
                if font is not None:
                    cell.font = font
                if align is not None:
                    cell.alignment = align
                if fill is not None:
                    cell.fill = fill
                if border is not None:
                    cell.border = border
                if number_format is not None:
                    cell.number_format = number_format
            except Exception:
                pass

        bold = (Font(bold=True) if Font else None)
        title_font = (Font(bold=True, size=16) if Font else None)
        h_font = (Font(bold=True, size=12) if Font else None)
        center = (Alignment(horizontal="center", vertical="center") if Alignment else None)
        left = (Alignment(horizontal="left", vertical="center") if Alignment else None)

        green_fill = (PatternFill("solid", fgColor="C6EFCE") if PatternFill else None)  # light green
        gray_fill = (PatternFill("solid", fgColor="F2F2F2") if PatternFill else None)

        thin = (Side(style="thin", color="D0D0D0") if Side else None)
        box_border = (Border(left=thin, right=thin, top=thin, bottom=thin) if Border and thin else None)

        # Build date range
        try:
            if str(day_str).strip().lower() == "all":
                start = datetime(year, month, 1, 0, 0, 0)
                if month == 12:
                    end = datetime(year + 1, 1, 1, 0, 0, 0)
                else:
                    end = datetime(year, month + 1, 1, 0, 0, 0)
                stamp = f"{year:04d}_{month:02d}"
            else:
                d = int(str(day_str).strip())
                start = datetime(year, month, d, 0, 0, 0)
                end = start + timedelta(days=1)
                stamp = f"{year:04d}_{month:02d}_{d:02d}"
        except Exception:
            messagebox.showerror("Invalid date", "Please choose a valid year, month, and day.")
            return

        if not _local_db_available_or_warn("Export / raw DB analytics"):
            return

        db_path = data_path("pos.db")
        if not os.path.exists(db_path):
            messagebox.showerror("Database missing", f"Could not find pos.db at:\n\n{db_path}")
            return

        # Read data
        try:
            conn = sqlite3.connect(db_path)
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()

            # Detect optional columns
            cur.execute("PRAGMA table_info(sales)")
            sales_cols = [r[1] for r in cur.fetchall()]
            has_receipt = ("receipt_code" in sales_cols)

            cur.execute("PRAGMA table_info(cash_shifts)")
            shift_cols = [r[1] for r in cur.fetchall()]
            has_shift_code = ("shift_code" in shift_cols)

            # Sales (raw)
            select_cols = [
                "s.id",
                "s.created_at",
                ("s.receipt_code" if has_receipt else "'' AS receipt_code"),
                "s.total_amount",
                "s.payment_method",
                "s.customer_name",
                "s.shift_id",
            ]
            cur.execute(f"""
                SELECT {", ".join(select_cols)}
                FROM sales s
                WHERE datetime(s.created_at) >= datetime(?)
                  AND datetime(s.created_at) < datetime(?)
                ORDER BY datetime(s.created_at) ASC
            """, (start.strftime("%Y-%m-%d %H:%M:%S"), end.strftime("%Y-%m-%d %H:%M:%S")))
            sales_rows = cur.fetchall()

            # Items (raw)
            cur.execute("""
                SELECT si.sale_id, si.product_id, si.name, si.price, si.qty, si.line_total
                FROM sale_items si
                JOIN sales s ON s.id = si.sale_id
                WHERE datetime(s.created_at) >= datetime(?)
                  AND datetime(s.created_at) < datetime(?)
                ORDER BY datetime(s.created_at) ASC, si.id ASC
            """, (start.strftime("%Y-%m-%d %H:%M:%S"), end.strftime("%Y-%m-%d %H:%M:%S")))
            items_rows = cur.fetchall()

            # Shifts for the period (by opened_at date range)
            # Note: we include shifts that open inside the date range, even if they close later.
            shift_select = [
                "cs.id",
                ("cs.shift_code" if has_shift_code else "'' AS shift_code"),
                "cs.opened_at",
                "cs.closed_at",
                "cs.opening_cash",
                "cs.closing_cash",
                "cs.notes",
                "e.name AS employee_name",
            ]
            cur.execute(f"""
                SELECT {", ".join(shift_select)}
                FROM cash_shifts cs
                LEFT JOIN employees e ON e.id = cs.employee_id
                WHERE datetime(cs.opened_at) >= datetime(?)
                  AND datetime(cs.opened_at) < datetime(?)
                ORDER BY datetime(cs.opened_at) ASC
            """, (start.strftime("%Y-%m-%d %H:%M:%S"), end.strftime("%Y-%m-%d %H:%M:%S")))
            shift_rows = cur.fetchall()

        except Exception as ex:
            try:
                conn.close()
            except Exception:
                pass
            messagebox.showerror("Export failed", f"Could not read sales data.\n\n{ex}")
            return
        finally:
            try:
                conn.close()
            except Exception:
                pass

        # Helpers to group rows
        def _day_of(dt_str: str) -> str:
            s = str(dt_str or "")
            return s[:10] if len(s) >= 10 else s

        def _time_of(dt_str: str) -> str:
            s = str(dt_str or "")
            return s[11:19] if len(s) >= 19 else s

        # Build groupings
        sales_by_day = {}
        for r in sales_rows:
            day = _day_of(r["created_at"])
            sales_by_day.setdefault(day, []).append(r)

        shifts_by_day = {}
        for sh in shift_rows:
            day = _day_of(sh["opened_at"])
            shifts_by_day.setdefault(day, []).append(sh)

        # days to create sheets for: any day that has sales or shifts
        all_days = sorted(set(list(sales_by_day.keys()) + list(shifts_by_day.keys())))

        # Totals
        total_revenue = sum(float(r["total_amount"] or 0) for r in sales_rows)
        total_orders = len(sales_rows)
        total_items = sum(int(r["qty"] or 0) for r in items_rows)
        avg_order = (total_revenue / total_orders) if total_orders else 0.0

        wb = Workbook()

        # ---------------- Summary sheet ----------------
        ws_sum = wb.active
        ws_sum.title = "Summary"
        ws_sum.append(["Report Period", f"{start.strftime('%Y-%m-%d')} to {end.strftime('%Y-%m-%d')}"])
        ws_sum.append(["Total Orders", int(total_orders)])
        ws_sum.append(["Total Items Sold", int(total_items)])
        ws_sum.append(["Total Revenue", float(total_revenue)])
        ws_sum.append(["Average Order Value", float(avg_order)])
        ws_sum.append([])
        ws_sum.append(["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

        try:
            ws_sum["A1"].font = h_font
            ws_sum["A2"].font = bold
            ws_sum["A3"].font = bold
            ws_sum["A4"].font = bold
            ws_sum["A5"].font = bold
        except Exception:
            pass

        # ---------------- Raw tables (for reliability + charts) ----------------
        ws_raw_sales = wb.create_sheet("RawSales")
        ws_raw_sales.append(["Sale ID", "Created At", "Receipt", "Total", "Payment", "Customer", "Shift ID"])
        for r in sales_rows:
            ws_raw_sales.append([
                int(r["id"]),
                str(r["created_at"]),
                str(r["receipt_code"] or ""),
                float(r["total_amount"] or 0),
                str(r["payment_method"] or ""),
                str(r["customer_name"] or ""),
                (int(r["shift_id"]) if r["shift_id"] is not None else None),
            ])

        ws_raw_items = wb.create_sheet("RawItems")
        ws_raw_items.append(["Sale ID", "Product ID", "Name", "Price", "Qty", "Line Total"])
        for r in items_rows:
            ws_raw_items.append([
                int(r["sale_id"]),
                (int(r["product_id"]) if r["product_id"] is not None else None),
                str(r["name"] or ""),
                float(r["price"] or 0),
                int(r["qty"] or 0),
                float(r["line_total"] or 0),
            ])

        # ---------------- Per-day sheets (your main view) ----------------
        def _autosize(ws):
            for col_cells in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col_cells[0].column)
                for c in col_cells:
                    v = "" if c.value is None else str(c.value)
                    if len(v) > max_len:
                        max_len = len(v)
                ws.column_dimensions[col_letter].width = min(48, max(10, max_len + 2))

        def _write_day_sheet(day: str):
            ws = wb.create_sheet(day)
            ws.sheet_view.showGridLines = True

            # Title row
            ws.merge_cells("A1:F1")
            ws["A1"] = f"Daily Report: {day}"
            _style(ws["A1"], font=title_font, align=left)

            # Big daily total box (green) in the "green box" area: put it at H1:J3 visually
            # Keep columns A-F for tables, and use H-J for the total box.
            ws.merge_cells("H1:J1")
            ws["H1"] = "DAILY TOTAL"
            _style(ws["H1"], font=h_font, align=center, fill=green_fill, border=box_border)

            ws.merge_cells("H2:J3")
            daily_total = 0.0
            for r in sales_by_day.get(day, []):
                daily_total += float(r["total_amount"] or 0)
            ws["H2"] = float(daily_total)
            _style(ws["H2"], font=Font(bold=True, size=22) if Font else None, align=center, fill=green_fill,
                   border=box_border, number_format="#,##0.00")

            # A small info block under title
            row = 3
            ws["A3"] = "Orders"
            ws["B3"] = len(sales_by_day.get(day, []))
            ws["A4"] = "Shifts"
            ws["B4"] = len(shifts_by_day.get(day, []))
            _style(ws["A3"], font=bold)
            _style(ws["A4"], font=bold)

            row = 6

            # Shifts ordered by opened_at
            day_shifts = list(shifts_by_day.get(day, []))
            day_shifts.sort(key=lambda s: str(s["opened_at"] or ""))

            # Build sales mapping by shift for the day
            day_sales = list(sales_by_day.get(day, []))
            sales_by_shift = {}
            unassigned = []
            for sr in day_sales:
                sid = sr["shift_id"]
                if sid is None:
                    unassigned.append(sr)
                else:
                    sales_by_shift.setdefault(int(sid), []).append(sr)

            def _shift_display_code(sh):
                code = str(sh["shift_code"] or "").strip()
                if code:
                    return code
                return str(int(sh["id"]))

            def _shift_cash_sales(shift_id: int):
                total = 0.0
                for sr in sales_by_shift.get(int(shift_id), []):
                    try:
                        total += max(0.0, float(sr["cash_paid"] or 0.0))
                    except Exception:
                        if str(sr["payment_method"] or "").upper() == "CASH":
                            total += float(sr["total_amount"] or 0)
                return float(total)

            # Reusable sales table header
            sales_headers = ["Time", "Sale ID", "Receipt", "Total", "Payment", "Customer"]
            for idx, h in enumerate(sales_headers, start=1):
                cell = ws.cell(row=row, column=idx, value=h)
                _style(cell, font=bold, fill=gray_fill, border=box_border)

            # Move down and write each shift block
            row += 1
            for sh in day_shifts:
                shift_id = int(sh["id"])
                shift_code = _shift_display_code(sh)
                employee = str(sh["employee_name"] or "").strip() or "System"
                opened_at = str(sh["opened_at"] or "")
                closed_at = str(sh["closed_at"] or "")
                opening_cash = float(sh["opening_cash"] or 0)
                closing_cash = (float(sh["closing_cash"]) if sh["closing_cash"] is not None else None)
                notes = str(sh["notes"] or "")

                cash_sales = _shift_cash_sales(shift_id)
                expected = opening_cash + cash_sales
                diff = (closing_cash - expected) if closing_cash is not None else None

                # Shift header (clean, single row)
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                ws.cell(row=row, column=1,
                        value=f"Shift {shift_code}  |  Employee: {employee}").font = h_font if h_font else None
                row += 1

                # Shift details grid (2 rows)
                labels = [
                    ("Opened", opened_at, "Closed", closed_at),
                    ("Opening cash", opening_cash, "Closing cash", (closing_cash if closing_cash is not None else "")),
                    ("Cash sales", cash_sales, "Expected cash", expected),
                    ("Difference", (diff if diff is not None else ""), "Notes", notes),
                ]
                for (l1, v1, l2, v2) in labels:
                    ws.cell(row=row, column=1, value=l1)
                    ws.cell(row=row, column=2, value=v1)
                    ws.cell(row=row, column=3, value=l2)
                    ws.cell(row=row, column=4, value=v2)
                    for c in range(1, 5):
                        _style(ws.cell(row=row, column=c), border=box_border)
                    _style(ws.cell(row=row, column=1), font=bold)
                    _style(ws.cell(row=row, column=3), font=bold)
                    row += 1

                # Sales rows for this shift
                shift_sales = sales_by_shift.get(shift_id, [])
                shift_sales.sort(key=lambda s: str(s["created_at"] or ""))

                if not shift_sales:
                    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                    ws.cell(row=row, column=1, value="No sales in this shift.").font = bold if bold else None
                    row += 2
                else:
                    for sr in shift_sales:
                        ws.append([
                            _time_of(sr["created_at"]),
                            int(sr["id"]),
                            str(sr["receipt_code"] or ""),
                            float(sr["total_amount"] or 0),
                            str(sr["payment_method"] or ""),
                            str(sr["customer_name"] or ""),
                        ])
                        # style border on last appended row
                        for c in range(1, 7):
                            _style(ws.cell(row=ws.max_row, column=c), border=box_border)
                    row = ws.max_row + 2

            # Unassigned sales section
            if unassigned:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                ws.cell(row=row, column=1, value="Unassigned sales (no shift)").font = h_font if h_font else None
                row += 1
                for sr in unassigned:
                    ws.append([
                        _time_of(sr["created_at"]),
                        int(sr["id"]),
                        str(sr["receipt_code"] or ""),
                        float(sr["total_amount"] or 0),
                        str(sr["payment_method"] or ""),
                        str(sr["customer_name"] or ""),
                    ])
                    for c in range(1, 7):
                        _style(ws.cell(row=ws.max_row, column=c), border=box_border)

            _autosize(ws)
            return ws

        for d in all_days:
            _write_day_sheet(d)

        # ---------------- Charts sheet (kept simple) ----------------
        ws_charts = wb.create_sheet("Charts")
        ws_charts.append(["Daily Totals", ""])
        ws_charts.append(["Day", "Revenue"])
        for d in all_days:
            ws_charts.append([d, float(sum(float(r["total_amount"] or 0) for r in sales_by_day.get(d, [])))])

        # Line chart: daily revenue
        try:
            lc = LineChart()
            lc.title = "Daily Revenue"
            lc.y_axis.title = "Revenue"
            lc.x_axis.title = "Day"
            data = Reference(ws_charts, min_col=2, min_row=2, max_row=1 + len(all_days))
            cats = Reference(ws_charts, min_col=1, min_row=3, max_row=2 + len(all_days))
            lc.add_data(data, titles_from_data=True)
            lc.set_categories(cats)
            ws_charts.add_chart(lc, "D2")
        except Exception:
            pass

        # Auto width for readability on raw sheets too
        for ws in [ws_sum, ws_raw_sales, ws_raw_items, ws_charts]:
            _autosize(ws)

        # Save with timestamp so Excel-open files never block overwrites
        stamp2 = datetime.now().strftime("%Y_%m_%d_%H%M%S")
        out_path = os.path.join(self._reports_folder(), f"sales_report_{stamp}_{stamp2}.xlsx")
        try:
            wb.save(out_path)
            if hasattr(self, "report_status_lbl"):
                self.report_status_lbl.config(text=f"Saved: {out_path}")
            messagebox.showinfo("Report generated", f"Saved:\n\n{out_path}")
        except Exception as ex:
            messagebox.showerror("Export failed", f"Could not save Excel file.\n\n{ex}")

    def _export_inventory_report_excel(self):
        """Export current product list (including barcodes) to an Excel file."""
        imported = self._lazy_import_openpyxl()
        if not imported:
            return
        Workbook, get_column_letter, LineChart, BarChart, Reference = imported

        if not _local_db_available_or_warn("Export / raw DB analytics"):
            return

        db_path = data_path("pos.db")
        if not os.path.exists(db_path):
            messagebox.showerror("Database missing", f"Could not find pos.db at:\n\n{db_path}")
            return

        try:
            conn = sqlite3.connect(db_path)
            cur = conn.cursor()
            cur.execute("PRAGMA table_info(products)")
            cols = [r[1] for r in cur.fetchall()]

            wanted = ["id", "barcode", "name", "category", "brand", "location", "sell_price", "stock_qty", "low_stock_level",
                      "created_at"]
            select_cols = [c for c in wanted if c in cols]

            where = ""
            if "is_deleted" in cols:
                where = " WHERE is_deleted = 0"

            if not select_cols:
                messagebox.showerror("Export failed", "Could not read product columns from database.")
                return

            cur.execute(f"SELECT {', '.join(select_cols)} FROM products{where} ORDER BY name ASC")
            rows = cur.fetchall()

        except Exception as ex:
            messagebox.showerror("Export failed", f"Could not read product data.\n\n{ex}")
            return
        finally:
            try:
                conn.close()
            except Exception:
                pass

        wb = Workbook()
        ws = wb.active
        ws.title = "Products"

        headers = []
        for c in select_cols:
            if c == "sell_price":
                headers.append("Sell Price")
            elif c == "stock_qty":
                headers.append("Stock Qty")
            elif c == "low_stock_level":
                headers.append("Low Stock Level")
            elif c == "created_at":
                headers.append("Created At")
            else:
                headers.append(c.replace("_", " ").title())

        ws.append(headers + ["Stock Status"])

        # Determine column positions
        col_index = {c: idx for idx, c in enumerate(select_cols, start=1)}
        for r in rows:
            r_list = list(r)
            stock = 0
            low = 0
            if "stock_qty" in col_index:
                try:
                    stock = int(r[col_index["stock_qty"] - 1] or 0)
                except Exception:
                    stock = 0
            if "low_stock_level" in col_index:
                try:
                    low = int(r[col_index["low_stock_level"] - 1] or 0)
                except Exception:
                    low = 0
            status = "LOW" if (low > 0 and stock <= low) else "OK"
            ws.append(list(r_list) + [status])

        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for c in col_cells:
                v = "" if c.value is None else str(c.value)
                if len(v) > max_len:
                    max_len = len(v)
            ws.column_dimensions[col_letter].width = min(48, max(10, max_len + 2))

        stamp = datetime.now().strftime("%Y_%m_%d")
        out_path = os.path.join(self._reports_folder(), f"inventory_report_{stamp}.xlsx")
        try:
            wb.save(out_path)
            if hasattr(self, "report_status_lbl"):
                self.report_status_lbl.config(text=f"Saved: {out_path}")
            messagebox.showinfo("Report generated", f"Saved:\n\n{out_path}")
        except Exception as ex:
            messagebox.showerror("Export failed", f"Could not save Excel file.\n\n{ex}")


class ShiftsPage(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg=UI.CONTENT_BG)
        self.employee_var = tk.StringVar(value="")
        self.opening_var = tk.StringVar(value="0")
        self.closing_var = tk.StringVar(value="0")
        self.date_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        # Date picker (dropdowns)
        today = datetime.now()
        self.sales_year = tk.StringVar(value=str(today.year))
        self.sales_month = tk.StringVar(value=f"{today.month:02d}")
        self.sales_day = tk.StringVar(value=f"{today.day:02d}")
        self.reprint_var = tk.StringVar(value="")  # Sale ID / receipt scan for reprint
        self.show_voided_var = tk.BooleanVar(value=False)
        self._last_lookup_sale_id = None
        self._build()

    def _build(self):
        scroll = VScrollableFrame(self, bg=UI.CONTENT_BG)
        scroll.pack(fill="both", expand=True)
        wrap = tk.Frame(scroll.inner, bg=UI.CONTENT_BG)
        wrap.pack(fill="both", expand=True, padx=(10 if UI.COMPACT else 18), pady=(10 if UI.COMPACT else 18))

        header = Card(wrap, padx=18, pady=14)
        header.pack(fill="x")
        header_row = tk.Frame(header.inner, bg=UI.CARD)
        header_row.pack(fill="x")
        HeaderBar(
            header_row,
            "Cash Drawer",
            "Open shift per employee, see expected cash, browse sales day by day."
        ).pack(side="left", fill="x", expand=True)
        GhostButton(header_row, "Fullscreen", self.toggle_drawer_fullscreen).pack(side="right", padx=(10, 0))

        body = tk.Frame(wrap, bg=UI.CONTENT_BG)
        body.pack(fill="both", expand=True, pady=((8 if UI.COMPACT else 14), 0))
        body.grid_columnconfigure(0, weight=2)
        body.grid_columnconfigure(1, weight=3)
        body.grid_rowconfigure(0, weight=1)

        left = Card(body, padx=14, pady=14)
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 6 if UI.COMPACT else 10))
        right = Card(body, padx=14, pady=14)
        right.grid(row=0, column=1, sticky="nsew", padx=((6 if UI.COMPACT else 10), 0))

        tk.Label(left.inner, text="Register (Shift)", font=UI.FONT_LG, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w")

        emp_row = tk.Frame(left.inner, bg=UI.CARD)
        emp_row.pack(fill="x", pady=(10, 6))
        tk.Label(emp_row, text="Employee", bg=UI.CARD, fg="#334155", width=12, anchor="w").pack(side="left")

        self.emp_cb = ttk.Combobox(emp_row, textvariable=self.employee_var, values=[], state="readonly", width=18)
        self.emp_cb.pack(side="left", padx=(0, 8))
        self._preferred_employee = ""
        self.emp_cb.bind("<<ComboboxSelected>>", self._remember_employee_selection)

        GhostButton(emp_row, "Add employee", self.add_employee_popup).pack(side="left")
        GhostButton(emp_row, "Remove employee", self.remove_employee_popup).pack(side="left", padx=(8, 0))

        btn_row = tk.Frame(left.inner, bg=UI.CARD)
        btn_row.pack(fill="x", pady=(10, 0))
        PrimaryButton(btn_row, "Open Register", self.open_shift_clicked).pack(side="left")
        GhostButton(btn_row, "Refresh", self.refresh_all).pack(side="left", padx=10)
        tk.Frame(left.inner, bg=UI.BORDER, height=1).pack(fill="x", pady=12)

        self.shift_status = tk.Label(left.inner, text="", bg=UI.CARD, fg=UI.TEXT, font=("Segoe UI", 11 if UI.COMPACT else 14, "bold"))
        self.shift_status.pack(anchor="w", pady=(0, 8))

        self.shift_info = tk.Label(left.inner, text="", bg=UI.CARD, fg="#334155", justify="left",
                                   font=("Segoe UI", 9 if UI.COMPACT else 13))
        self.shift_info.pack(anchor="w")

        drawer_now_box = tk.Frame(left.inner, bg="#eef6ff", highlightthickness=1, highlightbackground="#bfdbfe")
        drawer_now_box.pack(fill="x", pady=(12, 0))
        tk.Label(
            drawer_now_box,
            text="Should be in register now",
            bg="#eef6ff",
            fg=UI.MUTED,
            font=UI.FONT_SM,
        ).pack(anchor="w", padx=12, pady=(10, 0))
        self.drawer_now_value_lbl = tk.Label(
            drawer_now_box,
            text="-",
            bg="#eef6ff",
            fg=UI.PRIMARY,
            font=("Segoe UI", 16 if UI.COMPACT else 24, "bold"),
        )
        self.drawer_now_value_lbl.pack(anchor="w", padx=12, pady=(2, 0))
        self.drawer_opening_lbl = tk.Label(
            drawer_now_box,
            text="",
            bg="#eef6ff",
            fg="#0f172a",
            font=("Segoe UI", 10 if UI.COMPACT else 12, "bold"),
            justify="left",
        )
        self.drawer_opening_lbl.pack(anchor="w", padx=12, pady=(2, 2))
        self.drawer_now_sub_lbl = tk.Label(
            drawer_now_box,
            text="",
            bg="#eef6ff",
            fg=UI.MUTED,
            justify="left",
            wraplength=460,
            font=UI.FONT_SM,
        )
        self.drawer_now_sub_lbl.pack(anchor="w", fill="x", padx=12, pady=(0, 10))

        self.drawer_metric_labels = {}
        metric_grid = tk.Frame(left.inner, bg=UI.CARD)
        metric_grid.pack(fill="x", pady=(10, 0))
        for c in range(3):
            metric_grid.grid_columnconfigure(c, weight=1, uniform="drawer_metric")

        metric_defs = [
            ("revenue", "Register Sales"),
            ("daily_sales", "Daily Sales"),
            ("cash_sales", "Register Cash"),
            ("cash_in", "Cash In"),
            ("cash_out", "Cash Out"),
            ("refunds", "Refunds"),
            ("orders", "Register Orders"),
            ("daily_orders", "Daily Orders"),
            ("daily_cash", "Daily Cash"),
        ]
        for idx, (key, title) in enumerate(metric_defs):
            cell = tk.Frame(metric_grid, bg="#f8fafc", highlightthickness=1, highlightbackground=UI.BORDER)
            cell.grid(row=idx // 3, column=idx % 3, sticky="nsew", padx=(0 if idx % 3 == 0 else 8, 0), pady=(0, 8))
            tk.Label(cell, text=title, bg="#f8fafc", fg=UI.MUTED, font=UI.FONT_SM).pack(anchor="w", padx=10, pady=(8, 0))
            value_lbl = tk.Label(cell, text="-", bg="#f8fafc", fg=UI.TEXT, font=("Segoe UI", 13, "bold"))
            value_lbl.pack(anchor="w", padx=10, pady=(2, 8))
            self.drawer_metric_labels[key] = value_lbl

        self.drawer_formula_lbl = tk.Label(
            left.inner,
            text="",
            bg=UI.CARD,
            fg=UI.MUTED,
            justify="left",
            wraplength=460,
            font=UI.FONT_SM,
        )
        self.drawer_formula_lbl.pack(anchor="w", fill="x", pady=(0, 8))

        movement_row = tk.Frame(left.inner, bg=UI.CARD)
        movement_row.pack(anchor="w", pady=(14, 0))
        DangerButton(movement_row, "Close Register", self.close_shift_clicked).pack(side="left")
        GhostButton(movement_row, "Add Cash In", self.cash_in_clicked).pack(side="left", padx=(8, 0))
        GhostButton(movement_row, "Take Cash Out", self.cash_out_clicked).pack(side="left", padx=(8, 0))

        tk.Frame(left.inner, bg=UI.BORDER, height=1).pack(fill="x", pady=12)
        tk.Label(left.inner, text="Current Shift Activity", font=UI.FONT_MD, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w")
        activity_wrap = tk.Frame(left.inner, bg=UI.CARD)
        activity_wrap.pack(fill="both", expand=True, pady=(8, 0))
        act_cols = ("time", "type", "amount", "drawer", "note")
        self.shift_activity_tree = ttk.Treeview(activity_wrap, columns=act_cols, show="headings", height=7)
        self.shift_activity_tree.heading("time", text="Time")
        self.shift_activity_tree.heading("type", text="Activity")
        self.shift_activity_tree.heading("amount", text="Amount")
        self.shift_activity_tree.heading("drawer", text="Drawer")
        self.shift_activity_tree.heading("note", text="Note")
        self.shift_activity_tree.column("time", width=78, anchor="center")
        self.shift_activity_tree.column("type", width=110)
        self.shift_activity_tree.column("amount", width=90, anchor="e")
        self.shift_activity_tree.column("drawer", width=90, anchor="e")
        self.shift_activity_tree.column("note", width=180)
        act_scroll = ttk.Scrollbar(activity_wrap, orient="vertical", command=self.shift_activity_tree.yview)
        self.shift_activity_tree.configure(yscrollcommand=act_scroll.set)
        self.shift_activity_tree.pack(side="left", fill="both", expand=True)
        act_scroll.pack(side="right", fill="y")

        tk.Label(right.inner, text="Daily Sales History", font=UI.FONT_LG, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w")

        top = tk.Frame(right.inner, bg=UI.CARD)
        top.pack(fill="x", pady=(10, 10))
        tk.Label(top, text="Date", bg=UI.CARD, fg="#334155").pack(side="left")
        years = [str(y) for y in range(datetime.now().year - 3, datetime.now().year + 1)]
        months = [f"{m:02d}" for m in range(1, 13)]
        days = [f"{d:02d}" for d in range(1, 32)]
        ttk.Combobox(top, textvariable=self.sales_year, values=years, width=6, state="readonly").pack(side="left",
                                                                                                      padx=(8, 6))
        ttk.Combobox(top, textvariable=self.sales_month, values=months, width=4, state="readonly").pack(side="left",
                                                                                                        padx=(0, 6))
        ttk.Combobox(top, textvariable=self.sales_day, values=days, width=4, state="readonly").pack(side="left",
                                                                                                    padx=(0, 10))
        GhostButton(top, "Load", self.load_day).pack(side="left")
        GhostButton(top, "Today", self.set_today).pack(side="left", padx=8)
        DangerButton(top, "Void Sale", self.delete_selected_sale).pack(side="left")
        ttk.Checkbutton(
            top, text="Show voided", variable=self.show_voided_var,
            command=lambda: self.load_day(silent=True),
        ).pack(side="left", padx=(8, 0))

        rep = tk.Frame(right.inner, bg=UI.CARD)
        rep.pack(fill="x", pady=(0, 10))
        tk.Label(rep, text="Sale ID / Receipt Scan", bg=UI.CARD, fg="#334155").pack(side="left")
        ent = tk.Entry(rep, textvariable=self.reprint_var, bd=1, relief="solid", width=28)
        ent.pack(side="left", padx=8)
        PrimaryButton(rep, "Reprint Receipt", self.reprint_receipt_clicked).pack(side="left")
        GhostButton(rep, "Gift Receipt", self.reprint_gift_receipt_clicked).pack(side="left", padx=(8, 0))
        GhostButton(rep, "View Details", self.view_details_clicked).pack(side="left", padx=8)
        GhostButton(rep, "Find", self.lookup_sale_clicked).pack(side="left")
        GhostButton(rep, "Search All", self.open_sales_search).pack(side="left", padx=(8, 0))
        ent.bind("<Return>", lambda e: self.lookup_sale_clicked())

        cols = ("receipt", "time", "total", "pay", "shift")
        self.sales_tree = ttk.Treeview(right.inner, columns=cols, show="headings", height=10)
        self.sales_tree.heading("receipt", text="Sale ID")
        self.sales_tree.heading("time", text="Time")
        self.sales_tree.heading("total", text="Total")
        self.sales_tree.heading("pay", text="Pay")
        self.sales_tree.heading("shift", text="Shift")
        self.sales_tree.column("receipt", width=80, anchor="center")
        self.sales_tree.column("time", width=140)
        self.sales_tree.column("total", width=110, anchor="e")
        self.sales_tree.column("pay", width=80, anchor="center")
        self.sales_tree.column("shift", width=70, anchor="center")
        self.sales_tree.pack(fill="both", expand=True)
        self.sales_tree.tag_configure("voided", foreground="#991b1b", background="#fee2e2")

        self.sales_tree.bind("<Double-1>", self.open_sale_detail)

        self.day_total_lbl = tk.Label(
            right.inner,
            text="",
            bg=UI.CARD,
            fg=UI.TEXT,
            font=("Segoe UI", 9 if UI.COMPACT else 11, "bold"),
            justify="left",
            anchor="w",
            wraplength=520,
        )
        self.day_total_lbl.pack(fill="x", anchor="w", pady=(10, 0))

        def _wrap_day_total(_evt=None):
            try:
                self.day_total_lbl.configure(wraplength=max(260, right.inner.winfo_width() - 12))
            except Exception:
                pass

        right.inner.bind("<Configure>", _wrap_day_total, add="+")
        self.after(100, _wrap_day_total)

        def _wrap_shift_layout(_evt=None):
            try:
                w = body.winfo_width()
                if w and w < (1150 if UI.COMPACT else 900):
                    body.grid_columnconfigure(0, weight=1)
                    body.grid_columnconfigure(1, weight=1)
                    body.grid_rowconfigure(0, weight=0)
                    body.grid_rowconfigure(1, weight=1)
                    left.grid_configure(row=0, column=0, sticky="ew", padx=0, pady=(0, 10))
                    right.grid_configure(row=1, column=0, sticky="nsew", padx=0, pady=0)
                else:
                    body.grid_columnconfigure(0, weight=2)
                    body.grid_columnconfigure(1, weight=3)
                    body.grid_rowconfigure(0, weight=1)
                    body.grid_rowconfigure(1, weight=0)
                    left.grid_configure(row=0, column=0, sticky="nsew", padx=(0, 10), pady=0)
                    right.grid_configure(row=0, column=1, sticky="nsew", padx=(10, 0), pady=0)
            except Exception:
                pass

        body.bind("<Configure>", _wrap_shift_layout)
        self.after(80, _wrap_shift_layout)

        self.refresh_all()
        self.load_day()

    def toggle_drawer_fullscreen(self):
        top = self.winfo_toplevel()
        if hasattr(top, "toggle_fullscreen"):
            return top.toggle_fullscreen()
        return "break"

    def _open_seasonal_sale_manager(self):
        return _seasonal_sale_manager_window(self)

    def _open_bundle_offer_manager(self):
        return _bundle_offer_manager_window(self)

    def refresh_employees(self):
        emps = list_employees(True)
        names = [e["name"] for e in emps]
        self.emp_cb["values"] = names
        current = self.employee_var.get().strip()
        if current and current in names:
            self._preferred_employee = current
            return
        preferred = str(getattr(self, "_preferred_employee", "") or "").strip()
        self.employee_var.set(preferred if preferred in names else (names[0] if names else ""))

    def _remember_employee_selection(self, _event=None):
        selected = self.employee_var.get().strip()
        if selected:
            self._preferred_employee = selected

    def _set_drawer_metric(self, key: str, value: str, fg=None):
        lbl = getattr(self, "drawer_metric_labels", {}).get(key)
        if not lbl:
            return
        try:
            lbl.config(text=str(value), fg=(fg or UI.TEXT))
        except Exception:
            pass

    def _reset_drawer_session_view(self):
        for key in getattr(self, "drawer_metric_labels", {}).keys():
            self._set_drawer_metric(key, "-")
        try:
            self.drawer_now_value_lbl.config(text="-")
            self.drawer_opening_lbl.config(text="")
            self.drawer_now_sub_lbl.config(text="")
        except Exception:
            pass
        try:
            self.drawer_formula_lbl.config(text="")
        except Exception:
            pass
        try:
            self.shift_activity_tree.delete(*self.shift_activity_tree.get_children())
            self.shift_activity_tree.insert("", tk.END, values=("", "No open register", "", "", "Open a register to begin tracking cash."))
        except Exception:
            pass

    def _activity_time_label(self, created_at):
        text = str(created_at or "")
        try:
            return datetime.strptime(text, "%Y-%m-%d %H:%M:%S").strftime("%I:%M %p").lstrip("0")
        except Exception:
            return text[11:16] if len(text) >= 16 else text

    def _open_shift_days(self, sh) -> list[str]:
        days = []
        opened = str(row_get(sh, "opened_at", "") or "")[:10]
        today = datetime.now().strftime("%Y-%m-%d")
        for d in (opened, today, self.date_var.get().strip()):
            if d and d not in days:
                try:
                    datetime.strptime(d, "%Y-%m-%d")
                    days.append(d)
                except Exception:
                    pass
        return days or [today]

    def _selected_sales_day(self) -> str:
        day = str(self.date_var.get() or "").strip()
        if not day:
            day = f"{self.sales_year.get()}-{self.sales_month.get()}-{self.sales_day.get()}"
        try:
            datetime.strptime(day, "%Y-%m-%d")
            return day
        except Exception:
            return datetime.now().strftime("%Y-%m-%d")

    def _shift_days_before(self, sh, selected_day: str) -> list[str]:
        opened = str(row_get(sh, "opened_at", "") or "")[:10]
        try:
            cur = datetime.strptime(opened, "%Y-%m-%d").date()
            end = datetime.strptime(selected_day, "%Y-%m-%d").date()
        except Exception:
            return []
        days = []
        while cur < end:
            days.append(cur.strftime("%Y-%m-%d"))
            cur += timedelta(days=1)
        return days

    def _sale_credit_used_amount(self, sale_row) -> float:
        credit_used = row_get(sale_row, "store_credit_used", None)
        if credit_used is not None:
            try:
                return max(0.0, float(credit_used or 0.0))
            except Exception:
                pass
        notes = str(row_get(sale_row, "notes", "") or row_get(sale_row, "note", "") or "")
        for part in notes.split(";"):
            part = part.strip()
            if part.startswith("EXCHANGE_CREDIT_APPLIED="):
                try:
                    return max(0.0, float(part.split("=", 1)[1] or 0.0))
                except Exception:
                    return 0.0
            if part.startswith("BON_CREDIT_APPLIED="):
                try:
                    return max(0.0, float(part.split("=", 1)[1] or 0.0))
                except Exception:
                    return 0.0
        return 0.0

    def _sale_cash_amount(self, sale_row) -> float:
        pm = str(row_get(sale_row, "payment_method", "") or "").strip().upper()
        if pm in ("EXCHANGE", "STORE_CREDIT", "CARD", "DEBIT", "CREDIT_CARD", "WHISH"):
            return 0.0
        cash_paid = row_get(sale_row, "cash_paid", None)
        if cash_paid is not None:
            try:
                return max(0.0, float(cash_paid or 0.0))
            except Exception:
                pass
        gross = row_get(sale_row, "total_sales", None)
        if gross is None:
            gross = row_get(sale_row, "net_sales", None)
        if gross is None:
            gross = row_get(sale_row, "total_amount", 0.0)
        try:
            return max(0.0, float(gross or 0.0) - self._sale_credit_used_amount(sale_row))
        except Exception:
            return 0.0

    def _sale_credit_used_amount(self, sale_row) -> float:
        credit_used = row_get(sale_row, "store_credit_used", None)
        if credit_used is not None:
            try:
                return max(0.0, float(credit_used or 0.0))
            except Exception:
                pass
        notes = str(row_get(sale_row, "notes", "") or row_get(sale_row, "note", "") or "")
        for part in notes.split(";"):
            part = part.strip()
            if part.startswith("EXCHANGE_CREDIT_APPLIED=") or part.startswith("BON_CREDIT_APPLIED="):
                try:
                    return max(0.0, float(part.split("=", 1)[1] or 0.0))
                except Exception:
                    return 0.0
        return 0.0

    def _sale_new_money_amount(self, sale_row) -> float:
        pm = str(row_get(sale_row, "payment_method", "") or "").strip().upper()
        if pm in ("EXCHANGE", "STORE_CREDIT"):
            return 0.0
        try:
            total_amount = float(row_get(sale_row, "total_amount", 0.0) or 0.0)
        except Exception:
            total_amount = 0.0
        try:
            gross = float(row_get(sale_row, "total_sales", 0.0) or 0.0)
        except Exception:
            gross = 0.0
        credit = self._sale_credit_used_amount(sale_row)
        if gross > 0 and credit > 0 and abs(total_amount - gross) < 0.005:
            return max(0.0, gross - credit)
        return max(0.0, total_amount)

    def _sale_merchandise_amount(self, sale_row) -> float:
        for key in ("total_sales", "net_sales"):
            val = row_get(sale_row, key, None)
            if val is not None:
                try:
                    return max(0.0, float(val or 0.0))
                except Exception:
                    pass
        try:
            total_amount = float(row_get(sale_row, "total_amount", 0.0) or 0.0)
        except Exception:
            total_amount = 0.0
        return max(0.0, total_amount + self._sale_credit_used_amount(sale_row))

    def _sale_new_money_amount(self, sale_row) -> float:
        pm = str(row_get(sale_row, "payment_method", "") or "").strip().upper()
        if pm in ("EXCHANGE", "STORE_CREDIT"):
            return 0.0
        try:
            total_amount = float(row_get(sale_row, "total_amount", 0.0) or 0.0)
        except Exception:
            total_amount = 0.0
        try:
            gross = float(row_get(sale_row, "total_sales", 0.0) or 0.0)
        except Exception:
            gross = 0.0
        credit = self._sale_credit_used_amount(sale_row)
        if gross > 0 and credit > 0 and abs(total_amount - gross) < 0.005:
            return max(0.0, gross - credit)
        return max(0.0, total_amount)

    def _drawer_sales_totals_from_history(self, shift_id: int, days) -> dict:
        totals = {"orders": 0, "cash_sales": 0.0, "merchandise_sales": 0.0, "new_money_sales": 0.0}
        seen_sale_ids = set()
        for day in days:
            try:
                sales = list_sales_for_day(day, limit=2000) or []
            except Exception:
                continue
            for sale in sales:
                try:
                    if int(row_get(sale, "shift_id", 0) or 0) != int(shift_id):
                        continue
                except Exception:
                    continue
                sale_id = row_get(sale, "id", None)
                if sale_id is not None:
                    sale_key = str(sale_id)
                    if sale_key in seen_sale_ids:
                        continue
                    seen_sale_ids.add(sale_key)
                totals["orders"] += 1
                totals["cash_sales"] += self._sale_cash_amount(sale)
                totals["merchandise_sales"] += self._sale_merchandise_amount(sale)
                totals["new_money_sales"] += self._sale_new_money_amount(sale)
        totals["cash_sales"] = round(float(totals["cash_sales"]), 2)
        totals["merchandise_sales"] = round(float(totals["merchandise_sales"]), 2)
        totals["new_money_sales"] = round(float(totals["new_money_sales"]), 2)
        return totals

    def _drawer_day_sales_totals(self, days) -> dict:
        totals = {"orders": 0, "cash_sales": 0.0, "merchandise_sales": 0.0, "new_money_sales": 0.0}
        seen_sale_ids = set()
        for day in days:
            try:
                sales = list_sales_for_day(day, limit=3000) or []
            except Exception:
                continue
            for sale in sales:
                sale_id = row_get(sale, "id", None)
                if sale_id is not None:
                    sale_key = str(sale_id)
                    if sale_key in seen_sale_ids:
                        continue
                    seen_sale_ids.add(sale_key)
                totals["orders"] += 1
                totals["cash_sales"] += self._sale_cash_amount(sale)
                totals["merchandise_sales"] += self._sale_merchandise_amount(sale)
                totals["new_money_sales"] += self._sale_new_money_amount(sale)
        totals["cash_sales"] = round(float(totals["cash_sales"]), 2)
        totals["merchandise_sales"] = round(float(totals["merchandise_sales"]), 2)
        totals["new_money_sales"] = round(float(totals["new_money_sales"]), 2)
        return totals

    def _drawer_cash_movements_from_history(self, shift_id: int, days) -> dict:
        totals = {"cash_in_value": 0.0, "cash_in_count": 0, "cash_out_value": 0.0, "cash_out_count": 0}
        seen = set()
        for day in days:
            try:
                movements = list_cash_movements(day_str=day, limit=1000) or []
            except Exception:
                continue
            for movement in movements:
                try:
                    if int(row_get(movement, "shift_id", 0) or 0) != int(shift_id):
                        continue
                except Exception:
                    continue
                movement_id = row_get(movement, "id", None)
                if movement_id is not None:
                    key = str(movement_id)
                    if key in seen:
                        continue
                    seen.add(key)
                mtype = str(row_get(movement, "movement_type", "OUT") or "OUT").upper()
                try:
                    amount = float(row_get(movement, "amount_value", 0.0) or 0.0)
                except Exception:
                    amount = 0.0
                if mtype == "IN":
                    totals["cash_in_value"] += amount
                    totals["cash_in_count"] += 1
                elif mtype == "OUT":
                    totals["cash_out_value"] += amount
                    totals["cash_out_count"] += 1
        totals["cash_in_value"] = round(float(totals["cash_in_value"]), 2)
        totals["cash_out_value"] = round(float(totals["cash_out_value"]), 2)
        return totals

    def _dominant_shift_from_sales(self, sales, fallback_shift_id: int) -> int:
        counts = {}
        for sale in sales or []:
            try:
                sid = int(row_get(sale, "shift_id", 0) or 0)
            except Exception:
                sid = 0
            if sid <= 0:
                continue
            counts[sid] = counts.get(sid, 0) + 1
        if not counts:
            return int(fallback_shift_id or 0)
        return sorted(counts.items(), key=lambda item: (item[1], item[0]), reverse=True)[0][0]

    def _shift_row_for_id(self, shift_id: int, fallback=None):
        try:
            for sh in list_shifts(250) or []:
                try:
                    if int(row_get(sh, "id", 0) or 0) == int(shift_id):
                        return sh
                except Exception:
                    continue
        except Exception:
            pass
        if fallback is not None:
            try:
                if int(row_get(fallback, "id", 0) or 0) == int(shift_id):
                    return fallback
            except Exception:
                pass
        return {
            "id": int(shift_id or 0),
            "opened_at": f"{self._selected_sales_day()} 00:00:00",
            "closed_at": None,
            "opening_cash": 0.0,
            "opening_usd": 0.0,
            "opening_lbp": 0.0,
            "lbp_per_usd": get_lbp_per_usd(),
            "employee_name": row_get(fallback, "employee_name", "") if fallback is not None else "",
            "shift_code": str(shift_id or ""),
            "shift_seq": shift_id,
        }

    def _effective_drawer_state(self, fallback_shift=None) -> dict:
        fallback_shift = fallback_shift or get_open_shift() or {}
        try:
            fallback_sid = int(row_get(fallback_shift, "id", 0) or 0)
        except Exception:
            fallback_sid = 0
        selected_day = self._selected_sales_day()
        try:
            selected_sales = list_sales_for_day(selected_day, limit=2000) or []
        except Exception:
            selected_sales = []
        if fallback_sid > 0 and row_get(fallback_shift, "closed_at", None) is None:
            sid = fallback_sid
        else:
            sid = self._dominant_shift_from_sales(selected_sales, fallback_sid)
        sh = self._shift_row_for_id(sid, fallback_shift)
        try:
            summ = shift_summary(sid) or {}
        except Exception:
            summ = {}
        opening_value = float(summ.get("opening_cash", row_get(sh, "opening_cash", 0.0)) or 0.0)
        prior_days = self._shift_days_before(sh, selected_day)
        prior_sales = self._drawer_sales_totals_from_history(sid, prior_days)
        prior_movements = self._drawer_cash_movements_from_history(sid, prior_days)
        day_opening = round(
            opening_value
            + float(prior_sales.get("cash_sales", 0.0) or 0.0)
            + float(prior_movements.get("cash_in_value", 0.0) or 0.0)
            - float(prior_movements.get("cash_out_value", 0.0) or 0.0),
            2,
        )
        day_sales = self._drawer_sales_totals_from_history(sid, [selected_day])
        day_movements = self._drawer_cash_movements_from_history(sid, [selected_day])
        cash_refunds = float(summ.get("cash_refunds", 0.0) or 0.0)
        expected = round(
            day_opening
            + float(day_sales.get("cash_sales", 0.0) or 0.0)
            - cash_refunds
            + float(day_movements.get("cash_in_value", 0.0) or 0.0)
            - float(day_movements.get("cash_out_value", 0.0) or 0.0),
            2,
        )
        return {
            "sid": sid,
            "shift": sh,
            "summary": summ,
            "selected_day": selected_day,
            "day_opening": day_opening,
            "sales": day_sales,
            "movements": day_movements,
            "cash_refunds": cash_refunds,
            "expected": expected,
        }

    def _refresh_shift_activity(self, sh, day_opening=None):
        try:
            self.shift_activity_tree.delete(*self.shift_activity_tree.get_children())
        except Exception:
            return

        try:
            sid = int(row_get(sh, "id", 0) or 0)
        except Exception:
            sid = 0
        if sid <= 0:
            return

        selected_day = self._selected_sales_day()
        rows = []
        try:
            opening_delta = float(day_opening if day_opening is not None else row_get(sh, "opening_cash", 0.0) or 0.0)
        except Exception:
            opening_delta = 0.0
        rows.append({
            "created_at": f"{selected_day} 00:00:00",
            "type": "Day Opening" if day_opening is not None else "Register Open",
            "delta": opening_delta,
            "amount": money(opening_delta),
            "note": str(row_get(sh, "employee_name", "") or "Opening count"),
        })
        for day in [selected_day]:
            try:
                for sale in list_sales_for_day(day, limit=1500) or []:
                    try:
                        if int(row_get(sale, "shift_id", 0) or 0) != sid:
                            continue
                    except Exception:
                        continue
                    receipt = str(row_get(sale, "receipt_code", "") or row_get(sale, "id", "") or "")
                    amount = self._sale_cash_amount(sale)
                    rows.append({
                        "created_at": row_get(sale, "created_at", ""),
                        "type": f"Sale {receipt}",
                        "delta": amount,
                        "amount": money(amount),
                        "note": str(row_get(sale, "payment_method", "") or ""),
                    })
            except Exception:
                pass

        try:
            movements = list_cash_movements(day_str=selected_day, limit=1000) or []
        except Exception:
            movements = []
        for m in movements:
            try:
                if int(row_get(m, "shift_id", 0) or 0) != sid:
                    continue
            except Exception:
                continue
            mtype = str(row_get(m, "movement_type", "OUT") or "OUT").upper()
            amount = float(row_get(m, "amount_value", 0.0) or 0.0)
            delta = amount if mtype == "IN" else -amount
            rows.append({
                "created_at": row_get(m, "created_at", ""),
                "type": "Cash In" if mtype == "IN" else "Cash Out",
                "delta": delta,
                "amount": money(amount) if mtype == "IN" else f"-{money(amount)}",
                "note": str(row_get(m, "reason", "") or row_get(m, "employee_name", "") or ""),
            })

        rows.sort(key=lambda r: str(r.get("created_at") or ""))
        running = 0.0
        for r in rows:
            try:
                running += float(r.get("delta") or 0.0)
            except Exception:
                pass
            r["drawer"] = money(running)

        if not rows:
            self.shift_activity_tree.insert("", tk.END, values=("", "No activity yet", "", "", "Sales and cash in/out will appear here."))
            return
        for r in list(reversed(rows))[:120]:
            self.shift_activity_tree.insert("", tk.END, values=(
                self._activity_time_label(r.get("created_at")),
                r.get("type", ""),
                r.get("amount", ""),
                r.get("drawer", ""),
                r.get("note", ""),
            ))

    def refresh_all(self):
        try:
            self.refresh_employees()
        except Exception as e:
            messagebox.showerror("Employees", f"{type(e).__name__}: {e}")

        sh = get_open_shift()
        if not sh:
            self.shift_status.config(text="No open shift")
            self.shift_info.config(text="Open a register to track drawer cash, paid in/out, and close variance.")
            self._reset_drawer_session_view()
            return

        open_sid = int(row_get(sh, "id", 0) or 0)
        selected_day = self._selected_sales_day()
        sid = open_sid
        sh = self._shift_row_for_id(sid, sh)
        try:
            summ = shift_summary(sid) or {}
        except Exception:
            summ = {}
        expected = float(summ.get("expected_cash", 0.0) or 0.0)
        orders = int(summ.get("orders", 0) or 0)
        cash_sales = float(summ.get("cash_sales", 0.0) or 0.0)
        revenue_value = float(summ.get("net_revenue", summ.get("new_money_sales", 0.0)) or 0.0)
        cash_refunds = float(summ.get("cash_refunds", 0.0) or 0.0)
        emp_name = row_get(sh, "employee_name", "") or "Unassigned"
        shift_label = str(row_get(sh, "shift_code", "") or row_get(sh, "shift_seq", "") or sid).strip()

        status_prefix = "Open shift" if int(sid or 0) == int(open_sid or 0) else "Daily shift"
        self.shift_status.config(text=f"{status_prefix} {shift_label}  |  Employee: {emp_name}")
        rate = float(summ.get("lbp_per_usd", row_get(sh, "lbp_per_usd", 0)) or 0)
        opening_usd = float(summ.get("opening_usd", row_get(sh, "opening_usd", summ.get("opening_cash", row_get(sh, "opening_cash", 0)))) or 0)
        opening_lbp = float(summ.get("opening_lbp", row_get(sh, "opening_lbp", 0)) or 0)
        opening_value = float(summ.get("opening_cash", row_get(sh, "opening_cash", 0.0)) or 0.0)
        cash_out_value = float(summ.get("cash_out_value", 0.0) or 0.0)
        cash_out_count = int(summ.get("cash_out_count", 0) or 0)
        cash_in_value = float(summ.get("cash_in_value", 0.0) or 0.0)
        cash_in_count = int(summ.get("cash_in_count", 0) or 0)
        prior_days = self._shift_days_before(sh, selected_day)
        prior_sales = self._drawer_sales_totals_from_history(sid, prior_days)
        prior_movements = self._drawer_cash_movements_from_history(sid, prior_days)
        day_opening_value = round(
            opening_value
            + float(prior_sales.get("cash_sales", 0.0) or 0.0)
            + float(prior_movements.get("cash_in_value", 0.0) or 0.0)
            - float(prior_movements.get("cash_out_value", 0.0) or 0.0),
            2,
        )
        history_totals = self._drawer_sales_totals_from_history(sid, [selected_day])
        daily_totals = self._drawer_day_sales_totals([selected_day])
        day_movements = self._drawer_cash_movements_from_history(sid, [selected_day])
        cash_sales = float(history_totals.get("cash_sales", cash_sales) or 0.0)
        revenue_value = float(history_totals.get("new_money_sales", revenue_value) or 0.0)
        orders = int(history_totals.get("orders", orders) or 0)
        daily_sales_value = float(daily_totals.get("new_money_sales", 0.0) or 0.0)
        daily_cash_sales_value = float(daily_totals.get("cash_sales", 0.0) or 0.0)
        daily_orders = int(daily_totals.get("orders", 0) or 0)
        cash_in_value = float(day_movements.get("cash_in_value", 0.0) or 0.0)
        cash_in_count = int(day_movements.get("cash_in_count", 0) or 0)
        cash_out_value = float(day_movements.get("cash_out_value", 0.0) or 0.0)
        cash_out_count = int(day_movements.get("cash_out_count", 0) or 0)
        expected = round(day_opening_value + cash_sales - cash_refunds + cash_in_value - cash_out_value, 2)

        try:
            self.drawer_now_value_lbl.config(text=drawer_money(expected))
            self.drawer_opening_lbl.config(text=f"Opening drawer: {drawer_money(opening_usd)} + {lbp_money(opening_lbp)}")
            self.drawer_now_sub_lbl.config(
                text=(
                    f"Open {self._activity_time_label(row_get(sh, 'opened_at', ''))} | "
                    f"register {drawer_money(revenue_value)} | day {drawer_money(daily_sales_value)}"
                )
            )
        except Exception:
            pass

        self._set_drawer_metric("revenue", drawer_money(revenue_value), UI.SUCCESS if revenue_value > 0 else None)
        self._set_drawer_metric("daily_sales", drawer_money(daily_sales_value), UI.SUCCESS if daily_sales_value > 0 else None)
        self._set_drawer_metric("cash_sales", drawer_money(cash_sales), UI.SUCCESS if cash_sales > 0 else None)
        self._set_drawer_metric("cash_in", drawer_money(cash_in_value), UI.SUCCESS if cash_in_value > 0 else None)
        self._set_drawer_metric("cash_out", drawer_money(cash_out_value), UI.DANGER if cash_out_value > 0 else None)
        self._set_drawer_metric("refunds", drawer_money(cash_refunds), UI.DANGER if cash_refunds > 0 else None)
        self._set_drawer_metric("orders", str(orders))
        self._set_drawer_metric("daily_orders", str(daily_orders))
        self._set_drawer_metric("daily_cash", drawer_money(daily_cash_sales_value), UI.SUCCESS if daily_cash_sales_value > 0 else None)
        try:
            self.drawer_formula_lbl.config(
                text=(
                    f"Register now = day opening {drawer_money(day_opening_value)} + cash sales {drawer_money(cash_sales)} "
                    f"- cash refunds {drawer_money(cash_refunds)} + cash in {drawer_money(cash_in_value)} "
                    f"- cash out {drawer_money(cash_out_value)}."
                )
            )
        except Exception:
            pass

        lines = [
            f"Opened: {row_get(sh, 'opened_at', '')}",
            f"Opening drawer: {drawer_money(opening_usd)} + {lbp_money(opening_lbp)}"
            + (f"  @ {lbp_money(rate)}/$" if rate > 0 else ""),
            f"Day opening basis ({selected_day}): {drawer_money(day_opening_value)}",
            f"Current register: {drawer_money(revenue_value)} new money  |  {drawer_money(cash_sales)} cash  |  Orders: {orders}",
            f"Full day: {drawer_money(daily_sales_value)} new money  |  {drawer_money(daily_cash_sales_value)} cash  |  Orders: {daily_orders}",
        ]
        if cash_refunds > 0.005:
            lines.append(f"Cash refunds: {drawer_money(cash_refunds)}")
        if cash_out_value > 0.005:
            lines.append(f"Cash removed: {drawer_money(cash_out_value)}  |  Entries: {cash_out_count}")
        if cash_in_value > 0.005:
            lines.append(f"Cash added: {drawer_money(cash_in_value)}  |  Entries: {cash_in_count}")
        lines.append(f"Should be in register now: {drawer_money(expected)}")
        if summ.get("closing_cash") is not None:
            lines.append(
                f"Closed drawer: {drawer_money(float(summ.get('closing_usd') or 0))} + "
                f"{lbp_money(float(summ.get('closing_lbp') or 0))}"
            )
            lines.append(f"Difference by value: {drawer_money(float(summ.get('difference') or 0))}")
        self.shift_info.config(text="\n".join(lines))
        self._refresh_shift_activity(sh, day_opening=day_opening_value)

    def add_employee_popup(self):
        win = tk.Toplevel(self)
        win.title("Add employee")
        win.geometry("360x220")
        win.configure(bg=UI.CONTENT_BG)
        win.grab_set()

        box = Card(win, padx=16, pady=16)
        box.pack(fill="both", expand=True, padx=14, pady=14)

        tk.Label(box.inner, text="Employee name", bg=UI.CARD, fg=UI.TEXT).pack(anchor="w")
        name_e = tk.Entry(box.inner, bd=1, relief="solid")
        name_e.pack(anchor="w", fill="x", pady=(6, 10))
        name_e.focus()

        tk.Label(box.inner, text="PIN (optional)", bg=UI.CARD, fg=UI.TEXT).pack(anchor="w")
        pin_e = tk.Entry(box.inner, bd=1, relief="solid", show="*")
        pin_e.pack(anchor="w", fill="x", pady=(6, 10))

        def save():
            name = name_e.get().strip()
            pin = pin_e.get().strip()
            if not name:
                messagebox.showerror("Missing", "Employee name required.")
                return
            try:
                ensure_employee(name, pin)
            except Exception as e:
                messagebox.showerror("Error", f"{type(e).__name__}: {e}")
                return
            self._preferred_employee = name
            self.employee_var.set(name)
            win.destroy()
            self.refresh_all()

        btns = tk.Frame(box.inner, bg=UI.CARD)
        btns.pack(fill="x", pady=(10, 0))
        GhostButton(btns, "Cancel", win.destroy).pack(side="right")
        PrimaryButton(btns, "Save", save).pack(side="right", padx=(0, 10))

    def remove_employee_popup(self):
        emp = self.employee_var.get().strip()
        if not emp:
            messagebox.showinfo("Remove employee", "Select an employee first.")
            return

        sh = get_open_shift()
        if sh and (sh.get("employee_name") or "").strip() == emp and sh.get("closed_at") is None:
            messagebox.showerror("Cannot remove", "This employee has the current open shift. Close the shift first.")
            return

        if not messagebox.askyesno(
                "Remove employee",
                f"Remove '{emp}' from the active employees list?\n\nThis will not delete past shift/sales history."
        ):
            return

        try:
            ok = deactivate_employee(emp)
        except Exception as e:
            messagebox.showerror("Error", f"{type(e).__name__}: {e}")
            return

        if not ok:
            messagebox.showinfo("Not found", "Employee not found.")
            return

        self.refresh_employees()
        self.refresh_all()

    def _prompt_pin(self, emp_name: str, action: str):
        win = tk.Toplevel(self)
        win.title("PIN")
        win.geometry("320x190")
        win.configure(bg=UI.CONTENT_BG)
        win.grab_set()

        body = Card(win, padx=16, pady=14)
        body.pack(fill="both", expand=True, padx=12, pady=12)

        tk.Label(body.inner, text=f"Enter PIN for {emp_name} to {action}:", bg=UI.CARD, fg=UI.TEXT, wraplength=260,
                 justify="left").pack(anchor="w")
        e = tk.Entry(body.inner, bd=1, relief="solid", show="*")
        e.pack(fill="x", pady=(10, 0))
        e.focus()

        result = {"pin": None}

        def _ok():
            result["pin"] = (e.get() or "").strip()
            try:
                win.grab_release()
            except Exception:
                pass
            win.destroy()

        def _cancel():
            result["pin"] = None
            try:
                win.grab_release()
            except Exception:
                pass
            win.destroy()

        btns = tk.Frame(body.inner, bg=UI.CARD)
        btns.pack(fill="x", pady=(12, 0))
        GhostButton(btns, "Cancel", _cancel).pack(side="right")
        PrimaryButton(btns, "OK", _ok).pack(side="right", padx=(0, 10))

        win.bind("<Return>", lambda e2: _ok())
        win.bind("<Escape>", lambda e2: _cancel())
        win.protocol("WM_DELETE_WINDOW", _cancel)

        try:
            win.update_idletasks()
            x = self.winfo_rootx() + (self.winfo_width() // 2) - (win.winfo_width() // 2)
            y = self.winfo_rooty() + (self.winfo_height() // 2) - (win.winfo_height() // 2)
            win.geometry(f"+{max(10, x)}+{max(10, y)}")
        except Exception:
            pass

        self.wait_window(win)
        return result["pin"]

    def open_shift_clicked(self):
        sh = get_open_shift()
        if sh:
            messagebox.showinfo("Open shift", "There is already an open shift. Close it first.")
            return

        emp = self.employee_var.get().strip()
        if not emp:
            messagebox.showerror("Employee", "Select an employee to open the register.")
            return

        # Popup: opening cash + PIN
        win = tk.Toplevel(self)
        win.title("Open Register")
        win.geometry("620x460")
        win.configure(bg=UI.CONTENT_BG)
        win.grab_set()

        box = Card(win, padx=16, pady=16)
        box.pack(fill="both", expand=True, padx=14, pady=14)

        try:
            default_rate = int(round(float(get_lbp_per_usd() or 89500)))
        except Exception:
            default_rate = 89500
        try:
            last_closed = get_last_closed_shift() or {}
        except Exception:
            last_closed = {}
        try:
            expected_opening_value = float(row_get(last_closed, "closing_cash", 0.0) or 0.0)
            expected_opening_usd = float(row_get(last_closed, "closing_usd", expected_opening_value) or 0.0)
            expected_opening_lbp = float(row_get(last_closed, "closing_lbp", 0.0) or 0.0)
            last_closed_label = str(row_get(last_closed, "shift_code", "") or row_get(last_closed, "id", "") or "").strip()
            last_closed_at = str(row_get(last_closed, "closed_at", "") or "").strip()
        except Exception:
            expected_opening_value = 0.0
            expected_opening_usd = 0.0
            expected_opening_lbp = 0.0
            last_closed_label = ""
            last_closed_at = ""

        tk.Label(box.inner, text=f"Employee: {emp}", bg=UI.CARD, fg=UI.TEXT, font=UI.FONT_MD).pack(anchor="w")
        if last_closed:
            tk.Label(
                box.inner,
                text=(
                    f"Should be present from last close: {drawer_money(expected_opening_value)} "
                    f"({drawer_money(expected_opening_usd)} + {lbp_money(expected_opening_lbp)})"
                    + (f" from shift {last_closed_label}" if last_closed_label else "")
                    + (f", closed {last_closed_at}" if last_closed_at else "")
                ),
                bg=UI.CARD,
                fg=UI.MUTED,
                wraplength=540,
                justify="left",
            ).pack(anchor="w", pady=(6, 0))

        tk.Label(box.inner, text="Opening drawer count", bg=UI.CARD, fg=UI.TEXT).pack(anchor="w", pady=(10, 0))
        count_grid = tk.Frame(box.inner, bg=UI.CARD)
        count_grid.pack(fill="x", pady=(6, 10))
        count_grid.grid_columnconfigure(1, weight=1, minsize=330)

        tk.Label(count_grid, text="USD", bg=UI.CARD, fg="#334155", width=12, anchor="w").grid(row=0, column=0, sticky="w", pady=3)
        opening_usd_e = tk.Entry(count_grid, bd=1, relief="solid", font=("Segoe UI", 18), justify="right")
        opening_usd_e.grid(row=0, column=1, sticky="ew", pady=3)
        opening_usd_e.insert(0, self.opening_var.get().strip())

        tk.Label(count_grid, text="LBP", bg=UI.CARD, fg="#334155", width=12, anchor="w").grid(row=1, column=0, sticky="w", pady=3)
        opening_lbp_e = tk.Entry(count_grid, bd=1, relief="solid", font=("Segoe UI", 18), justify="right")
        opening_lbp_e.grid(row=1, column=1, sticky="ew", pady=3)
        opening_lbp_e.insert(0, "0")
        bind_lbp_grouping(opening_lbp_e)

        opening_check_var = tk.StringVar(value="")
        tk.Label(
            box.inner,
            textvariable=opening_check_var,
            bg=UI.CARD,
            fg=UI.TEXT,
            justify="left",
            wraplength=540,
        ).pack(anchor="w", pady=(0, 8))

        def _update_opening_check(*_):
            try:
                opening_usd_now = parse_whole_money_text(opening_usd_e.get())
                opening_lbp_now = parse_lbp_text(opening_lbp_e.get())
                opening_now = round(opening_usd_now + (opening_lbp_now / int(default_rate)))
                if last_closed:
                    diff_now = opening_now - expected_opening_value
                    opening_check_var.set(
                        f"Opening check: counted {drawer_money(opening_now)}. "
                        f"Difference from last close: {drawer_money(diff_now)}."
                    )
                else:
                    opening_check_var.set(f"Opening count: {drawer_money(opening_now)}.")
            except Exception:
                opening_check_var.set("")

        opening_usd_e.bind("<KeyRelease>", _update_opening_check)
        opening_lbp_e.bind("<KeyRelease>", _update_opening_check, add="+")
        _update_opening_check()

        tk.Label(
            box.inner,
            text=f"Exchange rate: {lbp_money(default_rate)} per $1 (change it in Settings).",
            bg=UI.CARD,
            fg=UI.MUTED,
        ).pack(anchor="w", pady=(0, 10))

        pin_required = bool(employee_pin_required(emp))
        pin_label = "Enter PIN to open register" if pin_required else "Set PIN (optional)"
        tk.Label(box.inner, text=pin_label, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w")
        pin_e = tk.Entry(box.inner, bd=1, relief="solid", show="*")
        pin_e.pack(anchor="w", fill="x", pady=(6, 10))

        hint = "PIN is required for this employee." if pin_required else "Leave blank to continue without setting a PIN."
        tk.Label(box.inner, text=hint, bg=UI.CARD, fg="#64748b", wraplength=330, justify="left").pack(anchor="w",
                                                                                                      pady=(0, 6))

        def _do_open():
            try:
                opening_usd = parse_whole_money_text(opening_usd_e.get())
                opening_lbp = parse_lbp_text(opening_lbp_e.get())
                rate = int(default_rate)
            except Exception:
                messagebox.showerror("Invalid", "USD and LBP must be whole numbers.")
                return
            if opening_usd < 0 or opening_lbp < 0 or rate <= 0:
                messagebox.showerror("Invalid", "USD/LBP cannot be negative and rate must be greater than 0.")
                return
            opening = round(opening_usd + (opening_lbp / rate))
            opening_diff = opening - expected_opening_value if last_closed else 0.0
            if last_closed and abs(opening_diff) > 0.005:
                if not messagebox.askyesno(
                    "Opening count difference",
                    (
                        f"Last close left {drawer_money(expected_opening_value)}.\n"
                        f"You counted {drawer_money(opening)}.\n"
                        f"Difference: {drawer_money(opening_diff)}.\n\n"
                        "Open the register with this count and record the difference?"
                    ),
                    parent=win,
                ):
                    return

            pin = (pin_e.get() or "").strip()

            # Always verify PIN against backend before opening.
            # - If employee has NO PIN set: backend returns True even when pin is blank.
            # - If employee HAS a PIN set: blank/wrong pin returns False.
            if not verify_employee_pin(emp, pin):
                if not pin:
                    messagebox.showerror("PIN required", "Please enter the employee PIN.")
                else:
                    messagebox.showerror("Wrong PIN", "Incorrect PIN.")
                return

            # If employee had no PIN yet and this is NOT a JOIN terminal, allow setting it now.
            try:
                role = connection_role()
            except Exception:
                role = ""

            if role != "JOIN" and pin and (not employee_pin_required(emp)):
                try:
                    ensure_employee(emp, pin)
                except Exception as e:
                    messagebox.showerror("Error", f"{type(e).__name__}: {e}")
                    return

            try:
                open_notes = ""
                if last_closed:
                    open_notes = (
                        f"Opening check: previous shift {last_closed_label or row_get(last_closed, 'id', '')} "
                        f"left {drawer_money(expected_opening_value)}; counted {drawer_money(opening)}; "
                        f"difference {drawer_money(opening_diff)}."
                    )
                sid = open_shift(
                    opening_cash=opening,
                    notes=open_notes,
                    employee_name=emp,
                    opening_usd=opening_usd,
                    opening_lbp=opening_lbp,
                    lbp_per_usd=rate,
                )
            except Exception as e:
                messagebox.showerror("Error", f"{type(e).__name__}: {e}")
                return

            self.opening_var.set(f"{opening:g}")
            win.destroy()
            try:
                opened_shift = get_open_shift() or {}
                shift_label = str(row_get(opened_shift, "shift_code", "") or row_get(opened_shift, "shift_seq", "") or sid).strip()
            except Exception:
                shift_label = str(sid)
            messagebox.showinfo("Shift opened", f"Shift {shift_label} opened.")
            self.refresh_all()


        btns = tk.Frame(box.inner, bg=UI.CARD)
        btns.pack(fill="x", pady=(6, 0))
        GhostButton(btns, "Cancel", win.destroy).pack(side="right")
        PrimaryButton(btns, "Continue", _do_open).pack(side="right", padx=(0, 10))

        opening_usd_e.focus()
        win.bind("<Return>", lambda e: _do_open())
        win.bind("<Escape>", lambda e: win.destroy())

    def cash_in_clicked(self):
        self._cash_movement_clicked("IN")

    def cash_out_clicked(self):
        self._cash_movement_clicked("OUT")

    def _cash_movement_clicked(self, movement_type="OUT"):
        movement_type = str(movement_type or "OUT").upper()
        is_in = (movement_type == "IN")
        title = "Add Cash In" if is_in else "Take Cash Out"
        action_text = "added to" if is_in else "taken out of"
        sh = get_open_shift()
        if not sh:
            messagebox.showinfo("No shift", f"Open a register before cash is {action_text} the drawer.")
            return

        sid = int(row_get(sh, "id", 0) or 0)
        selected_day = self._selected_sales_day()
        # Cash in/out must always attach to the real open register. Do not switch
        # to an earlier same-day register just because it has more sales.
        sh = self._shift_row_for_id(sid, sh)
        summ = shift_summary(sid) or {}
        prior_days = self._shift_days_before(sh, selected_day)
        prior_sales = self._drawer_sales_totals_from_history(sid, prior_days)
        prior_movements = self._drawer_cash_movements_from_history(sid, prior_days)
        day_movements = self._drawer_cash_movements_from_history(sid, [selected_day])
        day_opening_value = round(
            float(summ.get("opening_cash", 0.0) or 0.0)
            + float(prior_sales.get("cash_sales", 0.0) or 0.0)
            + float(prior_movements.get("cash_in_value", 0.0) or 0.0)
            - float(prior_movements.get("cash_out_value", 0.0) or 0.0),
            2,
        )
        history_totals = self._drawer_sales_totals_from_history(sid, [selected_day])
        corrected_expected = round(
            day_opening_value
            + float(history_totals.get("cash_sales", summ.get("cash_sales", 0.0)) or 0.0)
            - float(summ.get("cash_refunds", 0.0) or 0.0)
            + float(day_movements.get("cash_in_value", 0.0) or 0.0)
            - float(day_movements.get("cash_out_value", 0.0) or 0.0),
            2,
        )
        try:
            rate_default = int(round(float(summ.get("lbp_per_usd") or get_lbp_per_usd() or 89500)))
        except Exception:
            rate_default = 89500

        shift_emp = str(row_get(sh, "employee_name", "") or "").strip()
        try:
            employee_names = [str(e["name"]) for e in list_employees(True)]
        except Exception:
            employee_names = []
        if shift_emp and shift_emp not in employee_names:
            employee_names.insert(0, shift_emp)

        win = tk.Toplevel(self)
        win.title(title)
        win.geometry("560x420")
        win.configure(bg=UI.CONTENT_BG)
        win.grab_set()

        box = Card(win, padx=16, pady=16)
        box.pack(fill="both", expand=True, padx=14, pady=14)

        tk.Label(box.inner, text=f"Shift #{sid}", bg=UI.CARD, fg=UI.TEXT, font=UI.FONT_MD).pack(anchor="w")
        tk.Label(
            box.inner,
            text=f"Expected drawer value now: {drawer_money(corrected_expected)}",
            bg=UI.CARD,
            fg="#334155",
            font=("Segoe UI", 13, "bold"),
        ).pack(anchor="w", pady=(6, 10))

        count_grid = tk.Frame(box.inner, bg=UI.CARD)
        count_grid.pack(fill="x", pady=(0, 10))
        count_grid.grid_columnconfigure(1, weight=1, minsize=320)

        tk.Label(count_grid, text=("USD in" if is_in else "USD out"), bg=UI.CARD, fg="#334155", width=13, anchor="w").grid(row=0, column=0, sticky="w", pady=3)
        amount_usd_e = tk.Entry(count_grid, bd=1, relief="solid", font=("Segoe UI", 16), justify="right")
        amount_usd_e.grid(row=0, column=1, sticky="ew", pady=3)

        tk.Label(count_grid, text=("LBP in" if is_in else "LBP out"), bg=UI.CARD, fg="#334155", width=13, anchor="w").grid(row=1, column=0, sticky="w", pady=3)
        amount_lbp_e = tk.Entry(count_grid, bd=1, relief="solid", font=("Segoe UI", 16), justify="right")
        amount_lbp_e.grid(row=1, column=1, sticky="ew", pady=3)
        bind_lbp_grouping(amount_lbp_e)

        tk.Label(
            box.inner,
            text=f"Exchange rate: {lbp_money(rate_default)} per $1.",
            bg=UI.CARD,
            fg=UI.MUTED,
        ).pack(anchor="w", pady=(0, 10))

        emp_row = tk.Frame(box.inner, bg=UI.CARD)
        emp_row.pack(fill="x", pady=(0, 10))
        tk.Label(emp_row, text=("Added by" if is_in else "Taken by"), bg=UI.CARD, fg="#334155", width=13, anchor="w").pack(side="left")
        employee_var = tk.StringVar(value=shift_emp or (employee_names[0] if employee_names else ""))
        emp_cb = ttk.Combobox(emp_row, textvariable=employee_var, values=employee_names, state="readonly", width=24)
        emp_cb.pack(side="left")

        tk.Label(box.inner, text="Reason", bg=UI.CARD, fg="#334155").pack(anchor="w")
        reason_e = tk.Entry(box.inner, bd=1, relief="solid")
        reason_e.pack(fill="x", pady=(6, 0))

        status_var = tk.StringVar(value="")
        tk.Label(box.inner, textvariable=status_var, bg=UI.CARD, fg=UI.MUTED, wraplength=480, justify="left").pack(anchor="w", pady=(8, 0))

        def _update_movement_preview(*_):
            try:
                amount_usd = parse_whole_money_text(amount_usd_e.get())
                amount_lbp = parse_lbp_text(amount_lbp_e.get())
                rate = int(rate_default)
                movement_value = round(amount_usd + (amount_lbp / rate), 2)
                expected_value = float(corrected_expected or 0.0)
                after_value = expected_value + movement_value if is_in else expected_value - movement_value
                status_var.set(
                    f"Drawer now: {drawer_money(expected_value)}  ->  "
                    f"After this {'cash in' if is_in else 'cash out'}: {drawer_money(after_value)}"
                )
            except Exception:
                status_var.set("")

        amount_usd_e.bind("<KeyRelease>", _update_movement_preview)
        amount_lbp_e.bind("<KeyRelease>", _update_movement_preview, add="+")
        _update_movement_preview()

        def _do_save():
            try:
                amount_usd = parse_whole_money_text(amount_usd_e.get())
                amount_lbp = parse_lbp_text(amount_lbp_e.get())
                rate = int(rate_default)
            except Exception:
                messagebox.showerror("Invalid", "USD and LBP must be whole numbers.", parent=win)
                return
            if amount_usd < 0 or amount_lbp < 0 or rate <= 0:
                messagebox.showerror("Invalid", "Amounts cannot be negative and rate must be greater than 0.", parent=win)
                return
            if amount_usd <= 0 and amount_lbp <= 0:
                messagebox.showerror("Invalid", f"Enter the cash amount {action_text} the drawer.", parent=win)
                return

            reason = reason_e.get().strip()
            if not reason:
                messagebox.showerror("Reason", f"Enter why cash was {action_text} the drawer.", parent=win)
                return

            emp_name = employee_var.get().strip()
            if not emp_name:
                messagebox.showerror("Employee", f"Select who {action_text} the cash.", parent=win)
                return

            if employee_pin_required(emp_name):
                pin = simpledialog.askstring(
                    "PIN Required",
                    f"Enter PIN for {emp_name}:",
                    show="*",
                    parent=win,
                )
                if pin is None:
                    return
                if not verify_employee_pin(emp_name, pin):
                    messagebox.showerror("Wrong PIN", "Incorrect PIN.", parent=win)
                    return

            movement_value = round(amount_usd + (amount_lbp / rate), 2)
            expected_value = float(summ.get("expected_cash", 0.0) or 0.0)
            if (not is_in) and movement_value > expected_value + 0.005:
                if not messagebox.askyesno(
                    "Cash out",
                    "This cash-out is more than the current expected drawer value.\n\nRecord it anyway?",
                    parent=win,
                ):
                    return

            try:
                movement_id = record_cash_movement(
                    sid,
                    movement_type=("IN" if is_in else "OUT"),
                    amount_usd=amount_usd,
                    amount_lbp=amount_lbp,
                    reason=reason,
                    employee_name=emp_name,
                    notes="",
                    lbp_per_usd=rate,
                )
            except Exception as e:
                messagebox.showerror(title, f"{type(e).__name__}: {e}", parent=win)
                return

            win.destroy()
            messagebox.showinfo(f"{title} recorded", f"{title} #{movement_id} recorded.")
            self.refresh_all()
            try:
                self.load_day(silent=True)
            except Exception:
                pass
            try:
                top = self.winfo_toplevel()
                if getattr(top, "shift_history_page", None):
                    top.shift_history_page.load_day()
            except Exception:
                pass

        btns = tk.Frame(box.inner, bg=UI.CARD)
        btns.pack(fill="x", pady=(14, 0))
        GhostButton(btns, "Cancel", win.destroy).pack(side="right")
        PrimaryButton(btns, f"Record {'Cash In' if is_in else 'Cash Out'}", _do_save).pack(side="right", padx=(0, 10))

        amount_usd_e.focus()
        win.bind("<Return>", lambda e: _do_save())
        win.bind("<Escape>", lambda e: win.destroy())

    def close_shift_clicked(self):
        sh = get_open_shift()
        if not sh:
            messagebox.showinfo("No shift", "No open shift to close.")
            return

        drawer_state = self._effective_drawer_state(sh)
        sid = int(drawer_state.get("sid", row_get(sh, "id", 0)) or row_get(sh, "id", 0) or 0)
        sh = drawer_state.get("shift") or sh
        summ = drawer_state.get("summary") or {}
        try:
            rate_default = int(round(float(summ.get("lbp_per_usd") or get_lbp_per_usd() or 89500)))
        except Exception:
            rate_default = 89500

        expected_now = float(drawer_state.get("expected", summ.get("expected_cash", 0.0)) or 0.0)
        opening_value = float(summ.get("opening_cash", 0.0) or 0.0)
        day_opening_value = float(drawer_state.get("day_opening", opening_value) or 0.0)
        day_sales = drawer_state.get("sales") or {}
        day_movements = drawer_state.get("movements") or {}
        cash_sales_value = float(day_sales.get("cash_sales", summ.get("cash_sales", 0.0)) or 0.0)
        cash_refunds_value = float(drawer_state.get("cash_refunds", summ.get("cash_refunds", 0.0)) or 0.0)
        cash_in_value = float(day_movements.get("cash_in_value", 0.0) or 0.0)
        cash_in_count = int(day_movements.get("cash_in_count", 0) or 0)
        cash_out_value = float(day_movements.get("cash_out_value", 0.0) or 0.0)
        cash_out_count = int(day_movements.get("cash_out_count", 0) or 0)

        win = tk.Toplevel(self)
        win.title("Close Register")
        win.geometry("680x640")
        win.configure(bg=UI.CONTENT_BG)
        win.grab_set()

        box = Card(win, padx=16, pady=16)
        box.pack(fill="both", expand=True, padx=14, pady=14)
        shift_label = str(row_get(sh, "shift_code", "") or row_get(sh, "shift_seq", "") or sid).strip()

        tk.Label(box.inner, text=f"Close shift {shift_label}", bg=UI.CARD, fg=UI.TEXT, font=UI.FONT_MD).pack(anchor="w")
        tk.Label(
            box.inner,
            text=f"Expected drawer value: {drawer_money(expected_now)}",
            bg=UI.CARD,
            fg="#334155",
            font=("Segoe UI", 13, "bold"),
        ).pack(anchor="w", pady=(6, 8))

        formula = (
            f"Day opening {drawer_money(day_opening_value)} + cash sales {drawer_money(cash_sales_value)} "
            f"- cash refunds {drawer_money(cash_refunds_value)} "
            f"+ cash in {drawer_money(cash_in_value)} - cash out {drawer_money(cash_out_value)}"
        )
        tk.Label(
            box.inner,
            text=formula,
            bg=UI.CARD,
            fg=UI.MUTED,
            wraplength=560,
            justify="left",
        ).pack(anchor="w", pady=(0, 8))

        tk.Label(
            box.inner,
            text=f"Exchange rate: {lbp_money(rate_default)} per $1 (change it in Settings).",
            bg=UI.CARD,
            fg=UI.MUTED,
        ).pack(anchor="w", pady=(0, 10))

        count_grid = tk.Frame(box.inner, bg=UI.CARD)
        count_grid.pack(fill="x", pady=(0, 10))
        count_grid.grid_columnconfigure(1, weight=1, minsize=330)

        tk.Label(box.inner, text="Count everything in the drawer before taking money out", bg=UI.CARD, fg=UI.TEXT).pack(anchor="w", pady=(2, 4))

        tk.Label(count_grid, text="USD counted", bg=UI.CARD, fg="#334155", width=14, anchor="w").grid(row=0, column=0, sticky="w", pady=3)
        closing_usd_e = tk.Entry(count_grid, bd=1, relief="solid", font=("Segoe UI", 18), justify="right")
        closing_usd_e.grid(row=0, column=1, sticky="ew", pady=3)
        closing_usd_e.insert(0, self.closing_var.get().strip() or f"{round(expected_now):g}")

        tk.Label(count_grid, text="LBP counted", bg=UI.CARD, fg="#334155", width=14, anchor="w").grid(row=1, column=0, sticky="w", pady=3)
        closing_lbp_e = tk.Entry(count_grid, bd=1, relief="solid", font=("Segoe UI", 18), justify="right")
        closing_lbp_e.grid(row=1, column=1, sticky="ew", pady=3)
        closing_lbp_e.insert(0, "")
        bind_lbp_grouping(closing_lbp_e)

        tk.Label(box.inner, text="Cash taken out at close", bg=UI.CARD, fg=UI.TEXT).pack(anchor="w", pady=(4, 4))
        take_grid = tk.Frame(box.inner, bg=UI.CARD)
        take_grid.pack(fill="x", pady=(0, 8))
        take_grid.grid_columnconfigure(1, weight=1, minsize=330)

        tk.Label(take_grid, text="USD taken", bg=UI.CARD, fg="#334155", width=14, anchor="w").grid(row=0, column=0, sticky="w", pady=3)
        take_usd_e = tk.Entry(take_grid, bd=1, relief="solid", font=("Segoe UI", 16), justify="right")
        take_usd_e.grid(row=0, column=1, sticky="ew", pady=3)
        take_usd_e.insert(0, "0")

        tk.Label(take_grid, text="LBP taken", bg=UI.CARD, fg="#334155", width=14, anchor="w").grid(row=1, column=0, sticky="w", pady=3)
        take_lbp_e = tk.Entry(take_grid, bd=1, relief="solid", font=("Segoe UI", 16), justify="right")
        take_lbp_e.grid(row=1, column=1, sticky="ew", pady=3)
        take_lbp_e.insert(0, "0")
        bind_lbp_grouping(take_lbp_e)

        close_live_var = tk.StringVar(value="")
        tk.Label(
            box.inner,
            textvariable=close_live_var,
            bg=UI.CARD,
            fg=UI.TEXT,
            justify="left",
            wraplength=560,
        ).pack(anchor="w", pady=(2, 8))

        def _update_close_preview(*_):
            try:
                counted_usd = parse_whole_money_text(closing_usd_e.get())
                counted_lbp = parse_lbp_text(closing_lbp_e.get())
                taken_usd = parse_whole_money_text(take_usd_e.get())
                taken_lbp = parse_lbp_text(take_lbp_e.get())
                rate = int(rate_default)
                counted_value = round(counted_usd + (counted_lbp / rate))
                taken_value = round(taken_usd + (taken_lbp / rate))
                left_usd = counted_usd - taken_usd
                left_lbp = counted_lbp - taken_lbp
                left_value = round(left_usd + (left_lbp / rate))
                expected_after_take = expected_now - taken_value
                diff_value = left_value - expected_after_take
                expected_usd = float(summ.get("expected_usd_cash", expected_now) or expected_now)
                expected_lbp = float(summ.get("expected_lbp_cash", summ.get("opening_lbp", 0.0)) or 0.0)
                close_live_var.set(
                    f"Counted before take-out: {drawer_money(counted_value)}\n"
                    f"Taken out at close: {drawer_money(taken_value)}\n"
                    f"Left for next opening: {drawer_money(left_value)}\n"
                    f"Expected after take-out: {drawer_money(expected_after_take)}\n"
                    f"Variance: {drawer_money(diff_value)}\n"
                    f"Counted vs expected: USD {drawer_money(counted_usd - expected_usd)}   "
                    f"{lbp_money(counted_lbp - expected_lbp)}"
                )
            except Exception:
                close_live_var.set("Enter the counted USD and LBP to see the closing difference.")

        closing_usd_e.bind("<KeyRelease>", _update_close_preview)
        closing_lbp_e.bind("<KeyRelease>", _update_close_preview, add="+")
        take_usd_e.bind("<KeyRelease>", _update_close_preview)
        take_lbp_e.bind("<KeyRelease>", _update_close_preview, add="+")
        _update_close_preview()

        def _do_close():
            try:
                counted_usd = parse_whole_money_text(closing_usd_e.get())
                counted_lbp = parse_lbp_text(closing_lbp_e.get())
                taken_usd = parse_whole_money_text(take_usd_e.get())
                taken_lbp = parse_lbp_text(take_lbp_e.get())
                rate = int(rate_default)
            except Exception:
                messagebox.showerror("Invalid", "USD and LBP must be whole numbers.", parent=win)
                return
            if counted_usd < 0 or counted_lbp < 0 or taken_usd < 0 or taken_lbp < 0 or rate <= 0:
                messagebox.showerror("Invalid", "USD/LBP cannot be negative and rate must be greater than 0.", parent=win)
                return
            if taken_usd > counted_usd + 0.005 or taken_lbp > counted_lbp + 0.005:
                messagebox.showerror("Invalid", "Cash taken out cannot be more than the cash counted.", parent=win)
                return

            counted = round(counted_usd + (counted_lbp / rate))
            taken = round(taken_usd + (taken_lbp / rate))
            left_usd = round(counted_usd - taken_usd)
            left_lbp = round(counted_lbp - taken_lbp)
            left_for_next_opening = round(left_usd + (left_lbp / rate))
            expected = expected_now
            expected_after_take = expected - taken
            diff = left_for_next_opening - expected_after_take

            expected_usd = float(summ.get("expected_usd_cash", expected) or expected)
            expected_lbp = float(summ.get("expected_lbp_cash", summ.get("opening_lbp", 0.0)) or 0.0)
            usd_diff = counted_usd - expected_usd
            lbp_diff = counted_lbp - expected_lbp
            lbp_diff_usd = lbp_diff / rate

            cash_in_line = ""
            if cash_in_value > 0.005:
                cash_in_line = f"Cash added during shift: {drawer_money(cash_in_value)} ({cash_in_count} entries)\n"
            cash_out_line = ""
            if cash_out_value > 0.005:
                cash_out_line = f"Cash removed during shift: {drawer_money(cash_out_value)} ({cash_out_count} entries)\n"

            msg = (
                "This will close the register and save the closing count.\n\n"
                f"Expected drawer value:\n"
                f"  Day opening {drawer_money(day_opening_value)} + cash sales {drawer_money(cash_sales_value)} "
                f"- cash refunds {drawer_money(cash_refunds_value)} "
                f"+ cash in {drawer_money(cash_in_value)} - cash out {drawer_money(cash_out_value)}\n"
                f"  = {drawer_money(expected)}\n\n"
                f"{cash_in_line}"
                f"{cash_out_line}"
                f"Counted before take-out: {drawer_money(counted)}\n"
                f"Cash taken out at close: {drawer_money(taken)}\n"
                f"Left for next opening: {drawer_money(left_for_next_opening)}\n"
                f"Expected after take-out: {drawer_money(expected_after_take)}\n"
                f"Variance after take-out: {drawer_money(diff)}\n\n"
                f"USD difference: {drawer_money(usd_diff)}\n"
                f"LBP difference: {lbp_money(lbp_diff)} ({drawer_money(lbp_diff_usd)})"
            )
            if usd_diff < -0.005 and lbp_diff_usd > 0.005:
                msg += "\n\nLBP increased, so it may be covering some missing USD."
            msg += "\n\nClose shift now?"

            if not messagebox.askyesno("Close shift", msg, parent=win):
                return

            try:
                close_shift_with_cash_takeout(
                    sid,
                    closing_cash=left_for_next_opening,
                    notes=(
                        f"Close count: counted {drawer_money(counted)}; "
                        f"taken at close {drawer_money(taken)}; "
                        f"left for next opening {drawer_money(left_for_next_opening)}; "
                        f"variance {drawer_money(diff)}."
                    ),
                    closing_usd=left_usd,
                    closing_lbp=left_lbp,
                    lbp_per_usd=rate,
                    takeout_usd=taken_usd,
                    takeout_lbp=taken_lbp,
                    employee_name=str(row_get(sh, "employee_name", "") or ""),
                    takeout_reason="End of day close cash removed",
                    takeout_notes=(
                        f"Cash taken at close. Counted {drawer_money(counted)}; "
                        f"left for next opening {drawer_money(left_for_next_opening)}."
                    ),
                )
            except Exception as e:
                messagebox.showerror("Error", f"{type(e).__name__}: {e}", parent=win)
                return

            self.closing_var.set(f"{left_for_next_opening:g}")
            try:
                backup_pos_db()
            except Exception:
                pass
            win.destroy()
            messagebox.showinfo("Shift closed", f"Shift {shift_label} closed.")
            self.refresh_all()
            try:
                is_connect = (backend_mode() == "connect")
                cfg = get_daily_report_email_config()
                email_ready = (
                    bool(cfg.get("enabled", True))
                    and bool(str(cfg.get("sender_email") or "").strip())
                    and bool(str(cfg.get("smtp_username") or "").strip())
                    and bool(str(cfg.get("smtp_password") or "").strip())
                )
                if is_connect or email_ready:
                    self.winfo_toplevel().send_daily_report_email_for_day(
                        datetime.now().strftime("%Y-%m-%d"),
                        source="close",
                        silent=False,
                    )
            except Exception as e:
                messagebox.showwarning("Daily report email", f"Could not send the close-register report.\n{e}")


        btns = tk.Frame(box.inner, bg=UI.CARD)
        btns.pack(fill="x", pady=(10, 0))
        GhostButton(btns, "Cancel", win.destroy).pack(side="right")
        PrimaryButton(btns, "Close Shift", _do_close).pack(side="right", padx=(0, 10))

        closing_usd_e.focus()
        win.bind("<Return>", lambda e: _do_close())
        win.bind("<Escape>", lambda e: win.destroy())

    def set_today(self):
        today = datetime.now()
        self.sales_year.set(str(today.year))
        self.sales_month.set(f"{today.month:02d}")
        self.sales_day.set(f"{today.day:02d}")
        self.date_var.set(today.strftime("%Y-%m-%d"))
        self.load_day()

    # FIXED: indentation + no stray top-level lines
    def load_day(self, silent=False):
        # Build date from dropdowns (no manual typing)
        day = f"{self.sales_year.get()}-{self.sales_month.get()}-{self.sales_day.get()}"
        self.date_var.set(day)
        try:
            datetime.strptime(day, "%Y-%m-%d")
        except Exception:
            if not silent:
                messagebox.showerror("Invalid", "Selected date is invalid.")
            return

        selected_sale_id = None
        try:
            sel = self.sales_tree.selection()
            if sel:
                selected_sale_id = sel[0]
        except Exception:
            selected_sale_id = None

        try:
            yview = self.sales_tree.yview()
        except Exception:
            yview = None

        self.sales_tree.delete(*self.sales_tree.get_children())

        rows = list_sales_for_day(day, limit=1000, include_voided=bool(self.show_voided_var.get()))

        # Drawer cash is new physical cash collected. Display amount is the sale amount
        # so Whish/Card rows do not look like $0 sales.
        net_total_sum = 0.0
        displayed_sales_sum = 0.0
        cash_in_sum = 0.0
        cash_out_sum = 0.0

        for r in rows:
            is_voided = bool(int(row_get(r, "is_voided", 0) or 0))
            pm = str(row_get(r, "payment_method", "") or "").strip().upper()
            cash_paid_val = row_get(r, "cash_paid", None)
            total_val = float(row_get(r, "total_amount", 0) or 0)
            if pm == "EXCHANGE":
                paid = 0.0
            elif pm in ("CARD", "DEBIT", "CREDIT_CARD", "WHISH"):
                # cash-only register: cards do not change the drawer
                paid = 0.0
            else:
                # Prefer stored cash_paid. If missing (older server/DB), compute it so exchanges
                # only add what was actually paid extra (cash movement).
                if cash_paid_val is None:
                    credit_used = row_get(r, "store_credit_used", None)
                    total_sales_val = row_get(r, "total_sales", None)
                    try:
                        if credit_used is not None and total_sales_val is not None:
                            paid = max(float(total_sales_val or 0) - float(credit_used or 0), 0.0)
                        elif credit_used is not None:
                            # If we only know the displayed total, assume it's merchandise total and subtract credit.
                            paid = max(float(total_val or 0) - float(credit_used or 0), 0.0)
                        else:
                            paid = float(total_val or 0)
                    except Exception:
                        paid = float(total_val or 0)
                else:
                    try:
                        paid = float(cash_paid_val or 0)
                    except Exception:
                        paid = float(total_val or 0)
            if not is_voided:
                net_total_sum += paid
            display_amount = sale_display_amount(r)
            if not is_voided:
                displayed_sales_sum += display_amount

            t = row_get(r, "created_at", "")
            try:
                dt = datetime.strptime(t, "%Y-%m-%d %H:%M:%S")
                t_disp = dt.strftime("%I:%M %p").lstrip("0")
            except Exception:
                t_disp = t

            receipt_code = (row_get(r, "receipt_code", "") or "").strip()
            if not receipt_code:
                receipt_code = str(int(row_get(r, "id", 0) or 0))
            if is_voided:
                receipt_code = f"VOID {receipt_code}"

            self.sales_tree.insert(
                "",
                tk.END,
                iid=str(int(row_get(r, "id", 0) or 0)),
                tags=(("voided",) if is_voided else ()),
                values=(
                    receipt_code,
                    t_disp,
                    money(display_amount),
                    payment_method_label(row_get(r, "payment_method", "") or ""),
                    (row_get(r, "shift_code", "") or row_get(r, "shift_id", "") or "")
                )
            )

        try:
            movements = list_cash_movements(day_str=day, limit=1000) or []
        except Exception:
            movements = []

        for m in reversed(movements):
            mtype = str(row_get(m, "movement_type", "OUT") or "OUT").upper()
            if mtype not in ("IN", "OUT"):
                continue
            amount = float(row_get(m, "amount_value", 0.0) or 0.0)
            if mtype == "IN":
                cash_in_sum += amount
            else:
                cash_out_sum += amount
            t = row_get(m, "created_at", "")
            try:
                dt = datetime.strptime(t, "%Y-%m-%d %H:%M:%S")
                t_disp = dt.strftime("%I:%M %p").lstrip("0")
            except Exception:
                t_disp = t
            movement_id = int(row_get(m, "id", 0) or 0)
            who = str(row_get(m, "employee_name", "") or "").strip()
            reason = str(row_get(m, "reason", "") or "").strip()
            label = f"Cash {'In' if mtype == 'IN' else 'Out'} #{movement_id}" if movement_id else f"Cash {'In' if mtype == 'IN' else 'Out'}"
            pay = mtype
            if who:
                pay = who[:18]
            if reason:
                label = f"{label}: {reason[:32]}"
            self.sales_tree.insert(
                "",
                tk.END,
                iid=f"{'cashin' if mtype == 'IN' else 'cashout'}_{movement_id}",
                values=(
                    label,
                    t_disp,
                    (money(amount) if mtype == "IN" else f"-{money(amount)}"),
                    pay,
                    (row_get(m, "shift_id", "") if row_get(m, "shift_id", None) is not None else ""),
                ),
            )

        try:
            if selected_sale_id is not None:
                sid = str(selected_sale_id)
                if sid in self.sales_tree.get_children():
                    self.sales_tree.selection_set(sid)
                    self.sales_tree.focus(sid)
        except Exception:
            pass

        try:
            if yview is not None:
                self.sales_tree.yview_moveto(yview[0])
        except Exception:
            pass

        try:
            drawer_state = self._effective_drawer_state(get_open_shift())
            cash_in_register = float(drawer_state.get("expected", 0.0) or 0.0)
        except Exception:
            cash_in_register = net_total_sum + cash_in_sum - cash_out_sum

        self.day_total_lbl.config(
            text=(
                f"Sales total: {money(displayed_sales_sum)}   |   Cash collected: {money(net_total_sum)}   |   Active sales: {sum(1 for r in rows if not bool(int(row_get(r, 'is_voided', 0) or 0)))}\n"
                f"Cash in: {money(cash_in_sum)}   |   Cash out: {money(cash_out_sum)}\n"
                f"Drawer net: {money(net_total_sum + cash_in_sum - cash_out_sum)}   |   "
                f"Cash in register: {money(cash_in_register)}"
            )
        )
        try:
            self.refresh_all()
        except Exception:
            pass

    def delete_selected_sale(self):
        sel = self.sales_tree.selection()
        if not sel:
            messagebox.showinfo("Select", "Select a sale first.")
            return
        vals = self.sales_tree.item(sel[0], "values")
        if not vals:
            return

        iid = str(sel[0])
        if not iid.isdigit():
            messagebox.showinfo("Void sale", "Cash drawer movement records cannot be voided from this list.")
            return

        sale_id = int(iid)
        receipt_code = vals[0] if len(vals) > 0 else str(sale_id)

        sale, _items = get_sale_detail(sale_id)
        if sale and bool(int(row_get(sale, "is_voided", 0) or 0)):
            messagebox.showinfo("Void sale", "This sale is already voided.")
            return

        reason = simpledialog.askstring(
            "Void Sale",
            f"Why are you voiding sale {receipt_code}?\n\nThe sale will remain visible in Show voided and its stock will be restored.",
            parent=self,
        )
        if reason is None:
            return
        reason = reason.strip()
        if not reason:
            messagebox.showerror("Void Sale", "Enter a reason so the void has a useful audit record.")
            return
        if not messagebox.askyesno("Confirm Void", f"Void sale {receipt_code} and restore its stock?", parent=self):
            return

        try:
            void_sale(sale_id, reason, self.employee_var.get().strip(), restore_stock=True)
        except Exception as e:
            messagebox.showerror("Error", f"{type(e).__name__}: {e}")
            return

        self.load_day(silent=True)
        self.refresh_all()
        messagebox.showinfo("Voided", f"Sale #{sale_id} was voided, kept in history, and its stock was restored.")

    # FIXED: indentation + no stray try/labels outside function
    def open_sale_detail(self, event=None):
        sel = self.sales_tree.selection()
        if not sel:
            return

        iid = str(sel[0])
        if iid.startswith("cashout_") or iid.startswith("cashin_"):
            self.open_cash_movement_detail(iid)
            return
        vals = self.sales_tree.item(iid, "values")
        sale_id = int(iid)
        receipt_code = vals[0] if vals and len(vals) > 0 else str(sale_id)

        sale, items = get_sale_detail_with_returns(sale_id)
        if not sale:
            messagebox.showerror("Missing", "Sale not found.")
            return

        win = tk.Toplevel(self)
        win.title(f"Sale {receipt_code}")
        win.geometry("600x520")
        win.configure(bg=UI.CONTENT_BG)
        win.grab_set()

        box = Card(win, padx=16, pady=16)
        box.pack(fill="both", expand=True, padx=14, pady=14)

        tk.Label(box.inner, text=f"Sale {receipt_code}", font=UI.FONT_LG, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w")
        if bool(int(row_get(sale, "is_voided", 0) or 0)):
            tk.Label(
                box.inner,
                text=(
                    f"VOIDED {row_get(sale, 'voided_at', '')} by {row_get(sale, 'voided_by', '') or 'Unassigned'}\n"
                    f"Reason: {row_get(sale, 'void_reason', '') or 'No reason recorded'}"
                ),
                bg="#fee2e2", fg="#991b1b", justify="left", anchor="w",
                font=("Segoe UI", 10, "bold"),
            ).pack(fill="x", pady=(6, 8), ipady=6, ipadx=8)

        # Calculate returned value from the actual return records. This matters
        # for exchange sales where store credit reduced the amount actually paid.
        returned_value = 0.0
        try:
            for ret in list_returns_for_sale(int(sale_id), include_voided=False) or []:
                returned_value += float(row_get(ret, "total_return_amount", 0.0) or 0.0)
        except Exception:
            returned_value = 0.0

        # gross_total should be total_sales (merchandise value) if available
        gross_total = float(row_get(sale, "total_sales", 0) or row_get(sale, "total_amount", 0) or 0)
        net_total = gross_total - returned_value

        tk.Label(
            box.inner,
            text=(
                f"Time: {row_get(sale, 'created_at', '')}   |   "
                f"Gross: {money(gross_total)}   |   "
                f"Returned: {money(returned_value)}   |   "
                f"Net: {money(net_total)}   |   "
                f"Pay: {payment_method_label(row_get(sale, 'payment_method', '') or '')}   |   "
                f"Shift: {row_get(sale, 'shift_id', '')}"
            ),
            bg=UI.CARD, fg="#334155"
        ).pack(anchor="w", pady=(6, 12))

        cols = ("name", "price", "qty", "returned", "line")
        tree = ttk.Treeview(box.inner, columns=cols, show="headings", height=14)
        tree.heading("name", text="Item")
        tree.heading("price", text="Price")
        tree.heading("qty", text="Qty")
        tree.heading("returned", text="Returned")
        tree.heading("line", text="Line total")
        tree.column("name", width=280)
        tree.column("price", width=90, anchor="e")
        tree.column("qty", width=70, anchor="center")
        tree.column("returned", width=80, anchor="center")
        tree.column("line", width=110, anchor="e")
        tree.pack(fill="both", expand=True)

        for it in items:
            tree.insert("", tk.END, values=(
                row_get(it, "name", "") or "",
                money(float(row_get(it, "price", 0) or 0)),
                int(row_get(it, "qty", 0) or 0),
                int(row_get(it, "returned_qty", 0) or 0),
                money(float(row_get(it, "line_total", 0) or 0))
            ))

        btns = tk.Frame(box.inner, bg=UI.CARD)
        btns.pack(fill="x", pady=(12, 0))
        GhostButton(btns, "Close", win.destroy).pack(side="right")

    def open_cash_movement_detail(self, iid: str):
        try:
            movement_id = int(str(iid).split("_", 1)[1])
        except Exception:
            return

        movement = None
        try:
            day = self.date_var.get().strip() or None
            for m in list_cash_movements(day_str=day, limit=2000) or []:
                if int(row_get(m, "id", 0) or 0) == movement_id:
                    movement = m
                    break
        except Exception:
            movement = None

        if not movement:
            messagebox.showerror("Cash out", "Cash-out record not found.")
            return

        win = tk.Toplevel(self)
        mtype = str(row_get(movement, "movement_type", "OUT") or "OUT").upper()
        movement_title = "Cash In" if mtype == "IN" else "Cash Out"

        win.title(f"{movement_title} #{movement_id}")
        win.geometry("480x320")
        win.configure(bg=UI.CONTENT_BG)
        win.grab_set()

        box = Card(win, padx=16, pady=16)
        box.pack(fill="both", expand=True, padx=14, pady=14)

        amount_usd = float(row_get(movement, "amount_usd", 0.0) or 0.0)
        amount_lbp = float(row_get(movement, "amount_lbp", 0.0) or 0.0)
        amount_value = float(row_get(movement, "amount_value", 0.0) or 0.0)
        lines = [
            f"{movement_title} #{movement_id}",
            f"Time: {row_get(movement, 'created_at', '')}",
            f"Shift: {row_get(movement, 'shift_id', '')}",
            f"{'Added by' if mtype == 'IN' else 'Taken by'}: {row_get(movement, 'employee_name', '') or 'Unassigned'}",
            f"USD: {drawer_money(amount_usd)}",
            f"LBP: {lbp_money(amount_lbp)}",
            f"Drawer value: {drawer_money(amount_value)}",
            f"Reason: {row_get(movement, 'reason', '')}",
        ]
        tk.Label(box.inner, text="\n".join(lines), bg=UI.CARD, fg=UI.TEXT, justify="left", wraplength=420).pack(anchor="w")

        btns = tk.Frame(box.inner, bg=UI.CARD)
        btns.pack(fill="x", pady=(14, 0))
        GhostButton(btns, "Close", win.destroy).pack(side="right")

    # ---------------- RETURNS / EXCHANGE PAGE ----------------

    # ---------------- Receipt Reprint (by Sale ID / Receipt Scan) ----------------
    def _resolve_sale_id_from_input(self, raw: str):
        s = str(raw or "").strip()
        if not s:
            return None
        # direct numeric sale id
        if s.isdigit():
            try:
                return int(s)
            except Exception:
                return None
        # Try scan resolver (supports formats like R-2025-12-29-0001, MASKPOS|..., etc.)
        try:
            sale, _items = get_sale_by_receipt_scan(s)
            if sale:
                sid = row_get(sale, "id", None)
                if sid is not None:
                    return int(sid)
        except Exception:
            pass
        # Fallback: first number group
        import re as _re
        m = _re.search(r"(\d+)", s)
        if m:
            try:
                return int(m.group(1))
            except Exception:
                return None
        return None

    def open_sales_search(self):
        win=tk.Toplevel(self); win.title("Search All Sales"); win.geometry("980x560"); win.minsize(720,430)
        win.configure(bg=UI.CONTENT_BG); win.transient(self.winfo_toplevel())
        card=Card(win,padx=14,pady=14); card.pack(fill="both",expand=True,padx=12,pady=12)
        HeaderBar(card.inner,"Search All Sales","Search every date by receipt, product, barcode, employee, or payment method.").pack(fill="x")
        top=tk.Frame(card.inner,bg=UI.CARD); top.pack(fill="x",pady=(10,8))
        query=tk.StringVar(value=""); include_voided=tk.BooleanVar(value=True)
        entry=tk.Entry(top,textvariable=query,bd=1,relief="solid"); entry.pack(side="left",fill="x",expand=True,ipady=4)
        ttk.Checkbutton(top,text="Include voided",variable=include_voided).pack(side="left",padx=10)
        cols=("receipt","date","total","pay","employee","status","items")
        tree=ttk.Treeview(card.inner,columns=cols,show="headings",height=16)
        for col,title,width in [("receipt","Receipt",85),("date","Date / time",145),("total","Total",85),("pay","Payment",90),("employee","Employee",100),("status","Status",65),("items","Items",300)]:
            tree.heading(col,text=title); tree.column(col,width=width,anchor=("e" if col=="total" else "w"))
        tree.pack(fill="both",expand=True); tree.tag_configure("voided",foreground="#991b1b",background="#fee2e2")
        rows_by_id={}
        def load():
            try: rows=search_sales(query.get().strip(),bool(include_voided.get()),500) or []
            except Exception as exc: messagebox.showerror("Sales search",str(exc),parent=win); return
            tree.delete(*tree.get_children()); rows_by_id.clear()
            for row in rows:
                sid=int(row_get(row,"id",0) or 0); rows_by_id[str(sid)]=row
                voided=bool(int(row_get(row,"is_voided",0) or 0))
                tree.insert("",tk.END,iid=str(sid),tags=(("voided",) if voided else ()),values=(row_get(row,"receipt_code","") or sid,row_get(row,"created_at",""),money(sale_display_amount(row)),payment_method_label(row_get(row,"payment_method","")),row_get(row,"employee_name",""),"VOID" if voided else "Active",row_get(row,"item_names","") or ""))
        def open_selected():
            sel=tree.selection()
            if not sel: messagebox.showinfo("Sales search","Select a sale first.",parent=win); return
            sid=sel[0]; row=rows_by_id.get(sid,{})
            created=str(row_get(row,"created_at","") or "")
            if len(created)>=10: self.date_var.set(created[:10])
            if bool(int(row_get(row,"is_voided",0) or 0)): self.show_voided_var.set(True)
            self.load_day(silent=True); self.reprint_var.set(str(row_get(row,"receipt_code","") or sid)); self._last_lookup_sale_id=int(sid)
            if sid in self.sales_tree.get_children(): self.sales_tree.selection_set(sid); self.sales_tree.focus(sid); self.sales_tree.see(sid)
            self.open_sale_detail()
        PrimaryButton(top,"Search",load).pack(side="left")
        btn=tk.Frame(card.inner,bg=UI.CARD); btn.pack(fill="x",pady=(8,0))
        PrimaryButton(btn,"View Selected",open_selected).pack(side="left"); GhostButton(btn,"Close",win.destroy).pack(side="right")
        entry.bind("<Return>",lambda _e: load()); tree.bind("<Double-1>",lambda _e: open_selected()); entry.focus_set(); load()

    def lookup_sale_clicked(self):
        raw = self.reprint_var.get().strip()
        sale_id = self._resolve_sale_id_from_input(raw)
        if not sale_id:
            messagebox.showerror("Not found", "Enter a valid Sale ID or scan a receipt.")
            return

        self._last_lookup_sale_id = int(sale_id)

        # If we can read the sale, auto-jump the day picker to that sale date
        try:
            sale, _items = get_sale_detail(sale_id)
            if sale:
                created = str(row_get(sale, "created_at", "") or "")
                if len(created) >= 10:
                    self.date_var.set(created[:10])
        except Exception:
            pass

        # Reload list and select row if present
        self.load_day(silent=True)
        sid = str(sale_id)
        try:
            if sid in self.sales_tree.get_children():
                self.sales_tree.selection_set(sid)
                self.sales_tree.focus(sid)
                self.sales_tree.see(sid)
        except Exception:
            pass

    def reprint_receipt_clicked(self):
        sale_id = None
        try:
            sel = self.sales_tree.selection()
            if sel:
                iid = str(sel[0])
                if iid.startswith("cashout_") or iid.startswith("cashin_"):
                    messagebox.showinfo("Reprint", "Cash drawer movement records do not have receipts to reprint.")
                    return
                if iid.isdigit():
                    sale_id = int(iid)
        except Exception:
            sale_id = None

        if sale_id is None:
            sale_id = self._resolve_sale_id_from_input(self.reprint_var.get())

        if sale_id is None:
            sale_id = self._last_lookup_sale_id

        if sale_id is None:
            messagebox.showinfo("Reprint", "Select a sale or enter a Sale ID / scan a receipt first.")
            return

        try:
            sale, items = get_sale_receipt_data(int(sale_id))
        except Exception as e:
            messagebox.showerror("Error", f"{type(e).__name__}: {e}")
            return

        if not sale:
            messagebox.showerror("Missing", "Sale not found.")
            return

        try:
            ok = bool(print_configured_receipt(get_store_name(), sale, items))
            if not ok:
                messagebox.showwarning("Print", "Receipt was not sent. Check Settings > Receipt Printer and use Test Print.")
        except Exception as e:
            messagebox.showerror("Print", f"{type(e).__name__}: {e}")

    def reprint_gift_receipt_clicked(self):
        sale_id = None
        sel = self.sales_tree.selection()
        if sel and str(sel[0]).isdigit():
            sale_id = int(sel[0])
        if sale_id is None:
            sale_id = self._resolve_sale_id_from_input(self.reprint_var.get()) or self._last_lookup_sale_id
        if sale_id is None:
            messagebox.showinfo("Gift receipt", "Select a sale or enter a Sale ID / receipt scan first.")
            return
        sale, items = get_sale_detail_with_returns(int(sale_id))
        if not sale:
            messagebox.showerror("Missing", "Sale not found.")
            return
        selected = self.winfo_toplevel().cashier_page._select_gift_items(items)
        if selected is None:
            return
        if not selected:
            messagebox.showinfo("Gift receipt", "Select at least one item.")
            return
        try:
            ok = bool(print_configured_gift_receipt(get_store_name(), sale, selected))
            if not ok:
                messagebox.showwarning("Print", "Gift receipt was not sent. Check Settings > Receipt Printer and use Test Print.")
        except Exception as exc:
            messagebox.showerror("Print", str(exc))

    def view_details_clicked(self):
        # Prefer currently selected row
        try:
            sel = self.sales_tree.selection()
            if sel:
                self.open_sale_detail()
                return
        except Exception:
            pass

        # Otherwise try to lookup + open
        raw = self.reprint_var.get().strip()
        sale_id = self._resolve_sale_id_from_input(raw) or self._last_lookup_sale_id
        if not sale_id:
            messagebox.showinfo("Details", "Enter a Sale ID / scan a receipt first.")
            return

        # Select it if it exists in the current day list then open
        self.reprint_var.set(str(raw or sale_id))
        self.lookup_sale_clicked()
        self.open_sale_detail()


class ShiftHistoryPage(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg=UI.CONTENT_BG)
        today = datetime.now()
        self.sales_year = tk.StringVar(value=str(today.year))
        self.sales_month = tk.StringVar(value=f"{today.month:02d}")
        self.sales_day = tk.StringVar(value=f"{today.day:02d}")
        self.summary_var = tk.StringVar(value="")
        self._build()

    def _build(self):
        scroll = VScrollableFrame(self, bg=UI.CONTENT_BG)
        scroll.pack(fill="both", expand=True)
        wrap = tk.Frame(scroll.inner, bg=UI.CONTENT_BG)
        wrap.pack(fill="both", expand=True, padx=(10 if UI.COMPACT else 18), pady=(10 if UI.COMPACT else 18))

        header = Card(wrap, padx=18, pady=14)
        header.pack(fill="x")
        HeaderBar(
            header.inner,
            "Shift History",
            "Day-by-day drawer opening, closing, register status, and sales by shift."
        ).pack(fill="x")

        body = Card(wrap, padx=14, pady=14)
        body.pack(fill="both", expand=True, pady=((8 if UI.COMPACT else 14), 0))

        top = tk.Frame(body.inner, bg=UI.CARD)
        top.pack(fill="x", pady=(0, 10))
        tk.Label(top, text="Date", bg=UI.CARD, fg="#334155").pack(side="left")
        years = [str(y) for y in range(datetime.now().year - 5, datetime.now().year + 1)]
        months = [f"{m:02d}" for m in range(1, 13)]
        days = [f"{d:02d}" for d in range(1, 32)]
        ttk.Combobox(top, textvariable=self.sales_year, values=years, width=6, state="readonly").pack(side="left",
                                                                                                      padx=(8, 6))
        ttk.Combobox(top, textvariable=self.sales_month, values=months, width=4, state="readonly").pack(side="left",
                                                                                                        padx=(0, 6))
        ttk.Combobox(top, textvariable=self.sales_day, values=days, width=4, state="readonly").pack(side="left",
                                                                                                    padx=(0, 10))
        GhostButton(top, "< Day", lambda: self.move_day(-1)).pack(side="left", padx=(0, 8))
        GhostButton(top, "Day >", lambda: self.move_day(1)).pack(side="left", padx=(0, 8))
        GhostButton(top, "Today", self.set_today).pack(side="left", padx=(0, 8))
        PrimaryButton(top, "Load", self.load_day).pack(side="left")

        tk.Label(body.inner, textvariable=self.summary_var, bg=UI.CARD, fg=UI.TEXT,
                 font=("Segoe UI", 14, "bold")).pack(anchor="w", pady=(0, 10))

        cols = ("time", "employee", "opening", "closing", "sales", "diff", "status")
        self.tree = ttk.Treeview(body.inner, columns=cols, show="tree headings", height=18)
        self.tree.heading("#0", text="Shift / Sale")
        self.tree.heading("time", text="Time")
        self.tree.heading("employee", text="Employee / Pay")
        self.tree.heading("opening", text="Opening")
        self.tree.heading("closing", text="Closing")
        self.tree.heading("sales", text="Sales")
        self.tree.heading("diff", text="Diff")
        self.tree.heading("status", text="Status")
        self.tree.column("#0", width=180, anchor="w")
        self.tree.column("time", width=170, anchor="w")
        self.tree.column("employee", width=140, anchor="w")
        self.tree.column("opening", width=105, anchor="e")
        self.tree.column("closing", width=105, anchor="e")
        self.tree.column("sales", width=105, anchor="e")
        self.tree.column("diff", width=90, anchor="e")
        self.tree.column("status", width=80, anchor="center")
        self.tree.pack(fill="both", expand=True)

        scroll = ttk.Scrollbar(body.inner, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scroll.set)
        scroll.place(relx=1.0, rely=0.19, relheight=0.78, anchor="ne")

        self.load_day()

    def _selected_day(self) -> str:
        day = f"{self.sales_year.get()}-{self.sales_month.get()}-{self.sales_day.get()}"
        datetime.strptime(day, "%Y-%m-%d")
        return day

    def set_today(self):
        today = datetime.now()
        self.sales_year.set(str(today.year))
        self.sales_month.set(f"{today.month:02d}")
        self.sales_day.set(f"{today.day:02d}")
        self.load_day()

    def move_day(self, delta: int):
        try:
            dt = datetime.strptime(self._selected_day(), "%Y-%m-%d").date() + timedelta(days=int(delta))
        except Exception:
            dt = date.today()
        self.sales_year.set(str(dt.year))
        self.sales_month.set(f"{dt.month:02d}")
        self.sales_day.set(f"{dt.day:02d}")
        self.load_day()

    def _parse_dt(self, value):
        try:
            return datetime.strptime(str(value or ""), "%Y-%m-%d %H:%M:%S")
        except Exception:
            return None

    def _time_range(self, opened, closed):
        odt = self._parse_dt(opened)
        cdt = self._parse_dt(closed)
        left = odt.strftime("%I:%M %p").lstrip("0") if odt else str(opened or "")[:16]
        right = cdt.strftime("%I:%M %p").lstrip("0") if cdt else "Open"
        return f"{left} - {right}"

    def _shift_matches_day(self, sh, day: str, sale_shift_ids: set[int]) -> bool:
        try:
            sid = int(row_get(sh, "id", 0) or 0)
            if sid in sale_shift_ids:
                return True
        except Exception:
            pass
        for key in ("opened_at", "closed_at"):
            value = str(row_get(sh, key, "") or "")
            if value.startswith(day):
                return True
        return False

    def _sale_credit_used_amount(self, sale_row) -> float:
        for key in ("store_credit_used", "exchange_credit_used", "bon_credit_used"):
            value = row_get(sale_row, key, None)
            if value is not None:
                try:
                    return max(0.0, float(value or 0.0))
                except Exception:
                    pass
        notes = str(row_get(sale_row, "notes", "") or "")
        for part in notes.replace(";", "|").split("|"):
            part = part.strip().upper()
            if part.startswith("EXCHANGE_CREDIT_APPLIED=") or part.startswith("BON_CREDIT_APPLIED="):
                try:
                    return max(0.0, float(part.split("=", 1)[1] or 0.0))
                except Exception:
                    return 0.0
        return 0.0

    def _sale_new_money_amount(self, sale_row) -> float:
        pm = str(row_get(sale_row, "payment_method", "") or "").strip().upper()
        if pm in ("EXCHANGE", "STORE_CREDIT"):
            return 0.0
        try:
            total_amount = float(row_get(sale_row, "total_amount", 0.0) or 0.0)
        except Exception:
            total_amount = 0.0
        try:
            gross = float(row_get(sale_row, "total_sales", 0.0) or 0.0)
        except Exception:
            gross = 0.0
        credit = self._sale_credit_used_amount(sale_row)
        if gross > 0 and credit > 0 and abs(total_amount - gross) < 0.005:
            return max(0.0, gross - credit)
        return max(0.0, total_amount)

    def _sale_cash_amount(self, sale_row) -> float:
        pm = str(row_get(sale_row, "payment_method", "") or "").strip().upper()
        if pm in ("EXCHANGE", "STORE_CREDIT", "CARD", "DEBIT", "CREDIT_CARD", "WHISH"):
            return 0.0
        cash_paid = row_get(sale_row, "cash_paid", None)
        if cash_paid is not None:
            try:
                return max(0.0, float(cash_paid or 0.0))
            except Exception:
                pass
        try:
            return max(0.0, float(row_get(sale_row, "total_amount", 0.0) or 0.0))
        except Exception:
            return 0.0

    def load_day(self):
        try:
            day = self._selected_day()
        except Exception:
            messagebox.showerror("Invalid", "Selected date is invalid.")
            return

        try:
            self.tree.delete(*self.tree.get_children())
        except Exception:
            pass

        try:
            sales = list_sales_for_day(day, limit=2000) or []
        except Exception:
            sales = []
        sales_by_shift = {}
        movements_by_shift = {}
        unassigned_movements = []
        unassigned = []
        daily_sales_total = 0.0
        daily_cash_sales_total = 0.0
        daily_cash_in_total = 0.0
        daily_cash_out_total = 0.0
        sale_shift_ids = set()

        for sale in sales:
            daily_sales_total += self._sale_new_money_amount(sale)
            daily_cash_sales_total += self._sale_cash_amount(sale)
            try:
                sid = int(row_get(sale, "shift_id", 0) or 0)
            except Exception:
                sid = 0
            if sid > 0:
                sale_shift_ids.add(sid)
                sales_by_shift.setdefault(sid, []).append(sale)
            else:
                unassigned.append(sale)

        try:
            movements = list_cash_movements(day_str=day, limit=2000) or []
        except Exception:
            movements = []
        for movement in movements:
            mtype = str(row_get(movement, "movement_type", "OUT") or "OUT").upper()
            if mtype not in ("IN", "OUT"):
                continue
            amount = float(row_get(movement, "amount_value", 0.0) or 0.0)
            if mtype == "IN":
                daily_cash_in_total += amount
            else:
                daily_cash_out_total += amount
            try:
                sid = int(row_get(movement, "shift_id", 0) or 0)
            except Exception:
                sid = 0
            if sid > 0:
                sale_shift_ids.add(sid)
                movements_by_shift.setdefault(sid, []).append(movement)
            else:
                unassigned_movements.append(movement)

        try:
            shifts = list_shifts(limit=1000) or []
        except Exception:
            shifts = []
        day_shifts = []
        seen = set()
        for sh in shifts:
            try:
                sid = int(row_get(sh, "id", 0) or 0)
            except Exception:
                sid = 0
            if sid <= 0 or sid in seen:
                continue
            if self._shift_matches_day(sh, day, sale_shift_ids):
                day_shifts.append(sh)
                seen.add(sid)

        day_shifts.sort(key=lambda r: str(row_get(r, "opened_at", "") or ""))

        day_opening_total = 0.0
        day_closing_total = 0.0
        closed_shift_count = 0

        for sh in day_shifts:
            sid = int(row_get(sh, "id", 0) or 0)
            try:
                summ = shift_summary(sid) or {}
            except Exception:
                summ = {}

            employee = str(row_get(sh, "employee_name", "") or "Unassigned")
            try:
                opening_raw = float(summ.get("opening_cash", row_get(sh, "opening_cash", 0)) or 0.0)
            except Exception:
                opening_raw = 0.0
            day_opening_total += opening_raw
            opening = drawer_money(opening_raw)
            closing_raw = summ.get("closing_cash", row_get(sh, "closing_cash", None))
            if closing_raw is not None:
                try:
                    day_closing_total += float(closing_raw or 0.0)
                    closed_shift_count += 1
                except Exception:
                    pass
            closing = drawer_money(closing_raw) if closing_raw is not None else ""
            shift_sales = sum(self._sale_new_money_amount(s) for s in sales_by_shift.get(sid, []))
            diff = summ.get("difference", None)
            diff_text = drawer_money(diff) if diff is not None else ""
            status = "OPEN"
            if closing_raw is not None:
                try:
                    status = "OK" if abs(float(diff or 0.0)) < 0.5 else "CHECK"
                except Exception:
                    status = "CHECK"

            shift_lbl = str(row_get(sh, "shift_code", "") or sid).strip()
            parent = self.tree.insert(
                "",
                tk.END,
                iid=f"shift_{sid}",
                text=f"Shift #{shift_lbl}",
                open=True,
                values=(
                    self._time_range(row_get(sh, "opened_at", ""), row_get(sh, "closed_at", "")),
                    employee,
                    opening,
                    closing,
                    drawer_money(shift_sales),
                    diff_text,
                    status,
                )
            )

            for sale in sales_by_shift.get(sid, []):
                self._insert_sale_row(parent, sale)
            for movement in movements_by_shift.get(sid, []):
                self._insert_cash_movement_row(parent, movement)

        if unassigned:
            parent = self.tree.insert(
                "",
                tk.END,
                text="Unassigned Sales",
                open=True,
                values=("", "", "", "", drawer_money(sum(self._sale_new_money_amount(s) for s in unassigned)), "", "CHECK")
            )
            for sale in unassigned:
                self._insert_sale_row(parent, sale)

        if unassigned_movements:
            unassigned_net = 0.0
            for m in unassigned_movements:
                amount = float(row_get(m, "amount_value", 0.0) or 0.0)
                if str(row_get(m, "movement_type", "OUT") or "OUT").upper() == "IN":
                    unassigned_net += amount
                else:
                    unassigned_net -= amount
            parent = self.tree.insert(
                "",
                tk.END,
                text="Unassigned Cash Movements",
                open=True,
                values=("", "", "", "", drawer_money(unassigned_net), "", "CHECK")
            )
            for movement in unassigned_movements:
                self._insert_cash_movement_row(parent, movement)

        closing_text = drawer_money(day_closing_total) if closed_shift_count > 0 else "Open"
        self.summary_var.set(
            f"{day}  |  New money: {drawer_money(daily_sales_total)}  |  "
            f"Cash sales: {drawer_money(daily_cash_sales_total)}  |  "
            f"Cash in: {drawer_money(daily_cash_in_total)}  |  "
            f"Cash out: {drawer_money(daily_cash_out_total)}  |  "
            f"Day opening: {drawer_money(day_opening_total)}  |  Day closing: {closing_text}  |  "
            f"Shifts: {len(day_shifts)}  |  Sales: {len(sales)}"
        )

    def _insert_sale_row(self, parent, sale):
        created = str(row_get(sale, "created_at", "") or "")
        try:
            t_disp = datetime.strptime(created, "%Y-%m-%d %H:%M:%S").strftime("%I:%M %p").lstrip("0")
        except Exception:
            t_disp = created[11:16] if len(created) >= 16 else created
        receipt = str(row_get(sale, "receipt_code", "") or row_get(sale, "id", "") or "")
        pm = payment_method_label(row_get(sale, "payment_method", "") or "")
        amount = sale_display_amount(sale)
        self.tree.insert(
            parent,
            tk.END,
            text=f"Sale {receipt}",
            values=(t_disp, pm, "", "", drawer_money(amount), "", "")
        )

    def _insert_cash_movement_row(self, parent, movement):
        created = str(row_get(movement, "created_at", "") or "")
        try:
            t_disp = datetime.strptime(created, "%Y-%m-%d %H:%M:%S").strftime("%I:%M %p").lstrip("0")
        except Exception:
            t_disp = created[11:16] if len(created) >= 16 else created
        amount = float(row_get(movement, "amount_value", 0.0) or 0.0)
        mtype = str(row_get(movement, "movement_type", "OUT") or "OUT").upper()
        who = str(row_get(movement, "employee_name", "") or "").strip()
        reason = str(row_get(movement, "reason", "") or "").strip()
        movement_id = int(row_get(movement, "id", 0) or 0)
        label = f"Cash {'In' if mtype == 'IN' else 'Out'} #{movement_id}"
        if reason:
            label = f"{label}: {reason[:42]}"
        self.tree.insert(
            parent,
            tk.END,
            text=label,
            values=(t_disp, who, "", "", (drawer_money(amount) if mtype == "IN" else f"-{drawer_money(amount)}"), "", mtype)
        )


class ReturnsPage(tk.Frame):
    def __init__(self, parent, cashier_page: CashierPage):
        super().__init__(parent, bg=UI.CONTENT_BG)
        self.cashier_page = cashier_page
        self.sale = None
        self.items = []
        self.return_map = {}
        self.credit_var = tk.StringVar(value="Credit: $0.00")
        self.status_var = tk.StringVar(value="")
        self._processing_return = False
        self._build()

    def _build(self):
        wrap = VScrollableFrame(self, bg=UI.CONTENT_BG)
        wrap.pack(fill="both", expand=True)
        content = tk.Frame(wrap.inner, bg=UI.CONTENT_BG)
        content.pack(fill="both", expand=True, padx=18, pady=18)

        header = Card(content, padx=18, pady=14)
        header.pack(fill="x")
        HeaderBar(header.inner, "Returns / Exchange",
                  "Scan the receipt barcode, select items, then create a bon or start an exchange.").pack(fill="x")

        scan_row = tk.Frame(header.inner, bg=UI.CARD)
        scan_row.pack(fill="x", pady=(12, 0))

        tk.Label(scan_row, text="Scan receipt barcode", font=UI.FONT_MD, bg=UI.CARD, fg=UI.TEXT).pack(side="left")
        self.scan_entry = tk.Entry(scan_row, font=("Segoe UI", 14), width=28, bd=1, relief="solid")
        self.scan_entry.pack(side="left", padx=12)
        self.scan_entry.bind("<Return>", lambda e: self.load_receipt())
        self.scan_entry.focus()

        GhostButton(scan_row, "Load", self.load_receipt).pack(side="left")
        GhostButton(scan_row, "Clear", self.clear).pack(side="left", padx=(10, 0))

        body = Card(content, padx=14, pady=14)
        body.pack(fill="both", expand=True, pady=(14, 0))

        top_info = tk.Frame(body.inner, bg=UI.CARD)
        top_info.pack(fill="x")
        self.sale_info_lbl = tk.Label(top_info, text="No receipt loaded.", bg=UI.CARD, fg=UI.MUTED)
        self.sale_info_lbl.pack(side="left")
        tk.Label(top_info, textvariable=self.credit_var, bg=UI.CARD, fg=UI.TEXT, font=("Segoe UI", 11, "bold")).pack(
            side="right")

        cols = ("item", "price", "remaining_qty", "return_qty")
        self.tree = ttk.Treeview(body.inner, columns=cols, show="headings", height=14, selectmode="browse")
        self.tree.heading("item", text="Item")
        self.tree.heading("price", text="Price")
        self.tree.heading("remaining_qty", text="Remaining")
        self.tree.heading("return_qty", text="Return Qty")

        self.tree.column("item", width=380, anchor="w")
        self.tree.column("price", width=90, anchor="e")
        self.tree.column("remaining_qty", width=90, anchor="center")
        self.tree.column("return_qty", width=90, anchor="center")
        self.tree.pack(fill="both", expand=True, pady=(12, 10))
        try:
            self.tree.tag_configure("returned", background="#fee2e2", foreground="#991b1b")
            self.tree.tag_configure("partial_return", background="#fff7ed", foreground="#9a3412")
        except Exception:
            pass

        self.tree.bind("<Double-1>", lambda e: self.toggle_full_return())

        btn_row = tk.Frame(body.inner, bg=UI.CARD)
        btn_row.pack(fill="x")

        # Left action
        GhostButton(btn_row, "Remove Selected", self.clear_selected).pack(side="left")

        # Right actions (next to each other)
        self.btn_confirm_return = PrimaryButton(btn_row, "Create Bon", lambda: self.confirm_return(as_exchange=False))
        self.btn_confirm_return.config(bg="#16a34a", activebackground="#15803d")  # green
        self.btn_confirm_return.pack(side="right")

        self.btn_setqty = PrimaryButton(btn_row, "Set Return Qty", self.set_return_qty_popup)
        self.btn_setqty.config(bg="#facc15", activebackground="#eab308", fg="#111827")  # yellow
        self.btn_setqty.pack(side="right", padx=(10, 10))

        self.btn_start_exchange = PrimaryButton(btn_row, "Start Exchange", self.start_exchange)
        self.btn_start_exchange.pack(side="right", padx=(10, 10))

        tk.Label(body.inner, textvariable=self.status_var, bg=UI.CARD, fg=UI.PRIMARY, font=UI.FONT_SM).pack(anchor="w",
                                                                                                            pady=(
                                                                                                            10, 0))
        # ----------------------------
        # Daily sales (for quick returns)
        # ----------------------------
        daily = Card(content, padx=14, pady=14)
        daily.pack(fill="both", expand=False, pady=(14, 0))

        top = tk.Frame(daily.inner, bg=UI.CARD)
        top.pack(fill="x")

        tk.Label(top, text="Daily Sales", font=UI.FONT_MD, bg=UI.CARD, fg=UI.TEXT).pack(side="left")

        # Clickable date selector (no manual typing)
        today = datetime.now()
        years = [str(y) for y in range(today.year - 2, today.year + 2)]
        months = [f"{i:02d}" for i in range(1, 13)]
        days = [f"{i:02d}" for i in range(1, 32)]

        self.daily_year = tk.StringVar(value=str(today.year))
        self.daily_month = tk.StringVar(value=f"{today.month:02d}")
        self.daily_day = tk.StringVar(value=f"{today.day:02d}")

        tk.Label(top, text="Date:", bg=UI.CARD, fg=UI.MUTED).pack(side="left", padx=(14, 6))
        ttk.Combobox(top, textvariable=self.daily_year, values=years, width=6, state="readonly").pack(side="left")
        ttk.Combobox(top, textvariable=self.daily_month, values=months, width=4, state="readonly").pack(side="left",
                                                                                                        padx=(6, 0))
        ttk.Combobox(top, textvariable=self.daily_day, values=days, width=4, state="readonly").pack(side="left",
                                                                                                    padx=(6, 0))

        GhostButton(top, "Load", self.load_daily_sales).pack(side="left", padx=(10, 0))
        GhostButton(top, "Today", lambda: self._set_daily_date_today()).pack(side="left", padx=(10, 0))

        self.day_total_lbl = tk.Label(daily.inner, text="", bg=UI.CARD, fg=UI.MUTED)
        self.day_total_lbl.pack(anchor="w", pady=(10, 6))

        cols2 = ("receipt", "time", "net", "pay", "shift")
        self.sales_tree = ttk.Treeview(daily.inner, columns=cols2, show="headings", height=9)
        self.sales_tree.heading("receipt", text="Receipt")
        self.sales_tree.heading("time", text="Time")
        self.sales_tree.heading("net", text="Total")
        self.sales_tree.heading("pay", text="Pay")
        self.sales_tree.heading("shift", text="Shift")

        self.sales_tree.column("receipt", width=120, anchor="w")
        self.sales_tree.column("time", width=170, anchor="w")
        self.sales_tree.column("net", width=110, anchor="e")
        self.sales_tree.column("pay", width=90, anchor="center")
        self.sales_tree.column("shift", width=70, anchor="center")

        self.sales_tree.pack(fill="x", pady=(0, 6))
        self.sales_tree.bind("<<TreeviewSelect>>", lambda e: self.fill_selected_daily_sale())
        self.sales_tree.bind("<ButtonRelease-1>", lambda e: self._daily_click())
        self.sales_tree.bind("<Double-1>", lambda e: self.load_selected_daily_sale())

        self._build_bons_card(content)

        # Initial load
        self.load_daily_sales(silent=True)
        self.load_bons(silent=True)

    def _build_bons_card(self, content):
        bons = Card(content, padx=14, pady=14)
        bons.pack(fill="both", expand=False, pady=(14, 0))

        top = tk.Frame(bons.inner, bg=UI.CARD)
        top.pack(fill="x")
        tk.Label(top, text="Recent Bons", font=UI.FONT_MD, bg=UI.CARD, fg=UI.TEXT).pack(side="left")

        self.bon_search_var = tk.StringVar(value="")
        self.bon_active_only_var = tk.BooleanVar(value=False)
        tk.Entry(top, textvariable=self.bon_search_var, font=UI.FONT_MD, width=24, bd=1, relief="solid").pack(
            side="left", padx=(14, 8))
        tk.Checkbutton(top, text="Active only", variable=self.bon_active_only_var, bg=UI.CARD).pack(side="left")
        GhostButton(top, "Load", self.load_bons).pack(side="left", padx=(8, 0))

        cols = ("code", "date", "remaining", "original", "status", "employee")
        self.bons_tree = ttk.Treeview(bons.inner, columns=cols, show="headings", height=7)
        self.bons_tree.heading("code", text="Bon")
        self.bons_tree.heading("date", text="Date")
        self.bons_tree.heading("remaining", text="Remaining")
        self.bons_tree.heading("original", text="Original")
        self.bons_tree.heading("status", text="Status")
        self.bons_tree.heading("employee", text="Employee")
        self.bons_tree.column("code", width=160, anchor="w")
        self.bons_tree.column("date", width=150, anchor="w")
        self.bons_tree.column("remaining", width=100, anchor="e")
        self.bons_tree.column("original", width=100, anchor="e")
        self.bons_tree.column("status", width=90, anchor="center")
        self.bons_tree.column("employee", width=130, anchor="w")
        self.bons_tree.pack(fill="x", pady=(10, 6))
        self.bons_tree.bind("<Double-1>", lambda _e: self.load_selected_bon_credit())

        btns = tk.Frame(bons.inner, bg=UI.CARD)
        btns.pack(fill="x")
        PrimaryButton(btns, "Load in Register", self.load_selected_bon_credit).pack(side="left")
        GhostButton(btns, "Reprint", self.reprint_selected_bon).pack(side="left", padx=(8, 0))
        DangerButton(btns, "Void Bon", self.void_selected_bon).pack(side="left", padx=(8, 0))

    def load_bons(self, silent: bool = False):
        if not hasattr(self, "bons_tree"):
            return
        try:
            rows = list_bons(
                query=self.bon_search_var.get(),
                active_only=bool(self.bon_active_only_var.get()),
                limit=200,
            ) or []
        except Exception as e:
            if not silent:
                messagebox.showerror("Bons", f"Could not load bons.\n{e}")
            return
        try:
            self.bons_tree.delete(*self.bons_tree.get_children())
        except Exception:
            pass
        for bon in rows:
            code = str(row_get(bon, "code", "") or "")
            self.bons_tree.insert("", tk.END, iid=code, values=(
                code,
                str(row_get(bon, "created_at", "") or ""),
                money(row_get(bon, "remaining_amount", 0.0) or 0.0),
                money(row_get(bon, "original_amount", 0.0) or 0.0),
                str(row_get(bon, "status", "") or ""),
                str(row_get(bon, "issued_by_name", "") or ""),
            ))

    def _selected_bon_code(self):
        try:
            sel = self.bons_tree.selection()
        except Exception:
            sel = None
        if not sel:
            messagebox.showinfo("Bon", "Select a bon first.")
            return ""
        return str(sel[0])

    def reprint_selected_bon(self):
        code = self._selected_bon_code()
        if not code:
            return
        try:
            bon = get_bon_by_code(code)
            ok = bool(print_configured_bon(get_store_name(), bon))
        except Exception as e:
            messagebox.showerror("Bon", f"Could not reprint bon.\n{e}")
            return
        if not ok:
            messagebox.showwarning("Bon", "Bon was not sent to the printer. Check Settings > Receipt Printer.")

    def void_selected_bon(self):
        code = self._selected_bon_code()
        if not code:
            return
        if not messagebox.askyesno("Void Bon", f"Void remaining balance for {code}?"):
            return
        try:
            void_bon(code, "Voided from Returns screen")
            self.load_bons(silent=True)
        except Exception as e:
            messagebox.showerror("Void Bon", str(e))

    def load_selected_bon_credit(self):
        code = self._selected_bon_code()
        if not code:
            return
        try:
            self.cashier_page.load_bon_credit(code)
            root = self.winfo_toplevel()
            if hasattr(root, "show_page"):
                root.show_page("CashierPage")
        except Exception as e:
            messagebox.showerror("Bon", f"Could not load bon in register.\n{e}")

    def _daily_date_str(self) -> str:
        y = (self.daily_year.get() or "").strip()
        m = (self.daily_month.get() or "").strip()
        d = (self.daily_day.get() or "").strip()
        if not (y and m and d):
            dt = datetime.now()
            return dt.strftime("%Y-%m-%d")
        return f"{y}-{m}-{d}"

    def _set_daily_date_today(self):
        dt = datetime.now()
        self.daily_year.set(str(dt.year))
        self.daily_month.set(f"{dt.month:02d}")
        self.daily_day.set(f"{dt.day:02d}")
        self.load_daily_sales(silent=True)

    def load_daily_sales(self, silent: bool = False):
        """Populate the Daily Sales list on the Returns page."""
        try:
            day = self._daily_date_str()
            rows = list_sales_for_day(day, limit=1000)
        except Exception as e:
            if not silent:
                messagebox.showerror("Error", f"{type(e).__name__}: {e}")
            return

        try:
            self.sales_tree.delete(*self.sales_tree.get_children())
        except Exception:
            pass

        net_total_sum = 0.0
        for r in rows or []:
            sale_id = int(row_get(r, "id", 0) or 0)
            receipt = (row_get(r, "receipt_code", "") or "").strip()
            if not receipt:
                receipt = str(sale_id)

            created_at = row_get(r, "created_at", "") or ""
            t_disp = created_at
            try:
                # show HH:MM if possible
                if " " in created_at:
                    t_disp = created_at
            except Exception:
                pass
            net = sale_display_amount(r)
            net_total_sum += net

            self.sales_tree.insert(
                "",
                tk.END,
                iid=str(sale_id),
                values=(
                    receipt,
                    t_disp,
                    money(net),
                    payment_method_label(row_get(r, "payment_method", "") or ""),
                    (row_get(r, "shift_code", "") or row_get(r, "shift_id", "") or "")
                )
            )

        self.day_total_lbl.config(text=f"Day total: {money(net_total_sum)}  |  Sales: {len(rows or [])}")

    def _daily_click(self):
        """Ensure click always fills scan box, even if selection doesn't change."""
        try:
            self.after(1, self.fill_selected_daily_sale)
        except Exception:
            try:
                self.fill_selected_daily_sale()
            except Exception:
                pass

    def fill_selected_daily_sale(self):
        """When you click a sale in Daily Sales, put a *resolvable* receipt code in the scan box (no auto-load).

        IMPORTANT:
          Your receipt "code" column (e.g. 0004) is NOT always the sale id.
          The scanner resolver supports: R-YYYY-MM-DD-0004.
        """
        try:
            sel = self.sales_tree.selection()
        except Exception:
            sel = None
        if not sel:
            return

        iid = sel[0]
        vals = self.sales_tree.item(iid, "values") or ()
        receipt_code = str(vals[0]).strip() if len(vals) > 0 else str(iid)

        # If it already looks like a full receipt scan, keep it
        if receipt_code.lower().startswith("r-"):
            scan_value = receipt_code
        else:
            # Build R-YYYY-MM-DD-XXXX using the row date (column "Time")
            date_part = ""
            if len(vals) > 1 and str(vals[1]).strip():
                # vals[1] example: 2026-01-03 12:20:14
                date_part = str(vals[1]).strip()[:10]
            if not date_part:
                date_part = datetime.now().strftime("%Y-%m-%d")

            scan_value = f"R-{date_part}-{receipt_code}"

        self.scan_entry.delete(0, tk.END)
        self.scan_entry.insert(0, scan_value)

    def load_selected_daily_sale(self):
        """Double-click a sale: fill the scan box then load it."""
        self.fill_selected_daily_sale()
        self.load_receipt()

    def on_scan_code(self, code: str):
        self.scan_entry.delete(0, tk.END)
        self.scan_entry.insert(0, code)
        self.load_receipt()

    def clear(self):
        self.sale = None
        self.items = []
        self.return_map = {}
        self.credit_var.set("Credit: $0.00")
        self.status_var.set("")
        self.sale_info_lbl.config(text="No receipt loaded.")
        try:
            self.tree.delete(*self.tree.get_children())
        except Exception:
            pass
        try:
            self.scan_entry.delete(0, tk.END)
            self.scan_entry.focus()
        except Exception:
            pass

    def load_receipt(self):
        code = self.scan_entry.get().strip()
        if not code:
            return
        sale, items = get_sale_by_receipt_scan(code)
        if not sale:
            messagebox.showerror("Not found", "Receipt not found. Scan again.")
            return

        self.sale = sale
        self.all_items = list(items or [])
        self.items = list(items or [])
        self.return_map = {}

        receipt_code = (sale.get("receipt_code") or "").strip() or str(sale["id"])
        self.sale_info_lbl.config(
            text=(
                f"Sale {receipt_code}   |   {sale['created_at']}   |   "
                f"Total: {money(sale['total_amount'])}   |   "
                f"Pay: {payment_method_label(sale.get('payment_method', ''))}"
            )
        )

        self.tree.delete(*self.tree.get_children())

        for it in self.items:
            sale_item_id = int(it["id"])
            self.return_map[sale_item_id] = 0
            sold_qty = self._sold_qty(it)
            returned_qty = self._returned_qty(it)
            remaining_qty = self._remaining_qty(it)
            tags = ()
            if sold_qty > 0 and remaining_qty <= 0:
                tags = ("returned",)
            elif returned_qty > 0:
                tags = ("partial_return",)
            self.tree.insert("", tk.END, iid=str(sale_item_id), values=(
                it["name"],
                money(self._return_unit_price(it)),
                remaining_qty,
                0
            ), tags=tags)

        self._recalc_credit()
        self.status_var.set("Select items to return. Double click a row to mark full return.")

    def _return_unit_price(self, it: dict) -> float:
        """Return unit price the customer actually paid (best effort).

        Exchange credit is a tender, not a discount. Use the recorded net line value
        so replacement merchandise keeps its returnable value across later exchanges.
        """
        # 1) explicit net/paid per-unit fields
        for k in (
            "paid_unit_price", "net_unit_price", "unit_net_price", "final_unit_price",
            "discounted_unit_price", "unit_price_net", "unit_paid_price"
        ):
            try:
                v = it.get(k, None)
                if v is not None and str(v) != "":
                    return float(v)
            except Exception:
                pass

        # qty
        try:
            qty = int(float(it.get("qty") or 0))
        except Exception:
            qty = 0
        qty = max(0, qty)

        # 2) net/paid line totals
        if qty > 0:
            for k in ("paid_line_total", "net_line_total", "final_line_total", "line_total", "total_paid"):
                try:
                    v = it.get(k, None)
                    if v is not None and str(v) != "":
                        return float(v) / qty
                except Exception:
                    pass

        # 3) base unit price with per-item discount percent if present
        try:
            price = float(it.get("unit_price") or it.get("price") or 0.0)
        except Exception:
            price = 0.0

        disc_pct = 0.0
        for k in ("discount_pct", "discount_percent", "disc_pct", "discount", "pct_discount"):
            try:
                v = it.get(k, None)
                if v is None or str(v) == "":
                    continue
                disc_pct = float(v)
                break
            except Exception:
                pass

        # allow fractional (0.2 == 20%)
        if 0 < disc_pct < 1:
            disc_pct = disc_pct * 100.0
        disc_pct = max(0.0, min(100.0, disc_pct))

        base = price * (1.0 - disc_pct / 100.0)

        return base

    def _qty_value(self, it: dict, key: str, default: int = 0) -> int:
        try:
            if isinstance(it, dict) and key in it and it.get(key) is not None:
                return max(0, int(float(it.get(key) or 0)))
        except Exception:
            pass
        try:
            return max(0, int(float(row_get(it, key, default) or default)))
        except Exception:
            return max(0, int(default or 0))

    def _sold_qty(self, it: dict) -> int:
        return self._qty_value(it, "qty", 0)

    def _returned_qty(self, it: dict) -> int:
        return self._qty_value(it, "returned_qty", 0)

    def _remaining_qty(self, it: dict) -> int:
        # Important: remaining_qty=0 is a real value. Do not use `or qty`.
        if isinstance(it, dict) and "remaining_qty" in it and it.get("remaining_qty") is not None:
            return self._qty_value(it, "remaining_qty", 0)
        return max(0, self._sold_qty(it) - self._returned_qty(it))


    def _recalc_credit(self):
        credit = 0.0
        for it in self.items:
            sid = int(it["id"])
            rq = int(self.return_map.get(sid, 0) or 0)
            if rq <= 0:
                continue
            remaining_qty = self._remaining_qty(it)
            rq = max(0, min(remaining_qty, rq))
            self.return_map[sid] = rq
            unit_price = float(self._return_unit_price(it) or 0.0)
            credit += unit_price * rq
        self.credit_var.set(f"Credit: {money(credit)}")
        return credit

    def _selected_item(self):
        sel = self.tree.selection()
        if not sel:
            return None
        sid = int(sel[0])
        for it in self.items:
            if int(it["id"]) == sid:
                return it
        return None

    def toggle_full_return(self):
        it = self._selected_item()
        if not it:
            return
        sid = int(it["id"])
        remaining_qty = self._remaining_qty(it)
        if remaining_qty <= 0:
            messagebox.showinfo("Already returned", "This item is already fully returned. Void/reset the return to make it returnable again.")
            return
        cur = int(self.return_map.get(sid, 0) or 0)
        new_qty = 0 if cur > 0 else remaining_qty
        self.return_map[sid] = new_qty
        self.tree.set(str(sid), "return_qty", str(new_qty))
        self._recalc_credit()

    def set_return_qty_popup(self):
        it = self._selected_item()
        if not it:
            messagebox.showinfo("Select", "Select an item first.")
            return
        sid = int(it["id"])
        remaining_qty = self._remaining_qty(it)
        if remaining_qty <= 0:
            messagebox.showinfo("Already returned", "This item is already fully returned. Void/reset the return to make it returnable again.")
            return

        win = tk.Toplevel(self)
        win.title("Return quantity")
        win.geometry("360x220")
        win.configure(bg=UI.CONTENT_BG)
        win.grab_set()

        box = Card(win, padx=16, pady=16)
        box.pack(fill="both", expand=True, padx=14, pady=14)

        tk.Label(box.inner, text=it["name"], font=UI.FONT_LG, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w")
        tk.Label(box.inner, text=f"Remaining returnable qty: {remaining_qty}", bg=UI.CARD, fg=UI.MUTED).pack(anchor="w",
                                                                                                             pady=(
                                                                                                             6, 12))

        e = tk.Entry(box.inner, font=("Segoe UI", 16), width=10, bd=1, relief="solid")
        e.insert(0, str(int(self.return_map.get(sid, 0) or 0)))
        e.pack(anchor="w")
        e.focus()

        def apply():
            try:
                q = int(float(e.get().strip() or "0"))
            except Exception:
                messagebox.showerror("Invalid", "Qty must be a number.")
                return
            q = max(0, min(remaining_qty, q))
            self.return_map[sid] = q
            self.tree.set(str(sid), "return_qty", str(q))
            self._recalc_credit()
            win.destroy()

        btns = tk.Frame(box.inner, bg=UI.CARD)
        btns.pack(fill="x", pady=(16, 0))
        GhostButton(btns, "Cancel", win.destroy).pack(side="right")
        PrimaryButton(btns, "Apply", apply).pack(side="right", padx=(0, 10))

    def clear_selected(self):
        it = self._selected_item()
        if not it:
            return
        sid = int(it["id"])
        self.return_map[sid] = 0
        self.tree.set(str(sid), "return_qty", "0")
        self._recalc_credit()

    def start_exchange(self):
        """Create exchange credit when items are selected, otherwise jump to cashier."""
        try:
            credit = float(self._recalc_credit() or 0.0)
        except Exception:
            credit = 0.0

        if credit > 0:
            return self.confirm_return(as_exchange=True)

        try:
            if float(getattr(self.cashier_page, "exchange_credit", 0.0) or 0.0) <= 0:
                messagebox.showinfo("No items", "Select items to return first.")
                return
        except Exception:
            messagebox.showinfo("No items", "Select items to return first.")
            return

        try:
            root = self.winfo_toplevel()
            if hasattr(root, "show_page"):
                root.show_page("CashierPage")
        except Exception:
            pass
        try:
            self.cashier_page.scan_entry.focus()
        except Exception:
            pass


    
    def confirm_return(self, as_exchange: bool = False):
            """Create a return record and issue either a digital bon or immediate exchange credit.

            - as_exchange=False: creates a printable bon for later store credit.
            - as_exchange=True : creates pending credit and jumps back to the register.
            """
            if not self.sale:
                messagebox.showinfo("Load receipt", "Scan a receipt first.")
                return

            # Prevent double-click/spam creating multiple credits
            if getattr(self, "_processing_return", False):
                return
            self._processing_return = True

            try:
                # Disable buttons while processing
                try:
                    if hasattr(self, "btn_confirm_return"):
                        self.btn_confirm_return.config(state="disabled")
                    if hasattr(self, "btn_start_exchange"):
                        self.btn_start_exchange.config(state="disabled")
                    if hasattr(self, "btn_setqty"):
                        self.btn_setqty.config(state="disabled")
                except Exception:
                    pass

                credit = self._recalc_credit()
                if credit <= 0:
                    messagebox.showinfo("No items", "Select at least one item to return.")
                    return

                returned_lines = []
                for it in self.items:
                    sid = int(it["id"])
                    rq = int(self.return_map.get(sid, 0) or 0)
                    if rq <= 0:
                        continue

                    remaining_qty = self._remaining_qty(it)
                    rq = max(0, min(remaining_qty, rq))
                    if rq <= 0:
                        continue

                    price = float(self._return_unit_price(it) or 0.0)
                    returned_lines.append({
                        "sale_item_id": sid,
                        "product_id": int(it["product_id"]) if it.get("product_id") is not None else None,
                        "name": it.get("name", ""),
                        "price": price,
                        "qty": rq,
                        "line_total": round(price * rq, 2)
                    })

                if not returned_lines:
                    messagebox.showinfo("No items", "Nothing to return (already fully returned).")
                    return

                try:
                    # Save the return against the ORIGINAL sale (the original sale is NOT modified).
                    rid, total_return = create_return(int(self.sale["id"]), returned_lines, notes=("Exchange" if as_exchange else "Return"))
                except Exception as e:
                    messagebox.showerror("Error", str(e))
                    return

                bon = None
                if as_exchange:
                    try:
                        self.cashier_page.add_exchange_credit(
                            float(total_return or 0.0),
                            origin_sale_id=int(self.sale["id"]),
                            return_id=int(rid),
                        )
                    except Exception:
                        pass
                else:
                    employee_name = ""
                    try:
                        sh = get_open_shift()
                        employee_name = str(row_get(sh, "employee_name", "") or "").strip()
                    except Exception:
                        employee_name = ""
                    if not employee_name:
                        employee_name = simpledialog.askstring("Bon employee", "Employee name for bon:", parent=self) or ""
                    try:
                        bon = create_bon(
                            int(rid),
                            issued_by_name=employee_name,
                            signature_text=employee_name,
                            notes="Created from return",
                        )
                    except Exception as e:
                        messagebox.showerror("Bon", f"Return #{rid} was saved, but bon could not be created.\n{e}")
                        return
                    printed = False
                    try:
                        printed = bool(print_configured_bon(get_store_name(), bon))
                    except Exception:
                        printed = False
                    if not printed:
                        messagebox.showwarning(
                            "Bon print failed",
                            "The bon was created, but it was not sent to the printer.\n\n"
                            "Check Settings > Receipt Printer."
                        )

                if as_exchange:
                    self.status_var.set(f"Exchange saved (Return #{rid}). Credit ready: {money(total_return)}")
                elif bon:
                    self.status_var.set(f"Bon {row_get(bon, 'code', '')} created for {money(total_return)} (Return #{rid}).")
                else:
                    self.status_var.set(f"Return saved (Return #{rid}). Credit ready: {money(total_return)}")
                self.credit_var.set(f"Credit: {money(total_return)}")

                # Refresh the receipt so remaining qty updates and returned items disappear.
                try:
                    self.load_receipt()
                except Exception:
                    pass
                try:
                    self.load_bons(silent=True)
                except Exception:
                    pass

                # Exchange: jump back to register immediately.
                if as_exchange:
                    try:
                        root = self.winfo_toplevel()
                        if hasattr(root, "show_page"):
                            root.show_page("CashierPage")
                    except Exception:
                        pass
                    try:
                        self.cashier_page.scan_entry.focus()
                    except Exception:
                        pass

            finally:
                self._processing_return = False
                try:
                    if hasattr(self, "btn_confirm_return"):
                        self.btn_confirm_return.config(state="normal")
                    if hasattr(self, "btn_start_exchange"):
                        self.btn_start_exchange.config(state="normal")
                    if hasattr(self, "btn_setqty"):
                        self.btn_setqty.config(state="normal")
                except Exception:
                    pass

# ---------------- OFFERS PAGE ----------------

class OffersPage(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg=UI.CONTENT_BG)
        self._build()

    def _build(self):
        scroll = VScrollableFrame(self, bg=UI.CONTENT_BG)
        scroll.pack(fill="both", expand=True)
        wrap = tk.Frame(scroll.inner, bg=UI.CONTENT_BG)
        wrap.pack(fill="both", expand=True, padx=(10 if UI.COMPACT else 18), pady=(10 if UI.COMPACT else 18))

        header = Card(wrap, padx=18, pady=18)
        header.pack(fill="x")
        HeaderBar(header.inner, "Offers", "Store credit bons, seasonal sales, bundle pricing, and rewards.").pack(fill="x")

        body = tk.Frame(wrap, bg=UI.CONTENT_BG)
        body.pack(fill="both", expand=True, pady=((8 if UI.COMPACT else 14), 0))
        body.grid_columnconfigure(0, weight=1)
        body.grid_columnconfigure(1, weight=1)

        seasonal = Card(body, padx=18, pady=18)
        seasonal.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        tk.Label(seasonal.inner, text="Seasonal Sale", font=UI.FONT_LG, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w")
        tk.Label(
            seasonal.inner,
            text="Set percentage discounts for selected products.",
            font=UI.FONT_MD,
            bg=UI.CARD,
            fg=UI.MUTED,
        ).pack(anchor="w", pady=(6, 14))
        PrimaryButton(seasonal.inner, "Manage Seasonal Sale", self._open_seasonal_sale_manager).pack(anchor="w")

        bundles = Card(body, padx=18, pady=18)
        bundles.grid(row=0, column=1, sticky="nsew", padx=(8, 0))
        tk.Label(bundles.inner, text="Bundle Offers", font=UI.FONT_LG, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w")
        tk.Label(
            bundles.inner,
            text="Create offers like 3 jeans for $25.",
            font=UI.FONT_MD,
            bg=UI.CARD,
            fg=UI.MUTED,
        ).pack(anchor="w", pady=(6, 14))
        PrimaryButton(bundles.inner, "Manage Bundle Offers", self._open_bundle_offer_manager).pack(anchor="w")

        warehouse = Card(body, padx=18, pady=18)
        warehouse.grid(row=1, column=0, sticky="nsew", pady=(14, 0), padx=(0, 8))
        tk.Label(warehouse.inner, text="Warehouse Paper", font=UI.FONT_LG, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w")
        tk.Label(
            warehouse.inner,
            text="Select products and print their locations with prices.",
            font=UI.FONT_MD,
            bg=UI.CARD,
            fg=UI.MUTED,
        ).pack(anchor="w", pady=(6, 14))
        PrimaryButton(warehouse.inner, "Make Warehouse Paper", self._open_warehouse_paper_manager).pack(anchor="w")

        wheel = Card(body, padx=18, pady=18)
        wheel.grid(row=1, column=1, sticky="nsew", pady=(14, 0), padx=(8, 0))
        tk.Label(wheel.inner, text="Spin Wheel", font=UI.FONT_LG, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w")
        tk.Label(
            wheel.inner,
            text="Configure weighted discounts, free items, and no-prize slices.",
            font=UI.FONT_MD,
            bg=UI.CARD,
            fg=UI.MUTED,
        ).pack(anchor="w", pady=(6, 14))
        PrimaryButton(wheel.inner, "Manage Spin Wheel", self._open_spin_wheel_manager).pack(anchor="w")

        bon = Card(body, padx=18, pady=18)
        bon.grid(row=2, column=0, columnspan=2, sticky="ew", pady=(14, 0))
        tk.Label(bon.inner, text="Store Credit Bon", font=UI.FONT_LG, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w")

        form = tk.Frame(bon.inner, bg=UI.CARD)
        form.pack(fill="x", pady=(10, 8))
        tk.Label(form, text="Amount", font=UI.FONT_MD, bg=UI.CARD, fg=UI.TEXT).grid(row=0, column=0, sticky="w")
        tk.Label(form, text="Person", font=UI.FONT_MD, bg=UI.CARD, fg=UI.TEXT).grid(row=0, column=1, sticky="w", padx=(12, 0))

        self.bon_amount_var = tk.StringVar(value="")
        self.bon_person_var = tk.StringVar(value="")
        tk.Entry(form, textvariable=self.bon_amount_var, font=("Segoe UI", 13), width=14, bd=1, relief="solid").grid(
            row=1, column=0, sticky="w", pady=(4, 0))
        self.bon_person_combo = ttk.Combobox(
            form,
            textvariable=self.bon_person_var,
            values=[],
            width=28,
            state="normal",
            font=("Segoe UI", 12),
        )
        self.bon_person_combo.grid(row=1, column=1, sticky="w", padx=(12, 0), pady=(4, 0))

        btns = tk.Frame(bon.inner, bg=UI.CARD)
        btns.pack(fill="x", pady=(6, 0))
        PrimaryButton(btns, "Create & Print Bon", self._create_manual_bon).pack(side="left")
        GhostButton(btns, "Refresh", self._refresh_manual_bon_card).pack(side="left", padx=(8, 0))

        self.bon_status_var = tk.StringVar(value="")
        tk.Label(bon.inner, textvariable=self.bon_status_var, font=UI.FONT_SM, bg=UI.CARD, fg=UI.PRIMARY).pack(
            anchor="w", pady=(8, 0))

        cols = ("code", "date", "remaining", "person", "status")
        self.offer_bons_tree = ttk.Treeview(bon.inner, columns=cols, show="headings", height=5)
        self.offer_bons_tree.heading("code", text="Bon")
        self.offer_bons_tree.heading("date", text="Date")
        self.offer_bons_tree.heading("remaining", text="Remaining")
        self.offer_bons_tree.heading("person", text="Person")
        self.offer_bons_tree.heading("status", text="Status")
        self.offer_bons_tree.column("code", width=155, anchor="w")
        self.offer_bons_tree.column("date", width=155, anchor="w")
        self.offer_bons_tree.column("remaining", width=100, anchor="e")
        self.offer_bons_tree.column("person", width=160, anchor="w")
        self.offer_bons_tree.column("status", width=90, anchor="center")
        self.offer_bons_tree.pack(fill="x", pady=(10, 6))
        self.offer_bons_tree.bind("<Double-1>", lambda _e: self._reprint_selected_offer_bon())

        recent_btns = tk.Frame(bon.inner, bg=UI.CARD)
        recent_btns.pack(fill="x")
        GhostButton(recent_btns, "Reprint Selected", self._reprint_selected_offer_bon).pack(side="left")

        credit = Card(body, padx=18, pady=18)
        credit.grid(row=3, column=0, columnspan=2, sticky="ew", pady=(14, 0))
        tk.Label(credit.inner, text="Pending Store Credit", font=UI.FONT_LG, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w")
        self.credit_status_var = tk.StringVar(value="")
        tk.Label(credit.inner, textvariable=self.credit_status_var, font=UI.FONT_MD, bg=UI.CARD, fg=UI.MUTED).pack(
            anchor="w", pady=(6, 14))
        DangerButton(credit.inner, "Clear Pending Store Credit", self._clear_pending_credit).pack(anchor="w")
        self._refresh_manual_bon_card()
        self.refresh_credit_status()

    def _open_seasonal_sale_manager(self):
        return _seasonal_sale_manager_window(self)

    def _open_bundle_offer_manager(self):
        return _bundle_offer_manager_window(self)

    def _open_spin_wheel_manager(self):
        return _spin_wheel_manager_window(self)

    def _open_warehouse_paper_manager(self):
        return _warehouse_paper_window(self)

    def _refresh_manual_bon_card(self):
        try:
            names = [str(row_get(e, "name", "") or "").strip() for e in (list_employees(True) or [])]
            names = [n for n in names if n]
        except Exception:
            names = []
        try:
            self.bon_person_combo.configure(values=names)
        except Exception:
            pass
        if not str(self.bon_person_var.get() or "").strip():
            try:
                shift = get_open_shift()
                shift_name = str(row_get(shift, "employee_name", "") or "").strip()
            except Exception:
                shift_name = ""
            if shift_name:
                self.bon_person_var.set(shift_name)
        self._refresh_offer_bons()

    def _parse_bon_amount(self):
        raw = str(self.bon_amount_var.get() or "").strip().replace("$", "").replace(",", "")
        try:
            amount = round(float(raw), 2)
        except Exception:
            raise ValueError("Enter a valid bon amount.")
        if amount <= 0:
            raise ValueError("Bon amount must be greater than zero.")
        return amount

    def _create_manual_bon(self):
        try:
            amount = self._parse_bon_amount()
            person = str(self.bon_person_var.get() or "").strip()
            if not person:
                raise ValueError("Enter the person who issued this bon.")
            bon = create_bon(
                None,
                issued_by_name=person,
                signature_text=person,
                notes="Manual store credit bon",
                amount=amount,
            )
            ok = bool(print_configured_bon(get_store_name(), bon))
        except Exception as e:
            messagebox.showerror("Bon", str(e))
            return
        code = str(row_get(bon, "code", "") or "")
        self.bon_status_var.set(f"Created {code} for {money(amount)}.")
        self.bon_amount_var.set("")
        self._refresh_offer_bons()
        if not ok:
            messagebox.showwarning("Bon", "Bon was created but was not sent to the printer. Check Settings > Receipt Printer.")

    def _refresh_offer_bons(self):
        if not hasattr(self, "offer_bons_tree"):
            return
        try:
            rows = list_bons(limit=25) or []
        except Exception:
            rows = []
        try:
            self.offer_bons_tree.delete(*self.offer_bons_tree.get_children())
        except Exception:
            pass
        for bon in rows:
            code = str(row_get(bon, "code", "") or "")
            if not code:
                continue
            self.offer_bons_tree.insert("", tk.END, iid=code, values=(
                code,
                str(row_get(bon, "created_at", "") or ""),
                money(row_get(bon, "remaining_amount", 0.0) or 0.0),
                str(row_get(bon, "issued_by_name", "") or ""),
                str(row_get(bon, "status", "") or ""),
            ))

    def _selected_offer_bon_code(self):
        try:
            sel = self.offer_bons_tree.selection()
        except Exception:
            sel = None
        if not sel:
            messagebox.showinfo("Bon", "Select a bon first.")
            return ""
        return str(sel[0])

    def _reprint_selected_offer_bon(self):
        code = self._selected_offer_bon_code()
        if not code:
            return
        try:
            bon = get_bon_by_code(code)
            ok = bool(print_configured_bon(get_store_name(), bon))
        except Exception as e:
            messagebox.showerror("Bon", f"Could not reprint bon.\n{e}")
            return
        if not ok:
            messagebox.showwarning("Bon", "Bon was not sent to the printer. Check Settings > Receipt Printer.")

    def refresh_credit_status(self):
        try:
            top = self.winfo_toplevel()
            amount = float(getattr(top.cashier_page, "exchange_credit", 0.0) or 0.0)
        except Exception:
            amount = 0.0
        if amount > 0.005:
            self.credit_status_var.set(f"Current pending credit: {money(amount)}")
        else:
            self.credit_status_var.set("No pending store credit.")

    def _clear_pending_credit(self):
        try:
            top = self.winfo_toplevel()
            top.cashier_page.clear_exchange_credit_prompt()
        except Exception:
            return
        self.refresh_credit_status()


# ---------------- SIMPLE PAGE ----------------

class SimplePage(tk.Frame):
    def __init__(self, parent, title, subtitle):
        super().__init__(parent, bg=UI.CONTENT_BG)
        wrap = tk.Frame(self, bg=UI.CONTENT_BG)
        wrap.pack(fill="both", expand=True, padx=18, pady=18)

        card = Card(wrap, padx=18, pady=18)
        card.pack(fill="x")
        HeaderBar(card.inner, title, subtitle).pack(fill="x")

        card2 = Card(wrap, padx=18, pady=18)
        card2.pack(fill="both", expand=True, pady=(14, 0))
        tk.Label(card2.inner, text="We will build this next.", font=UI.FONT_MD, bg=UI.CARD, fg=UI.MUTED).pack(
            anchor="w")


# ---------------- RUN ----------------


class SettingsPage(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg=UI.CONTENT_BG)
        self._build()

    def _build(self):
        wrap = VScrollableFrame(self, bg=UI.CONTENT_BG)
        wrap.pack(fill="both", expand=True)
        content = tk.Frame(wrap.inner, bg=UI.CONTENT_BG)
        content.pack(fill="both", expand=True, padx=18, pady=18)

        header = Card(content, padx=18, pady=18)
        header.pack(fill="x")
        HeaderBar(header.inner, "Settings", "Connection + receipt printer.").pack(fill="x")

        body = Card(content, padx=18, pady=18)
        body.pack(fill="both", expand=True, pady=(14, 0))

        # -------- Connection --------
        cfg = get_backend_config()
        saved_mode = cfg.get("mode", "standalone")
        self.mode_var = tk.StringVar(value=saved_mode)
        self.url_var = tk.StringVar(value=cfg.get("server_url", "http://127.0.0.1:8000"))
        self.port_var = tk.StringVar(value=str(cfg.get("host_port", 8000)))

        tk.Label(body.inner, text="Connection Mode", font=UI.FONT_MD, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w")

        rb = tk.Frame(body.inner, bg=UI.CARD)
        rb.pack(fill="x", pady=(8, 10))
        ttk.Radiobutton(rb, text="Standalone (local database on this PC)", value="standalone",
                        variable=self.mode_var).pack(anchor="w", pady=2)
        ttk.Radiobutton(rb, text="Host (main database; share on LAN and mirror to cloud)", value="host", variable=self.mode_var).pack(
            anchor="w", pady=2)
        ttk.Radiobutton(rb, text="Join (use the Host main database)", value="connect", variable=self.mode_var).pack(anchor="w",
                                                                                                            pady=2)
        ttk.Radiobutton(rb, text="Cloud (hosted sync with local offline cache)", value="cloud",
                        variable=self.mode_var).pack(anchor="w", pady=2)

        tk.Label(body.inner, text="Host Port (Host mode)", font=UI.FONT_SM, bg=UI.CARD, fg=UI.MUTED).pack(anchor="w",
                                                                                                          pady=(10, 2))
        ttk.Entry(body.inner, textvariable=self.port_var, width=12).pack(anchor="w")

        tk.Label(body.inner, text="Host URL (Join mode)", font=UI.FONT_SM, bg=UI.CARD, fg=UI.MUTED).pack(anchor="w",
                                                                                                         pady=(10, 2))

        join_row = tk.Frame(body.inner, bg=UI.CARD)
        join_row.pack(anchor="w", fill="x")

        url_entry = ttk.Entry(join_row, textvariable=self.url_var, width=42)
        url_entry.pack(side="left")

        self._found_hosts = []  # list of dicts from backend.discover_hosts
        self._host_display = tk.StringVar(value="")

        host_combo = ttk.Combobox(join_row, textvariable=self._host_display, width=32, state="readonly")
        host_combo.pack(side="left", padx=(8, 8))

        def _apply_selected_host(_evt=None):
            disp = (self._host_display.get() or "").strip()
            if not disp:
                return
            for h in self._found_hosts:
                if disp == h.get("display"):
                    self.url_var.set(str(h.get("url") or "").strip())
                    return

        host_combo.bind("<<ComboboxSelected>>", _apply_selected_host)

        def find_hosts():
            try:
                items = discover_hosts(timeout_sec=1.8)
            except Exception:
                items = []

            self._found_hosts = []
            displays = []
            for h in items:
                name = str(h.get("name") or "Host").strip()
                ip = str(h.get("ip") or h.get("from") or "").strip()
                try:
                    port = int(h.get("port") or 8000)
                except Exception:
                    port = 8000
                url = str(h.get("url") or "").strip()
                if not url:
                    continue
                disp = f"{name}  ({ip}:{port})"
                self._found_hosts.append({"display": disp, "url": url})
                displays.append(disp)

            host_combo["values"] = displays
            if displays:
                self._host_display.set(displays[0])
                _apply_selected_host()
            else:
                messagebox.showinfo(
                    "Find Hosts",
                    "No host found on the network.\n\nMake sure the Host PC is running in Host mode and both PCs are on the same Wi-Fi/LAN."
                )

        self._host_status = tk.StringVar(value="")

        ttk.Button(join_row, text="Refresh", command=find_hosts).pack(side="left")

        def search_hosts():
            # HTTP scan fallback (works when UDP broadcast is blocked)
            try:
                self._host_status.set("Searching network...")
                self.update_idletasks()
            except Exception:
                pass

            try:
                port = int(str(self.port_var.get() or "8000").strip() or "8000")
            except Exception:
                port = 8000

            try:
                items = discover_hosts_scan_http(port=port, timeout_sec=0.35, max_seconds=3.5)
            except Exception:
                items = []

            self._found_hosts = []
            displays = []
            for h in items:
                name = str(h.get("name") or "Host").strip()
                ip = str(h.get("ip") or h.get("from") or "").strip()
                try:
                    p = int(h.get("port") or port)
                except Exception:
                    p = port
                url = str(h.get("url") or f"http://{ip}:{p}").strip()

                disp = f"{name}  ({ip}:{p})"
                self._found_hosts.append({"display": disp, "url": url})
                displays.append(disp)

            host_combo["values"] = displays
            if displays:
                self._host_display.set(displays[0])
                _apply_selected_host()
                try:
                    self._host_status.set(f"Found {len(displays)} host(s).")
                except Exception:
                    pass
            else:
                try:
                    self._host_status.set("No hosts found.")
                except Exception:
                    pass
                messagebox.showinfo(
                    "Search Network",
                    "No host found.\n\nTip: if manual URL works, discovery is being blocked (usually Windows Firewall)."
                )

        ttk.Button(join_row, text="Search", command=search_hosts).pack(side="left", padx=(8, 0))
        tk.Label(body.inner, textvariable=self._host_status, bg=UI.CARD, fg=UI.MUTED).pack(anchor="w", pady=(6, 0))

        tk.Label(body.inner,
                 text="Tip: You can still type a URL manually if needed (example: http://192.168.1.10:8000).",
                 font=UI.FONT_SM, bg=UI.CARD, fg=UI.MUTED).pack(anchor="w", pady=(6, 0))

        cloud_row = tk.Frame(body.inner, bg=UI.CARD)
        cloud_row.pack(anchor="w", fill="x", pady=(10, 0))
        self.cloud_status_var = tk.StringVar(value="")

        def sync_cloud_now():
            try:
                self.cloud_status_var.set("Syncing with hosted cloud...")
                self.update_idletasks()
            except Exception:
                pass
            try:
                s = cloud_sync_now()
                online = s.get("online")
                pending = int(s.get("pending") or 0)
                uploaded = int(s.get("uploaded") or 0)
                downloaded = int(s.get("downloaded") or 0)
                seeded = int(s.get("seeded_products") or 0)
                products = int(s.get("cloud_products_applied") or 0)
                core = int(s.get("cloud_core_applied") or 0)
                if online:
                    if s.get("cloud_write_paused"):
                        reason = str(s.get("cloud_write_pause_reason") or "Cloud upload paused.")[:140]
                        self.cloud_status_var.set(f"Cloud read online, upload paused. Local Host DB still works. Pending {pending}. {reason}")
                    else:
                        self.cloud_status_var.set(f"Cloud online. Products pulled {products}, records pulled {core}, events uploaded {uploaded}, events applied {downloaded}, pending {pending}.")
                else:
                    err = str(s.get("last_error") or "Could not reach hosted cloud.")
                    self.cloud_status_var.set(f"Cloud offline. Local cache still works. Pending {pending}. {err[:120]}")
            except Exception as e:
                self.cloud_status_var.set(f"Cloud sync failed: {e}")

        ttk.Button(cloud_row, text="Sync Cloud Now", command=sync_cloud_now).pack(side="left")
        tk.Label(cloud_row, textvariable=self.cloud_status_var, bg=UI.CARD, fg=UI.MUTED, wraplength=620,
                 justify="left").pack(side="left", padx=(10, 0))
        if not supabase_emergency_enabled():
            self.cloud_status_var.set("Cloud configuration is missing on this PC. Cloud mode remains available but cannot sync yet.")

        hint = tk.Label(
            body.inner,
            text="After saving, the POS restarts. If you set Host, the server starts automatically on launch.",
            bg=UI.CARD, fg=UI.MUTED, justify="left", wraplength=720
        )
        hint.pack(anchor="w", pady=(12, 0))

        # -------- Store Name --------
        try:
            self.store_name_var = tk.StringVar(value=get_store_name() or "Keep Sports Wear")
        except Exception:
            self.store_name_var = tk.StringVar(value="Keep Sports Wear")

        tk.Label(body.inner, text="Store Name", font=UI.FONT_MD, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w", pady=(18, 6))
        sn_row = tk.Frame(body.inner, bg=UI.CARD)
        sn_row.pack(fill="x")
        ttk.Entry(sn_row, textvariable=self.store_name_var).pack(fill="x")

        # -------- Store Subtitle (Line 2) --------
        try:
            self.store_subtitle_var = tk.StringVar(value=get_store_subtitle() or "Sports Wear")
        except Exception:
            self.store_subtitle_var = tk.StringVar(value="Sports Wear")

        tk.Label(body.inner, text="Store Subtitle (Line 2)", font=UI.FONT_SM, bg=UI.CARD, fg=UI.MUTED).pack(anchor="w",
                                                                                                            pady=(8, 6))
        st_row = tk.Frame(body.inner, bg=UI.CARD)
        st_row.pack(fill="x")
        ttk.Entry(st_row, textvariable=self.store_subtitle_var).pack(fill="x")

        # -------- Currency --------
        try:
            self.lbp_rate_var = tk.StringVar(value=f"{float(get_lbp_per_usd() or 89500):.0f}")
        except Exception:
            self.lbp_rate_var = tk.StringVar(value="89500")

        tk.Label(body.inner, text="Exchange Rate", font=UI.FONT_MD, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w",
                                                                                                  pady=(18, 6))
        rate_row = tk.Frame(body.inner, bg=UI.CARD)
        rate_row.pack(fill="x")
        tk.Label(rate_row, text="LBP per $1", bg=UI.CARD, fg=UI.TEXT).pack(side="left")
        self.lbp_rate_entry = tk.Entry(rate_row, textvariable=self.lbp_rate_var, bd=1, relief="solid", width=18,
                                       font=("Segoe UI", 14))
        self.lbp_rate_entry.pack(side="left", padx=(10, 0))
        bind_lbp_grouping(self.lbp_rate_entry)
        tk.Label(rate_row, text="Used for opening/closing drawer counts.", bg=UI.CARD, fg=UI.MUTED).pack(
            side="left", padx=(10, 0))

        # -------- Printer -
        pr_cfg = get_printer_config()
        bc_cfg = get_barcode_printer_config()
        self.printer_var = tk.StringVar(value=pr_cfg.get("printer_name", ""))
        self.print_mode_var = tk.StringVar(value=pr_cfg.get("print_mode", "raw"))
        self.bc_printer_var = tk.StringVar(value=bc_cfg.get("barcode_printer_name", ""))
        self.bc_print_mode_var = tk.StringVar(value=bc_cfg.get("barcode_print_mode", "sumatra"))

        tk.Label(body.inner, text="Receipt Printer", font=UI.FONT_MD, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w",
                                                                                                   pady=(18, 6))

        pr_row = tk.Frame(body.inner, bg=UI.CARD)
        pr_row.pack(fill="x")
        tk.Label(pr_row, text="Printer:", bg=UI.CARD, fg=UI.TEXT).pack(side="left")

        self._printers_cache = list_printers()
        combo_state = "readonly" if self._printers_cache else "normal"
        self.printer_combo = ttk.Combobox(pr_row, values=self._printers_cache, textvariable=self.printer_var,
                                          state=combo_state, width=48)
        self.printer_combo.pack(side="left", padx=(8, 0))

        def refresh_printers():
            try:
                self._printers_cache = list_printers()
                self.printer_combo["values"] = self._printers_cache
                self.printer_combo.configure(state=("readonly" if self._printers_cache else "normal"))
            except Exception:
                pass

        ttk.Button(pr_row, text="Refresh", command=refresh_printers).pack(side="left", padx=(8, 0))

        mode_row = tk.Frame(body.inner, bg=UI.CARD)
        mode_row.pack(fill="x", pady=(8, 0))
        tk.Label(mode_row, text="Mode:", bg=UI.CARD, fg=UI.TEXT).pack(side="left")
        ttk.Radiobutton(mode_row, text="Direct (RAW)", value="raw", variable=self.print_mode_var).pack(side="left",
                                                                                                       padx=(8, 0))
        ttk.Radiobutton(mode_row, text="PDF (SumatraPDF)", value="sumatra", variable=self.print_mode_var).pack(
            side="left", padx=(8, 0))

        def test_print():
            try:
                set_printer_config(self.printer_var.get().strip(), self.print_mode_var.get().strip())
            except Exception:
                pass
            ok = False
            try:
                ok = bool(test_print_configured())
            except Exception:
                ok = False
            if ok:
                messagebox.showinfo("Test Print", "Sent a test print to the selected printer.")
            else:
                messagebox.showwarning("Test Print",
                                       "Could not print. Check printer name/drivers and SumatraPDF if using PDF mode.")

        ttk.Button(body.inner, text="Test Print", command=test_print).pack(anchor="w", pady=(10, 0))

        def clear_queue():
            prn = (self.printer_var.get() or "").strip()
            if not prn:
                messagebox.showwarning("Clear Queue", "Please select a receipt printer first.")
                return
            if not messagebox.askyesno("Clear Queue",
                                       "This will cancel ALL pending jobs for the selected printer. Continue?"):
                return
            ok = clear_printer_queue(prn)
            if ok:
                messagebox.showinfo("Clear Queue", "Printer queue cleared (or reset attempted).")
            else:
                messagebox.showwarning("Clear Queue",
                                       "Could not clear the printer queue. Try running as Administrator or clear it from Windows settings.")

        ttk.Button(body.inner, text="Clear Print Queue", command=clear_queue).pack(anchor="w", pady=(10, 0))

        def hard_reset_queue():
            prn = (self.printer_var.get() or "").strip()
            if not prn:
                messagebox.showwarning("Hard Reset", "Please select a receipt printer first.")
                return
            if not messagebox.askyesno(
                    "Hard Reset",
                    "This will force-reset Windows printing (spooler) and may require Administrator rights. Continue?",
            ):
                return
            ok = False
            try:
                ok = bool(hard_reset_printing(prn))
            except Exception:
                ok = False
            if ok:
                messagebox.showinfo("Hard Reset",
                                    "Hard reset completed. If the printer still prints later, power-cycle the printer (off 5 seconds) and reconnect.")
            else:
                messagebox.showwarning("Hard Reset",
                                       "Hard reset attempted but may have failed. Try running the app as Administrator, then retry.")

        ttk.Button(body.inner, text="Hard Reset Printing", command=hard_reset_queue).pack(anchor="w", pady=(6, 0))

        # -------- Barcode Labels Printer --------
        tk.Label(body.inner, text="Barcode Labels Printer", font=UI.FONT_SM, bg=UI.CARD, fg=UI.MUTED).pack(anchor="w",
                                                                                                           pady=(18, 2))

        bc_row = tk.Frame(body.inner, bg=UI.CARD)
        bc_row.pack(fill="x", anchor="w")

        bc_names = [""] + list_printers()
        bc_combo = ttk.Combobox(bc_row, values=bc_names, textvariable=self.bc_printer_var, width=46, state="readonly")
        bc_combo.pack(side="left")

        ttk.Label(bc_row, text="Mode:", background=UI.CARD).pack(side="left", padx=(12, 6))
        ttk.Combobox(bc_row, values=["sumatra", "open"], textvariable=self.bc_print_mode_var, width=12,
                     state="readonly").pack(side="left")

        def test_bc():
            try:
                set_barcode_printer_config(self.bc_printer_var.get().strip(), self.bc_print_mode_var.get().strip())
            except Exception:
                pass
            ok2 = False
            try:
                ok2 = bool(test_print_barcode_configured())
            except Exception:
                ok2 = False
            if ok2:
                messagebox.showinfo("Test Label", "Sent a test barcode label to the selected barcode printer.")
            else:
                messagebox.showwarning("Test Label",
                                       "Could not print the test label. Check printer name and SumatraPDF settings.")

        ttk.Button(body.inner, text="Test Label", command=test_bc).pack(anchor="w", pady=(10, 0))

        def clear_barcode_queue():
            prn = (self.bc_printer_var.get() or "").strip()
            if not prn:
                messagebox.showwarning("Clear Label Queue", "Please select a barcode printer first.")
                return
            if not messagebox.askyesno("Clear Label Queue", "Cancel pending label jobs for the barcode printer?"):
                return
            ok = False
            try:
                ok = bool(clear_printer_queue(prn))
            except Exception:
                ok = False
            if ok:
                messagebox.showinfo("Clear Label Queue", "Barcode printer queue cleared.")
            else:
                messagebox.showwarning("Clear Label Queue", "Could not clear the barcode printer queue.")

        ttk.Button(body.inner, text="Clear Label Queue", command=clear_barcode_queue).pack(anchor="w", pady=(6, 0))

        # -------- Daily Report Email --------
        tk.Label(body.inner, text="Daily Report Email", font=UI.FONT_MD, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w", pady=(18, 6))
        email_cfg = get_daily_report_email_config()
        self.report_email_enabled_var = tk.BooleanVar(value=bool(email_cfg.get("enabled", True)))
        self.report_email_recipients_var = tk.StringVar(value=", ".join(email_cfg.get("recipients") or []))
        self.report_email_sender_var = tk.StringVar(value=email_cfg.get("sender_email", ""))
        self.report_email_server_var = tk.StringVar(value=email_cfg.get("smtp_server", "smtp.gmail.com"))
        self.report_email_port_var = tk.StringVar(value=str(email_cfg.get("smtp_port", 587)))
        self.report_email_user_var = tk.StringVar(value=email_cfg.get("smtp_username", ""))
        self.report_email_password_var = tk.StringVar(value=email_cfg.get("smtp_password", ""))
        self.report_email_tls_var = tk.BooleanVar(value=bool(email_cfg.get("use_tls", True)))
        self.report_email_time_var = tk.StringVar(value=email_cfg.get("send_time", "19:50"))

        def _email_status_text():
            cfg = get_daily_report_email_config()
            send_time = str(cfg.get("send_time") or "19:50").strip() or "19:50"
            today = datetime.now().strftime("%Y-%m-%d")
            last_date = str(cfg.get("last_sent_date") or "").strip()
            last_at = str(cfg.get("last_sent_at") or "").strip()
            last_auto_date = str(cfg.get("last_auto_sent_date") or "").strip()
            last_auto_at = str(cfg.get("last_auto_sent_at") or "").strip()
            source = str(cfg.get("last_sent_source") or "").strip()
            source_label = {
                "schedule": "auto-send",
                "settings": "settings",
                "manual": "manual",
                "close": "close register",
            }.get(source, source)
            source_text = f" by {source_label}" if source_label else ""
            if last_date:
                if last_at:
                    last_text = f"Last emailed: {last_date} at {last_at}{source_text}."
                else:
                    last_text = f"Last emailed: {last_date}{source_text}."
            else:
                last_text = "Last emailed: never."
            if last_auto_date == today:
                auto_time_text = f" at {last_auto_at}" if last_auto_at else ""
                return f"{last_text} Auto-send at {send_time} will skip today because today's automatic report already sent{auto_time_text}."
            if last_date == today:
                return f"{last_text} Auto-send at {send_time} will still run because the automatic report has not sent yet."
            return f"{last_text} Next auto-send is {send_time} when the app is open."

        self.report_email_status_var = tk.StringVar(value=_email_status_text())

        email_grid = tk.Frame(body.inner, bg=UI.CARD)
        email_grid.pack(fill="x", anchor="w")
        email_grid.grid_columnconfigure(1, weight=1, minsize=360)

        ttk.Checkbutton(email_grid, text="Send daily Excel report", variable=self.report_email_enabled_var).grid(
            row=0, column=0, columnspan=2, sticky="w", pady=(0, 6)
        )

        def _email_row(r, label, var, show=None, width=48):
            tk.Label(email_grid, text=label, bg=UI.CARD, fg="#334155", width=18, anchor="w").grid(
                row=r, column=0, sticky="w", pady=3
            )
            e = ttk.Entry(email_grid, textvariable=var, width=width, show=show or "")
            e.grid(row=r, column=1, sticky="ew", pady=3)
            return e

        _email_row(1, "Recipients", self.report_email_recipients_var)
        _email_row(2, "Sender email", self.report_email_sender_var)
        _email_row(3, "SMTP server", self.report_email_server_var)
        smtp_row = tk.Frame(email_grid, bg=UI.CARD)
        smtp_row.grid(row=4, column=1, sticky="w", pady=3)
        ttk.Entry(smtp_row, textvariable=self.report_email_port_var, width=8).pack(side="left")
        ttk.Checkbutton(smtp_row, text="TLS", variable=self.report_email_tls_var).pack(side="left", padx=(12, 0))
        tk.Label(email_grid, text="SMTP port", bg=UI.CARD, fg="#334155", width=18, anchor="w").grid(row=4, column=0, sticky="w", pady=3)
        _email_row(5, "SMTP username", self.report_email_user_var)
        _email_row(6, "App password", self.report_email_password_var, show="*")
        _email_row(7, "Auto-send time", self.report_email_time_var, width=12)

        tk.Label(
            body.inner,
            text="Close Register emails immediately when email is configured. Auto-send runs once per day when the configured time has passed and the app is open. Time accepts 19:50 or 7:50 PM. Gmail requires a 16-character Google App Password, not the normal email password.",
            font=UI.FONT_SM,
            bg=UI.CARD,
            fg=UI.MUTED,
            wraplength=720,
            justify="left",
        ).pack(anchor="w", pady=(6, 0))
        tk.Label(
            body.inner,
            textvariable=self.report_email_status_var,
            font=UI.FONT_SM,
            bg=UI.CARD,
            fg=UI.MUTED,
            wraplength=720,
            justify="left",
        ).pack(anchor="w", pady=(4, 0))

        def _save_email_settings(show_message=True):
            try:
                port = int(str(self.report_email_port_var.get() or "587").strip())
            except Exception:
                port = 587
            try:
                set_daily_report_email_config(
                    enabled=bool(self.report_email_enabled_var.get()),
                    recipients=self.report_email_recipients_var.get(),
                    sender_email=self.report_email_sender_var.get(),
                    smtp_server=self.report_email_server_var.get(),
                    smtp_port=port,
                    smtp_username=self.report_email_user_var.get(),
                    smtp_password=self.report_email_password_var.get(),
                    use_tls=bool(self.report_email_tls_var.get()),
                    send_time=self.report_email_time_var.get(),
                )
                saved_cfg = get_daily_report_email_config()
                self.report_email_time_var.set(saved_cfg.get("send_time", "19:50"))
                self.report_email_status_var.set(_email_status_text())
                try:
                    top = self.winfo_toplevel()
                    top._daily_report_email_last_attempt = ""
                    top._daily_report_email_last_attempt_ts = 0.0
                except Exception:
                    pass
                if show_message:
                    messagebox.showinfo("Daily report email", "Email settings saved.")
                return True
            except Exception as e:
                if show_message:
                    messagebox.showwarning("Daily report email", f"Could not save email settings.\n{e}")
                return False

        email_btns = tk.Frame(body.inner, bg=UI.CARD)
        email_btns.pack(anchor="w", pady=(8, 0))
        ttk.Button(email_btns, text="Save Email Settings", command=_save_email_settings).pack(side="left")

        def _send_today_from_settings():
            if not _save_email_settings(show_message=False):
                return
            try:
                self.winfo_toplevel().send_daily_report_email_for_day(
                    datetime.now().strftime("%Y-%m-%d"),
                    source="settings",
                    silent=False,
                    force=True,
                )
                self.report_email_status_var.set(_email_status_text())
            except Exception as e:
                messagebox.showwarning("Daily report email", f"Could not send today's report.\n{e}")

        ttk.Button(email_btns, text="Email Today", command=_send_today_from_settings).pack(side="left", padx=(10, 0))

        # -------- Weekly Receipt Builder --------
        weekly_row = tk.Frame(body.inner, bg=UI.CARD)
        weekly_row.pack(fill="x", pady=(18, 0))
        tk.Label(weekly_row, text="Weekly Receipt Builder", font=UI.FONT_MD, bg=UI.CARD, fg=UI.TEXT).pack(side="left")
        PrimaryButton(weekly_row, "Open Builder", self._open_weekly_receipt_builder).pack(side="right")
        tk.Label(
            body.inner,
            text="Select real sales, then change only the printed time and printed amount for one custom combined receipt. This does not change reports, stock, or the database.",
            font=UI.FONT_SM,
            bg=UI.CARD,
            fg=UI.MUTED,
            wraplength=720,
            justify="left",
        ).pack(anchor="w", pady=(6, 0))

        # -------- Backups --------
        tk.Label(body.inner, text="Backups", font=UI.FONT_MD, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w", pady=(18, 6))

        backup_cfg = get_backup_config()
        self.backup_remote_var = tk.StringVar(value=backup_cfg.get("backup_rclone_remote", ""))
        self.backup_status_var = tk.StringVar(value="")

        remote_row = tk.Frame(body.inner, bg=UI.CARD)
        remote_row.pack(anchor="w", fill="x", pady=(0, 8))
        tk.Label(remote_row, text="Google Drive target:", bg=UI.CARD, fg=UI.TEXT).pack(side="left")
        ttk.Entry(remote_row, textvariable=self.backup_remote_var, width=44).pack(side="left", padx=(8, 0))

        def _refresh_backup_status():
            info = get_backup_config()
            remote = str(info.get("backup_rclone_remote") or "").strip()
            offsite = info.get("offsite") or {}
            if not remote:
                text = "Local snapshots are enabled. Google Drive upload is optional and currently off."
            elif not info.get("rclone_available"):
                text = "Google Drive target saved, but rclone is not installed or could not be found."
            elif offsite.get("ok"):
                text = f"Last Google Drive upload: {offsite.get('attempted_at', 'complete')}."
            elif offsite.get("message"):
                text = f"Last Google Drive upload needs attention: {str(offsite.get('message'))[:180]}"
            else:
                text = "Google Drive target saved. Click Backup Now to upload the first snapshot."
            self.backup_status_var.set(text)

        def _save_backup_target(show_message=True):
            try:
                set_backup_rclone_remote(self.backup_remote_var.get().strip())
                _refresh_backup_status()
                if show_message:
                    messagebox.showinfo("Backup", "Google Drive backup target saved.")
                return True
            except Exception as e:
                if show_message:
                    messagebox.showwarning("Backup", f"Could not save the backup target.\n{e}")
                return False

        ttk.Button(remote_row, text="Save Target", command=_save_backup_target).pack(side="left", padx=(8, 0))

        bkup_row = tk.Frame(body.inner, bg=UI.CARD)
        bkup_row.pack(anchor="w", fill="x")

        def _backup_now():
            try:
                if not _save_backup_target(show_message=False):
                    return
                ok = backup_pos_db()
                _refresh_backup_status()
                info = get_backup_config()
                offsite = info.get("offsite") or {}
                if not ok:
                    raise RuntimeError("Local snapshot failed.")
                if info.get("backup_rclone_remote") and not offsite.get("ok"):
                    messagebox.showwarning(
                        "Backup",
                        "Local backup created, but Google Drive upload needs attention.\n\n"
                        + str(offsite.get("message") or "Check the saved target and rclone setup.")
                    )
                else:
                    messagebox.showinfo("Backup", "Backup created successfully.")
            except Exception as e:
                _refresh_backup_status()
                messagebox.showwarning("Backup", f"Could not create backup.\n{e}")

        def _open_backups():
            try:
                open_backups_folder()
            except Exception:
                pass

        ttk.Button(bkup_row, text="Backup Now", command=_backup_now).pack(side="left")
        ttk.Button(bkup_row, text="Open Backups Folder", command=_open_backups).pack(side="left", padx=(10, 0))

        tk.Label(
            body.inner,
            textvariable=self.backup_status_var,
            font=UI.FONT_SM,
            bg=UI.CARD,
            fg=UI.MUTED,
            wraplength=720,
            justify="left",
        ).pack(anchor="w", pady=(6, 0))
        _refresh_backup_status()

        def _restore_backup():
            # Restore pos.db from a chosen backup file (use with care)
            try:
                from tkinter import filedialog
                import shutil, os
                # Try to restore from backups folder inside the app/base directory
                base_dir = BASE_DIR
                backups_dir = base_dir / "backups"
                initial_dir = str(backups_dir) if backups_dir.exists() else str(base_dir)
                path = filedialog.askopenfilename(
                    title="Select a backup database to restore",
                    initialdir=initial_dir,
                    filetypes=[("SQLite DB", "*.db"), ("All files", "*.*")]
                )
                if not path:
                    return

                if not messagebox.askyesno(
                    "Restore Backup",
                    "This will REPLACE your current pos.db with the selected backup.\n\n"
                    "Make sure no other PC is connected and you are not in the middle of a sale.\n\n"
                    "Continue?"
                ):
                    return

                src = Path(path)
                dst = base_dir / "pos.db"
                if not src.exists():
                    messagebox.showwarning("Restore Backup", "Selected file does not exist.")
                    return

                # Best effort: copy over and ask user to restart
                shutil.copy2(str(src), str(dst))
                messagebox.showinfo(
                    "Restore Backup",
                    "Backup restored successfully.\n\nPlease restart the app now."
                )
            except Exception as e:
                messagebox.showwarning("Restore Backup", f"Could not restore backup.\n{e}")

        ttk.Button(bkup_row, text="Restore Backup", command=_restore_backup).pack(side="left", padx=(10, 0))

        tk.Label(
            body.inner,
            text="Local auto-backup runs on startup, every six hours, when you close the register, and when you exit. Add a Google Drive target to upload the same verified snapshot off-site.",
            font=UI.FONT_SM,
            bg=UI.CARD,
            fg=UI.MUTED,
            wraplength=720,
            justify="left",
        ).pack(anchor="w", pady=(6, 0))

        # -------- App Updates --------
        tk.Label(body.inner, text="App Updates", font=UI.FONT_MD, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w", pady=(18, 6))
        update_row = tk.Frame(body.inner, bg=UI.CARD)
        update_row.pack(anchor="w", fill="x")
        self.update_status_var = tk.StringVar(value=f"Current version: {APP_VERSION}.")

        def _check_updates_clicked():
            try:
                self.winfo_toplevel().check_for_app_update(silent=False, status_var=self.update_status_var)
            except Exception as e:
                self.update_status_var.set(f"Update check failed: {e}")

        ttk.Button(update_row, text="Check for Updates", command=_check_updates_clicked).pack(side="left")
        tk.Label(
            update_row,
            textvariable=self.update_status_var,
            font=UI.FONT_SM,
            bg=UI.CARD,
            fg=UI.MUTED,
            wraplength=560,
            justify="left",
        ).pack(side="left", padx=(10, 0))
        tk.Label(
            body.inner,
            text="Updates come from GitHub Releases. Install this updater version manually once; future packaged releases can download here, close the POS, update app files, and reopen without touching your database or settings.",
            font=UI.FONT_SM,
            bg=UI.CARD,
            fg=UI.MUTED,
            wraplength=720,
            justify="left",
        ).pack(anchor="w", pady=(6, 0))

        # -------- Register Shifts Counter --------
        tk.Label(body.inner, text="Register (Shifts) Counter", font=UI.FONT_MD, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w", pady=(18, 6))

        def _reset_shift_num():
            if not messagebox.askyesno(
                "Reset Shift Counter",
                "This will reset the next register shift number back to 1.\n\n"
                "This does NOT delete any sales, shifts, or transaction history; "
                "it only resets the display number for future shifts.\n\n"
                "Are you sure you want to reset?"
            ):
                return
            ok = False
            try:
                ok = bool(reset_next_shift_number())
            except Exception as e:
                messagebox.showerror("Error", f"Failed to reset: {e}")
                return
            if ok:
                messagebox.showinfo("Reset Shift Counter", "Shift number counter reset to 1 successfully.")
            else:
                messagebox.showerror("Error", "Could not reset the shift number counter.")

        ttk.Button(body.inner, text="Reset Shift Counter to 1", command=_reset_shift_num).pack(anchor="w", pady=(6, 0))

        # -------- Save / Restart --------
        btns = tk.Frame(body.inner, bg=UI.CARD)
        btns.pack(fill="x", pady=(18, 0))

        def save_and_restart():
            mode = (self.mode_var.get() or "standalone").strip()
            url = (self.url_var.get() or "").strip()
            try:
                port = int((self.port_var.get() or "8000").strip())
            except Exception:
                port = 8000

            if mode == "host" and not (1 <= port <= 65535):
                messagebox.showerror("Port", "Please enter a valid port (1-65535).")
                return

            if mode == "connect" and not (url.startswith("http://") or url.startswith("https://")):
                messagebox.showerror("Host URL", "Host URL must start with http:// or https://")
                return

            if mode == "host":
                pwd = simpledialog.askstring(
                    "Protected mode",
                    "Enter password to save Host mode:",
                    show="*",
                    parent=self.winfo_toplevel(),
                )
                if not verify_mode_admin_password(pwd):
                    messagebox.showerror("Protected mode", "Wrong password.")
                    return

            try:
                set_printer_config(self.printer_var.get().strip(), self.print_mode_var.get().strip())
            except Exception:
                pass

            try:
                set_barcode_printer_config(self.bc_printer_var.get().strip(), self.bc_print_mode_var.get().strip())
            except Exception:
                pass
            try:
                _save_email_settings(show_message=False)
            except Exception:
                pass

            set_store_name(self.store_name_var.get())
            try:
                set_store_subtitle(self.store_subtitle_var.get())
            except Exception:
                pass
            try:
                rate = parse_lbp_text(self.lbp_rate_var.get())
                if rate <= 0:
                    raise ValueError()
                set_lbp_per_usd(rate)
                self.lbp_rate_var.set(f"{rate:,}")
            except Exception:
                messagebox.showerror("Exchange Rate", "Enter a valid whole-number LBP per $1 rate.")
                return
            set_backend_config(mode=mode, server_url=url, host_port=port)

            messagebox.showinfo("Restart", "Saved. The POS will restart now.")
            try:
                import subprocess as _sp, sys as _sys
                _release_single_instance()
                _sp.Popen([_sys.executable] + _sys.argv)
                os._exit(0)
            except Exception:
                pass
            try:
                self.winfo_toplevel().destroy()
            except Exception:
                pass

        ttk.Button(btns, text="Save + Restart POS", command=save_and_restart).pack(side="right")



    def _open_weekly_receipt_builder(self):
        today = date.today()
        start_default = today - timedelta(days=6)

        win = tk.Toplevel(self)
        win.title("Weekly Receipt Builder")
        win.geometry("980x700")
        win.minsize(900, 620)
        win.configure(bg=UI.CONTENT_BG)
        win.transient(self.winfo_toplevel())
        win.grab_set()

        outer = Card(win, padx=16, pady=16)
        outer.pack(fill="both", expand=True, padx=14, pady=14)

        HeaderBar(
            outer.inner,
            "Weekly Receipt Builder",
            "Choose past sales, then edit only the printed time and printed amount for one custom receipt.",
        ).pack(fill="x")

        top = tk.Frame(outer.inner, bg=UI.CARD)
        top.pack(fill="x", pady=(14, 10))

        start_var = tk.StringVar(value=start_default.strftime("%Y-%m-%d"))
        end_var = tk.StringVar(value=today.strftime("%Y-%m-%d"))

        tk.Label(top, text="From:", bg=UI.CARD, fg=UI.TEXT).pack(side="left")
        ttk.Entry(top, textvariable=start_var, width=14).pack(side="left", padx=(6, 12))
        tk.Label(top, text="To:", bg=UI.CARD, fg=UI.TEXT).pack(side="left")
        ttk.Entry(top, textvariable=end_var, width=14).pack(side="left", padx=(6, 12))

        body = tk.Frame(outer.inner, bg=UI.CARD)
        body.pack(fill="both", expand=True)
        body.grid_columnconfigure(0, weight=3)
        body.grid_columnconfigure(1, weight=2)
        body.grid_rowconfigure(0, weight=1)

        left_card = Card(body, padx=10, pady=10)
        left_card.grid(row=0, column=0, sticky="nsew", padx=(0, 10))

        cols = ("sel", "date", "time", "receipt", "total", "payment")
        tree = ttk.Treeview(left_card.inner, columns=cols, show="headings", height=18, selectmode="browse")
        tree.heading("sel", text="Use")
        tree.heading("date", text="Date")
        tree.heading("time", text="Time")
        tree.heading("receipt", text="Receipt")
        tree.heading("total", text="Amount")
        tree.heading("payment", text="Pay")
        tree.column("sel", width=48, anchor="center")
        tree.column("date", width=110, anchor="center")
        tree.column("time", width=90, anchor="center")
        tree.column("receipt", width=110, anchor="center")
        tree.column("total", width=90, anchor="e")
        tree.column("payment", width=90, anchor="center")
        tree.pack(side="left", fill="both", expand=True)
        ysb = ttk.Scrollbar(left_card.inner, orient="vertical", command=tree.yview)
        ysb.pack(side="right", fill="y")
        tree.configure(yscrollcommand=ysb.set)

        right_card = Card(body, padx=12, pady=12)
        right_card.grid(row=0, column=1, sticky="nsew")

        rows = []
        row_map = {}
        current_iid = {"value": None}

        selected_var = tk.StringVar(value="Selected sale: none")
        time_var = tk.StringVar(value="")
        amount_var = tk.StringVar(value="")
        status_var = tk.StringVar(value="Load sales to begin.")

        tk.Label(right_card.inner, textvariable=selected_var, bg=UI.CARD, fg=UI.TEXT, font=UI.FONT_MD).pack(anchor="w")
        tk.Label(right_card.inner, text="Printed time", bg=UI.CARD, fg=UI.TEXT).pack(anchor="w", pady=(14, 2))
        ttk.Entry(right_card.inner, textvariable=time_var, width=18).pack(anchor="w")
        tk.Label(right_card.inner, text="Printed amount", bg=UI.CARD, fg=UI.TEXT).pack(anchor="w", pady=(12, 2))
        ttk.Entry(right_card.inner, textvariable=amount_var, width=18).pack(anchor="w")

        tk.Label(
            right_card.inner,
            text="Tip: use time like 10:15 AM and amount like 18 or 18.00. Only this special receipt changes, not the real sale.",
            bg=UI.CARD,
            fg=UI.MUTED,
            justify="left",
            wraplength=280,
        ).pack(anchor="w", pady=(12, 0))

        def _parse_dt(raw):
            s = str(raw or "").strip()
            if not s:
                return None
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
                try:
                    return datetime.strptime(s[:19], fmt)
                except Exception:
                    pass
            try:
                return datetime.fromisoformat(s.replace("Z", ""))
            except Exception:
                return None

        def _refresh_tree():
            for iid in tree.get_children():
                tree.delete(iid)
            for idx, row in enumerate(rows):
                iid = f"sale_{idx}"
                row_map[iid] = row
                tree.insert(
                    "",
                    "end",
                    iid=iid,
                    values=(
                        "✓" if row.get("selected") else "",
                        row.get("date_label", ""),
                        row.get("printed_time", ""),
                        row.get("receipt_code", ""),
                        money(row.get("printed_amount", 0.0)),
                        row.get("payment_method", ""),
                    ),
                )
            try:
                status_var.set(f"Loaded {len(rows)} sale(s). Selected: {sum(1 for r in rows if r.get('selected'))}.")
            except Exception:
                pass

        def _load_current_into_editor(iid=None):
            iid = iid or current_iid.get("value")
            row = row_map.get(iid or "")
            if not row:
                selected_var.set("Selected sale: none")
                time_var.set("")
                amount_var.set("")
                return
            current_iid["value"] = iid
            selected_var.set(f"Selected sale: {row.get('receipt_code', '')} on {row.get('date_label', '')}")
            time_var.set(str(row.get("printed_time", "") or ""))
            amount_var.set(f"{float(row.get('printed_amount', 0.0)):.2f}")

        def _apply_edit_to_current():
            iid = current_iid.get("value")
            row = row_map.get(iid or "")
            if not row:
                return
            printed_time = str(time_var.get() or "").strip()
            if not printed_time:
                messagebox.showwarning("Printed Time", "Enter a printed time.", parent=win)
                return
            try:
                printed_amount = round(float(str(amount_var.get() or "0").strip()), 2)
            except Exception:
                messagebox.showwarning("Printed Amount", "Enter a valid number for the printed amount.", parent=win)
                return
            row["printed_time"] = printed_time
            row["printed_amount"] = printed_amount
            row["selected"] = True
            _refresh_tree()
            try:
                tree.selection_set(iid)
                tree.focus(iid)
            except Exception:
                pass

        def _toggle_selected(iid=None):
            iid = iid or (tree.selection()[0] if tree.selection() else None)
            row = row_map.get(iid or "")
            if not row:
                return
            row["selected"] = not bool(row.get("selected"))
            _refresh_tree()
            _load_current_into_editor(iid)
            try:
                tree.selection_set(iid)
                tree.focus(iid)
            except Exception:
                pass

        def _on_tree_select(_evt=None):
            sel = tree.selection()
            if sel:
                _load_current_into_editor(sel[0])

        tree.bind("<<TreeviewSelect>>", _on_tree_select)
        tree.bind("<Double-1>", lambda _e: _toggle_selected())

        def _load_sales():
            s0 = str(start_var.get() or "").strip()
            s1 = str(end_var.get() or "").strip()
            try:
                d0 = datetime.strptime(s0, "%Y-%m-%d").date()
                d1 = datetime.strptime(s1, "%Y-%m-%d").date()
            except Exception:
                messagebox.showwarning("Dates", "Use dates like 2026-03-16.", parent=win)
                return
            if d1 < d0:
                d0, d1 = d1, d0
            if (d1 - d0).days > 31:
                messagebox.showwarning("Date Range", "Keep the range to 31 days or less.", parent=win)
                return

            rows.clear()
            row_map.clear()
            cur = d0
            while cur <= d1:
                try:
                    day_sales = list_sales_for_day(cur.strftime("%Y-%m-%d"), limit=500) or []
                except Exception:
                    day_sales = []
                for sale in day_sales:
                    dt = _parse_dt(row_get(sale, "created_at", ""))
                    date_label = dt.strftime("%Y-%m-%d") if dt else str(row_get(sale, "created_at", "") or "")[:10]
                    time_label = fmt_time_ampm(dt) if dt else str(row_get(sale, "created_at", "") or "")[11:16]
                    receipt_code = str(row_get(sale, "receipt_code", "") or row_get(sale, "id", "")).strip()
                    try:
                        amount = float(row_get(sale, "total_amount", row_get(sale, "total_sales", 0.0)) or 0.0)
                    except Exception:
                        amount = 0.0
                    rows.append({
                        "sale_id": row_get(sale, "id", 0),
                        "receipt_code": receipt_code,
                        "date_label": date_label,
                        "created_at": str(row_get(sale, "created_at", "") or ""),
                        "original_time": time_label,
                        "printed_time": time_label or "",
                        "original_amount": amount,
                        "printed_amount": amount,
                        "payment_method": str(row_get(sale, "payment_method", "") or ""),
                        "selected": False,
                    })
                cur += timedelta(days=1)

            rows.sort(key=lambda r: (r.get("created_at", ""), str(r.get("receipt_code", ""))))
            _refresh_tree()
            if rows:
                first = tree.get_children()[0]
                tree.selection_set(first)
                tree.focus(first)
                _load_current_into_editor(first)
            else:
                selected_var.set("Selected sale: none")
                time_var.set("")
                amount_var.set("")
                status_var.set("No sales found in that range.")

        btn_row = tk.Frame(right_card.inner, bg=UI.CARD)
        btn_row.pack(anchor="w", pady=(14, 0))
        ttk.Button(btn_row, text="Apply Edit", command=_apply_edit_to_current).pack(side="left")
        ttk.Button(btn_row, text="Toggle Use", command=_toggle_selected).pack(side="left", padx=(8, 0))

        quick_row = tk.Frame(right_card.inner, bg=UI.CARD)
        quick_row.pack(anchor="w", pady=(10, 0))

        def _select_all():
            for row in rows:
                row["selected"] = True
            _refresh_tree()

        def _clear_selected():
            for row in rows:
                row["selected"] = False
            _refresh_tree()

        ttk.Button(quick_row, text="Select All", command=_select_all).pack(side="left")
        ttk.Button(quick_row, text="Clear All", command=_clear_selected).pack(side="left", padx=(8, 0))

        def _print_weekly_receipt():
            selected = [dict(r) for r in rows if r.get("selected")]
            if not selected:
                messagebox.showwarning("Weekly Receipt", "Select at least one sale first.", parent=win)
                return

            def _time_sort_value(raw_time, fallback_created=""):
                s = str(raw_time or "").strip()
                if s:
                    for fmt in ("%I:%M %p", "%I:%M%p", "%H:%M", "%H:%M:%S"):
                        try:
                            dt = datetime.strptime(s.upper(), fmt)
                            return dt.hour * 60 + dt.minute
                        except Exception:
                            pass
                dt = _parse_dt(fallback_created)
                if dt:
                    return dt.hour * 60 + dt.minute
                return 24 * 60

            for r in selected:
                if not str(r.get("printed_time", "")).strip():
                    messagebox.showwarning("Weekly Receipt", "Every selected sale needs a printed time.", parent=win)
                    return
                try:
                    r["printed_amount"] = round(float(r.get("printed_amount", 0.0) or 0.0), 2)
                except Exception:
                    messagebox.showwarning("Weekly Receipt", "One of the selected printed amounts is invalid.", parent=win)
                    return

            selected.sort(key=lambda r: (
                str(r.get("date_label", "") or ""),
                _time_sort_value(r.get("printed_time", ""), r.get("created_at", "")),
                str(r.get("created_at", "") or ""),
                str(r.get("receipt_code", "") or ""),
            ))

            day_counts = {}
            for r in selected:
                day_key = str(r.get("date_label", "") or "")
                day_counts[day_key] = int(day_counts.get(day_key, 0)) + 1
                r["print_receipt_no"] = day_counts[day_key]
                r["printed_receipt_code"] = str(day_counts[day_key])

                item_rows = []
                try:
                    _sale, sale_items = get_sale_receipt_data(int(r.get("sale_id") or 0))
                except Exception:
                    sale_items = []
                for it in list(sale_items or []):
                    try:
                        qty = int(float(row_get(it, "qty", 0) or 0))
                    except Exception:
                        qty = 0
                    name = str(row_get(it, "name", "") or "").strip()
                    if not name:
                        continue
                    item_rows.append({
                        "qty": max(1, qty),
                        "name": name,
                    })
                r["items"] = item_rows

            ok = False
            try:
                ok = bool(print_configured_weekly_receipt(get_store_name(), selected, title="Weekly Receipt"))
            except Exception:
                ok = False
            if ok:
                messagebox.showinfo("Weekly Receipt", "Weekly receipt sent to the configured printer.", parent=win)
                return
            try:
                pdf_path = str(create_weekly_selection_receipt_pdf(get_store_name(), selected, title="Weekly Receipt"))
                opened = open_pdf_in_chrome(pdf_path)
                if opened:
                    messagebox.showwarning("Weekly Receipt", "Could not send directly to printer, so the PDF was opened instead.", parent=win)
                else:
                    messagebox.showwarning("Weekly Receipt", "Could not print or open the weekly receipt PDF.", parent=win)
            except Exception as e:
                messagebox.showwarning("Weekly Receipt", f"Could not print the weekly receipt.\n{e}", parent=win)

        ttk.Button(right_card.inner, text="Print Weekly Receipt", command=_print_weekly_receipt).pack(anchor="w", fill="x", pady=(14, 0))

        tk.Label(right_card.inner, textvariable=status_var, bg=UI.CARD, fg=UI.MUTED, justify="left", wraplength=280).pack(anchor="w", pady=(14, 0))

        footer = tk.Frame(outer.inner, bg=UI.CARD)
        footer.pack(fill="x", pady=(12, 0))

        ttk.Button(top, text="Load Sales", command=_load_sales).pack(side="left", padx=(8, 0))
        ttk.Button(footer, text="Close", command=win.destroy).pack(side="right")

        _load_sales()

    def _open_seasonal_sale_manager(self):
        """Open the Seasonal Sale manager window."""
        return _seasonal_sale_manager_window(self)

# ---------------- SERVER AUTO-START (Host mode) ----------------


def _spin_wheel_manager_window(parent):
    win = tk.Toplevel(parent)
    win.title("Spin Wheel Prizes")
    win.geometry("900x560")
    win.minsize(760, 480)
    win.configure(bg=UI.CONTENT_BG)
    win.grab_set()

    outer = Card(win, padx=16, pady=16)
    outer.pack(fill="both", expand=True, padx=14, pady=14)
    HeaderBar(outer.inner, "Spin Wheel Prizes", "Odds are calculated from enabled weights. A zero weight keeps a prize saved but inactive.").pack(fill="x")

    table_frame = tk.Frame(outer.inner, bg=UI.CARD)
    table_frame.pack(fill="both", expand=True, pady=(14, 10))
    cols = ("label", "type", "reward", "weight", "chance", "enabled")
    tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=14)
    for key, title, width in (
        ("label", "Label", 240), ("type", "Type", 120), ("reward", "Reward", 240),
        ("weight", "Weight", 80), ("chance", "Odds", 80), ("enabled", "Enabled", 80),
    ):
        tree.heading(key, text=title)
        tree.column(key, width=width, anchor=("center" if key in ("weight", "chance", "enabled") else "w"))
    scroll = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=scroll.set)
    tree.pack(side="left", fill="both", expand=True)
    scroll.pack(side="right", fill="y")

    prizes = list(get_spin_wheel_prizes())
    summary_var = tk.StringVar(value="")

    def reward_text(prize):
        kind = prize.get("type")
        if kind == "discount":
            return f"{float(prize.get('value') or 0):g}% off order"
        if kind == "free_item":
            barcode = str(prize.get("barcode") or "")
            row = None
            for candidate in barcode_candidates(barcode):
                row = find_product_by_barcode(candidate)
                if row:
                    break
            if not row:
                return f"Missing item ({barcode})"
            stock = int(row_get(row, "stock_qty") or 0)
            suffix = "OUT OF STOCK" if stock <= 0 else f"{stock} in stock"
            return f"{row['name']} ({barcode}) - {suffix}"
        return "No prize"

    def refresh():
        tree.delete(*tree.get_children())
        total_weight = sum(
            float(prize.get("weight") or 0)
            for prize in prizes
            if bool(prize.get("enabled", True)) and float(prize.get("weight") or 0) > 0
        )
        for i, prize in enumerate(prizes):
            weight = float(prize.get("weight") or 0)
            chance = (100.0 * weight / total_weight) if bool(prize.get("enabled", True)) and total_weight > 0 else 0.0
            tree.insert("", tk.END, iid=str(i), values=(
                prize.get("label", ""),
                str(prize.get("type") or "").replace("_", " ").title(),
                reward_text(prize),
                f"{weight:g}",
                f"{chance:.2f}%",
                "Yes" if bool(prize.get("enabled", True)) else "No",
            ))
        active_count = sum(
            1 for prize in prizes
            if bool(prize.get("enabled", True)) and float(prize.get("weight") or 0) > 0
        )
        summary_var.set(f"{active_count} active prize{'s' if active_count != 1 else ''}. "
                        "Cashiers only see free-item prizes while the item is in stock.")

    def save():
        try:
            set_spin_wheel_prizes(prizes)
            prizes[:] = list(get_spin_wheel_prizes())
            refresh()
            return True
        except Exception as e:
            messagebox.showerror("Could not save", f"Spin wheel prizes could not be saved.\n{e}", parent=win)
            return False

    def edit_popup(index=None):
        current = prizes[index] if index is not None else {}
        popup = tk.Toplevel(win)
        popup.title("Edit Wheel Prize" if index is not None else "Add Wheel Prize")
        popup.geometry("520x390")
        popup.configure(bg=UI.CONTENT_BG)
        popup.grab_set()

        box = Card(popup, padx=18, pady=18)
        box.pack(fill="both", expand=True, padx=14, pady=14)
        tk.Label(box.inner, text="Wheel Prize", font=UI.FONT_LG, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w")
        form = tk.Frame(box.inner, bg=UI.CARD)
        form.pack(fill="x", pady=(12, 0))

        def field(label, value=""):
            row = tk.Frame(form, bg=UI.CARD)
            row.pack(fill="x", pady=5)
            tk.Label(row, text=label, width=18, anchor="w", bg=UI.CARD, fg=UI.TEXT).pack(side="left")
            entry = tk.Entry(row, font=UI.FONT_MD)
            entry.insert(0, str(value))
            entry.pack(side="left", fill="x", expand=True)
            return entry

        label_e = field("Wheel label", current.get("label", ""))
        type_row = tk.Frame(form, bg=UI.CARD)
        type_row.pack(fill="x", pady=5)
        tk.Label(type_row, text="Prize type", width=18, anchor="w", bg=UI.CARD, fg=UI.TEXT).pack(side="left")
        type_var = tk.StringVar(value=current.get("type", "discount"))
        type_box = ttk.Combobox(type_row, textvariable=type_var, values=("discount", "free_item", "none"),
                                state="readonly")
        type_box.pack(side="left", fill="x", expand=True)
        value_e = field("Discount percent", current.get("value", ""))
        barcode_e = field("Free item barcode", current.get("barcode", ""))
        weight_e = field("Odds weight", current.get("weight", "1"))
        enabled_var = tk.BooleanVar(value=bool(current.get("enabled", True)))
        tk.Checkbutton(form, text="Enabled", variable=enabled_var, bg=UI.CARD).pack(anchor="w", pady=(6, 0))

        def sync_type_fields(_event=None):
            kind = type_var.get().strip()
            value_e.configure(state=("normal" if kind == "discount" else "disabled"))
            barcode_e.configure(state=("normal" if kind == "free_item" else "disabled"))

        type_box.bind("<<ComboboxSelected>>", sync_type_fields)
        sync_type_fields()

        def apply():
            label = label_e.get().strip()
            kind = type_var.get().strip()
            try:
                weight = float(weight_e.get().strip() or "0")
            except Exception:
                messagebox.showerror("Invalid prize", "Odds weight must be a number.", parent=popup)
                return
            if not label or weight < 0:
                messagebox.showerror("Invalid prize", "Enter a label and a weight of zero or more.", parent=popup)
                return
            prize = {"label": label, "type": kind, "weight": weight, "enabled": bool(enabled_var.get())}
            if kind == "discount":
                try:
                    prize["value"] = float(value_e.get().strip() or "0")
                except Exception:
                    prize["value"] = 0.0
                if prize["value"] <= 0 or prize["value"] > 100:
                    messagebox.showerror("Invalid discount", "Discount must be between 0 and 100.", parent=popup)
                    return
            elif kind == "free_item":
                barcode = barcode_e.get().strip()
                product = None
                for candidate in barcode_candidates(barcode):
                    product = find_product_by_barcode(candidate)
                    if product:
                        break
                if not product:
                    messagebox.showerror("Item not found", "Enter the barcode of an existing product.", parent=popup)
                    return
                prize["barcode"] = str(row_get(product, "barcode") or barcode)
                if bool(enabled_var.get()) and weight > 0 and int(row_get(product, "stock_qty") or 0) <= 0:
                    if not messagebox.askyesno(
                        "Item out of stock",
                        "This item is currently out of stock. Save the prize anyway?\n\n"
                        "Cashiers will not land on it until the item is restocked.",
                        parent=popup,
                    ):
                        return
            if index is None:
                prizes.append(prize)
            else:
                prizes[index] = prize
            if save():
                popup.destroy()

        buttons = tk.Frame(box.inner, bg=UI.CARD)
        buttons.pack(fill="x", pady=(14, 0))
        GhostButton(buttons, "Cancel", popup.destroy).pack(side="right")
        PrimaryButton(buttons, "Save Prize", apply).pack(side="right", padx=(0, 10))

    def selected_index():
        sel = tree.selection()
        return int(sel[0]) if sel else None

    def edit_selected():
        index = selected_index()
        if index is None:
            messagebox.showinfo("Select prize", "Select a prize first.", parent=win)
            return
        edit_popup(index)

    def remove_selected():
        index = selected_index()
        if index is None:
            messagebox.showinfo("Select prize", "Select a prize first.", parent=win)
            return
        if messagebox.askyesno("Remove prize", "Remove the selected wheel prize?", parent=win):
            prizes.pop(index)
            save()

    tk.Label(outer.inner, textvariable=summary_var, font=UI.FONT_SM, bg=UI.CARD, fg=UI.MUTED,
             justify="left").pack(anchor="w", pady=(0, 8))
    actions = tk.Frame(outer.inner, bg=UI.CARD)
    actions.pack(fill="x")
    PrimaryButton(actions, "Add Prize", lambda: edit_popup()).pack(side="left")
    GhostButton(actions, "Edit Selected", edit_selected).pack(side="left", padx=(8, 0))
    DangerButton(actions, "Remove Selected", remove_selected).pack(side="left", padx=(8, 0))
    GhostButton(actions, "Close", win.destroy).pack(side="right")
    tree.bind("<Double-1>", lambda _e: edit_selected())
    refresh()


def _warehouse_paper_window(parent):
    """Open the Warehouse Paper manager window."""
    win = tk.Toplevel(parent)
    win.title("Warehouse Paper")
    win.geometry("980x560")
    win.minsize(920, 520)
    win.configure(bg=UI.CONTENT_BG)
    win.grab_set()

    outer = Card(win, padx=16, pady=16)
    outer.pack(fill="both", expand=True, padx=14, pady=14)

    HeaderBar(
        outer.inner,
        "Warehouse Paper",
        "Select products and print their warehouse location / section with prices."
    ).pack(fill="x")

    content = tk.Frame(outer.inner, bg=UI.CARD)
    content.pack(fill="both", expand=True, pady=(14, 0))
    content.grid_columnconfigure(0, weight=1)
    content.grid_columnconfigure(1, weight=1)
    content.grid_rowconfigure(1, weight=1)

    top = tk.Frame(content, bg=UI.CARD)
    top.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 10))
    status_var = tk.StringVar(value="")
    tk.Label(top, textvariable=status_var, bg=UI.CARD, fg=UI.PRIMARY, font=UI.FONT_MD).pack(side="left")

    left = tk.Frame(content, bg=UI.CARD)
    left.grid(row=1, column=0, sticky="nsew", padx=(0, 10))
    tk.Label(left, text="Paper Items", font=UI.FONT_MD, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w", pady=(0, 6))

    paper_cols = ("barcode", "location", "price")
    paper_tree = ttk.Treeview(left, columns=paper_cols, show="tree headings", selectmode="extended", height=14)
    paper_tree.heading("#0", text="Product")
    paper_tree.heading("barcode", text="Barcode")
    paper_tree.heading("location", text="LOCATION / SECTION")
    paper_tree.heading("price", text="Price")
    paper_tree.column("#0", width=260, anchor="w")
    paper_tree.column("barcode", width=135, anchor="w")
    paper_tree.column("location", width=145, anchor="w")
    paper_tree.column("price", width=80, anchor="e")
    paper_scroll = ttk.Scrollbar(left, orient="vertical", command=paper_tree.yview)
    paper_tree.configure(yscrollcommand=paper_scroll.set)
    paper_tree.pack(side="left", fill="both", expand=True)
    paper_scroll.pack(side="right", fill="y")

    right = tk.Frame(content, bg=UI.CARD)
    right.grid(row=1, column=1, sticky="nsew")
    tk.Label(right, text="Products", font=UI.FONT_MD, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w", pady=(0, 6))

    search_row = tk.Frame(right, bg=UI.CARD)
    search_row.pack(fill="x", pady=(0, 8))
    search_var = tk.StringVar(value="")
    search_entry = ttk.Entry(search_row, textvariable=search_var)
    search_entry.pack(side="left", fill="x", expand=True)

    res_cols = ("barcode", "name", "location", "price")
    res_tree = ttk.Treeview(right, columns=res_cols, show="headings", selectmode="extended", height=12)
    res_tree.heading("barcode", text="Barcode")
    res_tree.heading("name", text="Name")
    res_tree.heading("location", text="LOCATION / SECTION")
    res_tree.heading("price", text="Price")
    res_tree.column("barcode", width=130, anchor="w")
    res_tree.column("name", width=250, anchor="w")
    res_tree.column("location", width=140, anchor="w")
    res_tree.column("price", width=80, anchor="e")
    res_scroll = ttk.Scrollbar(right, orient="vertical", command=res_tree.yview)
    res_tree.configure(yscrollcommand=res_scroll.set)
    res_tree.pack(side="left", fill="both", expand=True)
    res_scroll.pack(side="right", fill="y")

    selected_items: dict[str, dict] = {}

    def _toggle_select(tree: ttk.Treeview, event):
        iid = tree.identify_row(event.y)
        if not iid:
            return "break"
        if iid in set(tree.selection()):
            tree.selection_remove(iid)
        else:
            tree.selection_add(iid)
        return "break"

    res_tree.bind("<Button-1>", lambda e: _toggle_select(res_tree, e))
    paper_tree.bind("<Button-1>", lambda e: _toggle_select(paper_tree, e))

    def refresh_paper():
        for iid in paper_tree.get_children():
            paper_tree.delete(iid)
        for bc, item in sorted(selected_items.items(), key=lambda pair: str(pair[1].get("name", "")).lower()):
            try:
                price = money(float(item.get("price") or 0.0))
            except Exception:
                price = "$0.00"
            paper_tree.insert(
                "",
                tk.END,
                iid=bc,
                text=str(item.get("name") or bc),
                values=(bc, str(item.get("location") or ""), price),
            )
        status_var.set(f"{len(selected_items)} item(s) selected")

    def refresh_results():
        for iid in res_tree.get_children():
            res_tree.delete(iid)
        q = (search_var.get() or "").strip()
        try:
            rows = list_products(q)
        except Exception:
            rows = []
        for r in rows:
            try:
                p_cat = str(row_get(r, "category") or "").strip()
                if p_cat.lower() == "quick":
                    continue
                bc = str(row_get(r, "barcode") or "").strip()
                if not bc:
                    continue
                nm = str(row_get(r, "name") or "").strip()
                loc = str(row_get(r, "location") or "").strip()
                price_value = float(row_get(r, "sell_price") or 0.0)
                price = money(price_value)
            except Exception:
                continue
            res_tree.insert("", tk.END, iid=bc, values=(bc, nm, loc, price))

    def add_selected():
        sels = list(res_tree.selection())
        if not sels:
            messagebox.showinfo("Select", "Select one or more products first.")
            return
        missing = []
        for iid in sels:
            vals = res_tree.item(iid, "values") or ()
            if len(vals) < 4:
                continue
            bc, name, location, price_text = vals
            bc = str(bc or "").strip()
            location = str(location or "").strip()
            if not location:
                missing.append(str(name or bc))
                continue
            try:
                price_value = float(str(price_text).replace("$", "").replace(",", ""))
            except Exception:
                price_value = 0.0
            selected_items[bc] = {
                "barcode": bc,
                "name": str(name or "").strip(),
                "location": location,
                "price": price_value,
            }
        refresh_paper()
        if missing:
            preview = "\n".join(missing[:8])
            extra = "" if len(missing) <= 8 else f"\n...and {len(missing) - 8} more"
            messagebox.showwarning("Location required", f"Add a location / section before printing:\n\n{preview}{extra}")

    def remove_selected():
        sels = list(paper_tree.selection())
        for iid in sels:
            selected_items.pop(str(iid), None)
        refresh_paper()

    def clear_paper():
        selected_items.clear()
        refresh_paper()

    def select_all_results():
        for iid in res_tree.get_children():
            res_tree.selection_add(iid)

    def print_paper():
        items = list(selected_items.values())
        if not items:
            messagebox.showinfo("Select", "Add products to the paper first.")
            return
        missing = [str(item.get("name") or item.get("barcode") or "") for item in items if not str(item.get("location") or "").strip()]
        if missing:
            preview = "\n".join(missing[:8])
            extra = "" if len(missing) <= 8 else f"\n...and {len(missing) - 8} more"
            messagebox.showwarning("Location required", f"Add a location / section before printing:\n\n{preview}{extra}")
            return
        ok = False
        try:
            ok = bool(print_configured_warehouse_paper(get_store_name(), items, title="Warehouse Locations"))
        except Exception:
            ok = False
        if ok:
            messagebox.showinfo("Printed", f"Warehouse paper was sent for {len(items)} item(s).")
        else:
            messagebox.showwarning("Not printed", "The warehouse paper was not sent.\n\nCheck Settings > Printer and use Test Print.")

    GhostButton(search_row, "Search", refresh_results).pack(side="left", padx=(8, 0))
    GhostButton(search_row, "Clear", lambda: (search_var.set(""), refresh_results())).pack(side="left", padx=(8, 0))
    search_entry.bind("<Return>", lambda _e: refresh_results())

    btn_row = tk.Frame(content, bg=UI.CARD)
    btn_row.grid(row=2, column=0, columnspan=2, sticky="ew", pady=(10, 0))
    btn_row.grid_columnconfigure(0, weight=1)
    btn_row.grid_columnconfigure(1, weight=1)

    left_btns = tk.Frame(btn_row, bg=UI.CARD)
    left_btns.grid(row=0, column=0, sticky="w")
    PrimaryButton(left_btns, "Print Warehouse Paper", print_paper).pack(side="left", padx=(0, 8))
    GhostButton(left_btns, "Remove Selected", remove_selected).pack(side="left", padx=(0, 8))
    DangerButton(left_btns, "Clear Paper", clear_paper).pack(side="left")

    right_btns = tk.Frame(btn_row, bg=UI.CARD)
    right_btns.grid(row=0, column=1, sticky="e")
    GhostButton(right_btns, "Select ALL results", select_all_results).pack(side="left", padx=(0, 8))
    PrimaryButton(right_btns, "Add Selected to Paper", add_selected).pack(side="left")

    refresh_results()
    refresh_paper()


def _seasonal_sale_manager_window(parent):
    """Open the Seasonal Sale manager window (bulk add/remove sale items)."""
    win = tk.Toplevel(parent)
    win.title("Seasonal Sale")
    win.geometry("980x560")
    win.minsize(920, 520)
    win.configure(bg=UI.CONTENT_BG)
    win.grab_set()

    outer = Card(win, padx=16, pady=16)
    outer.pack(fill="both", expand=True, padx=14, pady=14)

    HeaderBar(
        outer.inner,
        "Seasonal Sale",
        "Bulk add/remove items on sale. Discounts apply automatically in Cashier when Seasonal Sale is enabled."
    ).pack(fill="x")

    content = tk.Frame(outer.inner, bg=UI.CARD)
    content.pack(fill="both", expand=True, pady=(14, 0))
    content.grid_columnconfigure(0, weight=1)
    content.grid_columnconfigure(1, weight=1)
    content.grid_rowconfigure(1, weight=1)

    # Top controls (enabled toggle)
    top = tk.Frame(content, bg=UI.CARD)
    top.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 10))
    top.grid_columnconfigure(1, weight=1)

    try:
        enabled_var = tk.BooleanVar(value=bool(get_seasonal_sale_enabled()))
    except Exception:
        enabled_var = tk.BooleanVar(value=False)

    def toggle_enabled():
        try:
            set_seasonal_sale_enabled(bool(enabled_var.get()))
        except Exception:
            pass

    tk.Checkbutton(
        top,
        text="Seasonal Sale Enabled (admin)",
        variable=enabled_var,
        bg=UI.CARD,
        command=toggle_enabled
    ).pack(side="left")

    # Left: current sale items grouped by percent
    left = tk.Frame(content, bg=UI.CARD)
    left.grid(row=1, column=0, sticky="nsew", padx=(0, 10))
    tk.Label(left, text="Current Sale Items", font=UI.FONT_MD, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w", pady=(0, 6))

    sale_cols = ("barcode", "price")
    sale_tree = ttk.Treeview(left, columns=sale_cols, show="tree headings", selectmode="extended", height=14)
    sale_tree.heading("#0", text="Item (grouped by %)")
    sale_tree.heading("barcode", text="Barcode")
    sale_tree.heading("price", text="Price")
    sale_tree.column("#0", width=360)
    sale_tree.column("barcode", width=150, anchor="w")
    sale_tree.column("price", width=90, anchor="e")

    sale_scroll = ttk.Scrollbar(left, orient="vertical", command=sale_tree.yview)
    sale_tree.configure(yscrollcommand=sale_scroll.set)
    sale_tree.pack(side="left", fill="both", expand=True)
    sale_scroll.pack(side="right", fill="y")

    # Right: add to sale
    right = tk.Frame(content, bg=UI.CARD)
    right.grid(row=1, column=1, sticky="nsew")
    tk.Label(right, text="Add Products to Sale", font=UI.FONT_MD, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w", pady=(0, 6))

    pct_row = tk.Frame(right, bg=UI.CARD)
    pct_row.pack(fill="x", pady=(0, 8))
    tk.Label(pct_row, text="Discount %", bg=UI.CARD, fg=UI.TEXT).pack(side="left")
    pct_var = tk.StringVar(value="30")
    pct_box = ttk.Combobox(pct_row, textvariable=pct_var, values=["20", "30", "50"], width=8)
    pct_box.pack(side="left", padx=(8, 0))
    tk.Label(pct_row, text="(you can type custom)", bg=UI.CARD, fg=UI.MUTED).pack(side="left", padx=(8, 0))

    search_row = tk.Frame(right, bg=UI.CARD)
    search_row.pack(fill="x", pady=(0, 8))
    search_var = tk.StringVar(value="")
    search_entry = ttk.Entry(search_row, textvariable=search_var)
    search_entry.pack(side="left", fill="x", expand=True)

    res_cols = ("barcode", "name", "price")
    res_tree = ttk.Treeview(right, columns=res_cols, show="headings", selectmode="extended", height=12)
    res_tree.heading("barcode", text="Barcode")
    res_tree.heading("name", text="Name")
    res_tree.heading("price", text="Price")
    res_tree.column("barcode", width=140, anchor="w")
    res_tree.column("name", width=320, anchor="w")
    res_tree.column("price", width=90, anchor="e")

    res_scroll = ttk.Scrollbar(right, orient="vertical", command=res_tree.yview)
    res_tree.configure(yscrollcommand=res_scroll.set)
    res_tree.pack(side="left", fill="both", expand=True)
    res_scroll.pack(side="right", fill="y")

    # Make single-click behave like checkboxes (toggle without needing Ctrl)
    def _toggle_select(tree: ttk.Treeview, event):
        iid = tree.identify_row(event.y)
        if not iid:
            return "break"
        current = set(tree.selection())
        if iid in current:
            tree.selection_remove(iid)
        else:
            tree.selection_add(iid)
        # Don't let Tk clear other selections
        return "break"

    res_tree.bind("<Button-1>", lambda e: _toggle_select(res_tree, e))
    sale_tree.bind("<Button-1>", lambda e: _toggle_select(sale_tree, e))

    # Buttons row
    btn_row = tk.Frame(content, bg=UI.CARD)
    btn_row.grid(row=2, column=0, columnspan=2, sticky="ew", pady=(10, 0))
    btn_row.grid_columnconfigure(0, weight=1)
    btn_row.grid_columnconfigure(1, weight=1)

    # Helpers
    def parse_pct() -> float:
        try:
            s = str(pct_var.get() or "").strip().replace("%", "")
            return max(0.0, min(100.0, float(s or 0.0)))
        except Exception:
            return 0.0

    def refresh_current():
        for iid in sale_tree.get_children():
            sale_tree.delete(iid)

        try:
            sale_map = get_seasonal_sales_map()
        except Exception:
            sale_map = {}

        groups = {}
        for bc, pct in (sale_map or {}).items():
            try:
                p = float(pct)
            except Exception:
                p = 0.0
            p = round(max(0.0, min(100.0, p)), 2)
            if p <= 0:
                continue
            groups.setdefault(p, []).append(str(bc))

        for pct in sorted(groups.keys(), reverse=True):
            parent = sale_tree.insert("", tk.END, text=f"{pct:.0f}% OFF", open=True, values=("", ""))
            for bc in sorted(groups[pct]):
                disp_name = bc
                disp_price = ""
                try:
                    r = find_product_by_barcode(bc)
                    if r:
                        disp_name = str(r.get("name") or bc)
                        try:
                            disp_price = money(float(r.get("sell_price") or 0.0))
                        except Exception:
                            disp_price = ""
                except Exception:
                    pass
                sale_tree.insert(parent, tk.END, text=disp_name, values=(bc, disp_price))

    def refresh_results():
        for iid in res_tree.get_children():
            res_tree.delete(iid)

        q = (search_var.get() or "").strip()
        try:
            rows = list_products(q)
        except Exception:
            rows = []

        for r in rows:
            try:
                p_cat = str(row_get(r, "category") or "").strip()
                if p_cat.lower() == "quick":
                    continue
                bc = str(row_get(r, "barcode") or "")
                nm = str(row_get(r, "name") or "")
                pr = money(float(row_get(r, "sell_price") or 0.0))
            except Exception:
                continue
            res_tree.insert("", tk.END, values=(bc, nm, pr))

    def add_selected():
        pct = parse_pct()
        if pct <= 0:
            messagebox.showerror("Discount", "Enter a discount percent (e.g. 20, 30, 50).")
            return
        sels = list(res_tree.selection())
        if not sels:
            return
        for iid in sels:
            vals = res_tree.item(iid, "values") or ()
            bc = str(vals[0] if len(vals) > 0 else "").strip()
            if not bc:
                continue
            try:
                set_seasonal_sale_item(bc, pct)
            except Exception:
                pass
        refresh_current()

    def remove_selected_items():
        sels = list(sale_tree.selection())
        if not sels:
            return
        for iid in sels:
            parent = sale_tree.parent(iid)
            # if a group header is selected, skip (use remove_groups)
            if parent == "":
                continue
            vals = sale_tree.item(iid, "values") or ()
            bc = str(vals[0] if len(vals) > 0 else "").strip()
            if not bc:
                continue
            try:
                remove_seasonal_sale_item(bc)
            except Exception:
                pass
        refresh_current()

    def remove_groups():
        sels = list(sale_tree.selection())
        if not sels:
            return

        group_iids = []
        for iid in sels:
            gid = iid if sale_tree.parent(iid) == "" else sale_tree.parent(iid)
            if gid and gid not in group_iids:
                group_iids.append(gid)

        try:
            sale_map = get_seasonal_sales_map()
        except Exception:
            sale_map = {}

        for gid in group_iids:
            text = str(sale_tree.item(gid, "text") or "")
            pct = 0.0
            try:
                pct = float(text.split("%")[0].strip())
            except Exception:
                pct = 0.0
            if pct <= 0:
                continue
            for bc, p in list((sale_map or {}).items()):
                try:
                    if abs(float(p) - pct) < 0.01:
                        remove_seasonal_sale_item(bc)
                except Exception:
                    continue

        refresh_current()

    def clear_all():
        # No loud popups (only confirm)
        if not messagebox.askyesno("Clear all", "Remove ALL seasonal sale items?"):
            return
        try:
            sale_map = get_seasonal_sales_map()
        except Exception:
            sale_map = {}
        for bc in list((sale_map or {}).keys()):
            try:
                remove_seasonal_sale_item(str(bc))
            except Exception:
                pass
        refresh_current()

    def select_all_results():
        for iid in res_tree.get_children():
            res_tree.selection_add(iid)

    def clear_results_selection():
        res_tree.selection_remove(*res_tree.selection())

    def select_all_sale_items():
        # select all child rows (items), not group headers
        for gid in sale_tree.get_children():
            for iid in sale_tree.get_children(gid):
                sale_tree.selection_add(iid)

    # Bind search
    def _do_search():
        refresh_results()
    GhostButton(search_row, "Search", _do_search).pack(side="left", padx=(8, 0))
    search_entry.bind("<Return>", lambda e: _do_search())

    # Left-side actions
    left_btns = tk.Frame(btn_row, bg=UI.CARD)
    left_btns.grid(row=0, column=0, sticky="w")
    GhostButton(left_btns, "Select ALL sale items", select_all_sale_items).pack(side="left", padx=(0, 8))
    PrimaryButton(left_btns, "Remove Selected Items", remove_selected_items).pack(side="left", padx=(0, 8))
    GhostButton(left_btns, "Remove Selected % Group(s)", remove_groups).pack(side="left", padx=(0, 8))
    DangerButton(left_btns, "Clear ALL", clear_all).pack(side="left")

    # Right-side actions
    right_btns = tk.Frame(btn_row, bg=UI.CARD)
    right_btns.grid(row=0, column=1, sticky="e")
    GhostButton(right_btns, "Select ALL results", select_all_results).pack(side="left", padx=(0, 8))
    GhostButton(right_btns, "Clear selection", clear_results_selection).pack(side="left", padx=(0, 8))
    PrimaryButton(right_btns, "Add Selected to Sale", add_selected).pack(side="left")

    refresh_current()
    refresh_results()


def _bundle_offer_manager_window(parent):
    """Open the Bundle Offers manager window (example: 3 items for $25)."""
    win = tk.Toplevel(parent)
    win.title("Bundle Offers")
    win.geometry("980x560")
    win.minsize(920, 520)
    win.configure(bg=UI.CONTENT_BG)
    win.grab_set()

    outer = Card(win, padx=16, pady=16)
    outer.pack(fill="both", expand=True, padx=14, pady=14)

    HeaderBar(
        outer.inner,
        "Bundle Offers",
        "Create offers like 3 jeans for $25. The best automatic discount is applied in Cash Register."
    ).pack(fill="x")

    content = tk.Frame(outer.inner, bg=UI.CARD)
    content.pack(fill="both", expand=True, pady=(14, 0))
    content.grid_columnconfigure(0, weight=1)
    content.grid_columnconfigure(1, weight=1)
    content.grid_rowconfigure(1, weight=1)

    top = tk.Frame(content, bg=UI.CARD)
    top.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 10))

    try:
        enabled_var = tk.BooleanVar(value=bool(get_bundle_offers_enabled()))
    except Exception:
        enabled_var = tk.BooleanVar(value=True)

    def toggle_enabled():
        try:
            set_bundle_offers_enabled(bool(enabled_var.get()))
        except Exception:
            pass

    tk.Checkbutton(
        top,
        text="Bundle Offers Enabled",
        variable=enabled_var,
        bg=UI.CARD,
        command=toggle_enabled
    ).pack(side="left")

    left = tk.Frame(content, bg=UI.CARD)
    left.grid(row=1, column=0, sticky="nsew", padx=(0, 10))
    tk.Label(left, text="Current Bundle Offers", font=UI.FONT_MD, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w", pady=(0, 6))

    offer_cols = ("barcode", "qty", "offer", "unit")
    offer_tree = ttk.Treeview(left, columns=offer_cols, show="tree headings", selectmode="extended", height=14)
    offer_tree.heading("#0", text="Item")
    offer_tree.heading("barcode", text="Barcode")
    offer_tree.heading("qty", text="Qty")
    offer_tree.heading("offer", text="Offer Price")
    offer_tree.heading("unit", text="Unit Price")
    offer_tree.column("#0", width=300, anchor="w")
    offer_tree.column("barcode", width=140, anchor="w")
    offer_tree.column("qty", width=60, anchor="center")
    offer_tree.column("offer", width=90, anchor="e")
    offer_tree.column("unit", width=90, anchor="e")

    offer_scroll = ttk.Scrollbar(left, orient="vertical", command=offer_tree.yview)
    offer_tree.configure(yscrollcommand=offer_scroll.set)
    offer_tree.pack(side="left", fill="both", expand=True)
    offer_scroll.pack(side="right", fill="y")

    right = tk.Frame(content, bg=UI.CARD)
    right.grid(row=1, column=1, sticky="nsew")
    tk.Label(right, text="Add Products to Bundle Offer", font=UI.FONT_MD, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w", pady=(0, 6))

    form_row = tk.Frame(right, bg=UI.CARD)
    form_row.pack(fill="x", pady=(0, 8))
    tk.Label(form_row, text="Buy qty", bg=UI.CARD, fg=UI.TEXT).pack(side="left")
    qty_var = tk.StringVar(value="3")
    ttk.Entry(form_row, textvariable=qty_var, width=6).pack(side="left", padx=(8, 12))
    tk.Label(form_row, text="Offer price", bg=UI.CARD, fg=UI.TEXT).pack(side="left")
    price_var = tk.StringVar(value="25")
    ttk.Entry(form_row, textvariable=price_var, width=10).pack(side="left", padx=(8, 0))

    search_row = tk.Frame(right, bg=UI.CARD)
    search_row.pack(fill="x", pady=(0, 8))
    search_var = tk.StringVar(value="")
    search_entry = ttk.Entry(search_row, textvariable=search_var)
    search_entry.pack(side="left", fill="x", expand=True)

    res_cols = ("barcode", "name", "price")
    res_tree = ttk.Treeview(right, columns=res_cols, show="headings", selectmode="extended", height=12)
    res_tree.heading("barcode", text="Barcode")
    res_tree.heading("name", text="Name")
    res_tree.heading("price", text="Price")
    res_tree.column("barcode", width=140, anchor="w")
    res_tree.column("name", width=320, anchor="w")
    res_tree.column("price", width=90, anchor="e")

    res_scroll = ttk.Scrollbar(right, orient="vertical", command=res_tree.yview)
    res_tree.configure(yscrollcommand=res_scroll.set)
    res_tree.pack(side="left", fill="both", expand=True)
    res_scroll.pack(side="right", fill="y")

    btn_row = tk.Frame(content, bg=UI.CARD)
    btn_row.grid(row=2, column=0, columnspan=2, sticky="ew", pady=(10, 0))
    btn_row.grid_columnconfigure(0, weight=1)
    btn_row.grid_columnconfigure(1, weight=1)

    def _toggle_select(tree: ttk.Treeview, event):
        iid = tree.identify_row(event.y)
        if not iid:
            return "break"
        current = set(tree.selection())
        if iid in current:
            tree.selection_remove(iid)
        else:
            tree.selection_add(iid)
        return "break"

    res_tree.bind("<Button-1>", lambda e: _toggle_select(res_tree, e))
    offer_tree.bind("<Button-1>", lambda e: _toggle_select(offer_tree, e))

    def refresh_current():
        for iid in offer_tree.get_children():
            offer_tree.delete(iid)
        try:
            offer_map = get_bundle_offers_map()
        except Exception:
            offer_map = {}

        for bc in sorted((offer_map or {}).keys()):
            offer = offer_map.get(bc) or {}
            disp_name = bc
            disp_unit = ""
            try:
                r = find_product_by_barcode(bc)
                if r:
                    disp_name = str(r.get("name") or bc)
                    disp_unit = money(float(r.get("sell_price") or 0.0))
            except Exception:
                pass
            try:
                offer_tree.insert("", tk.END, text=disp_name, values=(
                    bc,
                    int(offer.get("qty") or 0),
                    money(float(offer.get("price") or 0.0)),
                    disp_unit,
                ))
            except Exception:
                pass

    def refresh_results():
        for iid in res_tree.get_children():
            res_tree.delete(iid)
        try:
            rows = list_products((search_var.get() or "").strip())
        except Exception:
            rows = []
        for r in rows:
            try:
                bc = str(row_get(r, "barcode") or "")
                nm = str(row_get(r, "name") or "")
                pr = money(float(row_get(r, "sell_price") or 0.0))
            except Exception:
                continue
            res_tree.insert("", tk.END, values=(bc, nm, pr))

    def parse_offer():
        try:
            qty = int(str(qty_var.get() or "").strip())
        except Exception:
            qty = 0
        try:
            offer_price = float(str(price_var.get() or "").strip().replace("$", ""))
        except Exception:
            offer_price = 0.0
        if qty < 2:
            messagebox.showerror("Bundle Offer", "Buy qty must be 2 or more.", parent=win)
            return None
        if offer_price <= 0:
            messagebox.showerror("Bundle Offer", "Offer price must be greater than 0.", parent=win)
            return None
        return qty, offer_price

    def add_selected():
        parsed = parse_offer()
        if not parsed:
            return
        qty, offer_price = parsed
        sels = list(res_tree.selection())
        if not sels:
            return
        for iid in sels:
            vals = res_tree.item(iid, "values") or ()
            bc = str(vals[0] if len(vals) > 0 else "").strip()
            if not bc:
                continue
            try:
                set_bundle_offer_item(bc, qty, offer_price)
            except Exception:
                pass
        refresh_current()

    def remove_selected():
        sels = list(offer_tree.selection())
        if not sels:
            return
        for iid in sels:
            vals = offer_tree.item(iid, "values") or ()
            bc = str(vals[0] if len(vals) > 0 else "").strip()
            if not bc:
                continue
            try:
                remove_bundle_offer_item(bc)
            except Exception:
                pass
        refresh_current()

    def clear_all():
        if not messagebox.askyesno("Clear all", "Remove ALL bundle offers?", parent=win):
            return
        try:
            clear_bundle_offers()
        except Exception:
            pass
        refresh_current()

    def select_all_results():
        for iid in res_tree.get_children():
            res_tree.selection_add(iid)

    def clear_results_selection():
        res_tree.selection_remove(*res_tree.selection())

    GhostButton(search_row, "Search", refresh_results).pack(side="left", padx=(8, 0))
    search_entry.bind("<Return>", lambda e: refresh_results())

    left_btns = tk.Frame(btn_row, bg=UI.CARD)
    left_btns.grid(row=0, column=0, sticky="w")
    PrimaryButton(left_btns, "Remove Selected Offers", remove_selected).pack(side="left", padx=(0, 8))
    DangerButton(left_btns, "Clear ALL", clear_all).pack(side="left")

    right_btns = tk.Frame(btn_row, bg=UI.CARD)
    right_btns.grid(row=0, column=1, sticky="e")
    GhostButton(right_btns, "Select ALL results", select_all_results).pack(side="left", padx=(0, 8))
    GhostButton(right_btns, "Clear selection", clear_results_selection).pack(side="left", padx=(0, 8))
    PrimaryButton(right_btns, "Add Selected Offer", add_selected).pack(side="left")

    refresh_current()
    refresh_results()

def _http_ok(url: str, timeout_s: float = 0.6) -> bool:
    """Return True if GET {url}/health responds with {"ok": True} (FastAPI server)."""
    try:
        import json
        import urllib.request
        health_url = url.rstrip("/") + "/health"
        req = urllib.request.Request(health_url, method="GET")
        with urllib.request.urlopen(req, timeout=timeout_s) as resp:
            body = resp.read().decode("utf-8", errors="ignore")
        try:
            data = json.loads(body)
        except Exception:
            return False
        return bool(data.get("ok")) is True
    except Exception:
        return False


def _find_server_py():
    """Find server.py next to this file (or inside PyInstaller bundle if present)."""
    try:
        import sys
        here = os.path.dirname(os.path.abspath(__file__))
        cand = os.path.join(here, "server.py")
        if os.path.exists(cand):
            return cand

        # PyInstaller onefile support (if you bundled server.py as a data file)
        meipass = getattr(sys, "_MEIPASS", None)
        if meipass:
            cand2 = os.path.join(meipass, "server.py")
            if os.path.exists(cand2):
                return cand2
    except Exception:
        return None


def _start_host_server_if_needed() -> None:
    """
    If POS is in Host mode, ensure the FastAPI server is running.

    - In normal Python runs, we can spawn server.py as a subprocess.
    - In a PyInstaller onefile EXE, spawning is unreliable because sys.executable is the EXE,
      so we run Uvicorn in-process in a background thread.
    """
    try:
        cfg = get_backend_config() or {}
        mode = (cfg.get("mode") or "").strip().lower()
        if mode != "host":
            return

        port = int(cfg.get("host_port") or 8000)
        base_url = f"http://127.0.0.1:{port}"

        # Already running
        if _http_ok(base_url):
            return

        # If frozen (EXE), run uvicorn in-process
        if getattr(sys, "frozen", False):
            try:
                import uvicorn  # type: ignore
                import server as server_mod  # server.py bundled as a module

                def _run():
                    try:
                        config = uvicorn.Config(
                            server_mod.app,
                            host="0.0.0.0",
                            port=port,
                            log_level="warning",
                            access_log=False,
                        )
                        uvicorn.Server(config).run()
                    except Exception:
                        pass

                t = threading.Thread(target=_run, daemon=True)
                t.start()

                # Wait briefly for it to come up
                for _ in range(25):
                    if _http_ok(base_url):
                        return
                    time.sleep(0.2)
                return
            except Exception:
                # If uvicorn isn't available for some reason, fall back to subprocess attempt below.
                pass

        # Normal python (non-frozen): spawn server.py
        server_py = _find_server_py()
        if not server_py:
            return

        python_exe = sys.executable

        # On Windows, DETACHED_PROCESS prevents a console window from popping up.
        creationflags = 0
        startupinfo = None
        if os.name == "nt":
            creationflags = getattr(subprocess, "DETACHED_PROCESS", 0) | getattr(subprocess, "CREATE_NEW_PROCESS_GROUP",
                                                                                 0)
            try:
                startupinfo = subprocess.STARTUPINFO()
                startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
            except Exception:
                startupinfo = None

        env = os.environ.copy()
        # Force the server to use the SAME data folder as the app (so pos.db is unified).
        env["MASKPOS_DATA_DIR"] = BASE_DIR

        subprocess.Popen(
            [python_exe, server_py, "--host", "0.0.0.0", "--port", str(port)],
            cwd=BASE_DIR,
            env=env,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            stdin=subprocess.DEVNULL,
            creationflags=creationflags,
            startupinfo=startupinfo,
        )

        # Wait briefly for it to come up
        for _ in range(25):
            if _http_ok(base_url):
                return
            time.sleep(0.2)

    except Exception:
        # Never block app startup because of server issues
        return


class DataHealthPage(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg=UI.CONTENT_BG)
        self._build()
        self.refresh()

    def _build(self):
        scroll = VScrollableFrame(self, bg=UI.CONTENT_BG)
        scroll.pack(fill="both", expand=True)
        wrap = tk.Frame(scroll.inner, bg=UI.CONTENT_BG)
        wrap.pack(fill="both", expand=True, padx=(10 if UI.COMPACT else 18), pady=(10 if UI.COMPACT else 18))

        header = Card(wrap, padx=18, pady=14)
        header.pack(fill="x")
        HeaderBar(header.inner, "Data Health & Tools", "Audit catalog anomalies, repair broken links, and import clean CSV data.").pack(fill="x")

        stats_frame = tk.Frame(wrap, bg=UI.CONTENT_BG)
        stats_frame.pack(fill="x", pady=12)
        for i in range(5):
            stats_frame.grid_columnconfigure(i, weight=1)

        self.stat_cards = {}
        
        def make_stat_card(parent, title, key, col, color):
            card = Card(parent, padx=12, pady=12)
            card.grid(row=0, column=col, sticky="nsew", padx=(0 if col == 0 else 6, 0 if col == 4 else 6))
            
            lbl_title = tk.Label(card.inner, text=title, font=UI.FONT_SM, bg=UI.CARD, fg=UI.MUTED)
            lbl_title.pack(anchor="w")
            
            val_var = tk.StringVar(value="0")
            lbl_val = tk.Label(card.inner, textvariable=val_var, font=("Segoe UI", 20, "bold"), bg=UI.CARD, fg=color)
            lbl_val.pack(anchor="w", pady=(4, 0))
            
            card.inner.bind("<Button-1>", lambda e: self.select_issue_type(key))
            lbl_title.bind("<Button-1>", lambda e: self.select_issue_type(key))
            lbl_val.bind("<Button-1>", lambda e: self.select_issue_type(key))
            
            self.stat_cards[key] = (val_var, card)
            
        make_stat_card(stats_frame, "Negative Stock", "negative_stock", 0, "#ef4444")
        make_stat_card(stats_frame, "Missing Category", "missing_category", 1, "#f59e0b")
        make_stat_card(stats_frame, "Missing Location", "missing_location", 2, "#3b82f6")
        make_stat_card(stats_frame, "Duplicate Names", "duplicate_names", 3, "#8b5cf6")
        make_stat_card(stats_frame, "Broken History Links", "broken_links", 4, "#ec4899")

        self.issue_section = Card(wrap, padx=14, pady=14)
        self.issue_section.pack(fill="both", expand=True, pady=(6, 0))

        sec_title_row = tk.Frame(self.issue_section.inner, bg=UI.CARD)
        sec_title_row.pack(fill="x", pady=(0, 10))
        
        self.issue_title_lbl = tk.Label(sec_title_row, text="Audit Details", font=UI.FONT_LG, bg=UI.CARD, fg=UI.TEXT)
        self.issue_title_lbl.pack(side="left")

        table_frame = tk.Frame(self.issue_section.inner, bg=UI.CARD)
        table_frame.pack(fill="both", expand=True)

        cols = ("id", "name", "barcode", "category", "brand", "location", "stock", "low_stock")
        self.issue_tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=8, selectmode="extended")
        
        for col, title, width in [
            ("id", "ID", 60), ("name", "Name", 200), ("barcode", "Barcode", 130),
            ("category", "Category", 120), ("brand", "Brand", 100), ("location", "Location", 100),
            ("stock", "Stock", 80), ("low_stock", "Low stock", 80)
        ]:
            self.issue_tree.heading(col, text=title)
            self.issue_tree.column(col, width=width, anchor="center" if col in ("stock", "low_stock") else "w")

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.issue_tree.yview)
        self.issue_tree.configure(yscrollcommand=vsb.set)
        self.issue_tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        self.controls_panel = tk.Frame(self.issue_section.inner, bg=UI.CARD)
        self.controls_panel.pack(fill="x", pady=(12, 0))

        self.bulk_edit_frame = tk.Frame(self.controls_panel, bg=UI.CARD)
        self.repair_frame = tk.Frame(self.controls_panel, bg=UI.CARD)

        tk.Label(self.bulk_edit_frame, text="Bulk Edit Selected:", font=UI.FONT_MD, bg=UI.CARD, fg=UI.TEXT).grid(row=0, column=0, sticky="w", pady=(0, 6), columnspan=4)
        
        tk.Label(self.bulk_edit_frame, text="Category", bg=UI.CARD, fg="#334155").grid(row=1, column=0, sticky="w", padx=(0, 4))
        self.bulk_cat_cb = ttk.Combobox(self.bulk_edit_frame, width=16, values=[], state="normal")
        self.bulk_cat_cb.grid(row=1, column=1, sticky="w", padx=(0, 12))
        
        tk.Label(self.bulk_edit_frame, text="Location", bg=UI.CARD, fg="#334155").grid(row=1, column=2, sticky="w", padx=(0, 4))
        self.bulk_loc_e = tk.Entry(self.bulk_edit_frame, width=14, bd=1, relief="solid")
        self.bulk_loc_e.grid(row=1, column=3, sticky="w", padx=(0, 12))
        
        tk.Label(self.bulk_edit_frame, text="Low stock level", bg=UI.CARD, fg="#334155").grid(row=1, column=4, sticky="w", padx=(0, 4))
        self.bulk_low_e = tk.Entry(self.bulk_edit_frame, width=8, bd=1, relief="solid")
        self.bulk_low_e.grid(row=1, column=5, sticky="w", padx=(0, 12))

        tk.Label(self.bulk_edit_frame, text="Brand", bg=UI.CARD, fg="#334155").grid(row=1, column=6, sticky="w", padx=(0, 4))
        self.bulk_brand_e = tk.Entry(self.bulk_edit_frame, width=14, bd=1, relief="solid")
        self.bulk_brand_e.grid(row=1, column=7, sticky="w", padx=(0, 12))

        PrimaryButton(self.bulk_edit_frame, "Apply Bulk Edit", self.apply_bulk_edit).grid(row=1, column=8, sticky="w", padx=(10, 0))

        tk.Label(self.repair_frame, text="Broken History Link Repair Tools:", font=UI.FONT_MD, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w", pady=(0, 6))
        GhostButton(self.repair_frame, "Map Selected to Valid Product", self.open_map_repair_dialog).pack(side="left")
        GhostButton(self.repair_frame, "Recreate and Repair Link", self.open_recreate_repair_dialog).pack(side="left", padx=(10, 0))

        import_card = Card(wrap, padx=14, pady=14)
        import_card.pack(fill="x", pady=(14, 0))

        tk.Label(import_card.inner, text="Excel/CSV Product Importer", font=UI.FONT_LG, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w")
        tk.Label(import_card.inner, text="Upload sheet containing: Barcode, Name, Sell price, Cost price, Stock qty, Category, Brand, Location, Low stock.",
                 font=UI.FONT_SM, bg=UI.CARD, fg=UI.MUTED).pack(anchor="w", pady=(2, 10))
        
        btn_import_row = tk.Frame(import_card.inner, bg=UI.CARD)
        btn_import_row.pack(fill="x")
        PrimaryButton(btn_import_row, "Select Import File (CSV)", self.select_import_file).pack(side="left")
        self.lbl_import_status = tk.Label(btn_import_row, text="No file loaded.", bg=UI.CARD, fg=UI.MUTED)
        self.lbl_import_status.pack(side="left", padx=12)

        self.preview_frame = tk.Frame(import_card.inner, bg=UI.CARD)
        
        self.lbl_preview_title = tk.Label(self.preview_frame, text="Import Sheet Preview & Validation (Red rows indicate errors):", font=UI.FONT_MD, bg=UI.CARD, fg=UI.TEXT)
        self.lbl_preview_title.pack(anchor="w", pady=(10, 6))

        preview_table_frame = tk.Frame(self.preview_frame, bg=UI.CARD)
        preview_table_frame.pack(fill="x")
        
        self.preview_tree = ttk.Treeview(preview_table_frame, columns=("row", "barcode", "name", "price", "cost", "stock", "category", "brand", "location", "low_stock", "error"), show="headings", height=5)
        for col, title, width in [
            ("row", "Row", 40), ("barcode", "Barcode", 110), ("name", "Name", 150),
            ("price", "Sell price", 80), ("cost", "Cost price", 80), ("stock", "Stock", 65),
            ("category", "Category", 90), ("brand", "Brand", 80), ("location", "Location", 80),
            ("low_stock", "Low stock", 70), ("error", "Validation error", 180)
        ]:
            self.preview_tree.heading(col, text=title)
            self.preview_tree.column(col, width=width)
        
        preview_vsb = ttk.Scrollbar(preview_table_frame, orient="vertical", command=self.preview_tree.yview)
        self.preview_tree.configure(yscrollcommand=preview_vsb.set)
        self.preview_tree.pack(side="left", fill="x", expand=True)
        preview_vsb.pack(side="right", fill="y")
        
        self.btn_commit_import = PrimaryButton(self.preview_frame, "Commit Import Data", self.commit_import)
        self.btn_commit_import.pack(anchor="w", pady=(10, 0))

        self.loaded_import_items = []
        self.active_issue_type = None

    def refresh(self):
        stats = get_data_health_stats()
        for key, (var, card) in self.stat_cards.items():
            count = stats.get(key, 0)
            var.set(str(count))
            if self.active_issue_type == key:
                card.configure(highlightbackground=UI.PRIMARY, highlightthickness=2)
            else:
                card.configure(highlightbackground=UI.CARD, highlightthickness=0)
                
        try:
            cats = get_distinct_categories() or []
            self.bulk_cat_cb['values'] = cats
        except Exception:
            pass

        if self.active_issue_type:
            self.select_issue_type(self.active_issue_type)
        else:
            self.select_issue_type("negative_stock")

    def select_issue_type(self, key):
        self.active_issue_type = key
        for k, (var, card) in self.stat_cards.items():
            if k == key:
                card.configure(highlightbackground=UI.PRIMARY, highlightthickness=2)
            else:
                card.configure(highlightbackground=UI.CARD, highlightthickness=0)
                
        self.issue_tree.delete(*self.issue_tree.get_children())
        
        if key == "broken_links":
            self.issue_tree.heading("id", text="Item ID")
            self.issue_tree.heading("name", text="Description")
            self.issue_tree.heading("barcode", text="Barcode")
            self.issue_tree.heading("category", text="Category")
            self.issue_tree.heading("brand", text="Brand")
            self.issue_tree.heading("location", text="Location")
            self.issue_tree.heading("stock", text="Occurrences")
            self.issue_tree.heading("low_stock", text="Linked ID")
            
            issues = list_health_issues("broken_links")
            for r in issues:
                self.issue_tree.insert("", tk.END, iid=str(row_get(r, "item_ids", "")), values=(
                    "MULTIPLE",
                    row_get(r, "name", ""),
                    row_get(r, "product_barcode", ""),
                    "", "", "",
                    row_get(r, "occurrence_count", 0),
                    "BROKEN"
                ))
            self.issue_title_lbl.config(text="Audit: Broken History Links (Sale items whose product records were deleted)")
            
            self.bulk_edit_frame.pack_forget()
            self.repair_frame.pack(fill="x")
        else:
            self.issue_tree.heading("id", text="ID")
            self.issue_tree.heading("name", text="Product name")
            self.issue_tree.heading("barcode", text="Barcode")
            self.issue_tree.heading("category", text="Category")
            self.issue_tree.heading("brand", text="Brand")
            self.issue_tree.heading("location", text="Location")
            self.issue_tree.heading("stock", text="Stock")
            self.issue_tree.heading("low_stock", text="Low stock")
            
            issues = list_health_issues(key)
            for r in issues:
                self.issue_tree.insert("", tk.END, iid=str(row_get(r, "id", "")), values=(
                    row_get(r, "id", ""),
                    row_get(r, "name", ""),
                    row_get(r, "barcode", ""),
                    row_get(r, "category", ""),
                    row_get(r, "brand", ""),
                    row_get(r, "location", ""),
                    row_get(r, "stock_qty", 0),
                    row_get(r, "low_stock_level", 0)
                ))
            
            titles = {
                "negative_stock": "Audit: Products with Negative Stock Levels",
                "missing_category": "Audit: Products missing Category classification",
                "missing_location": "Audit: Products missing Location location / section assignment",
                "duplicate_names": "Audit: Product records sharing duplicate Name identifiers"
            }
            self.issue_title_lbl.config(text=titles.get(key, "Audit Details"))
            
            self.repair_frame.pack_forget()
            self.bulk_edit_frame.pack(fill="x")

    def apply_bulk_edit(self):
        sel = self.issue_tree.selection()
        if not sel:
            messagebox.showwarning("Select", "Select one or more items in the table first.")
            return
            
        category = self.bulk_cat_cb.get().strip() or None
        location = self.bulk_loc_e.get().strip() or None
        brand = self.bulk_brand_e.get().strip() or None
        
        low_stock = None
        low_val = self.bulk_low_e.get().strip()
        if low_val:
            try:
                low_stock = int(low_val)
            except Exception:
                messagebox.showerror("Invalid", "Low stock level must be a number.")
                return
                
        if category is None and location is None and low_stock is None and brand is None:
            messagebox.showinfo("Incomplete", "Fill at least one field to edit in bulk.")
            return
            
        product_ids = [int(iid) for iid in sel]
        ok = bulk_update_products(product_ids, category=category, location=location, low_stock=low_stock, brand=brand)
        if ok:
            messagebox.showinfo("Success", f"Bulk updated {len(product_ids)} products successfully.")
            self.bulk_cat_cb.set("")
            self.bulk_loc_e.delete(0, tk.END)
            self.bulk_low_e.delete(0, tk.END)
            self.bulk_brand_e.delete(0, tk.END)
            self.refresh()
        else:
            messagebox.showerror("Error", "Failed to apply bulk edit.")

    def open_map_repair_dialog(self):
        sel = self.issue_tree.selection()
        if not sel:
            messagebox.showwarning("Select", "Select a broken sale item group to map.")
            return
            
        item_ids_str = sel[0]
        vals = self.issue_tree.item(item_ids_str, "values")
        broken_name = vals[1]
        
        win = tk.Toplevel(self)
        win.title("Repair Link: Map to Product")
        win.geometry("500x200")
        win.configure(bg=UI.CONTENT_BG)
        win.transient(self.winfo_toplevel())
        win.grab_set()
        
        card = Card(win, padx=14, pady=14)
        card.pack(fill="both", expand=True, padx=10, pady=10)
        
        tk.Label(card.inner, text=f"Mapping Sale Items: {broken_name}", font=UI.FONT_MD, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w")
        tk.Label(card.inner, text="Select target valid product catalog record:", font=UI.FONT_SM, bg=UI.CARD, fg=UI.MUTED).pack(anchor="w", pady=(2, 10))
        
        try:
            prods = list_products() or []
        except Exception:
            prods = []
        options = []
        prod_map = {}
        for p in prods:
            display = f"{row_get(p, 'name', '')} ({row_get(p, 'barcode', '')})"
            options.append(display)
            prod_map[display] = int(row_get(p, 'id', 0))
            
        cb_prod = ttk.Combobox(card.inner, values=options, width=45, state="readonly")
        cb_prod.pack(anchor="w", pady=10)
        if options:
            cb_prod.set(options[0])
            
        def do_repair():
            target_display = cb_prod.get()
            target_id = prod_map.get(target_display)
            if not target_id:
                messagebox.showerror("Invalid", "Pick a target product.")
                return
            ok = repair_broken_product_links(item_ids_str, target_id)
            if ok:
                messagebox.showinfo("Repaired", "Broken links successfully updated to target product.", parent=win)
                win.destroy()
                self.refresh()
            else:
                messagebox.showerror("Failed", "Failed to repair product links.", parent=win)
                
        btn_row = tk.Frame(card.inner, bg=UI.CARD)
        btn_row.pack(anchor="w", pady=(10, 0))
        PrimaryButton(btn_row, "Confirm & Repair", do_repair).pack(side="left")
        GhostButton(btn_row, "Cancel", win.destroy).pack(side="left", padx=(10, 0))

    def open_recreate_repair_dialog(self):
        sel = self.issue_tree.selection()
        if not sel:
            messagebox.showwarning("Select", "Select a broken sale item group to recreate.")
            return
            
        item_ids_str = sel[0]
        vals = self.issue_tree.item(item_ids_str, "values")
        broken_name = vals[1]
        broken_barcode = vals[2]
        
        win = tk.Toplevel(self)
        win.title("Repair Link: Recreate Product")
        win.geometry("540x360")
        win.configure(bg=UI.CONTENT_BG)
        win.transient(self.winfo_toplevel())
        win.grab_set()
        
        card = Card(win, padx=14, pady=14)
        card.pack(fill="both", expand=True, padx=10, pady=10)
        
        tk.Label(card.inner, text="Recreate Catalog Product Record", font=UI.FONT_MD, bg=UI.CARD, fg=UI.TEXT).pack(anchor="w")
        tk.Label(card.inner, text=f"Will recreate item and link all occurrences of '{broken_name}'", font=UI.FONT_SM, bg=UI.CARD, fg=UI.MUTED).pack(anchor="w", pady=(2, 10))
        
        fields = tk.Frame(card.inner, bg=UI.CARD)
        fields.pack(fill="x", pady=6)
        
        def field(parent, label, row, default_val=""):
            tk.Label(parent, text=label, bg=UI.CARD, fg="#334155", width=14, anchor="w").grid(row=row, column=0, sticky="w", pady=4)
            e = tk.Entry(parent, bd=1, relief="solid", width=32)
            e.insert(0, default_val)
            e.grid(row=row, column=1, sticky="w", pady=4)
            return e
            
        name_e = field(fields, "Product name", 0, broken_name)
        barcode_e = field(fields, "Barcode", 1, broken_barcode)
        price_e = field(fields, "Sell price", 2)
        cost_e = field(fields, "Cost price", 3)
        supplier_e = field(fields, "Supplier", 4)
        category_e = field(fields, "Category", 5)
        brand_e = field(fields, "Brand", 6)
        location_e = field(fields, "Location", 7)
        
        def do_recreate():
            try:
                name = name_e.get().strip()
                barcode = barcode_e.get().strip()
                price = float(price_e.get().strip())
                cost = float(cost_e.get().strip() or "0")
                supplier = supplier_e.get().strip()
                category = category_e.get().strip()
                brand = brand_e.get().strip()
                location = location_e.get().strip()
            except Exception:
                messagebox.showerror("Invalid", "Check fields. Sell price must be a number.", parent=win)
                return
                
            if not name or not barcode:
                messagebox.showerror("Invalid", "Name and Barcode are required.", parent=win)
                return
                
            ok = recreate_and_repair_product(
                name, barcode, price, cost, supplier, category, brand, location, item_ids_str
            )
            if ok:
                messagebox.showinfo("Success", "Catalog product recreated and historical transactions repaired.", parent=win)
                win.destroy()
                self.refresh()
            else:
                messagebox.showerror("Failed", "Failed to recreate and repair product link.", parent=win)
                
        btn_row = tk.Frame(card.inner, bg=UI.CARD)
        btn_row.pack(anchor="w", pady=(10, 0))
        PrimaryButton(btn_row, "Recreate & Confirm", do_recreate).pack(side="left")
        GhostButton(btn_row, "Cancel", win.destroy).pack(side="left", padx=(10, 0))

    def select_import_file(self):
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            title="Select CSV Import File",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if not path:
            return
            
        self.lbl_import_status.config(text=f"Loading: {Path(path).name}", fg=UI.PRIMARY)
        self.preview_tree.delete(*self.preview_tree.get_children())
        self.loaded_import_items = []
        
        try:
            import csv
            with open(path, "r", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                
                headers = [h.strip().lower() for h in (reader.fieldnames or [])]
                required = {"name", "sell_price"}
                present = set(headers)
                missing = required - present
                if missing:
                    messagebox.showerror("Invalid Format", f"CSV is missing required headers: {', '.join(missing)}")
                    self.lbl_import_status.config(text="Error reading file.", fg="#ef4444")
                    return
                    
                row_idx = 1
                for row in reader:
                    row_idx += 1
                    raw_name = row.get("name", "").strip()
                    raw_barcode = row.get("barcode", "").strip()
                    raw_price = row.get("sell_price", "").strip()
                    raw_cost = row.get("cost_price", "").strip() or "0"
                    raw_stock = row.get("stock_qty", "").strip() or "0"
                    raw_category = row.get("category", "").strip()
                    raw_brand = row.get("brand", "").strip()
                    raw_location = row.get("location", "").strip()
                    raw_low = row.get("low_stock_level", "").strip() or "0"
                    
                    errors = []
                    if not raw_name:
                        errors.append("Missing name")
                    try:
                        price = float(raw_price)
                        if price < 0: errors.append("Sell price negative")
                    except Exception:
                        errors.append("Invalid sell_price number")
                        price = 0.0
                    try:
                        cost = float(raw_cost)
                        if cost < 0: errors.append("Cost price negative")
                    except Exception:
                        errors.append("Invalid cost_price number")
                        cost = 0.0
                    try:
                        stock = int(float(raw_stock))
                    except Exception:
                        errors.append("Invalid stock_qty integer")
                        stock = 0
                    try:
                        low = int(float(raw_low))
                    except Exception:
                        errors.append("Invalid low_stock integer")
                        low = 0
                        
                    err_msg = "; ".join(errors)
                    iid = self.preview_tree.insert("", tk.END, values=(
                        row_idx, raw_barcode, raw_name, raw_price, raw_cost, raw_stock, raw_category, raw_brand, raw_location, raw_low, err_msg or "Valid"
                    ))
                    
                    if errors:
                        self.preview_tree.tag_configure("error_row", background="#fecaca", foreground="#991b1b")
                        self.preview_tree.item(iid, tags=("error_row",))
                        
                    self.loaded_import_items.append({
                        "name": raw_name,
                        "barcode": raw_barcode or None,
                        "sell_price": price,
                        "cost_price": cost,
                        "stock_qty": stock,
                        "category": raw_category,
                        "brand": raw_brand,
                        "location": raw_location,
                        "low_stock_level": low,
                        "has_errors": bool(errors)
                    })
                    
            self.lbl_import_status.config(text=f"Loaded: {Path(path).name} ({len(self.loaded_import_items)} products)", fg=UI.SUCCESS)
            self.preview_frame.pack(fill="x", pady=(10, 0))
        except Exception as e:
            messagebox.showerror("Error", f"Failed to parse CSV file:\n{e}")
            self.lbl_import_status.config(text="Error reading file.", fg="#ef4444")

    def commit_import(self):
        if not self.loaded_import_items:
            return
            
        has_errors = any(i["has_errors"] for i in self.loaded_import_items)
        if has_errors:
            if not messagebox.askyesno("Errors Found", "Some products have validation errors (highlighted in red). Do you want to skip those rows and import the valid ones?"):
                return
                
        imported_count = 0
        for item in self.loaded_import_items:
            if item["has_errors"]:
                continue
            try:
                add_product(
                    name=item["name"],
                    category=item["category"],
                    brand=item["brand"],
                    sell_price=item["sell_price"],
                    stock_qty=item["stock_qty"],
                    low_stock_level=item["low_stock_level"],
                    barcode=item["barcode"],
                    location=item["location"],
                    cost_price=item["cost_price"],
                    supplier=""
                )
                imported_count += 1
            except Exception:
                pass
                
        messagebox.showinfo("Import Complete", f"Successfully imported {imported_count} products.")
        self.preview_frame.pack_forget()
        self.lbl_import_status.config(text="No file loaded.", fg=UI.MUTED)
        self.loaded_import_items = []
        self.refresh()


if __name__ == "__main__":
    _acquire_single_instance()
    # _start_host_server_if_needed()  # server is started by backend_init() when mode == 'host'
    backend_init(APP_TITLE, interactive=True)

    app = MaskPOS()
    try:
        app.mainloop()
    finally:
        _release_single_instance()
