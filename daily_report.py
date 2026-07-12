from __future__ import annotations

import os
import sqlite3
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path


DEFAULT_DAILY_REPORT_RECIPIENTS = [
    "dani123khoueiry@gmail.com",
    "assaadmask@gmail.com",
]


def report_date_bounds(year: int, month: int, day_str: str):
    if str(day_str).strip().lower() == "all":
        start = datetime(int(year), int(month), 1, 0, 0, 0)
        if int(month) == 12:
            end = datetime(int(year) + 1, 1, 1, 0, 0, 0)
        else:
            end = datetime(int(year), int(month) + 1, 1, 0, 0, 0)
        stamp = f"{int(year):04d}_{int(month):02d}"
    else:
        day = int(str(day_str).strip())
        start = datetime(int(year), int(month), day, 0, 0, 0)
        end = start + timedelta(days=1)
        stamp = f"{int(year):04d}_{int(month):02d}_{day:02d}"
    return start, end, stamp


def _table_cols(cur, table: str) -> set[str]:
    try:
        cur.execute(f"PRAGMA table_info({table})")
        return {str(r[1]) for r in cur.fetchall()}
    except Exception:
        return set()


def _expr(cols: set[str], alias: str, col: str, default_sql: str):
    return f"{alias}.{col}" if col in cols else default_sql


def _first_expr(cols: set[str], alias: str, names, default_sql: str):
    for name in names:
        if name in cols:
            return f"{alias}.{name}"
    return default_sql


def _money_total_expr(sales_cols: set[str]) -> str:
    amount = _expr(sales_cols, "s", "total_amount", "0")
    credit = _expr(sales_cols, "s", "store_credit_used", "0")
    if "total_sales" in sales_cols:
        return f"COALESCE(NULLIF(s.total_sales, 0), COALESCE({amount}, 0) + COALESCE({credit}, 0), 0)"
    return f"COALESCE({amount}, 0) + COALESCE({credit}, 0)"


def _cash_paid_expr(sales_cols: set[str]) -> str:
    pm = _expr(sales_cols, "s", "payment_method", "''")
    if "cash_paid" in sales_cols:
        cash = "COALESCE(s.cash_paid, 0)"
    else:
        amount = _expr(sales_cols, "s", "total_amount", "0")
        cash = f"COALESCE({amount}, 0)"
    non_cash = "'EXCHANGE','STORE_CREDIT','CARD','DEBIT','CREDIT_CARD','WHISH'"
    return f"CASE WHEN UPPER(COALESCE({pm}, '')) IN ({non_cash}) THEN 0 ELSE {cash} END"


def _sale_new_money(sale: dict) -> float:
    pm = str(sale.get("payment_method") or "").strip().upper()
    if pm in ("EXCHANGE", "STORE_CREDIT"):
        return 0.0
    try:
        amount = float(sale.get("amount_due", sale.get("total_amount", 0.0)) or 0.0)
    except Exception:
        amount = 0.0
    try:
        merch = float(sale.get("merch_total") or 0.0)
    except Exception:
        merch = 0.0
    try:
        credit = float(sale.get("store_credit_used") or 0.0)
    except Exception:
        credit = 0.0
    if merch > 0 and credit > 0 and abs(amount - merch) < 0.005:
        return max(0.0, merch - credit)
    return max(0.0, amount)


def _day_of(value) -> str:
    text = str(value or "")
    return text[:10] if len(text) >= 10 else text


def _time_of(value) -> str:
    text = str(value or "")
    if len(text) >= 19:
        return text[11:19]
    return text


def _as_float(value, default: float = 0.0) -> float:
    try:
        return float(value or 0.0)
    except Exception:
        return default


def _as_int(value, default: int = 0) -> int:
    try:
        return int(value or 0)
    except Exception:
        return default


def _fetch_rows(db_path: str, start: datetime, end: datetime) -> dict:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        cur = conn.cursor()
        start_sql = start.strftime("%Y-%m-%d %H:%M:%S")
        end_sql = end.strftime("%Y-%m-%d %H:%M:%S")

        sales_cols = _table_cols(cur, "sales")
        item_cols = _table_cols(cur, "sale_items")
        product_cols = _table_cols(cur, "products")
        shift_cols = _table_cols(cur, "cash_shifts")
        movement_cols = _table_cols(cur, "cash_movements")
        return_cols = _table_cols(cur, "returns")
        return_item_cols = _table_cols(cur, "return_items")
        sql_empty = "''"
        sql_out = "'OUT'"

        created_expr = _expr(sales_cols, "s", "created_at", sql_empty)
        sales_void_filter = "AND COALESCE(s.is_voided, 0) = 0" if "is_voided" in sales_cols else ""
        sales_select = [
            "s.id AS id",
            f"{created_expr} AS created_at",
            f"{_expr(sales_cols, 's', 'receipt_code', sql_empty)} AS receipt_code",
            f"{_expr(sales_cols, 's', 'payment_method', sql_empty)} AS payment_method",
            f"{_expr(sales_cols, 's', 'customer_name', sql_empty)} AS customer_name",
            f"{_expr(sales_cols, 's', 'shift_id', 'NULL')} AS shift_id",
            f"{_money_total_expr(sales_cols)} AS merch_total",
            f"{_expr(sales_cols, 's', 'total_amount', '0')} AS amount_due",
            f"{_cash_paid_expr(sales_cols)} AS cash_paid",
            f"{_expr(sales_cols, 's', 'store_credit_used', '0')} AS store_credit_used",
            f"{_expr(sales_cols, 's', 'discount_total', _expr(sales_cols, 's', 'discount', '0'))} AS discount_total",
            f"{_expr(sales_cols, 's', 'subtotal', '0')} AS subtotal",
            f"{_expr(sales_cols, 's', 'notes', sql_empty)} AS notes",
        ]
        cur.execute(
            f"""
            SELECT {", ".join(sales_select)}
            FROM sales s
            WHERE datetime({created_expr}) >= datetime(?)
              AND datetime({created_expr}) < datetime(?)
              {sales_void_filter}
            ORDER BY datetime({created_expr}) ASC, s.id ASC
            """,
            (start_sql, end_sql),
        )
        sales_rows = [dict(r) for r in cur.fetchall()]

        item_created_expr = created_expr
        item_select = [
            "si.sale_id AS sale_id",
            f"{_expr(item_cols, 'si', 'product_id', 'NULL')} AS product_id",
            f"{_expr(item_cols, 'si', 'name', sql_empty)} AS name",
            f"{_first_expr(item_cols, 'si', ['price', 'unit_price', 'unit_price_used'], '0')} AS price",
            f"{_expr(item_cols, 'si', 'qty', '0')} AS qty",
            f"{_expr(item_cols, 'si', 'line_total', _expr(item_cols, 'si', 'subtotal', '0'))} AS line_total",
            f"{_expr(item_cols, 'si', 'gross_line_total', _expr(item_cols, 'si', 'line_total', '0'))} AS gross_line_total",
            f"{_expr(item_cols, 'si', 'discount_allocated', '0')} AS discount_allocated",
            f"{_expr(product_cols, 'p', 'barcode', sql_empty)} AS barcode",
            f"{_expr(product_cols, 'p', 'category', sql_empty)} AS category",
            f"{_expr(product_cols, 'p', 'brand', sql_empty)} AS brand",
            f"{_expr(product_cols, 'p', 'sell_price', '0')} AS current_price",
            f"{_expr(product_cols, 'p', 'stock_qty', '0')} AS current_stock",
        ]
        cur.execute(
            f"""
            SELECT {", ".join(item_select)}
            FROM sale_items si
            JOIN sales s ON s.id = si.sale_id
            LEFT JOIN products p ON p.id = si.product_id
            WHERE datetime({item_created_expr}) >= datetime(?)
              AND datetime({item_created_expr}) < datetime(?)
              {sales_void_filter}
            ORDER BY datetime({item_created_expr}) ASC, si.sale_id ASC, si.id ASC
            """,
            (start_sql, end_sql),
        )
        item_rows = [dict(r) for r in cur.fetchall()]

        shifts = []
        if shift_cols:
            opened_expr = _expr(shift_cols, "cs", "opened_at", sql_empty)
            closed_expr = _expr(shift_cols, "cs", "closed_at", "NULL")
            shift_select = [
                "cs.id AS id",
                f"{_expr(shift_cols, 'cs', 'shift_code', sql_empty)} AS shift_code",
                f"{opened_expr} AS opened_at",
                f"{closed_expr} AS closed_at",
                f"{_expr(shift_cols, 'cs', 'opening_cash', '0')} AS opening_cash",
                f"{_expr(shift_cols, 'cs', 'closing_cash', 'NULL')} AS closing_cash",
                f"{_expr(shift_cols, 'cs', 'opening_usd', _expr(shift_cols, 'cs', 'opening_cash', '0'))} AS opening_usd",
                f"{_expr(shift_cols, 'cs', 'opening_lbp', '0')} AS opening_lbp",
                f"{_expr(shift_cols, 'cs', 'closing_usd', 'NULL')} AS closing_usd",
                f"{_expr(shift_cols, 'cs', 'closing_lbp', 'NULL')} AS closing_lbp",
                f"{_expr(shift_cols, 'cs', 'lbp_per_usd', '0')} AS lbp_per_usd",
                f"{_expr(shift_cols, 'cs', 'notes', sql_empty)} AS notes",
                "COALESCE(e.name, '') AS employee_name",
            ]
            cur.execute(
                f"""
                SELECT {", ".join(shift_select)}
                FROM cash_shifts cs
                LEFT JOIN employees e ON e.id = cs.employee_id
                WHERE (
                    (datetime({opened_expr}) >= datetime(?) AND datetime({opened_expr}) < datetime(?))
                    OR ({closed_expr} IS NOT NULL AND datetime({closed_expr}) >= datetime(?) AND datetime({closed_expr}) < datetime(?))
                    OR (datetime({opened_expr}) < datetime(?) AND ({closed_expr} IS NULL OR datetime({closed_expr}) >= datetime(?)))
                )
                ORDER BY datetime({opened_expr}) ASC, cs.id ASC
                """,
                (start_sql, end_sql, start_sql, end_sql, end_sql, start_sql),
            )
            shifts = [dict(r) for r in cur.fetchall()]

        movements = []
        if movement_cols:
            mov_created_expr = _expr(movement_cols, "cm", "created_at", sql_empty)
            mov_select = [
                "cm.id AS id",
                f"{mov_created_expr} AS created_at",
                f"{_expr(movement_cols, 'cm', 'shift_id', 'NULL')} AS shift_id",
                f"{_expr(movement_cols, 'cm', 'movement_type', sql_out)} AS movement_type",
                f"{_expr(movement_cols, 'cm', 'amount_usd', '0')} AS amount_usd",
                f"{_expr(movement_cols, 'cm', 'amount_lbp', '0')} AS amount_lbp",
                f"{_expr(movement_cols, 'cm', 'lbp_per_usd', '0')} AS lbp_per_usd",
                f"{_expr(movement_cols, 'cm', 'amount_value', '0')} AS amount_value",
                f"{_expr(movement_cols, 'cm', 'reason', sql_empty)} AS reason",
                f"{_expr(movement_cols, 'cm', 'employee_name', sql_empty)} AS employee_name",
                f"{_expr(movement_cols, 'cm', 'notes', sql_empty)} AS notes",
            ]
            cur.execute(
                f"""
                SELECT {", ".join(mov_select)}
                FROM cash_movements cm
                WHERE datetime({mov_created_expr}) >= datetime(?)
                  AND datetime({mov_created_expr}) < datetime(?)
                ORDER BY datetime({mov_created_expr}) ASC, cm.id ASC
                """,
                (start_sql, end_sql),
            )
            movements = [dict(r) for r in cur.fetchall()]

        returns = []
        if return_cols:
            ret_created_expr = _expr(return_cols, "r", "created_at", sql_empty)
            ret_select = [
                "r.id AS id",
                f"{_expr(return_cols, 'r', 'original_sale_id', 'NULL')} AS original_sale_id",
                f"{ret_created_expr} AS created_at",
                f"{_expr(return_cols, 'r', 'total_return_amount', '0')} AS total_return_amount",
                f"{_expr(return_cols, 'r', 'cash_refund', '0')} AS cash_refund",
                f"{_expr(return_cols, 'r', 'credit_refund', '0')} AS credit_refund",
                f"{_expr(return_cols, 'r', 'shift_id', 'NULL')} AS shift_id",
                f"{_expr(return_cols, 'r', 'notes', sql_empty)} AS notes",
            ]
            void_filter = ""
            if "is_voided" in return_cols:
                void_filter = "AND COALESCE(r.is_voided, 0) = 0"
            cur.execute(
                f"""
                SELECT {", ".join(ret_select)}
                FROM returns r
                WHERE datetime({ret_created_expr}) >= datetime(?)
                  AND datetime({ret_created_expr}) < datetime(?)
                  {void_filter}
                ORDER BY datetime({ret_created_expr}) ASC, r.id ASC
                """,
                (start_sql, end_sql),
            )
            returns = [dict(r) for r in cur.fetchall()]

        return_items = []
        if return_cols and return_item_cols:
            ret_created_expr = _expr(return_cols, "r", "created_at", sql_empty)
            ret_item_select = [
                "ri.return_id AS return_id",
                f"{_expr(return_item_cols, 'ri', 'sale_item_id', 'NULL')} AS sale_item_id",
                f"{_expr(return_item_cols, 'ri', 'product_id', 'NULL')} AS product_id",
                f"{_expr(return_item_cols, 'ri', 'name', sql_empty)} AS name",
                f"{_expr(return_item_cols, 'ri', 'price', '0')} AS price",
                f"{_expr(return_item_cols, 'ri', 'qty', '0')} AS qty",
                f"{_expr(return_item_cols, 'ri', 'line_total', '0')} AS line_total",
                f"{ret_created_expr} AS created_at",
                f"{_expr(return_cols, 'r', 'original_sale_id', 'NULL')} AS original_sale_id",
                f"{_expr(return_cols, 'r', 'cash_refund', '0')} AS cash_refund",
                f"{_expr(return_cols, 'r', 'credit_refund', '0')} AS credit_refund",
                f"{_expr(return_cols, 'r', 'shift_id', 'NULL')} AS shift_id",
                f"{_expr(return_cols, 'r', 'notes', sql_empty)} AS notes",
                f"{_expr(product_cols, 'p', 'barcode', sql_empty)} AS barcode",
                f"{_expr(product_cols, 'p', 'category', sql_empty)} AS category",
                f"{_expr(product_cols, 'p', 'brand', sql_empty)} AS brand",
                f"{_expr(product_cols, 'p', 'stock_qty', '0')} AS current_stock",
            ]
            void_filter = ""
            if "is_voided" in return_cols:
                void_filter = "AND COALESCE(r.is_voided, 0) = 0"
            cur.execute(
                f"""
                SELECT {", ".join(ret_item_select)}
                FROM return_items ri
                JOIN returns r ON r.id = ri.return_id
                LEFT JOIN products p ON p.id = ri.product_id
                WHERE datetime({ret_created_expr}) >= datetime(?)
                  AND datetime({ret_created_expr}) < datetime(?)
                  {void_filter}
                ORDER BY datetime({ret_created_expr}) ASC, ri.return_id ASC, ri.id ASC
                """,
                (start_sql, end_sql),
            )
            return_items = [dict(r) for r in cur.fetchall()]

        return {
            "sales": sales_rows,
            "items": item_rows,
            "shifts": shifts,
            "movements": movements,
            "returns": returns,
            "return_items": return_items,
        }
    finally:
        conn.close()


def _last_shift_before_start(db_path: str, start: datetime) -> dict:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        cur = conn.cursor()
        shift_cols = _table_cols(cur, "cash_shifts")
        if not shift_cols:
            return {}
        start_sql = start.strftime("%Y-%m-%d %H:%M:%S")
        opened_expr = _expr(shift_cols, "cs", "opened_at", "''")
        closed_expr = _expr(shift_cols, "cs", "closed_at", "NULL")
        cur.execute(
            f"""
            SELECT cs.*, COALESCE(e.name, '') AS employee_name
            FROM cash_shifts cs
            LEFT JOIN employees e ON e.id = cs.employee_id
            WHERE {closed_expr} IS NOT NULL
              AND datetime({closed_expr}) < datetime(?)
            ORDER BY datetime({closed_expr}) DESC, cs.id DESC
            LIMIT 1
            """,
            (start_sql,),
        )
        row = cur.fetchone()
        return dict(row) if row else {}
    finally:
        conn.close()


def _next_shift_after(db_path: str, closed_at: str) -> dict:
    closed_at = str(closed_at or "").strip()
    if not closed_at:
        return {}
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        cur = conn.cursor()
        shift_cols = _table_cols(cur, "cash_shifts")
        if not shift_cols:
            return {}
        opened_expr = _expr(shift_cols, "cs", "opened_at", "''")
        cur.execute(
            f"""
            SELECT cs.*, COALESCE(e.name, '') AS employee_name
            FROM cash_shifts cs
            LEFT JOIN employees e ON e.id = cs.employee_id
            WHERE datetime({opened_expr}) > datetime(?)
            ORDER BY datetime({opened_expr}) ASC, cs.id ASC
            LIMIT 1
            """,
            (closed_at,),
        )
        row = cur.fetchone()
        return dict(row) if row else {}
    finally:
        conn.close()


def build_cash_drawer_pdf(db_path: str, reports_folder: str, year: int, month: int, day_str: str) -> dict:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    start, end, stamp = report_date_bounds(year, month, day_str)
    if not os.path.exists(db_path):
        raise FileNotFoundError(f"Could not find database: {db_path}")

    rows = _fetch_rows(db_path, start, end)
    sales = rows.get("sales", [])
    shifts = rows.get("shifts", [])
    movements = rows.get("movements", [])
    returns = rows.get("returns", [])

    sales_by_shift = defaultdict(list)
    sales_unassigned = []
    for sale in sales:
        sid = sale.get("shift_id")
        if sid in (None, ""):
            sales_unassigned.append(sale)
        else:
            sales_by_shift[int(sid)].append(sale)

    movements_by_shift = defaultdict(list)
    for movement in movements:
        sid = movement.get("shift_id")
        if sid not in (None, ""):
            movements_by_shift[int(sid)].append(movement)

    returns_by_shift = defaultdict(list)
    for ret in returns:
        sid = ret.get("shift_id")
        if sid not in (None, ""):
            returns_by_shift[int(sid)].append(ret)

    def money_fmt(value) -> str:
        try:
            return f"${float(value or 0.0):,.2f}"
        except Exception:
            return "$0.00"

    def lbp_fmt(value) -> str:
        try:
            return f"LBP {float(value or 0.0):,.0f}"
        except Exception:
            return "LBP 0"

    def shift_label(shift: dict) -> str:
        return str(shift.get("shift_code") or shift.get("id") or "").strip()

    def close_taken_for_shift(shift_id: int) -> float:
        total = 0.0
        for movement in movements_by_shift.get(int(shift_id), []):
            reason = str(movement.get("reason") or "").strip().lower()
            note = str(movement.get("notes") or "").strip().lower()
            if str(movement.get("movement_type") or "").upper() == "OUT" and (
                "end of day close" in reason or "cash taken at close" in note
            ):
                total += _as_float(movement.get("amount_value"))
        return total

    variance_tolerance = 4.0

    total_cash_sales = sum(_as_float(s.get("cash_paid")) for s in sales)
    total_revenue = sum(_sale_new_money(s) for s in sales) - sum(_as_float(r.get("cash_refund")) for r in returns)
    total_cash_in = sum(_as_float(m.get("amount_value")) for m in movements if str(m.get("movement_type") or "").upper() == "IN")
    total_cash_out = sum(_as_float(m.get("amount_value")) for m in movements if str(m.get("movement_type") or "").upper() == "OUT")
    total_cash_refunds = sum(_as_float(r.get("cash_refund")) for r in returns)
    total_close_taken = 0.0
    latest_closed_shift = None
    latest_open_shift = None
    total_variance = 0.0
    closed_shift_count = 0
    for shift in shifts:
        shift_id = int(shift.get("id") or 0)
        shift_sales = sales_by_shift.get(shift_id, [])
        shift_movements = movements_by_shift.get(shift_id, [])
        shift_returns = returns_by_shift.get(shift_id, [])
        shift_cash_sales = sum(_as_float(s.get("cash_paid")) for s in shift_sales)
        shift_cash_in = sum(_as_float(m.get("amount_value")) for m in shift_movements if str(m.get("movement_type") or "").upper() == "IN")
        shift_cash_out = sum(_as_float(m.get("amount_value")) for m in shift_movements if str(m.get("movement_type") or "").upper() == "OUT")
        shift_cash_refunds = sum(_as_float(r.get("cash_refund")) for r in shift_returns)
        shift_expected = _as_float(shift.get("opening_cash")) + shift_cash_sales + shift_cash_in - shift_cash_out - shift_cash_refunds
        if shift.get("closed_at"):
            total_close_taken += close_taken_for_shift(shift_id)
            if latest_closed_shift is None or str(shift.get("closed_at") or "") > str(latest_closed_shift.get("closed_at") or ""):
                latest_closed_shift = shift
            total_variance += _as_float(shift.get("closing_cash")) - shift_expected
            closed_shift_count += 1
        else:
            if latest_open_shift is None or str(shift.get("opened_at") or "") > str(latest_open_shift.get("opened_at") or ""):
                latest_open_shift = dict(shift)
                latest_open_shift["_expected_cash_now"] = shift_expected
    if latest_open_shift and (
        latest_closed_shift is None
        or str(latest_open_shift.get("opened_at") or "") >= str(latest_closed_shift.get("closed_at") or "")
    ):
        total_left = _as_float(latest_open_shift.get("_expected_cash_now"))
    elif latest_closed_shift:
        total_left = _as_float(latest_closed_shift.get("closing_cash"))
    else:
        total_left = 0.0
    adds_up = abs(total_variance) <= 0.005

    reports_dir = Path(reports_folder)
    reports_dir.mkdir(parents=True, exist_ok=True)
    out_path = reports_dir / f"cash_drawer_report_{stamp}_{datetime.now().strftime('%Y_%m_%d_%H%M%S')}.pdf"

    doc = SimpleDocTemplate(
        str(out_path),
        pagesize=letter,
        rightMargin=0.45 * inch,
        leftMargin=0.45 * inch,
        topMargin=0.45 * inch,
        bottomMargin=0.45 * inch,
    )
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="Small", parent=styles["Normal"], fontSize=8, leading=10))
    styles.add(ParagraphStyle(name="Muted", parent=styles["Normal"], fontSize=8, leading=10, textColor=colors.HexColor("#5b677a")))
    styles.add(ParagraphStyle(name="Section", parent=styles["Heading2"], fontSize=12, leading=14, spaceBefore=8, spaceAfter=4))
    styles.add(ParagraphStyle(name="HeroLabel", parent=styles["Normal"], fontSize=8, leading=10, textColor=colors.HexColor("#5b677a"), alignment=1))
    styles.add(ParagraphStyle(name="HeroValue", parent=styles["Normal"], fontSize=20, leading=23, textColor=colors.HexColor("#17324d"), alignment=1))
    styles.add(ParagraphStyle(name="HeroSales", parent=styles["HeroValue"], textColor=colors.HexColor("#16803C")))
    styles.add(ParagraphStyle(name="HeroDrawer", parent=styles["HeroValue"], textColor=colors.HexColor("#1D4ED8")))

    story = []
    title = "Mask POS Cash Drawer Report"
    period = f"{start.strftime('%Y-%m-%d')} to {(end - timedelta(days=1)).strftime('%Y-%m-%d')}"
    story.append(Paragraph(title, styles["Title"]))
    story.append(Paragraph(f"Period: {period} &nbsp;&nbsp; Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles["Muted"]))
    story.append(Spacer(1, 8))

    hero = [
        [
            Paragraph("TOTAL DAY SALES / REVENUE", styles["HeroLabel"]),
            Paragraph("LEFT OR CURRENT IN REGISTER", styles["HeroLabel"]),
        ],
        [
            Paragraph(money_fmt(total_revenue), styles["HeroSales"]),
            Paragraph(money_fmt(total_left), styles["HeroDrawer"]),
        ],
    ]
    hero_table = Table(hero, colWidths=[3.1 * inch, 3.1 * inch])
    hero_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#bfdbfe")),
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#eef6ff")),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#ECFDF3")),
        ("BACKGROUND", (1, 0), (1, -1), colors.HexColor("#EFF6FF")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(hero_table)
    story.append(Spacer(1, 8))

    previous = _last_shift_before_start(db_path, start)
    if previous:
        expected_start = _as_float(previous.get("closing_cash"))
        story.append(Paragraph(
            f"Previous closed register left {money_fmt(expected_start)} for the next opening "
            f"(shift {shift_label(previous)}, closed {_time_of(previous.get('closed_at'))} on {_day_of(previous.get('closed_at'))}).",
            styles["Small"],
        ))
        story.append(Spacer(1, 6))

    summary = [
        ["Day total revenue", money_fmt(total_revenue), "Register left/current", money_fmt(total_left)],
        ["Cash sales", money_fmt(total_cash_sales), "Cash added", money_fmt(total_cash_in)],
        ["Cash removed", money_fmt(total_cash_out), "Cash refunds", money_fmt(total_cash_refunds)],
        ["Taken at close", money_fmt(total_close_taken), "Register variance", money_fmt(total_variance)],
    ]
    summary_table = Table(summary, colWidths=[1.45 * inch, 1.25 * inch, 1.55 * inch, 1.25 * inch])
    summary_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("ALIGN", (3, 0), (3, -1), "RIGHT"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(summary_table)

    shift_rows = sorted(shifts, key=lambda s: str(s.get("opened_at") or ""))
    if not shift_rows:
        story.append(Spacer(1, 10))
        story.append(Paragraph("No register shift was opened or closed in this period.", styles["Small"]))

    for idx, shift in enumerate(shift_rows):
        if idx:
            story.append(PageBreak())
        shift_id = int(shift.get("id") or 0)
        shift_sales = sorted(sales_by_shift.get(shift_id, []), key=lambda s: str(s.get("created_at") or ""))
        shift_movements = sorted(movements_by_shift.get(shift_id, []), key=lambda m: str(m.get("created_at") or ""))
        shift_returns = sorted(returns_by_shift.get(shift_id, []), key=lambda r: str(r.get("created_at") or ""))

        shift_revenue = sum(_sale_new_money(s) for s in shift_sales) - sum(_as_float(r.get("cash_refund")) for r in shift_returns)
        cash_sales = sum(_as_float(s.get("cash_paid")) for s in shift_sales)
        cash_in = sum(_as_float(m.get("amount_value")) for m in shift_movements if str(m.get("movement_type") or "").upper() == "IN")
        cash_out = sum(_as_float(m.get("amount_value")) for m in shift_movements if str(m.get("movement_type") or "").upper() == "OUT")
        cash_refunds = sum(_as_float(r.get("cash_refund")) for r in shift_returns)
        opening = _as_float(shift.get("opening_cash"))
        opening_usd = _as_float(shift.get("opening_usd"))
        opening_lbp = _as_float(shift.get("opening_lbp"))
        closing = _as_float(shift.get("closing_cash")) if shift.get("closed_at") else None
        closing_usd = _as_float(shift.get("closing_usd")) if shift.get("closed_at") else None
        closing_lbp = _as_float(shift.get("closing_lbp")) if shift.get("closed_at") else None
        close_taken = close_taken_for_shift(shift_id)
        expected_before_close_take = opening + cash_sales + cash_in - (cash_out - close_taken) - cash_refunds
        expected_after_close_take = expected_before_close_take - close_taken
        variance = (closing - expected_after_close_take) if closing is not None else None

        story.append(Spacer(1, 10))
        story.append(Paragraph(f"Register Shift {shift_label(shift)}", styles["Section"]))
        details = [
            ["Employee", str(shift.get("employee_name") or "Unassigned"), "Opened", str(shift.get("opened_at") or "")],
            ["Opening cash", money_fmt(opening), "Closed", str(shift.get("closed_at") or "Open")],
            ["Opening USD", money_fmt(opening_usd), "Opening LBP", lbp_fmt(opening_lbp)],
            ["Closing USD", (money_fmt(closing_usd) if closing_usd is not None else "-"), "Closing LBP", (lbp_fmt(closing_lbp) if closing_lbp is not None else "-")],
            ["New money", money_fmt(shift_revenue), "Orders", str(len(shift_sales))],
            ["Cash sales", money_fmt(cash_sales), "Cash refunds", money_fmt(cash_refunds)],
            ["Cash added", money_fmt(cash_in), "Cash removed", money_fmt(cash_out)],
            ["Taken at close", money_fmt(close_taken), "Left for next opening", money_fmt(closing if closing is not None else expected_after_close_take)],
            ["Should be in register", money_fmt(expected_after_close_take), "Does it add up?", ("YES" if variance is None or abs(variance) <= 0.005 else f"NO {money_fmt(variance)}")],
        ]
        details_table = Table(details, colWidths=[1.45 * inch, 1.55 * inch, 1.45 * inch, 1.75 * inch])
        details_table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("ALIGN", (3, 0), (3, -1), "RIGHT"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(details_table)

        if shift.get("closed_at"):
            next_shift = _next_shift_after(db_path, str(shift.get("closed_at") or ""))
            if next_shift:
                next_opening = _as_float(next_shift.get("opening_cash"))
                next_diff = next_opening - (closing or 0.0)
                story.append(Spacer(1, 5))
                story.append(Paragraph(
                    f"Next opening check: shift {shift_label(next_shift)} opened with {money_fmt(next_opening)}. "
                    f"Difference from amount left: {money_fmt(next_diff)}.",
                    styles["Small"],
                ))
            else:
                story.append(Spacer(1, 5))
                story.append(Paragraph(
                    f"Next opening should start with {money_fmt(closing if closing is not None else expected_after_close_take)}.",
                    styles["Small"],
                ))

        story.append(Spacer(1, 8))
        story.append(Paragraph("Sales in this register", styles["Section"]))
        sale_data = [["Time", "Sale ID", "Receipt", "Cash Paid", "Total Sales", "Payment"]]
        for sale in shift_sales:
            sale_data.append([
                _time_of(sale.get("created_at")),
                str(sale.get("id") or ""),
                str(sale.get("receipt_code") or ""),
                money_fmt(sale.get("cash_paid")),
                money_fmt(sale.get("merch_total")),
                str(sale.get("payment_method") or ""),
            ])
        if len(sale_data) == 1:
            sale_data.append(["-", "-", "No sales", money_fmt(0), money_fmt(0), "-"])
        sale_table = Table(sale_data, colWidths=[0.75 * inch, 0.65 * inch, 1.35 * inch, 1.0 * inch, 1.0 * inch, 1.0 * inch], repeatRows=1)
        sale_table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#d9eaf7")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (3, 1), (4, -1), "RIGHT"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(sale_table)

        story.append(Spacer(1, 8))
        story.append(Paragraph("Cash movements", styles["Section"]))
        movement_data = [["Time", "Type", "Amount", "Employee", "Reason"]]
        for movement in shift_movements:
            mtype = str(movement.get("movement_type") or "OUT").upper()
            movement_data.append([
                _time_of(movement.get("created_at")),
                "Cash In" if mtype == "IN" else "Cash Out",
                money_fmt(movement.get("amount_value")),
                str(movement.get("employee_name") or ""),
                str(movement.get("reason") or movement.get("notes") or ""),
            ])
        for ret in shift_returns:
            if _as_float(ret.get("cash_refund")):
                movement_data.append([
                    _time_of(ret.get("created_at")),
                    "Cash Refund",
                    money_fmt(ret.get("cash_refund")),
                    "",
                    f"Return #{ret.get('id')} for sale {ret.get('original_sale_id')}",
                ])
        if len(movement_data) == 1:
            movement_data.append(["-", "-", money_fmt(0), "-", "No cash movements"])
        movement_table = Table(movement_data, colWidths=[0.75 * inch, 1.0 * inch, 1.0 * inch, 1.2 * inch, 2.75 * inch], repeatRows=1)
        movement_table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#d9eaf7")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (2, 1), (2, -1), "RIGHT"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(movement_table)

    if sales_unassigned:
        story.append(PageBreak())
        story.append(Paragraph("Sales Without Register Shift", styles["Section"]))
        data = [["Time", "Sale ID", "Receipt", "Cash Paid", "Total Sales", "Payment"]]
        for sale in sales_unassigned:
            data.append([
                _time_of(sale.get("created_at")),
                str(sale.get("id") or ""),
                str(sale.get("receipt_code") or ""),
                money_fmt(sale.get("cash_paid")),
                money_fmt(sale.get("merch_total")),
                str(sale.get("payment_method") or ""),
            ])
        table = Table(data, colWidths=[0.75 * inch, 0.65 * inch, 1.35 * inch, 1.0 * inch, 1.0 * inch, 1.0 * inch], repeatRows=1)
        table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#fce4d6")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (3, 1), (4, -1), "RIGHT"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ]))
        story.append(table)

    doc.build(story)
    return {
        "path": str(out_path),
        "start": start.strftime("%Y-%m-%d"),
        "end": (end - timedelta(days=1)).strftime("%Y-%m-%d"),
        "summary": {
            "total_revenue": round(total_revenue, 2),
            "cash_sales": round(total_cash_sales, 2),
            "cash_added": round(total_cash_in, 2),
            "cash_removed": round(total_cash_out, 2),
            "cash_refunds": round(total_cash_refunds, 2),
            "taken_at_close": round(total_close_taken, 2),
            "left_for_next_opening": round(total_left, 2),
            "variance": round(total_variance, 2),
            "balanced": bool(adds_up),
        },
    }


def build_cash_drawer_pdf(db_path: str, reports_folder: str, year: int = None, month: int = None, day_str: str = None, start_date = None, end_date = None, custom_stamp: str = "") -> dict:
    """Build a compact drawer PDF for email.

    Excel keeps the detailed sale/item breakdown. This PDF is intentionally
    manager-facing: important totals, shift health, and drawer variance only.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    if start_date is not None and end_date is not None:
        start = start_date
        end = end_date
        stamp = custom_stamp or f"range_{start.strftime('%Y%m%d')}_to_{end.strftime('%Y%m%d')}"
    else:
        start, end, stamp = report_date_bounds(year or datetime.now().year, month or datetime.now().month, day_str or "All")
    if not os.path.exists(db_path):
        raise FileNotFoundError(f"Could not find database: {db_path}")

    rows = _fetch_rows(db_path, start, end)
    sales = rows.get("sales", [])
    shifts = rows.get("shifts", [])
    movements = rows.get("movements", [])
    returns = rows.get("returns", [])

    sales_by_shift = defaultdict(list)
    sales_unassigned = []
    for sale in sales:
        sid = sale.get("shift_id")
        if sid in (None, ""):
            sales_unassigned.append(sale)
        else:
            sales_by_shift[int(sid)].append(sale)

    movements_by_shift = defaultdict(list)
    for movement in movements:
        sid = movement.get("shift_id")
        if sid not in (None, ""):
            movements_by_shift[int(sid)].append(movement)

    returns_by_shift = defaultdict(list)
    for ret in returns:
        sid = ret.get("shift_id")
        if sid not in (None, ""):
            returns_by_shift[int(sid)].append(ret)

    def money_fmt(value) -> str:
        try:
            return f"${float(value or 0.0):,.2f}"
        except Exception:
            return "$0.00"

    def lbp_fmt(value) -> str:
        try:
            return f"{float(value or 0.0):,.0f} LBP"
        except Exception:
            return "0 LBP"

    def shift_label(shift: dict) -> str:
        return str(shift.get("shift_code") or shift.get("id") or "").strip()

    def close_taken_for_shift(shift_id: int) -> float:
        total = 0.0
        for movement in movements_by_shift.get(int(shift_id), []):
            reason = str(movement.get("reason") or "").strip().lower()
            note = str(movement.get("notes") or "").strip().lower()
            if str(movement.get("movement_type") or "").upper() == "OUT" and (
                "end of day close" in reason or "cash taken at close" in note
            ):
                total += _as_float(movement.get("amount_value"))
        return total

    variance_tolerance = 4.0

    total_cash_sales = sum(_as_float(s.get("cash_paid")) for s in sales)
    total_store_credit = sum(_as_float(s.get("store_credit_used")) for s in sales)
    total_cash_refunds = sum(_as_float(r.get("cash_refund")) for r in returns)
    total_returns = sum(_as_float(r.get("total_return_amount")) for r in returns)
    total_revenue = sum(_sale_new_money(s) for s in sales) - total_cash_refunds
    total_cash_in = sum(_as_float(m.get("amount_value")) for m in movements if str(m.get("movement_type") or "").upper() == "IN")
    total_cash_out = sum(_as_float(m.get("amount_value")) for m in movements if str(m.get("movement_type") or "").upper() == "OUT")
    total_close_taken = 0.0
    total_variance = 0.0
    closed_shift_count = 0
    latest_closed_shift = None
    latest_open_shift = None
    shift_rows = []

    for shift in sorted(shifts, key=lambda s: str(s.get("opened_at") or "")):
        shift_id = int(shift.get("id") or 0)
        shift_sales = sales_by_shift.get(shift_id, [])
        shift_movements = movements_by_shift.get(shift_id, [])
        shift_returns = returns_by_shift.get(shift_id, [])

        cash_sales = sum(_as_float(s.get("cash_paid")) for s in shift_sales)
        new_money = sum(_sale_new_money(s) for s in shift_sales) - sum(_as_float(r.get("cash_refund")) for r in shift_returns)
        cash_in = sum(_as_float(m.get("amount_value")) for m in shift_movements if str(m.get("movement_type") or "").upper() == "IN")
        cash_out = sum(_as_float(m.get("amount_value")) for m in shift_movements if str(m.get("movement_type") or "").upper() == "OUT")
        refunds = sum(_as_float(r.get("cash_refund")) for r in shift_returns)
        opening = _as_float(shift.get("opening_cash"))
        expected = opening + cash_sales + cash_in - cash_out - refunds
        counted = None
        diff = None
        close_taken = 0.0
        status = "OPEN"
        if shift.get("closed_at"):
            counted = _as_float(shift.get("closing_cash"))
            diff = counted - expected
            total_variance += diff
            closed_shift_count += 1
            close_taken = close_taken_for_shift(shift_id)
            total_close_taken += close_taken
            status = "OK" if abs(diff) <= variance_tolerance else "CHECK"
            if latest_closed_shift is None or str(shift.get("closed_at") or "") > str(latest_closed_shift.get("closed_at") or ""):
                latest_closed_shift = shift
        else:
            counted = expected
            if latest_open_shift is None or str(shift.get("opened_at") or "") > str(latest_open_shift.get("opened_at") or ""):
                latest_open_shift = dict(shift)
                latest_open_shift["_expected_cash_now"] = expected

        shift_rows.append({
            "shift": shift_label(shift),
            "employee": str(shift.get("employee_name") or "Unassigned"),
            "time": f"{_time_of(shift.get('opened_at'))} - {_time_of(shift.get('closed_at')) if shift.get('closed_at') else 'Open'}",
            "orders": len(shift_sales),
            "new_money": new_money,
            "cash_sales": cash_sales,
            "cash_in": cash_in,
            "cash_out": cash_out,
            "refunds": refunds,
            "opening": opening,
            "opening_usd": _as_float(shift.get("opening_usd")),
            "opening_lbp": _as_float(shift.get("opening_lbp")),
            "closing_usd": _as_float(shift.get("closing_usd")) if shift.get("closed_at") else None,
            "closing_lbp": _as_float(shift.get("closing_lbp")) if shift.get("closed_at") else None,
            "should_be": expected,
            "counted": counted,
            "diff": diff,
            "close_taken": close_taken,
            "status": status,
        })

    if latest_open_shift and (
        latest_closed_shift is None
        or str(latest_open_shift.get("opened_at") or "") >= str(latest_closed_shift.get("closed_at") or "")
    ):
        total_left = _as_float(latest_open_shift.get("_expected_cash_now"))
        register_label = "Current open register"
    elif latest_closed_shift:
        total_left = _as_float(latest_closed_shift.get("closing_cash"))
        register_label = "Left after last close"
    else:
        total_left = 0.0
        register_label = "Register"

    balanced = abs(total_variance) <= variance_tolerance
    reports_dir = Path(reports_folder)
    reports_dir.mkdir(parents=True, exist_ok=True)
    out_path = reports_dir / f"cash_drawer_report_{stamp}_{datetime.now().strftime('%Y_%m_%d_%H%M%S')}.pdf"

    doc = SimpleDocTemplate(
        str(out_path),
        pagesize=landscape(letter),
        rightMargin=0.38 * inch,
        leftMargin=0.38 * inch,
        topMargin=0.35 * inch,
        bottomMargin=0.35 * inch,
    )
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="MutedSmall", parent=styles["Normal"], fontSize=7.5, leading=9, textColor=colors.HexColor("#64748b")))
    styles.add(ParagraphStyle(name="MetricLabel", parent=styles["Normal"], fontSize=8, leading=10, alignment=1, textColor=colors.HexColor("#475569")))
    styles.add(ParagraphStyle(name="MetricGreen", parent=styles["Normal"], fontSize=20, leading=23, alignment=1, textColor=colors.HexColor("#15803d")))
    styles.add(ParagraphStyle(name="MetricBlue", parent=styles["Normal"], fontSize=20, leading=23, alignment=1, textColor=colors.HexColor("#1d4ed8")))
    styles.add(ParagraphStyle(name="MetricRed", parent=styles["Normal"], fontSize=20, leading=23, alignment=1, textColor=colors.HexColor("#b91c1c")))
    styles.add(ParagraphStyle(name="SectionCompact", parent=styles["Heading2"], fontSize=11, leading=13, spaceBefore=8, spaceAfter=4))
    styles.add(ParagraphStyle(name="CellParagraph", parent=styles["Normal"], fontSize=6.5, leading=8))
    styles.add(ParagraphStyle(name="CellParagraphBold", parent=styles["Normal"], fontSize=6.5, leading=8, fontName="Helvetica-Bold"))

    story = []
    period = f"{start.strftime('%Y-%m-%d')} to {(end - timedelta(days=1)).strftime('%Y-%m-%d')}"
    story.append(Paragraph("Mask POS Cash Drawer Summary", styles["Title"]))
    story.append(Paragraph(f"Period: {period} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles["MutedSmall"]))
    story.append(Spacer(1, 8))

    variance_style = styles["MetricGreen"] if balanced else styles["MetricRed"]
    hero = [
        [
            Paragraph("TOTAL SALES / NEW MONEY", styles["MetricLabel"]),
            Paragraph("CASH COLLECTED", styles["MetricLabel"]),
            Paragraph(register_label.upper(), styles["MetricLabel"]),
            Paragraph("REGISTER VARIANCE", styles["MetricLabel"]),
        ],
        [
            Paragraph(money_fmt(total_revenue), styles["MetricGreen"]),
            Paragraph(money_fmt(total_cash_sales), styles["MetricGreen"]),
            Paragraph(money_fmt(total_left), styles["MetricBlue"]),
            Paragraph(money_fmt(total_variance), variance_style),
        ],
    ]
    hero_table = Table(hero, colWidths=[2.45 * inch, 2.2 * inch, 2.2 * inch, 2.2 * inch])
    hero_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#bfdbfe")),
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.append(hero_table)
    story.append(Spacer(1, 8))

    important = [
        ["Orders", str(len(sales)), "Store credit used", money_fmt(total_store_credit)],
        ["Cash in", money_fmt(total_cash_in), "Cash out", money_fmt(total_cash_out)],
        ["Returns total", money_fmt(total_returns), "Cash refunds", money_fmt(total_cash_refunds)],
        ["Taken at close", money_fmt(total_close_taken), "Balance status", "OK" if balanced else "CHECK REGISTER"],
    ]
    important_table = Table(important, colWidths=[1.35 * inch, 1.55 * inch, 1.55 * inch, 1.55 * inch])
    important_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#ffffff")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("ALIGN", (3, 0), (3, -1), "RIGHT"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(important_table)

    story.append(Paragraph("Register Shifts", styles["SectionCompact"]))
    shift_data = [[
        "Shift", "Employee", "Time", "Orders", "New Money", "Opening",
        "Cash +/-", "Should Be", "Counted/Now", "Diff", "Status"
    ]]
    for row in shift_rows:
        cash_pm = f"+{money_fmt(row['cash_in'])} / -{money_fmt(row['cash_out'])}"
        diff_text = "-" if row["diff"] is None else money_fmt(row["diff"])

        op_usd = row.get("opening_usd", 0.0)
        op_lbp = row.get("opening_lbp", 0.0)
        opening_p = Paragraph(f"<b>{money_fmt(op_usd)}</b><br/>{lbp_fmt(op_lbp)}", styles["CellParagraph"])

        cl_usd = row.get("closing_usd")
        cl_lbp = row.get("closing_lbp")
        if cl_usd is not None or cl_lbp is not None:
            counted_p = Paragraph(f"<b>{money_fmt(cl_usd or 0.0)}</b><br/>{lbp_fmt(cl_lbp or 0.0)}", styles["CellParagraph"])
        else:
            counted_p = Paragraph(f"<b>{money_fmt(row['counted'])}</b>", styles["CellParagraphBold"])

        shift_data.append([
            str(row["shift"]),
            str(row["employee"])[:18],
            str(row["time"]),
            str(row["orders"]),
            money_fmt(row["new_money"]),
            opening_p,
            cash_pm,
            money_fmt(row["should_be"]),
            counted_p,
            diff_text,
            str(row["status"]),
        ])
    if len(shift_data) == 1:
        shift_data.append(["-", "-", "-", "0", money_fmt(0), money_fmt(0), "+$0.00 / -$0.00", money_fmt(0), money_fmt(0), "-", "NO SHIFT"])

    shift_table = Table(
        shift_data,
        colWidths=[0.55 * inch, 1.25 * inch, 1.2 * inch, 0.48 * inch, 0.9 * inch, 0.85 * inch, 1.1 * inch, 0.9 * inch, 0.9 * inch, 0.75 * inch, 0.65 * inch],
        repeatRows=1,
    )
    shift_style = [
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dbeafe")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (3, 1), (9, -1), "RIGHT"),
        ("ALIGN", (10, 1), (10, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]
    for idx, row in enumerate(shift_rows, start=1):
        if row["status"] == "CHECK":
            shift_style.append(("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#fee2e2")))
        elif row["status"] == "OPEN":
            shift_style.append(("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#eff6ff")))
    shift_table.setStyle(TableStyle(shift_style))
    story.append(shift_table)

    story.append(Paragraph("Sales Data", styles["SectionCompact"]))
    sales_data = [["Sale ID", "Time", "Total", "Pay", "Shift"]]
    displayed_sales_sum = 0.0
    for sale in sorted(sales, key=lambda s: str(s.get("created_at") or ""), reverse=True):
        display_amount = _sale_new_money(sale)
        displayed_sales_sum += display_amount
        receipt_code = str(sale.get("receipt_code") or "").strip()
        if not receipt_code:
            try:
                receipt_code = f"{int(sale.get('id') or 0):04d}"
            except Exception:
                receipt_code = str(sale.get("id") or "")
        pm = str(sale.get("payment_method") or "").strip().upper()
        if pm in ("EXCHANGE", "STORE_CREDIT"):
            pay = "Exchange / Store Credit"
        elif pm in ("CARD", "DEBIT", "CREDIT_CARD"):
            pay = "Card"
        elif pm == "WHISH":
            pay = "Whish"
        else:
            pay = "Cash" if pm == "CASH" or not pm else pm.title()
        sales_data.append([
            receipt_code,
            _time_of(sale.get("created_at")),
            money_fmt(display_amount),
            pay,
            str(sale.get("shift_id") if sale.get("shift_id") is not None else ""),
        ])
    if len(sales_data) == 1:
        sales_data.append(["-", "-", money_fmt(0), "-", "-"])

    sales_table = Table(
        sales_data,
        colWidths=[0.9 * inch, 0.85 * inch, 1.0 * inch, 1.55 * inch, 0.55 * inch],
        repeatRows=1,
    )
    sales_table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e2e8f0")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (2, 1), (2, -1), "RIGHT"),
        ("ALIGN", (4, 1), (4, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 2.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
    ]))
    story.append(sales_table)

    footer_totals = [
        [
            f"Sales total: {money_fmt(displayed_sales_sum)}",
            f"Cash collected: {money_fmt(total_cash_sales)}",
            f"Sales: {len(sales)}",
        ],
        [
            f"Cash in: {money_fmt(total_cash_in)}",
            f"Cash out: {money_fmt(total_cash_out)}",
            "",
        ],
        [
            f"Drawer net: {money_fmt(total_cash_sales + total_cash_in - total_cash_out)}",
            f"Cash in register: {money_fmt(total_left)}",
            "",
        ],
    ]
    footer_table = Table(footer_totals, colWidths=[2.1 * inch, 2.2 * inch, 1.2 * inch])
    footer_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#0f172a")),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(Spacer(1, 5))
    story.append(footer_table)

    attention = []
    if not balanced:
        attention.append(f"Register variance needs checking: {money_fmt(total_variance)}. Differences within +/- {money_fmt(variance_tolerance)} are treated as OK.")
    if sales_unassigned:
        attention.append(f"{len(sales_unassigned)} sale(s) were not attached to a register shift.")
    if latest_open_shift:
        attention.append(f"Open register should currently hold about {money_fmt(total_left)}.")
    if not attention:
        attention.append("No drawer issues detected.")

    story.append(Spacer(1, 8))
    story.append(Paragraph("Attention", styles["SectionCompact"]))
    story.append(Paragraph("<br/>".join(attention), styles["MutedSmall"]))
    story.append(Spacer(1, 5))
    story.append(Paragraph("Product/item details, returns, and inventory remain in the attached Excel report.", styles["MutedSmall"]))

    doc.build(story)
    return {
        "path": str(out_path),
        "start": start.strftime("%Y-%m-%d"),
        "end": (end - timedelta(days=1)).strftime("%Y-%m-%d"),
        "summary": {
            "total_revenue": round(total_revenue, 2),
            "cash_sales": round(total_cash_sales, 2),
            "cash_added": round(total_cash_in, 2),
            "cash_removed": round(total_cash_out, 2),
            "cash_refunds": round(total_cash_refunds, 2),
            "taken_at_close": round(total_close_taken, 2),
            "left_for_next_opening": round(total_left, 2),
            "variance": round(total_variance, 2),
            "balanced": bool(balanced),
        },
    }


def _build_sales_report_excel_v2(db_path: str, reports_folder: str, year: int, month: int, day_str: str) -> dict:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    start, end, stamp = report_date_bounds(year, month, day_str)
    if not os.path.exists(db_path):
        raise FileNotFoundError(f"Could not find database: {db_path}")

    rows = _fetch_rows(db_path, start, end)
    sales = rows.get("sales", [])
    items = rows.get("items", [])
    shifts = rows.get("shifts", [])
    movements = rows.get("movements", [])
    returns = rows.get("returns", [])
    return_items = rows.get("return_items", [])

    sales_by_id = {int(s.get("id") or 0): s for s in sales}
    shifts_by_id = {int(s.get("id") or 0): s for s in shifts}
    returns_by_id = {int(r.get("id") or 0): r for r in returns}

    items_by_sale = defaultdict(list)
    for item in items:
        items_by_sale[int(item.get("sale_id") or 0)].append(item)

    return_items_by_return = defaultdict(list)
    for item in return_items:
        return_items_by_return[int(item.get("return_id") or 0)].append(item)

    sales_by_day = defaultdict(list)
    for sale in sales:
        sales_by_day[_day_of(sale.get("created_at"))].append(sale)

    items_by_day = defaultdict(list)
    for item in items:
        sale = sales_by_id.get(int(item.get("sale_id") or 0), {})
        items_by_day[_day_of(sale.get("created_at"))].append(item)

    shifts_by_day = defaultdict(list)
    for shift in shifts:
        shifts_by_day[_day_of(shift.get("opened_at"))].append(shift)

    movements_by_day = defaultdict(list)
    for movement in movements:
        movements_by_day[_day_of(movement.get("created_at"))].append(movement)

    returns_by_day = defaultdict(list)
    for ret in returns:
        returns_by_day[_day_of(ret.get("created_at"))].append(ret)

    return_items_by_day = defaultdict(list)
    for item in return_items:
        return_items_by_day[_day_of(item.get("created_at"))].append(item)

    days = set(sales_by_day.keys()) | set(shifts_by_day.keys()) | set(movements_by_day.keys()) | set(returns_by_day.keys())
    if str(day_str).strip().lower() != "all":
        days.add(start.strftime("%Y-%m-%d"))
    if not days:
        days.add(start.strftime("%Y-%m-%d"))
    days = sorted(d for d in days if d)

    def safe_div(numerator, denominator) -> float:
        denominator = float(denominator or 0.0)
        return (float(numerator or 0.0) / denominator) if denominator else 0.0

    def product_key(row: dict) -> str:
        pid = row.get("product_id")
        if pid not in (None, "", 0, "0"):
            return f"id:{pid}"
        barcode = str(row.get("barcode") or "").strip()
        if barcode:
            return f"barcode:{barcode}"
        return f"name:{str(row.get('name') or '').strip().lower()}"

    def product_record(row: dict) -> dict:
        return {
            "product_id": row.get("product_id"),
            "name": str(row.get("name") or "").strip() or "(Unnamed item)",
            "barcode": str(row.get("barcode") or "").strip(),
            "category": str(row.get("category") or "").strip(),
            "brand": str(row.get("brand") or "").strip(),
            "qty_sold": 0,
            "qty_returned": 0,
            "gross_revenue": 0.0,
            "discounts": 0.0,
            "sales_revenue": 0.0,
            "return_amount": 0.0,
            "current_stock": row.get("current_stock"),
        }

    product_totals = {}
    for item in items:
        key = product_key(item)
        rec = product_totals.setdefault(key, product_record(item))
        if not rec.get("barcode") and item.get("barcode"):
            rec["barcode"] = str(item.get("barcode") or "").strip()
        if not rec.get("category") and item.get("category"):
            rec["category"] = str(item.get("category") or "").strip()
        if not rec.get("brand") and item.get("brand"):
            rec["brand"] = str(item.get("brand") or "").strip()
        if item.get("current_stock") not in (None, ""):
            rec["current_stock"] = item.get("current_stock")
        qty = _as_int(item.get("qty"))
        gross = _as_float(item.get("gross_line_total"))
        if not gross:
            gross = _as_float(item.get("price")) * qty
        line = _as_float(item.get("line_total"))
        discount = _as_float(item.get("discount_allocated"))
        rec["qty_sold"] += qty
        rec["gross_revenue"] += gross
        rec["discounts"] += discount
        rec["sales_revenue"] += line

    for item in return_items:
        key = product_key(item)
        rec = product_totals.setdefault(key, product_record(item))
        if not rec.get("barcode") and item.get("barcode"):
            rec["barcode"] = str(item.get("barcode") or "").strip()
        if not rec.get("category") and item.get("category"):
            rec["category"] = str(item.get("category") or "").strip()
        if not rec.get("brand") and item.get("brand"):
            rec["brand"] = str(item.get("brand") or "").strip()
        if item.get("current_stock") not in (None, ""):
            rec["current_stock"] = item.get("current_stock")
        rec["qty_returned"] += _as_int(item.get("qty"))
        rec["return_amount"] += _as_float(item.get("line_total"))

    product_rows = []
    total_product_net_revenue = 0.0
    for rec in product_totals.values():
        rec["net_qty"] = int(rec["qty_sold"]) - int(rec["qty_returned"])
        rec["net_revenue"] = float(rec["sales_revenue"]) - float(rec["return_amount"])
        rec["avg_unit"] = safe_div(rec["sales_revenue"], rec["qty_sold"])
        total_product_net_revenue += rec["net_revenue"]
        product_rows.append(rec)

    product_rows.sort(key=lambda r: (-float(r.get("net_revenue") or 0.0), -int(r.get("qty_sold") or 0), str(r.get("name") or "").lower()))
    top_by_revenue = product_rows[:10]
    top_by_qty = sorted(product_rows, key=lambda r: (-int(r.get("qty_sold") or 0), -float(r.get("net_revenue") or 0.0), str(r.get("name") or "").lower()))[:10]

    total_merch = sum(_as_float(s.get("merch_total")) for s in sales)
    total_cash = sum(_as_float(s.get("cash_paid")) for s in sales)
    total_credit = sum(_as_float(s.get("store_credit_used")) for s in sales)
    total_items = sum(_as_int(i.get("qty")) for i in items)
    total_return_qty = sum(_as_int(i.get("qty")) for i in return_items)
    total_returns = sum(_as_float(r.get("total_return_amount")) for r in returns)
    total_cash_refunds = sum(_as_float(r.get("cash_refund")) for r in returns)
    total_cash_in = sum(_as_float(m.get("amount_value")) for m in movements if str(m.get("movement_type") or "").upper() == "IN")
    total_cash_out = sum(_as_float(m.get("amount_value")) for m in movements if str(m.get("movement_type") or "").upper() == "OUT")
    total_net_sales = total_merch - total_returns
    total_net_qty = total_items - total_return_qty

    def shift_label(shift_id) -> str:
        try:
            sid = int(shift_id or 0)
        except Exception:
            sid = 0
        if not sid:
            return ""
        shift = shifts_by_id.get(sid, {})
        code = str(shift.get("shift_code") or "").strip()
        return code or str(sid)

    def sale_item_summary(sale_id: int) -> str:
        parts = []
        for item in items_by_sale.get(int(sale_id or 0), []):
            qty = _as_int(item.get("qty"))
            name = str(item.get("name") or "").strip() or "Item"
            parts.append(f"{qty} x {name}")
        text = "; ".join(parts)
        return text if len(text) <= 240 else text[:237] + "..."

    def daily_summary(day: str) -> dict:
        day_sales = sales_by_day.get(day, [])
        day_items = items_by_day.get(day, [])
        day_returns = returns_by_day.get(day, [])
        day_return_items = return_items_by_day.get(day, [])
        day_movements = movements_by_day.get(day, [])
        cash_in = sum(_as_float(m.get("amount_value")) for m in day_movements if str(m.get("movement_type") or "").upper() == "IN")
        cash_out = sum(_as_float(m.get("amount_value")) for m in day_movements if str(m.get("movement_type") or "").upper() == "OUT")
        sales_total = sum(_as_float(s.get("merch_total")) for s in day_sales)
        returns_total = sum(_as_float(r.get("total_return_amount")) for r in day_returns)
        cash_refunds = sum(_as_float(r.get("cash_refund")) for r in day_returns)
        orders = len(day_sales)
        qty_sold = sum(_as_int(i.get("qty")) for i in day_items)
        qty_returned = sum(_as_int(i.get("qty")) for i in day_return_items)
        net_sales = sales_total - returns_total
        return {
            "orders": orders,
            "qty_sold": qty_sold,
            "qty_returned": qty_returned,
            "net_qty": qty_sold - qty_returned,
            "sales_revenue": sales_total,
            "returns": returns_total,
            "net_sales": net_sales,
            "cash": sum(_as_float(s.get("cash_paid")) for s in day_sales),
            "credit": sum(_as_float(s.get("store_credit_used")) for s in day_sales),
            "cash_refunds": cash_refunds,
            "cash_in": cash_in,
            "cash_out": cash_out,
            "drawer_net": sum(_as_float(s.get("cash_paid")) for s in day_sales) + cash_in - cash_out - cash_refunds,
            "aov": safe_div(net_sales, orders),
        }

    wb = Workbook()
    ws_dash = wb.active
    ws_dash.title = "Dashboard"

    dark_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    sub_fill = PatternFill("solid", fgColor="EAF3F8")
    total_fill = PatternFill("solid", fgColor="E2F0D9")
    warn_fill = PatternFill("solid", fgColor="FCE4D6")
    thin = Side(style="thin", color="D8DEE9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(bold=True, size=18, color="FFFFFF")
    subtitle_font = Font(size=10, color="5B677A")
    header_font = Font(bold=True, color="17324D")
    section_font = Font(bold=True, size=13, color="17324D")
    bold = Font(bold=True)
    money_fmt = '$#,##0.00;[Red]-$#,##0.00'
    int_fmt = '0'
    percent_fmt = '0.0%'
    table_style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)

    def style_header(ws, row: int, start_col: int = 1, end_col: int | None = None):
        end_col = end_col or ws.max_column
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    def style_body(ws, start_row: int = 2):
        for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    def autosize(ws):
        for col_cells in ws.columns:
            letter = get_column_letter(col_cells[0].column)
            max_len = 0
            for cell in col_cells:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(70, len(value)))
            ws.column_dimensions[letter].width = min(42, max(9, max_len + 2))

    def format_cols(ws, money_cols=(), int_cols=(), percent_cols=(), start_row: int = 2):
        for col in money_cols:
            for row in range(start_row, ws.max_row + 1):
                ws.cell(row=row, column=col).number_format = money_fmt
        for col in int_cols:
            for row in range(start_row, ws.max_row + 1):
                ws.cell(row=row, column=col).number_format = int_fmt
        for col in percent_cols:
            for row in range(start_row, ws.max_row + 1):
                ws.cell(row=row, column=col).number_format = percent_fmt

    def add_table(ws, name: str, start_row: int, start_col: int = 1):
        if ws.max_row <= start_row:
            return
        end_col = get_column_letter(ws.max_column)
        ref = f"{get_column_letter(start_col)}{start_row}:{end_col}{ws.max_row}"
        try:
            table = Table(displayName=name, ref=ref)
            table.tableStyleInfo = table_style
            ws.add_table(table)
        except Exception:
            ws.auto_filter.ref = ref

    def write_table(ws, start_row: int, start_col: int, headers: list[str], data_rows: list[list], table_name: str | None = None):
        for idx, header in enumerate(headers, start=start_col):
            ws.cell(row=start_row, column=idx, value=header)
        style_header(ws, start_row, start_col, start_col + len(headers) - 1)
        for offset, row_data in enumerate(data_rows, start=1):
            row_num = start_row + offset
            for idx, value in enumerate(row_data, start=start_col):
                ws.cell(row=row_num, column=idx, value=value)
        if data_rows:
            for row in range(start_row + 1, ws.max_row + 1):
                for col in range(start_col, start_col + len(headers)):
                    ws.cell(row=row, column=col).border = border
                    ws.cell(row=row, column=col).alignment = Alignment(vertical="center", wrap_text=True)
        if table_name:
            add_table(ws, table_name, start_row, start_col)
        return ws.max_row

    def section(ws, row: int, title: str):
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = section_font
        cell.fill = sub_fill
        cell.border = border
        return row + 1

    for ws in [ws_dash]:
        ws.sheet_view.showGridLines = False

    ws_dash.merge_cells("A1:H1")
    ws_dash["A1"] = "Mask POS Sales Dashboard"
    ws_dash["A1"].font = title_font
    ws_dash["A1"].fill = dark_fill
    ws_dash["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws_dash.row_dimensions[1].height = 30
    ws_dash["A2"] = f"Period: {start.strftime('%Y-%m-%d')} to {(end - timedelta(days=1)).strftime('%Y-%m-%d')}"
    ws_dash["A2"].font = subtitle_font
    ws_dash["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_dash["A3"].font = subtitle_font

    summary_rows = [
        ["Net sales", total_net_sales],
        ["Sales before returns", total_merch],
        ["Returns", total_returns],
        ["Orders", len(sales)],
        ["Qty sold", total_items],
        ["Qty returned", total_return_qty],
        ["Net qty", total_net_qty],
        ["Average order value", safe_div(total_net_sales, len(sales))],
        ["Cash collected", total_cash],
        ["Store credit used", total_credit],
        ["Cash refunds", total_cash_refunds],
        ["Cash added", total_cash_in],
        ["Cash removed", total_cash_out],
        ["Drawer net change", total_cash + total_cash_in - total_cash_out - total_cash_refunds],
    ]
    ws_dash["A5"] = "Business Summary"
    ws_dash["A5"].font = section_font
    write_table(ws_dash, 6, 1, ["Metric", "Value"], summary_rows, None)
    for row in range(7, 7 + len(summary_rows)):
        metric = str(ws_dash.cell(row=row, column=1).value or "")
        if any(word in metric.lower() for word in ["sales", "returns", "average", "cash", "credit", "drawer"]):
            ws_dash.cell(row=row, column=2).number_format = money_fmt
        else:
            ws_dash.cell(row=row, column=2).number_format = int_fmt
    for row in range(6, 7 + len(summary_rows)):
        for col in (1, 2):
            ws_dash.cell(row=row, column=col).border = border

    top_rev_rows = [
        [idx, rec["name"], rec["qty_sold"], rec["net_revenue"], safe_div(rec["net_revenue"], total_net_sales)]
        for idx, rec in enumerate(top_by_revenue, start=1)
    ]
    ws_dash["D5"] = "Top Sellers by Revenue"
    ws_dash["D5"].font = section_font
    write_table(ws_dash, 6, 4, ["Rank", "Product", "Qty", "Net Revenue", "% Sales"], top_rev_rows, None)
    for row in range(7, 7 + len(top_rev_rows)):
        ws_dash.cell(row=row, column=4).number_format = int_fmt
        ws_dash.cell(row=row, column=6).number_format = int_fmt
        ws_dash.cell(row=row, column=7).number_format = money_fmt
        ws_dash.cell(row=row, column=8).number_format = percent_fmt

    start_qty = max(19, 8 + len(top_rev_rows))
    ws_dash.cell(row=start_qty, column=4, value="Top Sellers by Quantity").font = section_font
    top_qty_rows = [
        [idx, rec["name"], rec["qty_sold"], rec["net_revenue"], safe_div(rec["sales_revenue"], rec["qty_sold"])]
        for idx, rec in enumerate(top_by_qty, start=1)
    ]
    write_table(ws_dash, start_qty + 1, 4, ["Rank", "Product", "Qty Sold", "Net Revenue", "Avg Unit"], top_qty_rows, None)
    for row in range(start_qty + 2, start_qty + 2 + len(top_qty_rows)):
        ws_dash.cell(row=row, column=4).number_format = int_fmt
        ws_dash.cell(row=row, column=6).number_format = int_fmt
        ws_dash.cell(row=row, column=7).number_format = money_fmt
        ws_dash.cell(row=row, column=8).number_format = money_fmt

    ws_prod = wb.create_sheet("Product Mix")
    prod_headers = [
        "Rank", "Product", "Barcode", "Category", "Brand", "Qty Sold", "Qty Returned",
        "Net Qty", "Sales Revenue", "Return Amount", "Net Revenue", "Avg Unit",
        "% of Net Sales", "Current Stock",
    ]
    prod_data = []
    for idx, rec in enumerate(product_rows, start=1):
        prod_data.append([
            idx,
            rec["name"],
            rec["barcode"],
            rec["category"],
            rec["brand"],
            int(rec["qty_sold"]),
            int(rec["qty_returned"]),
            int(rec["net_qty"]),
            float(rec["sales_revenue"]),
            float(rec["return_amount"]),
            float(rec["net_revenue"]),
            float(rec["avg_unit"]),
            safe_div(rec["net_revenue"], total_net_sales),
            _as_int(rec.get("current_stock"), 0),
        ])
    write_table(ws_prod, 1, 1, prod_headers, prod_data, "ProductMix")
    format_cols(ws_prod, money_cols=(9, 10, 11, 12), int_cols=(1, 6, 7, 8, 14), percent_cols=(13,))
    ws_prod.freeze_panes = "A2"

    ws_top = wb.create_sheet("Top Sellers")
    ws_top.sheet_view.showGridLines = False
    ws_top["A1"] = "Top Sellers by Revenue"
    ws_top["A1"].font = section_font
    rev_data = [[idx, r["name"], r["barcode"], r["qty_sold"], r["net_revenue"], safe_div(r["net_revenue"], total_net_sales)] for idx, r in enumerate(top_by_revenue, start=1)]
    write_table(ws_top, 2, 1, ["Rank", "Product", "Barcode", "Qty Sold", "Net Revenue", "% Sales"], rev_data, None)
    qty_start = max(15, 4 + len(rev_data))
    ws_top.cell(row=qty_start, column=1, value="Top Sellers by Quantity").font = section_font
    qty_data = [[idx, r["name"], r["barcode"], r["qty_sold"], r["net_revenue"], r["avg_unit"]] for idx, r in enumerate(top_by_qty, start=1)]
    write_table(ws_top, qty_start + 1, 1, ["Rank", "Product", "Barcode", "Qty Sold", "Net Revenue", "Avg Unit"], qty_data, None)
    format_cols(ws_top, money_cols=(5, 6), int_cols=(1, 4), percent_cols=(6,), start_row=3)
    for row in range(qty_start + 2, qty_start + 2 + len(qty_data)):
        ws_top.cell(row=row, column=5).number_format = money_fmt
        ws_top.cell(row=row, column=6).number_format = money_fmt

    ws_daily = wb.create_sheet("Daily Sales")
    daily_headers = [
        "Day", "Orders", "Qty Sold", "Qty Returned", "Net Qty", "Sales Revenue",
        "Return Amount", "Net Sales", "Cash Collected", "Store Credit", "Cash Refunds",
        "Cash Added", "Cash Removed", "Drawer Net", "AOV",
    ]
    daily_data = []
    for day in days:
        ds = daily_summary(day)
        daily_data.append([
            day, ds["orders"], ds["qty_sold"], ds["qty_returned"], ds["net_qty"],
            ds["sales_revenue"], ds["returns"], ds["net_sales"], ds["cash"], ds["credit"],
            ds["cash_refunds"], ds["cash_in"], ds["cash_out"], ds["drawer_net"], ds["aov"],
        ])
    write_table(ws_daily, 1, 1, daily_headers, daily_data, "DailySales")
    format_cols(ws_daily, money_cols=(6, 7, 8, 9, 10, 11, 12, 13, 14, 15), int_cols=(2, 3, 4, 5))
    ws_daily.freeze_panes = "A2"

    ws_sales = wb.create_sheet("Transactions")
    sale_headers = [
        "Sale ID", "Time", "Receipt", "Payment", "Items", "Qty", "Merchandise Sales",
        "Discount", "Cash Paid", "Store Credit", "Customer", "Shift", "Notes",
    ]
    sale_data = []
    for sale in sales:
        sale_id = int(sale.get("id") or 0)
        sale_items = items_by_sale.get(sale_id, [])
        sale_data.append([
            sale_id,
            str(sale.get("created_at") or ""),
            str(sale.get("receipt_code") or sale_id),
            str(sale.get("payment_method") or ""),
            sale_item_summary(sale_id),
            sum(_as_int(i.get("qty")) for i in sale_items),
            _as_float(sale.get("merch_total")),
            _as_float(sale.get("discount_total")),
            _as_float(sale.get("cash_paid")),
            _as_float(sale.get("store_credit_used")),
            str(sale.get("customer_name") or ""),
            shift_label(sale.get("shift_id")),
            str(sale.get("notes") or ""),
        ])
    write_table(ws_sales, 1, 1, sale_headers, sale_data, "Transactions")
    format_cols(ws_sales, money_cols=(7, 8, 9, 10), int_cols=(1, 6))
    ws_sales.freeze_panes = "A2"

    ws_lines = wb.create_sheet("Transaction Lines")
    line_headers = [
        "Sale ID", "Time", "Receipt", "Product", "Barcode", "Category", "Brand",
        "Qty", "Unit Price", "Gross Line", "Discount", "Line Revenue", "Payment", "Shift",
    ]
    line_data = []
    for item in items:
        sale = sales_by_id.get(int(item.get("sale_id") or 0), {})
        qty = _as_int(item.get("qty"))
        gross = _as_float(item.get("gross_line_total")) or (_as_float(item.get("price")) * qty)
        line_data.append([
            item.get("sale_id"),
            str(sale.get("created_at") or ""),
            str(sale.get("receipt_code") or item.get("sale_id") or ""),
            str(item.get("name") or ""),
            str(item.get("barcode") or ""),
            str(item.get("category") or ""),
            str(item.get("brand") or ""),
            qty,
            _as_float(item.get("price")),
            gross,
            _as_float(item.get("discount_allocated")),
            _as_float(item.get("line_total")),
            str(sale.get("payment_method") or ""),
            shift_label(sale.get("shift_id")),
        ])
    write_table(ws_lines, 1, 1, line_headers, line_data, "TransactionLines")
    format_cols(ws_lines, money_cols=(9, 10, 11, 12), int_cols=(1, 8))
    ws_lines.freeze_panes = "A2"

    ws_drawer = wb.create_sheet("Cash Drawer")
    drawer_headers = [
        "Date", "Time", "Type", "Shift", "Employee", "Receipt / Ref", "Cash In",
        "Cash Out", "Counted Drawer", "Expected After Event", "Variance", "Notes",
    ]
    events = []
    for shift in shifts:
        sid = int(shift.get("id") or 0)
        events.append({
            "time": str(shift.get("opened_at") or ""),
            "type": "OPEN",
            "shift_id": sid,
            "employee": str(shift.get("employee_name") or ""),
            "ref": shift_label(sid),
            "cash_in": 0.0,
            "cash_out": 0.0,
            "counted": _as_float(shift.get("opening_cash")),
            "notes": "Opening drawer",
        })
        if shift.get("closed_at"):
            events.append({
                "time": str(shift.get("closed_at") or ""),
                "type": "CLOSE",
                "shift_id": sid,
                "employee": str(shift.get("employee_name") or ""),
                "ref": shift_label(sid),
                "cash_in": 0.0,
                "cash_out": 0.0,
                "counted": _as_float(shift.get("closing_cash")),
                "notes": str(shift.get("notes") or "Closing drawer"),
            })
    for sale in sales:
        cash_paid = _as_float(sale.get("cash_paid"))
        if cash_paid:
            sid = int(sale.get("shift_id") or 0)
            events.append({
                "time": str(sale.get("created_at") or ""),
                "type": "CASH SALE",
                "shift_id": sid,
                "employee": str(shifts_by_id.get(sid, {}).get("employee_name") or ""),
                "ref": str(sale.get("receipt_code") or sale.get("id") or ""),
                "cash_in": cash_paid,
                "cash_out": 0.0,
                "counted": None,
                "notes": str(sale.get("payment_method") or ""),
            })
    for movement in movements:
        mtype = str(movement.get("movement_type") or "OUT").upper()
        amount = _as_float(movement.get("amount_value"))
        events.append({
            "time": str(movement.get("created_at") or ""),
            "type": "CASH IN" if mtype == "IN" else "CASH OUT",
            "shift_id": int(movement.get("shift_id") or 0),
            "employee": str(movement.get("employee_name") or ""),
            "ref": str(movement.get("id") or ""),
            "cash_in": amount if mtype == "IN" else 0.0,
            "cash_out": amount if mtype != "IN" else 0.0,
            "counted": None,
            "notes": str(movement.get("reason") or movement.get("notes") or ""),
        })
    for ret in returns:
        cash_refund = _as_float(ret.get("cash_refund"))
        if cash_refund:
            events.append({
                "time": str(ret.get("created_at") or ""),
                "type": "CASH REFUND",
                "shift_id": int(ret.get("shift_id") or 0),
                "employee": str(shifts_by_id.get(int(ret.get("shift_id") or 0), {}).get("employee_name") or ""),
                "ref": f"Return {ret.get('id')}",
                "cash_in": 0.0,
                "cash_out": cash_refund,
                "counted": None,
                "notes": str(ret.get("notes") or ""),
            })

    event_order = {"OPEN": 0, "CASH SALE": 1, "CASH IN": 2, "CASH REFUND": 3, "CASH OUT": 4, "CLOSE": 9}
    events.sort(key=lambda e: (str(e.get("time") or ""), event_order.get(str(e.get("type") or ""), 5)))
    running = defaultdict(float)
    drawer_data = []
    for event in events:
        sid = int(event.get("shift_id") or 0)
        kind = str(event.get("type") or "")
        counted = event.get("counted")
        variance = None
        if kind == "OPEN":
            running[sid] = _as_float(counted)
            expected = running[sid]
        elif kind == "CLOSE":
            expected = running[sid]
            variance = _as_float(counted) - expected
        else:
            running[sid] += _as_float(event.get("cash_in")) - _as_float(event.get("cash_out"))
            expected = running[sid]
        event_time = str(event.get("time") or "")
        drawer_data.append([
            _day_of(event_time),
            _time_of(event_time),
            kind,
            shift_label(sid) if sid else "",
            event.get("employee") or "",
            event.get("ref") or "",
            _as_float(event.get("cash_in")),
            _as_float(event.get("cash_out")),
            counted if counted is not None else "",
            expected,
            variance if variance is not None else "",
            event.get("notes") or "",
        ])
    write_table(ws_drawer, 1, 1, drawer_headers, drawer_data, "CashDrawer")
    format_cols(ws_drawer, money_cols=(7, 8, 9, 10, 11))
    for row in range(2, ws_drawer.max_row + 1):
        if str(ws_drawer.cell(row=row, column=3).value or "") == "CLOSE":
            for col in range(1, ws_drawer.max_column + 1):
                ws_drawer.cell(row=row, column=col).fill = warn_fill
    ws_drawer.freeze_panes = "A2"

    ws_returns = wb.create_sheet("Returns")
    return_headers = [
        "Return ID", "Time", "Original Sale", "Product", "Barcode", "Qty",
        "Return Amount", "Cash Refund", "Credit Refund", "Shift", "Notes",
    ]
    return_data = []
    for item in return_items:
        ret = returns_by_id.get(int(item.get("return_id") or 0), {})
        return_data.append([
            item.get("return_id"),
            str(item.get("created_at") or ""),
            ret.get("original_sale_id") or item.get("original_sale_id") or "",
            str(item.get("name") or ""),
            str(item.get("barcode") or ""),
            _as_int(item.get("qty")),
            _as_float(item.get("line_total")),
            _as_float(ret.get("cash_refund")),
            _as_float(ret.get("credit_refund")),
            shift_label(ret.get("shift_id") or item.get("shift_id")),
            str(ret.get("notes") or item.get("notes") or ""),
        ])
    write_table(ws_returns, 1, 1, return_headers, return_data, "Returns")
    format_cols(ws_returns, money_cols=(7, 8, 9), int_cols=(1, 6))
    ws_returns.freeze_panes = "A2"

    try:
        if ws_prod.max_row > 1:
            chart = BarChart()
            chart.title = "Top Sellers by Net Revenue"
            chart.y_axis.title = "Net revenue"
            chart.x_axis.title = "Product"
            max_row = min(ws_prod.max_row, 11)
            data = Reference(ws_prod, min_col=11, min_row=1, max_row=max_row)
            cats = Reference(ws_prod, min_col=2, min_row=2, max_row=max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 8
            chart.width = 15
            ws_dash.add_chart(chart, "J4")
        if ws_daily.max_row > 1:
            daily_chart = BarChart()
            daily_chart.title = "Net Sales by Day"
            daily_chart.y_axis.title = "Net sales"
            daily_chart.x_axis.title = "Day"
            data = Reference(ws_daily, min_col=8, min_row=1, max_row=ws_daily.max_row)
            cats = Reference(ws_daily, min_col=1, min_row=2, max_row=ws_daily.max_row)
            daily_chart.add_data(data, titles_from_data=True)
            daily_chart.set_categories(cats)
            daily_chart.height = 7
            daily_chart.width = 15
            ws_dash.add_chart(daily_chart, "J21")
    except Exception:
        pass

    for ws in wb.worksheets:
        autosize(ws)
        if ws.max_row > 1:
            ws.auto_filter.ref = ws.dimensions

    ws_dash.column_dimensions["A"].width = 24
    ws_dash.column_dimensions["B"].width = 16
    ws_dash.column_dimensions["D"].width = 8
    ws_dash.column_dimensions["E"].width = 34
    ws_dash.column_dimensions["F"].width = 12
    ws_dash.column_dimensions["G"].width = 15
    ws_dash.column_dimensions["H"].width = 12

    reports_dir = Path(reports_folder)
    reports_dir.mkdir(parents=True, exist_ok=True)
    out_path = reports_dir / f"sales_report_{stamp}_{datetime.now().strftime('%Y_%m_%d_%H%M%S')}.xlsx"
    wb.save(out_path)

    return {
        "path": str(out_path),
        "start": start.strftime("%Y-%m-%d"),
        "end": (end - timedelta(days=1)).strftime("%Y-%m-%d"),
        "summary": {
            "orders": len(sales),
            "items_sold": total_items,
            "items_returned": total_return_qty,
            "net_items_sold": total_net_qty,
            "merchandise_sales": round(total_merch, 2),
            "net_sales": round(total_net_sales, 2),
            "cash_collected": round(total_cash, 2),
            "store_credit_used": round(total_credit, 2),
            "returns": round(total_returns, 2),
            "cash_refunds": round(total_cash_refunds, 2),
            "cash_added": round(total_cash_in, 2),
            "cash_removed": round(total_cash_out, 2),
            "drawer_net_change": round(total_cash + total_cash_in - total_cash_out - total_cash_refunds, 2),
            "top_seller_revenue": (top_by_revenue[0]["name"] if top_by_revenue else ""),
            "top_seller_qty": (top_by_qty[0]["name"] if top_by_qty else ""),
        },
    }


def build_sales_report_excel(db_path: str, reports_folder: str, year: int, month: int, day_str: str) -> dict:
    return _build_sales_report_excel_v2(db_path, reports_folder, year, month, day_str)

    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    start, end, stamp = report_date_bounds(year, month, day_str)
    if not os.path.exists(db_path):
        raise FileNotFoundError(f"Could not find database: {db_path}")

    rows = _fetch_rows(db_path, start, end)
    sales = rows["sales"]
    items = rows["items"]
    shifts = rows["shifts"]
    movements = rows["movements"]
    returns = rows["returns"]

    sales_by_id = {int(s["id"]): s for s in sales}
    items_by_sale = defaultdict(list)
    for item in items:
        items_by_sale[int(item.get("sale_id") or 0)].append(item)

    sales_by_day = defaultdict(list)
    for sale in sales:
        sales_by_day[_day_of(sale.get("created_at"))].append(sale)

    shifts_by_day = defaultdict(list)
    for shift in shifts:
        shifts_by_day[_day_of(shift.get("opened_at"))].append(shift)

    movements_by_day = defaultdict(list)
    for movement in movements:
        movements_by_day[_day_of(movement.get("created_at"))].append(movement)

    returns_by_day = defaultdict(list)
    for ret in returns:
        returns_by_day[_day_of(ret.get("created_at"))].append(ret)

    days = set(sales_by_day.keys()) | set(shifts_by_day.keys()) | set(movements_by_day.keys()) | set(returns_by_day.keys())
    if str(day_str).strip().lower() != "all":
        days.add(start.strftime("%Y-%m-%d"))
    if not days:
        days.add(start.strftime("%Y-%m-%d"))
    days = sorted(d for d in days if d)

    product_totals = {}
    for item in items:
        key = str(item.get("barcode") or item.get("product_id") or item.get("name") or "").strip()
        if not key:
            key = f"sale-{item.get('sale_id')}-{item.get('name')}"
        rec = product_totals.setdefault(
            key,
            {
                "name": str(item.get("name") or ""),
                "barcode": str(item.get("barcode") or ""),
                "category": str(item.get("category") or ""),
                "brand": str(item.get("brand") or ""),
                "qty": 0,
                "revenue": 0.0,
            },
        )
        rec["qty"] += _as_int(item.get("qty"))
        rec["revenue"] += _as_float(item.get("line_total"))

    total_merch = sum(_as_float(s.get("merch_total")) for s in sales)
    total_cash = sum(_as_float(s.get("cash_paid")) for s in sales)
    total_credit = sum(_as_float(s.get("store_credit_used")) for s in sales)
    total_items = sum(_as_int(i.get("qty")) for i in items)
    total_returns = sum(_as_float(r.get("total_return_amount")) for r in returns)
    total_cash_in = sum(_as_float(m.get("amount_value")) for m in movements if str(m.get("movement_type") or "").upper() == "IN")
    total_cash_out = sum(_as_float(m.get("amount_value")) for m in movements if str(m.get("movement_type") or "").upper() == "OUT")

    wb = Workbook()

    header_fill = PatternFill("solid", fgColor="E7EEF8")
    total_fill = PatternFill("solid", fgColor="DDEFE5")
    warn_fill = PatternFill("solid", fgColor="FCE4D6")
    thin = Side(style="thin", color="D8DEE9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(bold=True, size=16)
    h_font = Font(bold=True, size=12)
    bold = Font(bold=True)
    money_fmt = '#,##0.00'
    int_fmt = '0'

    def style_header(ws, row: int):
        for cell in ws[row]:
            cell.font = bold
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    def style_row(ws, row: int):
        for cell in ws[row]:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    def autosize(ws):
        for col_cells in ws.columns:
            letter = get_column_letter(col_cells[0].column)
            max_len = 0
            for cell in col_cells:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
            ws.column_dimensions[letter].width = min(54, max(10, max_len + 2))

    def apply_money(ws, cols):
        for col in cols:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col).number_format = money_fmt

    def daily_summary(day: str) -> dict:
        day_sales = sales_by_day.get(day, [])
        day_movements = movements_by_day.get(day, [])
        day_returns = returns_by_day.get(day, [])
        cash_in = sum(_as_float(m.get("amount_value")) for m in day_movements if str(m.get("movement_type") or "").upper() == "IN")
        cash_out = sum(_as_float(m.get("amount_value")) for m in day_movements if str(m.get("movement_type") or "").upper() == "OUT")
        return {
            "orders": len(day_sales),
            "items": sum(_as_int(i.get("qty")) for s in day_sales for i in items_by_sale.get(int(s.get("id") or 0), [])),
            "merch": sum(_as_float(s.get("merch_total")) for s in day_sales),
            "cash": sum(_as_float(s.get("cash_paid")) for s in day_sales),
            "credit": sum(_as_float(s.get("store_credit_used")) for s in day_sales),
            "returns": sum(_as_float(r.get("total_return_amount")) for r in day_returns),
            "cash_in": cash_in,
            "cash_out": cash_out,
            "drawer_net": sum(_as_float(s.get("cash_paid")) for s in day_sales) + cash_in - cash_out,
        }

    ws = wb.active
    ws.title = "Summary"
    ws.append(["Mask POS Sales Report"])
    ws["A1"].font = title_font
    ws.append(["Period", f"{start.strftime('%Y-%m-%d')} to {(end - timedelta(days=1)).strftime('%Y-%m-%d')}"])
    ws.append(["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append([])
    ws.append(["Metric", "Value"])
    style_header(ws, 5)
    summary_rows = [
        ("Orders", len(sales)),
        ("Items Sold", total_items),
        ("Merchandise Sales", total_merch),
        ("Store Credit Used", total_credit),
        ("Cash Collected", total_cash),
        ("Returns", total_returns),
        ("Net Merchandise Sales", total_merch - total_returns),
        ("Cash Added To Drawer", total_cash_in),
        ("Cash Removed From Drawer", total_cash_out),
        ("Drawer Net Change", total_cash + total_cash_in - total_cash_out),
        ("Closed Shifts", sum(1 for s in shifts if s.get("closed_at"))),
        ("Open Shifts In Period", sum(1 for s in shifts if not s.get("closed_at"))),
    ]
    for label, value in summary_rows:
        ws.append([label, value])
        style_row(ws, ws.max_row)
        if isinstance(value, float):
            ws.cell(row=ws.max_row, column=2).number_format = money_fmt
    autosize(ws)

    ws_prod = wb.create_sheet("Product Sales")
    ws_prod.append(["Product", "Barcode", "Category", "Brand", "Qty Sold", "Revenue", "Average Unit"])
    style_header(ws_prod, 1)
    for rec in sorted(product_totals.values(), key=lambda x: (-float(x["revenue"]), str(x["name"]).lower())):
        qty = _as_int(rec.get("qty"))
        revenue = _as_float(rec.get("revenue"))
        ws_prod.append([
            rec.get("name", ""),
            rec.get("barcode", ""),
            rec.get("category", ""),
            rec.get("brand", ""),
            qty,
            revenue,
            (revenue / qty if qty else 0.0),
        ])
        style_row(ws_prod, ws_prod.max_row)
    apply_money(ws_prod, [6, 7])
    ws_prod.freeze_panes = "A2"
    ws_prod.auto_filter.ref = ws_prod.dimensions
    autosize(ws_prod)

    for day in days:
        ws_day = wb.create_sheet(day)
        ws_day.append([f"Daily Sales: {day}"])
        ws_day["A1"].font = title_font

        ds = daily_summary(day)
        ws_day.append([])
        ws_day.append(["Daily Total", ds["merch"], "Cash Collected", ds["cash"], "Drawer Net", ds["drawer_net"]])
        for cell in ws_day[3]:
            cell.font = h_font
            cell.fill = total_fill
            cell.border = border
        for col in (2, 4, 6):
            ws_day.cell(row=3, column=col).number_format = money_fmt

        ws_day.append(["Orders", ds["orders"], "Items", ds["items"], "Returns", ds["returns"]])
        style_row(ws_day, 4)
        ws_day.cell(row=4, column=6).number_format = money_fmt
        ws_day.append(["Cash Added", ds["cash_in"], "Cash Removed", ds["cash_out"], "Store Credit", ds["credit"]])
        style_row(ws_day, 5)
        for col in (2, 4, 6):
            ws_day.cell(row=5, column=col).number_format = money_fmt

        ws_day.append([])
        ws_day.append(["Sales And Items"])
        ws_day[7][0].font = h_font
        sales_header_row = ws_day.max_row + 1
        ws_day.append([
            "Time", "Receipt", "Sale ID", "Product", "Barcode", "Qty", "Unit Price",
            "Line Revenue", "Payment", "Cash Paid", "Store Credit", "Shift",
        ])
        style_header(ws_day, sales_header_row)

        for sale in sales_by_day.get(day, []):
            sale_id = int(sale.get("id") or 0)
            sale_items = items_by_sale.get(sale_id, [])
            if not sale_items:
                sale_items = [{"name": "(No item rows)", "qty": 0, "price": 0, "line_total": 0, "barcode": ""}]
            first = True
            for item in sale_items:
                ws_day.append([
                    _time_of(sale.get("created_at")) if first else "",
                    str(sale.get("receipt_code") or sale_id) if first else "",
                    sale_id if first else "",
                    str(item.get("name") or ""),
                    str(item.get("barcode") or ""),
                    _as_int(item.get("qty")),
                    _as_float(item.get("price")),
                    _as_float(item.get("line_total")),
                    str(sale.get("payment_method") or "") if first else "",
                    _as_float(sale.get("cash_paid")) if first else "",
                    _as_float(sale.get("store_credit_used")) if first else "",
                    sale.get("shift_id") if first else "",
                ])
                style_row(ws_day, ws_day.max_row)
                for col in (7, 8, 10, 11):
                    ws_day.cell(row=ws_day.max_row, column=col).number_format = money_fmt
                first = False

        ws_day.append([])
        ws_day.append(["Cash Drawer Activity"])
        ws_day[ws_day.max_row][0].font = h_font
        drawer_header_row = ws_day.max_row + 1
        ws_day.append(["Time", "Type", "Shift", "Employee", "USD", "LBP", "Drawer Value", "Reason / Notes"])
        style_header(ws_day, drawer_header_row)

        activity = []
        for shift in shifts_by_day.get(day, []):
            activity.append({
                "time": shift.get("opened_at"),
                "type": "OPEN",
                "shift": shift.get("shift_code") or shift.get("id"),
                "employee": shift.get("employee_name") or "",
                "usd": _as_float(shift.get("opening_usd")),
                "lbp": _as_float(shift.get("opening_lbp")),
                "value": _as_float(shift.get("opening_cash")),
                "notes": "Opening drawer",
            })
            if shift.get("closed_at"):
                closing_value = shift.get("closing_cash")
                activity.append({
                    "time": shift.get("closed_at"),
                    "type": "CLOSE",
                    "shift": shift.get("shift_code") or shift.get("id"),
                    "employee": shift.get("employee_name") or "",
                    "usd": _as_float(shift.get("closing_usd")),
                    "lbp": _as_float(shift.get("closing_lbp")),
                    "value": _as_float(closing_value),
                    "notes": shift.get("notes") or "Closing drawer",
                })
        for movement in movements_by_day.get(day, []):
            mtype = str(movement.get("movement_type") or "OUT").upper()
            activity.append({
                "time": movement.get("created_at"),
                "type": "CASH IN" if mtype == "IN" else "CASH OUT",
                "shift": movement.get("shift_id") or "",
                "employee": movement.get("employee_name") or "",
                "usd": _as_float(movement.get("amount_usd")),
                "lbp": _as_float(movement.get("amount_lbp")),
                "value": _as_float(movement.get("amount_value")),
                "notes": movement.get("reason") or movement.get("notes") or "",
            })
        activity.sort(key=lambda x: str(x.get("time") or ""))
        for entry in activity:
            ws_day.append([
                _time_of(entry.get("time")),
                entry.get("type"),
                entry.get("shift"),
                entry.get("employee"),
                entry.get("usd"),
                entry.get("lbp"),
                entry.get("value"),
                entry.get("notes"),
            ])
            style_row(ws_day, ws_day.max_row)
            for col in (5, 7):
                ws_day.cell(row=ws_day.max_row, column=col).number_format = money_fmt
            ws_day.cell(row=ws_day.max_row, column=6).number_format = '#,##0'
            if entry.get("type") == "CLOSE":
                for cell in ws_day[ws_day.max_row]:
                    cell.fill = warn_fill

        ws_day.freeze_panes = "A8"
        autosize(ws_day)

    ws_sales = wb.create_sheet("Raw Sales")
    ws_sales.append([
        "Sale ID", "Created At", "Receipt", "Merchandise Total", "Amount Due",
        "Cash Paid", "Store Credit", "Discount", "Payment", "Customer", "Shift", "Notes",
    ])
    style_header(ws_sales, 1)
    for sale in sales:
        ws_sales.append([
            int(sale.get("id") or 0),
            str(sale.get("created_at") or ""),
            str(sale.get("receipt_code") or ""),
            _as_float(sale.get("merch_total")),
            _as_float(sale.get("amount_due")),
            _as_float(sale.get("cash_paid")),
            _as_float(sale.get("store_credit_used")),
            _as_float(sale.get("discount_total")),
            str(sale.get("payment_method") or ""),
            str(sale.get("customer_name") or ""),
            sale.get("shift_id"),
            str(sale.get("notes") or ""),
        ])
        style_row(ws_sales, ws_sales.max_row)
    apply_money(ws_sales, [4, 5, 6, 7, 8])
    ws_sales.freeze_panes = "A2"
    ws_sales.auto_filter.ref = ws_sales.dimensions
    autosize(ws_sales)

    ws_items = wb.create_sheet("Raw Items")
    ws_items.append(["Sale ID", "Product ID", "Product", "Barcode", "Category", "Brand", "Qty", "Unit Price", "Line Revenue", "Discount Allocated"])
    style_header(ws_items, 1)
    for item in items:
        ws_items.append([
            item.get("sale_id"),
            item.get("product_id"),
            str(item.get("name") or ""),
            str(item.get("barcode") or ""),
            str(item.get("category") or ""),
            str(item.get("brand") or ""),
            _as_int(item.get("qty")),
            _as_float(item.get("price")),
            _as_float(item.get("line_total")),
            _as_float(item.get("discount_allocated")),
        ])
        style_row(ws_items, ws_items.max_row)
    apply_money(ws_items, [8, 9, 10])
    ws_items.freeze_panes = "A2"
    ws_items.auto_filter.ref = ws_items.dimensions
    autosize(ws_items)

    ws_drawer = wb.create_sheet("Cash Drawer")
    ws_drawer.append(["Time", "Type", "Shift", "Employee", "USD", "LBP", "Drawer Value", "Reason / Notes"])
    style_header(ws_drawer, 1)
    for day in days:
        for shift in shifts_by_day.get(day, []):
            ws_drawer.append([
                str(shift.get("opened_at") or ""),
                "OPEN",
                shift.get("shift_code") or shift.get("id"),
                shift.get("employee_name") or "",
                _as_float(shift.get("opening_usd")),
                _as_float(shift.get("opening_lbp")),
                _as_float(shift.get("opening_cash")),
                "Opening drawer",
            ])
            style_row(ws_drawer, ws_drawer.max_row)
            if shift.get("closed_at"):
                ws_drawer.append([
                    str(shift.get("closed_at") or ""),
                    "CLOSE",
                    shift.get("shift_code") or shift.get("id"),
                    shift.get("employee_name") or "",
                    _as_float(shift.get("closing_usd")),
                    _as_float(shift.get("closing_lbp")),
                    _as_float(shift.get("closing_cash")),
                    shift.get("notes") or "Closing drawer",
                ])
                style_row(ws_drawer, ws_drawer.max_row)
        for movement in movements_by_day.get(day, []):
            mtype = str(movement.get("movement_type") or "OUT").upper()
            ws_drawer.append([
                str(movement.get("created_at") or ""),
                "CASH IN" if mtype == "IN" else "CASH OUT",
                movement.get("shift_id") or "",
                movement.get("employee_name") or "",
                _as_float(movement.get("amount_usd")),
                _as_float(movement.get("amount_lbp")),
                _as_float(movement.get("amount_value")),
                movement.get("reason") or movement.get("notes") or "",
            ])
            style_row(ws_drawer, ws_drawer.max_row)
    apply_money(ws_drawer, [5, 7])
    for row in range(2, ws_drawer.max_row + 1):
        ws_drawer.cell(row=row, column=6).number_format = '#,##0'
    ws_drawer.freeze_panes = "A2"
    ws_drawer.auto_filter.ref = ws_drawer.dimensions
    autosize(ws_drawer)

    ws_chart = wb.create_sheet("Charts")
    ws_chart.append(["Day", "Merchandise Sales", "Cash Collected"])
    style_header(ws_chart, 1)
    for day in days:
        ds = daily_summary(day)
        ws_chart.append([day, ds["merch"], ds["cash"]])
        style_row(ws_chart, ws_chart.max_row)
    apply_money(ws_chart, [2, 3])
    try:
        chart = BarChart()
        chart.title = "Daily Sales"
        chart.y_axis.title = "Amount"
        chart.x_axis.title = "Day"
        data = Reference(ws_chart, min_col=2, max_col=3, min_row=1, max_row=ws_chart.max_row)
        cats = Reference(ws_chart, min_col=1, min_row=2, max_row=ws_chart.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 16
        ws_chart.add_chart(chart, "E2")
    except Exception:
        pass
    autosize(ws_chart)

    reports_dir = Path(reports_folder)
    reports_dir.mkdir(parents=True, exist_ok=True)
    out_path = reports_dir / f"sales_report_{stamp}_{datetime.now().strftime('%Y_%m_%d_%H%M%S')}.xlsx"
    wb.save(out_path)

    return {
        "path": str(out_path),
        "start": start.strftime("%Y-%m-%d"),
        "end": (end - timedelta(days=1)).strftime("%Y-%m-%d"),
        "summary": {
            "orders": len(sales),
            "items_sold": total_items,
            "merchandise_sales": round(total_merch, 2),
            "cash_collected": round(total_cash, 2),
            "store_credit_used": round(total_credit, 2),
            "returns": round(total_returns, 2),
            "cash_added": round(total_cash_in, 2),
            "cash_removed": round(total_cash_out, 2),
            "drawer_net_change": round(total_cash + total_cash_in - total_cash_out, 2),
        },
    }
